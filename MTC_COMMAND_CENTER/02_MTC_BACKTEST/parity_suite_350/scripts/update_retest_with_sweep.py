#!/usr/bin/env python3
"""
Update 'Retest Candidates' sheet in XLSX using Python sweep results.

Categorizes all 71 no-effect cases into:
1. PYTHON_EFFECT: Parameter has effect in Python → TV retest with specific values
2. PYTHON_NO_EFFECT: No effect even in Python → structural no-effect, skip
3. NOT_IN_PYTHON: Parameter not in Python config → TV-only test needed
4. SIZING_EXPECTED: Position sizing → expected no-effect on trade count
5. SL_LOW_TRADES: Only 2 TV trades → inconclusive
"""

import sys, io, csv, json
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
NO_EFFECT_CSV = BASE / "compare_runs" / "true_no_effect_cases.csv"
SWEEP_CSV = BASE / "compare_runs" / "sweep_results.csv"

# =====================================================================
# Load sweep results
# =====================================================================
sweep_data = {}  # seq -> {value -> trade_count, has_effect, ...}
with open(SWEEP_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        seq = int(row["seq"])
        if seq not in sweep_data:
            sweep_data[seq] = {
                "tv_input_name": row["tv_input_name"],
                "config_path": row["config_path"],
                "has_effect": row["has_effect"] == "True",
                "values": {},
            }
        tc = row["trade_count"]
        try:
            tc = int(tc)
        except ValueError:
            pass
        sweep_data[seq]["values"][row["value"]] = tc

print(f"Loaded sweep data for {len(sweep_data)} sequences")

# =====================================================================
# Find best discriminating values for effect sequences
# =====================================================================
def get_best_test_values(seq_data):
    """Return (suggested_value, trade_count, reason) for the most discriminating value."""
    vals = seq_data["values"]
    numeric = {k: v for k, v in vals.items() if isinstance(v, int)}
    if len(set(numeric.values())) <= 1:
        return None

    counts = list(numeric.values())
    most_common = max(set(counts), key=counts.count)

    # Find value that produces most different trade count from most common
    best_val = None
    best_diff = 0
    best_tc = 0
    for val_str, tc in numeric.items():
        diff = abs(tc - most_common)
        if diff > best_diff:
            best_diff = diff
            best_val = val_str
            best_tc = tc

    return (best_val, best_tc, f"Python: {best_tc} trades (vs {most_common} default, Δ={best_diff})")


# =====================================================================
# Load no-effect cases
# =====================================================================
no_effect = {}
with open(NO_EFFECT_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        ro = row["run_order"].strip()
        no_effect[ro] = {
            "case_id": row["case_id"].strip(),
            "seq": int(row["seq"].strip()),
            "primary_name": row["primary_name"].strip(),
            "primary_value": row["primary_value"].strip(),
            "tv_trade_count": int(row["tv_trade_count"]),
            "reason": row["reason"].strip(),
        }

print(f"Loaded {len(no_effect)} no-effect cases")

# =====================================================================
# Classify each case
# =====================================================================
SIZING_ROS = {"098","099","101","102","104","105","117","118","183","184","188","189"}
SL_LOW_ROS = {"125","126","127","128","155","156","157","158","159"}
PYRAMID_ROS = {"048", "049"}

# Map from seq to unmapped status
UNMAPPED_SEQS = {36, 96, 102, 221, 224, 225}

def classify_case(ro, info):
    """Return (category, suggested_tv_value, python_note)."""
    seq = info["seq"]

    # Sizing expected
    if ro in SIZING_ROS:
        return "SIZING_EXPECTED", "", "Position sizing doesn't change trade count (expected)"

    # SL low trades
    if ro in SL_LOW_ROS:
        return "SL_LOW_TRADES", "", "Only 2 TV trades, inconclusive"

    # Pyramid
    if ro in PYRAMID_ROS:
        return "SIZING_EXPECTED", "", "Pyramiding=1 enforced, max_pyramid has no effect"

    # Not in Python
    if seq in UNMAPPED_SEQS:
        return "NOT_IN_PYTHON", "", "Parameter not implemented in Python engine, TV-only test"

    # Check sweep results
    if seq in sweep_data:
        sd = sweep_data[seq]
        if sd["has_effect"]:
            best = get_best_test_values(sd)
            if best:
                val, tc, reason = best
                return "PYTHON_EFFECT", val, reason
        else:
            return "PYTHON_NO_EFFECT", "", "No effect across all sweep values in Python"

    return "UNKNOWN", "", "No sweep data available"


# =====================================================================
# Open XLSX and rebuild sheet
# =====================================================================
wb = openpyxl.load_workbook(XLSX_PATH)
ws_cases = wb["Cases"]

# Get headers
case_headers = {}
orig_headers = []
for cell in ws_cases[1]:
    if cell.value:
        case_headers[cell.value] = cell.column
        orig_headers.append(cell.value)

# Delete old sheet
sheet_name = "Retest Candidates"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
    print("Deleted old Retest Candidates sheet")

ws = wb.create_sheet(sheet_name)

# Extra columns
extra_headers = [
    "category",
    "tv_trade_count",
    "suggested_tv_value",
    "python_note",
    "action_needed",
]

all_headers = orig_headers + extra_headers

# Category styles
CAT_FILLS = {
    "PYTHON_EFFECT":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
    "NOT_IN_PYTHON":    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow
    "PYTHON_NO_EFFECT": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # light gray
    "SIZING_EXPECTED":  PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # light gray
    "SL_LOW_TRADES":    PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),  # light pink
    "UNKNOWN":          PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}

ACTION_MAP = {
    "PYTHON_EFFECT":    "TV_RETEST: Use suggested value",
    "NOT_IN_PYTHON":    "TV_ONLY: Test in TradingView manually",
    "PYTHON_NO_EFFECT": "SKIP: Structurally no effect",
    "SIZING_EXPECTED":  "SKIP: Expected behavior",
    "SL_LOW_TRADES":    "SKIP: Too few trades",
    "UNKNOWN":          "INVESTIGATE",
}

# Header style
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

for ci, h in enumerate(all_headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Category sort order
CAT_ORDER = {
    "PYTHON_EFFECT": 1,
    "NOT_IN_PYTHON": 2,
    "PYTHON_NO_EFFECT": 3,
    "SL_LOW_TRADES": 4,
    "SIZING_EXPECTED": 5,
    "UNKNOWN": 6,
}

# Collect rows
rows = []
for row_idx in range(2, ws_cases.max_row + 1):
    ro = ws_cases.cell(row=row_idx, column=case_headers["run_order"]).value
    if ro is None:
        continue
    ro_str = str(ro).zfill(3)
    if ro_str not in no_effect:
        continue

    info = no_effect[ro_str]
    cat, sugg_val, py_note = classify_case(ro_str, info)
    action = ACTION_MAP.get(cat, "INVESTIGATE")

    row_data = []
    for h in orig_headers:
        val = ws_cases.cell(row=row_idx, column=case_headers[h]).value
        row_data.append(val)

    row_data.append(cat)
    row_data.append(info["tv_trade_count"])
    row_data.append(sugg_val)
    row_data.append(py_note)
    row_data.append(action)

    rows.append((cat, row_data))

# Sort by category then run_order
rows.sort(key=lambda r: (CAT_ORDER.get(r[0], 99), r[1][0] or 0))

# Write rows
for ri, (cat, row_data) in enumerate(rows, 2):
    fill = CAT_FILLS.get(cat)
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if fill:
            cell.fill = fill
        cell.border = thin_border
        # Bold for action column
        if ci == len(all_headers):
            if cat == "PYTHON_EFFECT":
                cell.font = Font(bold=True, color="006100")
            elif cat == "NOT_IN_PYTHON":
                cell.font = Font(bold=True, color="996600")

# Column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 55
for ci in range(4, len(orig_headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20
# Extra columns
for ci, w in enumerate([20, 14, 16, 60, 35], len(orig_headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(all_headers))}{len(rows)+1}"

# Save
wb.save(XLSX_PATH)
wb.close()

# Summary
from collections import Counter
cats = Counter(r[0] for r in rows)

print(f"\nCreated '{sheet_name}' sheet with {len(rows)} cases")
print(f"\nBY CATEGORY:")
for cat in ["PYTHON_EFFECT", "NOT_IN_PYTHON", "PYTHON_NO_EFFECT", "SL_LOW_TRADES", "SIZING_EXPECTED"]:
    print(f"  {cat:22s} {cats.get(cat, 0):3d} cases  ({ACTION_MAP.get(cat, '')})")
print(f"  {'TOTAL':22s} {sum(cats.values()):3d} cases")

# Detail for PYTHON_EFFECT
print(f"\nPYTHON_EFFECT details (TV retest recommended):")
for cat, row_data in rows:
    if cat == "PYTHON_EFFECT":
        ro = row_data[0]
        ro_str = str(ro).zfill(3)
        info = no_effect[ro_str]
        _, sugg_val, py_note = classify_case(ro_str, info)
        print(f"  RO {ro_str}: {info['primary_name']} = {info['primary_value']} | "
              f"Suggested TV value: {sugg_val} | {py_note}")

# Detail for NOT_IN_PYTHON
print(f"\nNOT_IN_PYTHON details (TV-only test needed):")
for cat, row_data in rows:
    if cat == "NOT_IN_PYTHON":
        ro = row_data[0]
        ro_str = str(ro).zfill(3)
        info = no_effect[ro_str]
        print(f"  RO {ro_str}: {info['primary_name']} = {info['primary_value']} (seq {info['seq']})")
