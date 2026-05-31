#!/usr/bin/env python3
"""
Update XLSX tracker with completed results + find new downloads needing parity test.

Steps:
1. Read XLSX tracker
2. Scan tv_manual_inputs/ for folders with XLSX files (= downloaded)
3. Update tv_xlsx_status for newly discovered downloads
4. Fill setup_check (M) and compare_status (N) from CSV results
5. Save updated XLSX
6. Report new cases needing parity test
"""

import csv
import shutil
from pathlib import Path
from datetime import datetime
import openpyxl


BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
TV_INPUTS = BASE / "tv_manual_inputs"
COMPARE_RUNS = BASE / "compare_runs"
SETUP_CSV = COMPARE_RUNS / "all_setup_check.csv"
PARITY_CSV = COMPARE_RUNS / "all_parity_results.csv"
DIAG_FIELDS = [
    "clip_strict_status",
    "raw_strict_status",
    "early_trade_end_candidate",
    "gap_days",
    "clip_tv_trades",
    "clip_py_trades",
    "raw_tv_trades",
    "raw_py_trades",
]


def load_csv_dict(csv_path, key_col="case_id"):
    """Load CSV into dict keyed by case_id."""
    result = {}
    if not csv_path.exists():
        return result
    with open(csv_path, "r") as f:
        for row in csv.DictReader(f):
            result[row[key_col]] = row
    return result


def scan_tv_inputs():
    """Scan tv_manual_inputs for folders that contain XLSX files."""
    downloaded = {}  # case_id -> (run_order, folder_name, xlsx_file)
    if not TV_INPUTS.exists():
        return downloaded

    for folder in sorted(TV_INPUTS.iterdir()):
        if not folder.is_dir():
            continue
        name = folder.name
        # Parse run_order from folder name: "001_parity_core_005_..."
        parts = name.split("_", 1)
        if len(parts) < 2:
            continue
        try:
            run_order = int(parts[0])
        except ValueError:
            continue
        case_id = parts[1]

        # Check if folder has an XLSX file
        xlsx_files = list(folder.glob("*.xlsx"))
        # Filter out temp files (starting with ~$)
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
        if xlsx_files:
            downloaded[case_id] = (run_order, name, xlsx_files[0].name)

    return downloaded


def main():
    # --- Load data ---
    setup_results = load_csv_dict(SETUP_CSV)
    parity_results = load_csv_dict(PARITY_CSV)
    downloaded_cases = scan_tv_inputs()

    print(f"Setup check results: {len(setup_results)} entries")
    print(f"Parity results: {len(parity_results)} entries")
    print(f"Downloaded folders with XLSX: {len(downloaded_cases)}")
    print()

    # --- Backup XLSX ---
    backup = XLSX_PATH.parent / f"CASE_SETUP_GUIDE_BACKUP_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(XLSX_PATH, backup)
    print(f"Backup: {backup.name}")

    # --- Open XLSX ---
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Cases"]

    # Build header map
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.value] = cell.column
    for field in DIAG_FIELDS:
        if field not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = field
            headers[field] = new_col

    print(f"Columns found: {list(headers.keys())}")

    col_run_order = headers["run_order"]
    col_case_id = headers["case_id"]
    col_tv_xlsx_status = headers["tv_xlsx_status"]
    col_tv_xlsx_file = headers.get("tv_xlsx_file")
    col_tv_download_date = headers.get("tv_download_date")
    col_setup_check = headers["setup_check"]
    col_compare_status = headers["compare_status"]
    col_clip_status = headers.get("clip_strict_status")
    col_raw_status = headers.get("raw_strict_status")
    col_early = headers.get("early_trade_end_candidate")
    col_gap = headers.get("gap_days")
    col_clip_tv = headers.get("clip_tv_trades")
    col_clip_py = headers.get("clip_py_trades")
    col_raw_tv = headers.get("raw_tv_trades")
    col_raw_py = headers.get("raw_py_trades")
    col_notes = headers.get("notes")

    # --- Process rows ---
    updated_status = 0
    updated_setup = 0
    updated_parity = 0
    new_downloads = []
    already_tested = []

    for row_idx in range(2, ws.max_row + 1):
        case_id = ws.cell(row=row_idx, column=col_case_id).value
        run_order = ws.cell(row=row_idx, column=col_run_order).value
        current_status = ws.cell(row=row_idx, column=col_tv_xlsx_status).value or ""

        if not case_id:
            continue

        # --- Update tv_xlsx_status for newly discovered downloads ---
        if case_id in downloaded_cases and current_status not in ("DOWNLOADED", "baseline"):
            ro, folder, xlsx_name = downloaded_cases[case_id]
            ws.cell(row=row_idx, column=col_tv_xlsx_status).value = "DOWNLOADED"
            if col_tv_xlsx_file:
                ws.cell(row=row_idx, column=col_tv_xlsx_file).value = xlsx_name
            if col_tv_download_date:
                ws.cell(row=row_idx, column=col_tv_download_date).value = datetime.now().strftime("%Y-%m-%d")
            updated_status += 1

        # --- Fill setup_check from CSV ---
        if case_id in setup_results:
            status_val = setup_results[case_id]["status"]
            # Only fill for cases that had a meaningful result (not NO_FILE for reuse cases)
            if status_val != "NO_FILE":
                ws.cell(row=row_idx, column=col_setup_check).value = status_val
                updated_setup += 1

        # --- Fill compare_status from CSV ---
        if case_id in parity_results:
            compare_val = parity_results[case_id]["compare_status"]
            parity_notes = parity_results[case_id].get("notes", "")
            ws.cell(row=row_idx, column=col_compare_status).value = compare_val
            if col_clip_status:
                ws.cell(row=row_idx, column=col_clip_status).value = parity_results[case_id].get("clip_strict_status", "")
            if col_raw_status:
                ws.cell(row=row_idx, column=col_raw_status).value = parity_results[case_id].get("raw_strict_status", "")
            if col_early:
                ws.cell(row=row_idx, column=col_early).value = parity_results[case_id].get("early_trade_end_candidate", "")
            if col_gap:
                ws.cell(row=row_idx, column=col_gap).value = parity_results[case_id].get("gap_days", "")
            if col_clip_tv:
                ws.cell(row=row_idx, column=col_clip_tv).value = parity_results[case_id].get("clip_tv_trades", "")
            if col_clip_py:
                ws.cell(row=row_idx, column=col_clip_py).value = parity_results[case_id].get("clip_py_trades", "")
            if col_raw_tv:
                ws.cell(row=row_idx, column=col_raw_tv).value = parity_results[case_id].get("raw_tv_trades", "")
            if col_raw_py:
                ws.cell(row=row_idx, column=col_raw_py).value = parity_results[case_id].get("raw_py_trades", "")
            if col_notes and parity_notes:
                existing_notes = ws.cell(row=row_idx, column=col_notes).value or ""
                if parity_notes not in existing_notes:
                    new_note = f"{existing_notes}; {parity_notes}" if existing_notes else parity_notes
                    ws.cell(row=row_idx, column=col_notes).value = new_note
            updated_parity += 1

        # --- Identify new downloads not yet tested ---
        has_parity = case_id in parity_results
        is_downloaded = case_id in downloaded_cases or current_status == "DOWNLOADED"
        is_skip = current_status in ("SKIP", "baseline", "baseline;resuse")

        if is_downloaded and not has_parity and not is_skip:
            new_downloads.append((run_order, case_id))
        elif has_parity:
            already_tested.append((run_order, case_id))

    # --- Save ---
    wb.save(XLSX_PATH)
    wb.close()

    # --- Report ---
    print()
    print("=" * 70)
    print("UPDATE SUMMARY")
    print("=" * 70)
    print(f"  tv_xlsx_status updated:  {updated_status}")
    print(f"  setup_check filled:      {updated_setup}")
    print(f"  compare_status filled:   {updated_parity}")
    print(f"  Already tested:          {len(already_tested)}")
    print()

    print("=" * 70)
    print(f"NEW DOWNLOADS NEEDING PARITY TEST: {len(new_downloads)}")
    print("=" * 70)
    for ro, cid in sorted(new_downloads, key=lambda x: int(x[0]) if str(x[0]).isdigit() else 0):
        print(f"  [{str(ro).zfill(3)}] {cid}")

    # Save new downloads list for next step
    new_csv = COMPARE_RUNS / "new_downloads_pending.csv"
    with open(new_csv, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["run_order", "case_id"])
        for ro, cid in sorted(new_downloads, key=lambda x: int(x[0]) if str(x[0]).isdigit() else 0):
            writer.writerow([ro, cid])
    print(f"\nSaved to: {new_csv}")


if __name__ == "__main__":
    main()
