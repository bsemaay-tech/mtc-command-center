#!/usr/bin/env python
"""
Build parity policy view from tracker workbook:
- Raw strict view (compare_status canonical)
- Clip-overlap effective view (reclassify mismatch cases where clip=PASS raw=FAIL)
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--suite-root", default="parity_suite_350")
    p.add_argument("--xlsx", default="CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx")
    p.add_argument("--out-csv", default="compare_runs/parity_policy_view_latest.csv")
    p.add_argument("--out-md", default="compare_runs/parity_policy_view_latest.md")
    return p.parse_args()


def _resolve(base: Path, path_like: str) -> Path:
    p = Path(path_like)
    return p if p.is_absolute() else base / p


def is_clip_only_candidate(notes: str) -> bool:
    n = (notes or "").upper()
    return ("CLIP=PASS" in n and "RAW=FAIL" in n) or ("TV_EARLY_TRADE_END_CANDIDATE" in n)


def main() -> int:
    args = parse_args()
    here = Path(__file__).resolve()
    mtc_root = here.parents[2]
    suite_root = _resolve(mtc_root, args.suite_root)

    xlsx = _resolve(suite_root, args.xlsx)
    out_csv = _resolve(suite_root, args.out_csv)
    out_md = _resolve(suite_root, args.out_md)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_md.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Cases"]
    headers = {}
    for i, c in enumerate(ws[1], start=1):
        if c.value:
            headers[str(c.value)] = i

    req = ["compare_status", "notes"]
    for r in req:
        if r not in headers:
            wb.close()
            raise SystemExit(f"missing header: {r}")

    total = 0
    skip = 0
    pass_raw = 0
    mismatch_raw = 0
    other = 0
    clip_only_candidates = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        status = str(row[headers["compare_status"] - 1] or "").strip()
        notes = str(row[headers["notes"] - 1] or "")
        if status == "":
            skip += 1
            continue
        if status in {"PASS", "PASS(reuse)"}:
            pass_raw += 1
            continue
        if status == "MISMATCH":
            mismatch_raw += 1
            if is_clip_only_candidate(notes):
                clip_only_candidates += 1
            continue
        other += 1
    wb.close()

    executable = total - skip
    pass_clip_effective = pass_raw + clip_only_candidates
    mismatch_clip_effective = mismatch_raw - clip_only_candidates

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "value"])
        w.writerow(["total_cases", total])
        w.writerow(["skip", skip])
        w.writerow(["executable_cases", executable])
        w.writerow(["raw_pass", pass_raw])
        w.writerow(["raw_mismatch", mismatch_raw])
        w.writerow(["clip_only_candidates", clip_only_candidates])
        w.writerow(["clip_effective_pass", pass_clip_effective])
        w.writerow(["clip_effective_mismatch", mismatch_clip_effective])
        w.writerow(["other_status", other])
        w.writerow(["raw_pass_rate_executable_pct", f"{(100.0 * pass_raw / executable) if executable else 0.0:.2f}"])
        w.writerow(["clip_effective_pass_rate_executable_pct", f"{(100.0 * pass_clip_effective / executable) if executable else 0.0:.2f}"])

    md = []
    md.append("# Parity Policy View (Latest)")
    md.append("")
    md.append("## Base Counts")
    md.append(f"- total_cases: `{total}`")
    md.append(f"- skip: `{skip}`")
    md.append(f"- executable_cases: `{executable}`")
    md.append("")
    md.append("## Raw Strict View")
    md.append(f"- pass: `{pass_raw}`")
    md.append(f"- mismatch: `{mismatch_raw}`")
    md.append(f"- pass_rate_executable: `{(100.0 * pass_raw / executable) if executable else 0.0:.2f}%`")
    md.append("")
    md.append("## Clip-Overlap Effective View")
    md.append(f"- clip_only_candidates: `{clip_only_candidates}`")
    md.append(f"- effective_pass: `{pass_clip_effective}`")
    md.append(f"- effective_mismatch: `{mismatch_clip_effective}`")
    md.append(f"- pass_rate_executable: `{(100.0 * pass_clip_effective / executable) if executable else 0.0:.2f}%`")
    md.append("")
    md.append("## Note")
    md.append("- `clip_only_candidate` means notes indicate `clip=PASS` and `raw=FAIL` (or explicit `TV_EARLY_TRADE_END_CANDIDATE`).")
    out_md.write_text("\n".join(md), encoding="utf-8")

    print(f"total={total} executable={executable} raw_pass={pass_raw} raw_mismatch={mismatch_raw} clip_candidates={clip_only_candidates}")
    print(f"csv={out_csv}")
    print(f"md={out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
