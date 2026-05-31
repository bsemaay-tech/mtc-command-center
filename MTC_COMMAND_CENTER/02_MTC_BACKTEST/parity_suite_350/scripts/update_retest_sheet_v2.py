#!/usr/bin/env python3
"""
Recreate 'Retest Candidates' sheet using TRUE TV trade counts from
compare_runs/true_no_effect_cases.csv (71 verified no-effect cases).
"""

import sys, io, csv, re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
CSV_PATH = BASE / "compare_runs" / "true_no_effect_cases.csv"

# =====================================================================
# LOAD TRUE NO-EFFECT CASES
# =====================================================================
no_effect = {}
with open(CSV_PATH, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        ro = row["run_order"].strip()
        no_effect[ro] = {
            "case_id": row["case_id"].strip(),
            "seq": row["seq"].strip(),
            "primary_name": row["primary_name"].strip(),
            "primary_value": row["primary_value"].strip(),
            "tv_trade_count": int(row["tv_trade_count"]),
            "reason": row["reason"].strip(),
        }

print(f"Loaded {len(no_effect)} TRUE no-effect cases from CSV")

# =====================================================================
# ISSUE GROUP CLASSIFICATION
# =====================================================================
RF_ROS = {
    "066","067","069","070","071","072",
    "076","077","078","082","083","084",
    "085","086","087","088","089","090",
    "091","092","093","094","095","096",
}

SIZING_ROS = {
    "098","099","101","102","104","105",
    "117","118","183","184","188","189",
}

SL_LOW_ROS = {
    "125","126","127","128",
    "155","156","157","158","159",
}

MACD_ROS = {
    "362","363","373","374","385","386",
}

MCGINLEY_ROS = {"230","232","236","237"}

CONFIRM_ROS = {
    "424","425","430","431","432","433","434",
}

# The rest: Pyramid, Range Agg, Max MAE, TP ATR Length
OTHER_ROS = {"048","049","167","169","253","254","326","327","328"}


def get_issue_group(ro):
    if ro in RF_ROS:
        return "RF_SUB_PARAMS"
    if ro in SIZING_ROS:
        return "SIZING_EXPECTED"
    if ro in SL_LOW_ROS:
        return "SL_LOW_TRADES"
    if ro in MACD_ROS:
        return "MACD_SUB_PARAMS"
    if ro in MCGINLEY_ROS:
        return "MCGINLEY_NO_EFFECT"
    if ro in CONFIRM_ROS:
        return "CONFIRM_NO_EFFECT"
    return "OTHER_NO_EFFECT"


GROUP_ORDER = {
    "RF_SUB_PARAMS": 1,
    "MACD_SUB_PARAMS": 2,
    "MCGINLEY_NO_EFFECT": 3,
    "CONFIRM_NO_EFFECT": 4,
    "SL_LOW_TRADES": 5,
    "SIZING_EXPECTED": 6,
    "OTHER_NO_EFFECT": 7,
}

GROUP_DESCRIPTIONS = {
    "RF_SUB_PARAMS": "Range Filter sub-parameters have no effect (RF likely not active in current signal mode)",
    "MACD_SUB_PARAMS": "MACD sub-parameters (Mode/Signal/Distance) have no effect",
    "MCGINLEY_NO_EFFECT": "McGinley Length/HTF Timeframe have no effect",
    "CONFIRM_NO_EFFECT": "Swing confirmation parameters have no effect",
    "SL_LOW_TRADES": "SL parameters produce only 2 TV trades - too few to distinguish",
    "SIZING_EXPECTED": "Position sizing params don't change trade count (expected behavior)",
    "OTHER_NO_EFFECT": "Other parameters (Pyramid, Range Agg, MAE, TP ATR) have no effect",
}

# =====================================================================
# SUGGESTED ALTERNATIVE VALUES (for cases worth retesting)
# =====================================================================
suggestions = {
    # RF Group - extreme values to try to force different filtering
    "066": ("40", "Extreme RSI length → very different smoothing"),
    "067": ("5", "Minimum RSI → very reactive signals"),
    "069": ("On", "BB filter needs extreme BB params to show effect"),
    "070": ("Off", "Expected if RF doesn't use BB in current mode"),
    "071": ("Off", "May not affect if data is same TF"),
    "072": ("On", "May not affect if data is same TF"),
    "076": ("10", "Minimum range threshold → very permissive"),
    "077": ("30", "Maximum range threshold → very restrictive"),
    "078": ("10", "At minimum boundary"),
    "082": ("80", "Maximum → almost nothing qualifies as range"),
    "083": ("50", "Minimum → very permissive range detection"),
    "084": ("50", "Minimum ge=50 → maximally permissive"),
    "085": ("10", "Extreme low → almost never oversold"),
    "086": ("40", "Maximum → very broad oversold zone"),
    "087": ("10", "Minimum ge=10"),
    "088": ("60", "Minimum ge=60 → wide overbought zone"),
    "089": ("90", "Maximum → only extreme overbought"),
    "090": ("60", "Minimum overbought"),
    "091": ("50", "Maximum → very smooth BB"),
    "092": ("10", "Minimum → very reactive BB"),
    "093": ("50", "Maximum BB length"),
    "094": ("1.0", "Tightest BB → most signals filtered"),
    "095": ("1.0", "Tightest BB for comparison"),
    "096": ("3.0", "Widest BB → least filtered"),
    # MACD Group
    "362": ("", "Mode STANDARD vs PPO_NORM → expected if MACD filter not decisive"),
    "363": ("", "Mode PPO_NORM vs STANDARD → expected if MACD filter not decisive"),
    "373": ("1", "Minimum signal → fastest MACD crossover"),
    "374": ("30", "Very slow signal → delays crossovers significantly"),
    "385": ("5.0", "Large distance gate → should filter many crosses"),
    "386": ("10.0", "Very large distance → filters most crosses"),
    # McGinley
    "230": ("50", "Very long McGinley → much smoother trend"),
    "232": ("50", "Very long McGinley"),
    "236": ("1440", "Daily HTF → much higher timeframe"),
    "237": ("1440", "Daily HTF"),
    # Confirmation
    "424": ("CANCEL_BOTH", "Cancel both should reduce entries"),
    "425": ("CANCEL_BOTH", "Cancel both → different handling"),
    "430": ("50", "Large buffer → price must break far beyond swing"),
    "431": ("100", "Very large buffer → significant movement needed"),
    "432": ("50", "Buffer 50 ticks should definitely filter some"),
    "433": ("0.5", "Very tight distance → only closest swings valid"),
    "434": ("0.1", "Extremely tight → almost no swing qualifies"),
    # SL low trades - need different SL config to get more trades
    "125": ("", "Swing+ATR vs SWING_ATR are aliases → 2 TV trades too few"),
    "126": ("", "SWING_ATR vs Swing+ATR are aliases → 2 TV trades too few"),
    "127": ("", "Only 2 TV trades → inconclusive"),
    "128": ("", "Only 2 TV trades → inconclusive"),
    "155": ("", "Only 2 TV trades → inconclusive"),
    "156": ("", "Only 2 TV trades → inconclusive"),
    "157": ("", "Only 2 TV trades → inconclusive"),
    "158": ("", "Only 2 TV trades → inconclusive"),
    "159": ("", "Only 2 TV trades → inconclusive"),
    # Sizing - expected behavior
    "098": ("", "Sizing doesn't change trade count (expected)"),
    "099": ("", "Sizing doesn't change trade count (expected)"),
    "101": ("", "Sizing doesn't change trade count (expected)"),
    "102": ("", "Sizing doesn't change trade count (expected)"),
    "104": ("", "Sizing doesn't change trade count (expected)"),
    "105": ("", "Sizing doesn't change trade count (expected)"),
    "117": ("", "Leverage cap doesn't change trade count (expected)"),
    "118": ("", "Leverage cap doesn't change trade count (expected)"),
    "183": ("", "TP1 close % doesn't change trade count (expected)"),
    "184": ("", "TP1 close % doesn't change trade count (expected)"),
    "188": ("", "Trailing ATR length doesn't change trade count (expected)"),
    "189": ("", "Trailing ATR length doesn't change trade count (expected)"),
    # Other
    "048": ("", "Pyramid=1 enforced → extra positions never open"),
    "049": ("", "Pyramid=1 enforced → extra positions never open"),
    "167": ("7", "Short TP ATR → should change TP distances"),
    "169": ("7", "Short TP ATR → should change TP distances"),
    "253": ("COUNT", "Only 2 modes (AND/COUNT) → may aggregate identically"),
    "254": ("AND", "Only 2 modes → may aggregate identically"),
    "326": ("0.1", "0.1% MAE → should force close nearly all trades early"),
    "327": ("0.1", "Minimum MAE → should force close most trades"),
    "328": ("0.1", "Minimum MAE for maximum impact"),
}

# =====================================================================
# OPEN XLSX AND REBUILD SHEET
# =====================================================================
wb = openpyxl.load_workbook(XLSX_PATH)
ws_cases = wb["Cases"]

# Get headers from Cases sheet
case_headers = {}
for cell in ws_cases[1]:
    if cell.value:
        case_headers[cell.value] = cell.column

orig_headers = []
for cell in ws_cases[1]:
    if cell.value:
        orig_headers.append(cell.value)

# Delete old sheet if exists
sheet_name = "Retest Candidates"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
    print("Deleted old Retest Candidates sheet")

ws_retest = wb.create_sheet(sheet_name)

# Extra columns
extra_headers = [
    "issue_group",
    "tv_trade_count",
    "suggested_value",
    "no_effect_reason",
    "suggestion_note",
    "retest_priority",
]

all_headers = orig_headers + extra_headers

# Header styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
thin_border = Border(
    bottom=Side(style="thin", color="999999"),
)

for col_idx, h in enumerate(all_headers, 1):
    cell = ws_retest.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Collect matching rows
retest_rows = []
for row_idx in range(2, ws_cases.max_row + 1):
    ro = ws_cases.cell(row=row_idx, column=case_headers["run_order"]).value
    if ro is None:
        continue
    ro_str = str(ro).zfill(3)

    if ro_str not in no_effect:
        continue

    info = no_effect[ro_str]
    group = get_issue_group(ro_str)

    # Determine retest priority
    if group == "SIZING_EXPECTED":
        priority = "LOW"      # expected behavior, no action needed
    elif group == "SL_LOW_TRADES":
        priority = "LOW"      # inconclusive due to low trades
    elif group == "OTHER_NO_EFFECT" and ro_str in {"048", "049"}:
        priority = "LOW"      # pyramid enforced
    elif group in ("RF_SUB_PARAMS", "MACD_SUB_PARAMS", "MCGINLEY_NO_EFFECT", "CONFIRM_NO_EFFECT"):
        priority = "HIGH"     # investigation / retest needed
    else:
        priority = "MEDIUM"

    # Get suggested value and note
    sugg_val, sugg_note = suggestions.get(ro_str, ("", ""))

    # Copy all original columns
    row_data = []
    for h in orig_headers:
        val = ws_cases.cell(row=row_idx, column=case_headers[h]).value
        row_data.append(val)

    # Add extra columns
    row_data.append(group)
    row_data.append(info["tv_trade_count"])
    row_data.append(sugg_val)
    row_data.append(info["reason"])
    row_data.append(sugg_note)
    row_data.append(priority)

    retest_rows.append(row_data)

# Sort by group then run_order
retest_rows.sort(key=lambda r: (GROUP_ORDER.get(r[len(orig_headers)], 99), r[0] or 0))

# Group fills
group_fills = {
    "RF_SUB_PARAMS":      PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "MACD_SUB_PARAMS":    PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
    "MCGINLEY_NO_EFFECT": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "CONFIRM_NO_EFFECT":  PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid"),
    "SL_LOW_TRADES":      PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    "SIZING_EXPECTED":    PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    "OTHER_NO_EFFECT":    PatternFill(start_color="F0E0FF", end_color="F0E0FF", fill_type="solid"),
}

priority_fonts = {
    "HIGH":   Font(bold=True, color="CC0000"),
    "MEDIUM": Font(bold=False, color="996600"),
    "LOW":    Font(bold=False, color="999999"),
}

# Write rows
for r_idx, row_data in enumerate(retest_rows, 2):
    group = row_data[len(orig_headers)]
    fill = group_fills.get(group)
    priority = row_data[-1]
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_retest.cell(row=r_idx, column=c_idx, value=val)
        if fill:
            cell.fill = fill
        cell.border = thin_border
        # Priority column gets special font
        if c_idx == len(all_headers):
            cell.font = priority_fonts.get(priority, Font())
            cell.alignment = Alignment(horizontal="center")

# Column widths
col_widths = {
    "run_order": 10, "package": 10, "case_id": 55, "case_folder": 55,
    "tv_preset_name": 55, "primary_change": 40, "tv_ui_actions": 60,
    "dependency_note": 50, "expected_trade_behavior": 12, "tv_xlsx_status": 20,
    "tv_xlsx_file": 45, "tv_download_date": 15, "setup_check": 12,
    "compare_status": 14, "notes": 25, "exec_conflict_tag": 15,
    "exec_conflict_ready": 15,
    "issue_group": 26, "tv_trade_count": 16, "suggested_value": 16,
    "no_effect_reason": 55, "suggestion_note": 55, "retest_priority": 14,
}
for col_idx, h in enumerate(all_headers, 1):
    ws_retest.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(h, 15)

# Freeze panes
ws_retest.freeze_panes = "A2"

# Auto-filter
ws_retest.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(all_headers))}{len(retest_rows)+1}"

# Save
wb.save(XLSX_PATH)
wb.close()

# Summary
print(f"\nCreated '{sheet_name}' sheet with {len(retest_rows)} cases")
print()
from collections import Counter
groups = Counter(r[len(orig_headers)] for r in retest_rows)
priorities = Counter(r[-1] for r in retest_rows)

print("BY GROUP:")
for g, c in sorted(groups.items(), key=lambda x: GROUP_ORDER.get(x[0], 99)):
    desc = GROUP_DESCRIPTIONS.get(g, "")
    print(f"  {g:26s} {c:3d} cases  ({desc})")
print(f"  {'TOTAL':26s} {sum(groups.values()):3d} cases")

print("\nBY PRIORITY:")
for p in ["HIGH", "MEDIUM", "LOW"]:
    print(f"  {p:10s} {priorities.get(p, 0):3d} cases")
print(f"  {'TOTAL':10s} {sum(priorities.values()):3d} cases")
