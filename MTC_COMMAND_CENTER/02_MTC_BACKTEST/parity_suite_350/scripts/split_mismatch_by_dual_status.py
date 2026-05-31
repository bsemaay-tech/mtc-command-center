#!/usr/bin/env python
"""
Split mismatch cases into:
1) EARLY_TRADE_END_CANDIDATE
2) TRUE_LOGIC_MISMATCH

Source of truth: CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx (Cases sheet)
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--suite-root",
        default="parity_suite_350",
        help="Path to parity_suite_350 root (relative to mtc_backtest or absolute)",
    )
    p.add_argument(
        "--xlsx",
        default="CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx",
        help="Workbook path relative to suite-root (or absolute)",
    )
    p.add_argument(
        "--out-csv",
        default="compare_runs/mismatch_split_latest.csv",
        help="Output CSV relative to suite-root (or absolute)",
    )
    p.add_argument(
        "--out-md",
        default="compare_runs/mismatch_split_latest.md",
        help="Output markdown summary relative to suite-root (or absolute)",
    )
    return p.parse_args()


def _resolve(base: Path, p: str) -> Path:
    q = Path(p)
    return q if q.is_absolute() else base / q


def classify(notes: str) -> str:
    n = (notes or "").upper()
    if "TV_EARLY_TRADE_END_CANDIDATE" in n:
        return "EARLY_TRADE_END_CANDIDATE"
    if "CLIP=PASS" in n and "RAW=FAIL" in n:
        return "EARLY_TRADE_END_CANDIDATE"
    return "TRUE_LOGIC_MISMATCH"


def main() -> int:
    args = parse_args()
    here = Path(__file__).resolve()
    mtc_root = here.parents[2]
    suite_root = _resolve(mtc_root, args.suite_root)

    xlsx = _resolve(suite_root, args.xlsx)
    out_csv = _resolve(suite_root, args.out_csv)
    out_md = _resolve(suite_root, args.out_md)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_md.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Cases"]
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value)] = idx

    req = ["run_order", "case_id", "compare_status", "notes"]
    for r in req:
        if r not in headers:
            wb.close()
            raise SystemExit(f"missing header: {r}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        compare_status = row[headers["compare_status"] - 1]
        if str(compare_status).strip().upper() != "MISMATCH":
            continue
        run_order = row[headers["run_order"] - 1]
        case_id = row[headers["case_id"] - 1]
        notes = row[headers["notes"] - 1] or ""
        bucket = classify(str(notes))
        rows.append(
            {
                "run_order": run_order,
                "case_id": case_id,
                "bucket": bucket,
                "notes": str(notes),
            }
        )
    wb.close()

    rows.sort(key=lambda r: (int(r["run_order"]) if str(r["run_order"]).isdigit() else 999999, str(r["case_id"])))

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["run_order", "case_id", "bucket", "notes"])
        w.writeheader()
        w.writerows(rows)

    early = [r for r in rows if r["bucket"] == "EARLY_TRADE_END_CANDIDATE"]
    true = [r for r in rows if r["bucket"] == "TRUE_LOGIC_MISMATCH"]

    md = []
    md.append("# Mismatch Split (Dual Status)")
    md.append("")
    md.append(f"- total_mismatch: `{len(rows)}`")
    md.append(f"- early_trade_end_candidate: `{len(early)}`")
    md.append(f"- true_logic_mismatch: `{len(true)}`")
    md.append("")
    md.append("## EARLY_TRADE_END_CANDIDATE")
    if early:
        for r in early:
            md.append(f"- [{int(r['run_order']):03d}] `{r['case_id']}`")
    else:
        md.append("- (none)")
    md.append("")
    md.append("## TRUE_LOGIC_MISMATCH")
    if true:
        for r in true:
            md.append(f"- [{int(r['run_order']):03d}] `{r['case_id']}`")
    else:
        md.append("- (none)")

    out_md.write_text("\n".join(md), encoding="utf-8")

    print(f"mismatch_total={len(rows)} early={len(early)} true={len(true)}")
    print(f"csv={out_csv}")
    print(f"md={out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
