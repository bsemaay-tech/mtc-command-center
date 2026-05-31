#!/usr/bin/env python3
"""
Post-cut lifecycle triage for mismatch blocker cases.

For each case:
- reads TV XLSX (List of trades) and finds last TV exit timestamp
- reads latest Python trades/signals debug CSV
- reports Python trades after TV cutoff
- reports first post-cut entry context from debug signals (entry diagnostics, margin lock)
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd


@dataclass
class CaseTriage:
    run_order: int
    case_id: str
    tv_xlsx: Path
    py_trades_csv: Path
    py_signals_csv: Path | None
    case_json: Path


def _latest(path_glob: Iterable[Path]) -> Path | None:
    files = sorted(path_glob, key=lambda p: p.stat().st_mtime)
    return files[-1] if files else None


def _tv_last_exit_utc(tv_xlsx: Path, tv_tz: str) -> tuple[pd.Timestamp | None, int]:
    wb = openpyxl.load_workbook(tv_xlsx, read_only=True, data_only=True)
    if "List of trades" not in wb.sheetnames:
        wb.close()
        return None, 0
    ws = wb["List of trades"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return None, 0
    tv = pd.DataFrame(rows[1:], columns=rows[0])
    if "Type" not in tv.columns or "Date and time" not in tv.columns:
        return None, 0
    exits = tv[tv["Type"].astype(str).str.contains("Exit", case=False, na=False)].copy()
    if exits.empty:
        return None, int(tv.get("Trade #", pd.Series(dtype=float)).nunique(dropna=True))
    dt = pd.to_datetime(exits["Date and time"], errors="coerce")
    dt = dt.dt.tz_localize(tv_tz, ambiguous="NaT", nonexistent="shift_forward").dt.tz_convert("UTC")
    return dt.max(), int(tv.get("Trade #", pd.Series(dtype=float)).nunique(dropna=True))


def _load_py_trades(py_csv: Path) -> pd.DataFrame:
    py = pd.read_csv(py_csv)
    for c in ("entry_timestamp", "exit_timestamp"):
        if c in py.columns:
            py[c] = pd.to_datetime(py[c], utc=True, errors="coerce")
    dedup_cols = [c for c in ["entry_timestamp", "exit_timestamp", "side", "entry_price", "exit_price", "qty", "reason"] if c in py.columns]
    if dedup_cols:
        py = py.drop_duplicates(subset=dedup_cols).copy()
    return py


def _load_py_signals(py_sig_csv: Path | None) -> pd.DataFrame | None:
    if py_sig_csv is None:
        return None
    s = pd.read_csv(py_sig_csv, low_memory=False)
    if "timestamp" in s.columns:
        s["timestamp"] = pd.to_datetime(s["timestamp"], utc=True, errors="coerce")
    return s


def _summarize_case(ct: CaseTriage, out_dir: Path) -> dict[str, str]:
    with open(ct.case_json, "r", encoding="utf-8") as f:
        payload = json.load(f)
    tv_tz = str(payload.get("tv_tz", "Europe/London"))
    case_end_raw = payload.get("end_date")
    case_end = pd.to_datetime(case_end_raw, utc=True, errors="coerce") if case_end_raw else pd.NaT

    last_tv_exit, tv_unique_trades = _tv_last_exit_utc(ct.tv_xlsx, tv_tz)
    py = _load_py_trades(ct.py_trades_csv)
    py_last_exit = py["exit_timestamp"].max() if "exit_timestamp" in py.columns and not py.empty else pd.NaT
    gap_days = ""
    if pd.notna(case_end) and last_tv_exit is not None and pd.notna(last_tv_exit):
        gap_days = f"{(case_end - last_tv_exit).total_seconds() / 86400.0:.2f}"

    post = py[py["entry_timestamp"] > last_tv_exit].copy() if (last_tv_exit is not None and "entry_timestamp" in py.columns) else py.iloc[0:0]
    post_count = len(post)
    first_post_entry = post["entry_timestamp"].min() if post_count else pd.NaT
    first_post_exit = post["exit_timestamp"].iloc[post["entry_timestamp"].idxmin()] if post_count and "exit_timestamp" in post.columns else pd.NaT

    sig = _load_py_signals(ct.py_signals_csv)
    sig_ctx = {}
    if sig is not None and post_count and pd.notna(first_post_entry):
        row = sig[sig["timestamp"] == first_post_entry]
        if row.empty:
            row = sig[(sig["timestamp"] >= first_post_entry - pd.Timedelta(minutes=15)) & (sig["timestamp"] <= first_post_entry + pd.Timedelta(minutes=15))]
            row = row.head(1)
        if not row.empty:
            r = row.iloc[0]
            for c in [
                "entry_diag_reason",
                "finalLongEntry",
                "finalShortEntry",
                "margin_call_lock",
                "exited_this_bar",
                "exitReason",
                "entry_diag_equity_now",
                "entry_diag_margin_required",
                "entry_diag_post_notional",
            ]:
                if c in row.columns:
                    sig_ctx[c] = str(r[c])

    post_reason_counts = post["reason"].value_counts().to_dict() if post_count and "reason" in post.columns else {}
    min_notional = ""
    max_notional = ""
    if post_count and "entry_price" in post.columns and "qty" in post.columns:
        notional = post["entry_price"] * post["qty"]
        min_notional = f"{float(notional.min()):.4f}"
        max_notional = f"{float(notional.max()):.4f}"

    return {
        "run_order": str(ct.run_order),
        "case_id": ct.case_id,
        "tv_xlsx": ct.tv_xlsx.name,
        "tv_unique_trades": str(tv_unique_trades),
        "tv_last_exit_utc": "" if last_tv_exit is None else str(last_tv_exit),
        "case_end_utc": "" if pd.isna(case_end) else str(case_end),
        "gap_days": gap_days,
        "py_trades_total": str(len(py)),
        "py_last_exit_utc": "" if pd.isna(py_last_exit) else str(py_last_exit),
        "py_post_cut_trades": str(post_count),
        "py_first_post_entry_utc": "" if pd.isna(first_post_entry) else str(first_post_entry),
        "py_first_post_exit_utc": "" if pd.isna(first_post_exit) else str(first_post_exit),
        "py_post_reason_counts": json.dumps(post_reason_counts, ensure_ascii=True),
        "py_post_notional_min": min_notional,
        "py_post_notional_max": max_notional,
        "py_signal_context": json.dumps(sig_ctx, ensure_ascii=True),
    }


def _resolve_cases(suite_root: Path, case_ids: list[str]) -> list[CaseTriage]:
    tracker = suite_root / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
    wb = openpyxl.load_workbook(tracker, read_only=True, data_only=True)
    ws = wb["Cases"]
    hdr = {str(c.value): c.column for c in ws[1] if c.value}
    by_case: dict[str, tuple[int, str]] = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr["case_id"]).value
        if not cid:
            continue
        ro = int(ws.cell(r, hdr["run_order"]).value)
        folder = ws.cell(r, hdr["case_folder"]).value
        by_case[str(cid)] = (ro, str(folder))
    wb.close()

    out: list[CaseTriage] = []
    for cid in case_ids:
        if cid not in by_case:
            continue
        ro, folder = by_case[cid]
        case_folder = suite_root / "tv_manual_inputs" / folder
        tv_xlsx = _latest([p for p in case_folder.glob("*.xlsx") if not p.name.startswith("~$")])
        if tv_xlsx is None:
            continue
        dbg = suite_root.parent / "debug" / "parity_suite_350" / cid
        py_trades = _latest(dbg.glob("debug_python_trades_*.csv"))
        if py_trades is None:
            continue
        py_signals = _latest(dbg.glob("debug_python_signals_*.csv"))
        case_json = suite_root / "cases" / f"{cid}.json"
        if not case_json.exists():
            continue
        out.append(
            CaseTriage(
                run_order=ro,
                case_id=cid,
                tv_xlsx=tv_xlsx,
                py_trades_csv=py_trades,
                py_signals_csv=py_signals,
                case_json=case_json,
            )
        )
    return sorted(out, key=lambda x: x.run_order)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--suite-root", default=r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
    ap.add_argument("--cases", nargs="+", default=["parity_bnd_211_swing_right_bars_v03", "parity_bnd_217_dynamic_update_mode_v02"])
    args = ap.parse_args()

    suite_root = Path(args.suite_root)
    out_dir = suite_root / "compare_runs"
    out_dir.mkdir(parents=True, exist_ok=True)

    triage_cases = _resolve_cases(suite_root, list(args.cases))
    rows = [_summarize_case(c, out_dir) for c in triage_cases]

    out_csv = out_dir / "post_cut_lifecycle_triage.csv"
    out_md = out_dir / "post_cut_lifecycle_triage.md"

    if rows:
        df = pd.DataFrame(rows)
        df.to_csv(out_csv, index=False)

        lines = ["# Post-cut Lifecycle Triage", ""]
        for r in rows:
            lines.append(f"## [{int(r['run_order']):03d}] {r['case_id']}")
            lines.append(f"- tv_xlsx: `{r['tv_xlsx']}`")
            lines.append(f"- tv_unique_trades: `{r['tv_unique_trades']}`")
            lines.append(f"- tv_last_exit_utc: `{r['tv_last_exit_utc']}`")
            lines.append(f"- case_end_utc: `{r['case_end_utc']}`")
            lines.append(f"- gap_days: `{r['gap_days']}`")
            lines.append(f"- py_trades_total: `{r['py_trades_total']}`")
            lines.append(f"- py_last_exit_utc: `{r['py_last_exit_utc']}`")
            lines.append(f"- py_post_cut_trades: `{r['py_post_cut_trades']}`")
            lines.append(f"- py_first_post_entry_utc: `{r['py_first_post_entry_utc']}`")
            lines.append(f"- py_first_post_exit_utc: `{r['py_first_post_exit_utc']}`")
            lines.append(f"- py_post_reason_counts: `{r['py_post_reason_counts']}`")
            lines.append(f"- py_post_notional_min/max: `{r['py_post_notional_min']} / {r['py_post_notional_max']}`")
            lines.append(f"- py_signal_context_first_post_entry: `{r['py_signal_context']}`")
            lines.append("")
        out_md.write_text("\n".join(lines), encoding="utf-8")
    else:
        out_md.write_text("# Post-cut Lifecycle Triage\n\nNo cases resolved.\n", encoding="utf-8")

    print(f"cases={len(rows)}")
    print(f"csv={out_csv}")
    print(f"md={out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

