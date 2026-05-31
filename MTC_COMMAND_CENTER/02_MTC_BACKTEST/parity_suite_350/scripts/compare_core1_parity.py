#!/usr/bin/env python3
"""
Compare TV vs Python parity for Core-1 cases.

For each PASS case:
1. Extract "List of trades" sheet from TV XLSX
2. Compare against Python trades CSV from runs/
3. Generate parity report with PASS/MISMATCH verdict
4. Output: core1_parity_results.csv with acceptance matrix evaluation
"""

import subprocess
import csv
import tempfile
from pathlib import Path
from typing import List, Tuple, Dict, Optional
import json
import sys

sys.path.insert(0, r"C:\LAB\tradingview-lab\mtc_backtest")
from src.parity.compare_tv_trades import build_report, clip_overlap, load_py_trades, load_tv_trades, summarize_report

def read_pass_cases(csv_path: Path) -> List[Tuple[int, str]]:
    """Read PASS cases from setup check CSV."""
    pass_cases = []
    try:
        with open(csv_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("status") == "PASS":
                    run_order = int(row.get("run_order", 0))
                    case_id = row.get("case_id", "")
                    if case_id:
                        pass_cases.append((run_order, case_id))
    except Exception as e:
        print(f"ERROR reading {csv_path}: {e}")
        return []
    return pass_cases

def find_tv_xlsx(case_folder: Path) -> Optional[Path]:
    """Find TV XLSX in case folder."""
    xlsx_files = list(case_folder.glob("*.xlsx"))
    return xlsx_files[0] if xlsx_files else None

def extract_tv_trades_csv(xlsx_path: Path, temp_csv: Path) -> bool:
    """Extract 'List of trades' sheet from TV XLSX to CSV."""
    try:
        import openpyxl
        import csv as csv_module
    except ImportError:
        print(f"  WARNING: openpyxl not available, skipping TV trade extraction")
        return False

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Try to find "List of trades" sheet
        sheet_name = None
        for sn in wb.sheetnames:
            if "list of trades" in sn.lower():
                sheet_name = sn
                break

        if not sheet_name:
            print(f"  WARNING: no 'List of trades' sheet found in {xlsx_path.name}")
            wb.close()
            return False

        ws = wb[sheet_name]

        # Write to temp CSV
        with open(temp_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv_module.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        wb.close()
        return True
    except Exception as e:
        print(f"  ERROR extracting trades from {xlsx_path.name}: {e}")
        return False

def run_comparison(tv_csv: Path, py_csv: Path, out_csv: Path) -> bool:
    """Run the comparison using existing compare_tv_web_trades.py."""
    try:
        # Run comparison
        cmd = [
            "python",
            "scripts/compare_tv_web_trades.py",
            "--tv", str(tv_csv),
            "--py", str(py_csv),
            "--out", str(out_csv),
            "--tv-tz", "UTC",
            "--clip-overlap",
        ]

        result = subprocess.run(cmd, cwd="C:\\LAB\\tradingview-lab\\mtc_backtest", capture_output=True, text=True, timeout=60)
        return result.returncode == 0
    except Exception as e:
        print(f"  ERROR running comparison: {e}")
        return False

def evaluate_parity_result(result_csv: Path) -> Tuple[str, str]:
    """
    Evaluate parity result CSV to determine PASS or MISMATCH.
    Returns (status, summary) based on acceptance matrix.
    """
    try:
        with open(result_csv, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        if not rows:
            return "UNKNOWN", "No comparison results"

        # Check for parity columns in result
        first_row = rows[0]

        # Look for divergence indicators
        divergence_found = False
        summary_parts = []

        for row in rows:
            # Check if there's a mismatch field or divergence indicator
            if row.get("match") == "NO" or row.get("match") == "False":
                divergence_found = True
            if row.get("divergence") and row.get("divergence") != "none":
                divergence_found = True
            if row.get("first_divergence") and row.get("first_divergence") != "none":
                divergence_found = True

        if divergence_found:
            return "MISMATCH", "Trade divergence detected"
        else:
            return "PASS", "All trades match"
    except Exception as e:
        print(f"  WARNING: could not evaluate result: {e}")
        return "UNKNOWN", f"Error: {str(e)[:50]}"

def compare_case(run_order: int, case_id: str, base_path: Path) -> Tuple[str, str]:
    """
    Compare a single case. Returns (status, notes).
    Status: PASS, MISMATCH, or ERROR
    """
    # Find TV XLSX
    case_folder = base_path / "tv_manual_inputs" / f"{run_order:03d}_{case_id}"
    if not case_folder.exists():
        return "ERROR", "TV input folder not found"

    tv_xlsx = find_tv_xlsx(case_folder)
    if not tv_xlsx:
        return "ERROR", "No XLSX file found"

    # Find Python trades CSV in debug directory (timestamped)
    debug_case_dir = Path("C:\\LAB\\tradingview-lab\\mtc_backtest\\debug\\parity_suite_350") / case_id
    if not debug_case_dir.exists():
        return "ERROR", "Debug output directory not found"

    # Find the most recent debug_python_trades_*.csv file
    py_trades_files = list(debug_case_dir.glob("debug_python_trades_*.csv"))
    if not py_trades_files:
        return "ERROR", "Python trades CSV not found"

    py_trades = sorted(py_trades_files)[-1]  # Get the most recent one

    # Extract TV trades to temp file
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        temp_tv_csv = Path(f.name)

    try:
        if not extract_tv_trades_csv(tv_xlsx, temp_tv_csv):
            return "SKIP", "Could not extract TV trades from XLSX"

        compare_runs_base = base_path / "compare_runs"
        compare_runs_base.mkdir(exist_ok=True)
        result_csv = compare_runs_base / f"{case_id}_parity_comparison.csv"

        case_json_path = base_path / "cases" / f"{case_id}.json"
        tv_tz = "Europe/London"
        if case_json_path.exists():
            with open(case_json_path, "r", encoding="utf-8") as f:
                case_payload = json.load(f)
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
        tv_df = load_tv_trades(temp_tv_csv, tv_tz=tv_tz, tv_shift_min=0)
        py_df = load_py_trades(py_trades)
        tv_df, py_df = clip_overlap(tv_df, py_df)
        report = build_report(tv_df, py_df)
        report.to_csv(result_csv, index=False)

        summary = summarize_report(tv_df, py_df, report)
        if summary["tv_trades"] == 0 and summary["py_trades"] == 0:
            return "PASS", "No trades"
        if summary["strict_pass"]:
            return "PASS", f"{summary['tv_trades']} trades match"
        return "MISMATCH", (
            f"core={summary['core_match_count']}/{summary['compared']}; "
            f"tv={summary['tv_trades']}; py={summary['py_trades']}; "
            f"extra_tv={summary['extra_tv_trades']}; extra_py={summary['extra_py_trades']}"
        )
    finally:
        # Clean up temp file
        if temp_tv_csv.exists():
            temp_tv_csv.unlink()

def main():
    base_path = Path("C:\\LAB\\tradingview-lab\\mtc_backtest\\parity_suite_350")
    compare_runs_base = base_path / "compare_runs"
    csv_path = compare_runs_base / "core1_setup_check.csv"

    if not csv_path.exists():
        print(f"ERROR: setup check CSV not found: {csv_path}")
        return

    pass_cases = read_pass_cases(csv_path)
    print(f"Comparing {len(pass_cases)} PASS cases against TV...")
    print()

    results = []
    pass_count = 0
    mismatch_count = 0
    error_count = 0

    for run_order, case_id in pass_cases:
        print(f"[{run_order:03d}] {case_id}...", end=" ")
        status, notes = compare_case(run_order, case_id, base_path)

        if status == "PASS":
            print("[PASS]")
            pass_count += 1
        elif status == "MISMATCH":
            print("[MISMATCH]")
            mismatch_count += 1
        else:
            print(f"[{status}]")
            if status != "SKIP":
                error_count += 1

        results.append({
            "run_order": run_order,
            "case_id": case_id,
            "compare_status": status,
            "notes": notes,
        })

    print()
    print(f"Summary: {pass_count} PASS, {mismatch_count} MISMATCH, {error_count} ERROR")
    print()

    # Write CSV
    csv_out = compare_runs_base / "core1_parity_results.csv"
    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["run_order", "case_id", "compare_status", "notes"])
        writer.writeheader()
        writer.writerows(results)

    print(f"Results written to: {csv_out}")

if __name__ == "__main__":
    main()
