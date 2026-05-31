#!/usr/bin/env python
"""
Build a user-friendly CASE_SETUP_GUIDE.xlsx for manual TradingView XLSX tracking.

The guide is generated from manifests/cases_manifest_all.csv and can preserve
user-entered tracking fields when regenerated.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


MANIFEST_HEADERS = [
    "run_order",
    "pack",
    "case_id",
    "case_json",
    "tv_preset_name",
    "enabled",
    "expected_trade_behavior",
    "primary_change",
    "ui_actions",
    "depends_on",
    "parent_required",
    "canonical_config_hash",
    "semantic_fingerprint",
    "symbol",
    "timeframe",
    "start_date",
    "end_date",
    "notes",
]


GUIDE_HEADERS = [
    "run_order",
    "package",
    "case_id",
    "case_folder",
    "tv_preset_name",
    "primary_change",
    "tv_ui_actions",
    "dependency_note",
    "expected_trade_behavior",
    "tv_xlsx_status",
    "tv_xlsx_file",
    "tv_download_date",
    "setup_check",
    "compare_status",
    "clip_strict_status",
    "raw_strict_status",
    "early_trade_end_candidate",
    "gap_days",
    "clip_tv_trades",
    "clip_py_trades",
    "raw_tv_trades",
    "raw_py_trades",
    "notes",
]


PRESERVED_COLUMNS = [
    "tv_xlsx_status",
    "tv_xlsx_file",
    "tv_download_date",
    "setup_check",
    "compare_status",
    "clip_strict_status",
    "raw_strict_status",
    "early_trade_end_candidate",
    "gap_days",
    "clip_tv_trades",
    "clip_py_trades",
    "raw_tv_trades",
    "raw_py_trades",
    "notes",
]


UI_LABELS = {
    "trade.allow_flip": "Allow Position Flips (LongShort same bar)",
    "trade.allow_same_bar_reentry": "Allow Position Flips (LongShort same bar)",
    "trade.entry_mode": "Entry Mode",
    "signal_mode": "Signal Mode",
    "trade.exit_on_filter_block": "Exit if Selected Filters Block While In Position",
    "trade.exit_on_opposite_signal": "Exit on Opposite Signal",
    "trade.use_regime_lock": "Regime Lock",
    "supertrend.use_ha": "Use HA for Supertrend",
    "supertrend.use_wicks": "[ST] Use Wicks",
    "supertrend.atr_len": "[ST] ATR Len",
    "supertrend.factor": "[ST] Factor",
    "range_filter.use_bb_filter": "[RF] Use BB Filter (range mode)",
    "filters.use_ma_filter": "Use MA Filter",
    "filters.use_ma_slope_filter": "Use MA Slope Filter",
    "filters.use_mcginley_filter": "Use McGinley Filter",
    "filters.use_htf_trend_filter": "Use HTF Trend Filter",
    "filters.htf_trend_timeframe": "[HTF] Trend Timeframe",
    "filters.use_volume_filter": "Use Volume Participation Filter",
    "filters.use_atr_vol_filter": "Use ATR Volatility Floor",
    "filters.use_range_filters": "Use Range Filters (Entry Pause)",
    "filters.use_range_regime_filter": "Use Range Regime Filter",
    "filters.use_macd_filter": "Enable MACD filter hub",
    "filters.macd_gate_mode": "Mode",
    "risk.use_daily_loss_limit": "Use Daily Loss Limit?",
    "risk.max_daily_loss_percent": "Max Daily Loss (% of equity)",
    "stop_loss.use_sl": "Use Stop Loss",
    "stop_loss.mode": "SL Mode",
    "stop_loss.atr_len": "SL ATR Length",
    "stop_loss.atr_mult": "SL ATR Multiplier",
    "take_profit.use_tp": "Use Take Profit",
    "take_profit.mode": "TP Mode (single TP)",
    "multi_tp.use_multi_tp": "Use Multi TP (2 TPs, requires Take Profit = ON)",
    "multi_tp.tp1_rr": "TP1 at R multiple",
    "multi_tp.tp1_pct": "TP1 Close % of position",
    "multi_tp.tp2_rr": "TP2 at R multiple",
    "break_even.use_break_even": "Use Break-Even?",
    "break_even.rr": "BE trigger (R multiple)",
    "trailing.use_trailing": "Use Trailing Stop",
    "trailing.start_r": "Start After (R multiple)",
    "trailing.dist_r": "Trail Distance (R multiple)",
    "exit_filter_block.exit_on_ma_block": "Exit on MA Filter Block",
    "exit_filter_block.exit_on_ma_slope_block": "Exit on MA Slope Filter Block",
    "exit_filter_block.exit_on_mcginley_block": "Exit on McGinley Filter Block",
    "exit_filter_block.exit_on_htf_trend_block": "Exit on HTF Trend Filter Block",
    "exit_filter_block.exit_on_vol_part_block": "Exit on Volume Filter Block",
    "exit_filter_block.exit_on_vol_block": "Exit on Volume Filter Block",
    "exit_filter_block.exit_on_atr_vol_block": "Exit on ATR Vol Filter Block",
    "exit_filter_block.exit_on_range_block": "Exit on Range Filter Block",
    "time_stop.enabled": "Use Time Stop (Position Duration Exit)",
    "time_stop.bars": "Time Stop Bars",
    "time_stop.condition": "Time Stop Condition",
    "time_stop.exit_end_of_day": "Exit End Of Day",
    "time_stop.exit_end_of_week": "Exit End Of Week",
    "time_stop.eod": "Exit at End of Day",
    "time_stop.eow": "Exit at End of Week",
    "guards.use_dd_guard": "Use Max Drawdown Guard",
    "guards.dd_guard_pct": "Max Drawdown %",
    "guards.use_consec_loss_guard": "Use Consecutive Loss Halt",
    "guards.consec_loss_max": "Max Consecutive Losses",
    "guards.use_cooldown_guard": "Use Trade Cooldown",
    "guards.use_eq_curve_guard": "Use Equity Curve Filter",
    "guards.use_mae_guard": "Use MAE Guard (in-trade)",
    "guards.use_guard_recovery": "Use Guard Recovery (auto-resume after block)",
    "guards.guard_recovery_mode": "Recovery Mode",
}


def _label_for_path(path: str) -> str:
    return UI_LABELS.get(path, path)


def _format_ui_value(value: Any) -> str:
    if isinstance(value, bool):
        return "ON" if value else "OFF"
    if value is None:
        return "EMPTY"
    return str(value)


def _action_sort_key(path: str) -> tuple[int, str]:
    groups = [
        "signal_mode",
        "trade.",
        "filters.",
        "exit_filter_block.",
        "risk.",
        "stop_loss.",
        "take_profit.",
        "multi_tp.",
        "break_even.",
        "trailing.",
        "time_stop.",
        "guards.",
        "confirmation.",
    ]
    for i, prefix in enumerate(groups):
        if path == prefix or path.startswith(prefix):
            return (i, path)
    return (len(groups), path)


def normalize_key(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).strip().lower()


def parse_int(raw: Any, default: int = 0) -> int:
    try:
        return int(str(raw).strip())
    except Exception:
        return default


def read_manifest(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        for h in MANIFEST_HEADERS:
            row.setdefault(h, "")
    rows.sort(key=lambda r: parse_int(r.get("run_order", 0), 0))
    return rows


def ensure_manifest(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_HEADERS)
        writer.writeheader()


def read_existing_tracking(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path)
    except Exception:
        return {}
    if "Cases" not in wb.sheetnames:
        return {}

    ws = wb["Cases"]
    if ws.max_row < 2:
        return {}
    header_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        header_map[normalize_key(ws.cell(row=1, column=c).value)] = c

    if "case_id" not in header_map:
        return {}

    result: dict[str, dict[str, str]] = {}
    for r in range(2, ws.max_row + 1):
        case_id = ws.cell(row=r, column=header_map["case_id"]).value
        cid = str(case_id).strip() if case_id is not None else ""
        if not cid:
            continue
        row_map: dict[str, str] = {}
        for field in PRESERVED_COLUMNS:
            col = header_map.get(field)
            if col is None:
                continue
            value = ws.cell(row=r, column=col).value
            row_map[field] = "" if value is None else str(value)
        result[cid] = row_map
    return result


def make_case_folder(run_order: int, case_id: str) -> str:
    if run_order > 0:
        return f"{run_order:03d}_{case_id}"
    return case_id


def compose_dependency_note(row: dict[str, str]) -> str:
    depends_on = row.get("depends_on", "").strip()
    parent_required = row.get("parent_required", "").strip()
    if depends_on and parent_required:
        return f"depends_on={depends_on} | parent_required={parent_required}"
    if depends_on:
        return f"depends_on={depends_on}"
    if parent_required:
        return f"parent_required={parent_required}"
    return ""


def try_load_optimizer_module(suite_root: Path):
    scripts_dir = suite_root / "scripts"
    if not scripts_dir.exists():
        return None
    sys.path.insert(0, str(scripts_dir))
    try:
        import optimize_ui_coverage_case_set as opt  # type: ignore

        return opt
    except Exception:
        return None


def load_case_config_map(suite_root: Path, manifest_rows: list[dict[str, str]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for row in manifest_rows:
        case_id = row.get("case_id", "").strip()
        case_json_rel = row.get("case_json", "").strip()
        if not case_id or not case_json_rel:
            continue
        case_json_path = (suite_root / case_json_rel).resolve()
        if not case_json_path.exists():
            continue
        try:
            obj = json.loads(case_json_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        cfg = obj.get("config", {})
        if isinstance(cfg, dict):
            out[case_id] = cfg
    return out


def pick_baseline_case_id(manifest_rows: list[dict[str, str]]) -> str:
    ordered = sorted(manifest_rows, key=lambda r: parse_int(r.get("run_order", 0), 0))
    preferred = "parity_core_005_enable_long_trades_v01"
    for row in ordered:
        cid = row.get("case_id", "").strip()
        if cid == preferred:
            return cid
    for row in ordered:
        cid = row.get("case_id", "").strip()
        if cid and "baseline" in cid.lower():
            return cid
    return ordered[0].get("case_id", "").strip() if ordered else ""


def _collect_case_active_changes(
    suite_root: Path,
    manifest_rows: list[dict[str, str]],
) -> tuple[str, dict[str, dict[str, Any]], dict[str, list[tuple[str, Any]]]]:
    opt = try_load_optimizer_module(suite_root)
    cfg_map = load_case_config_map(suite_root, manifest_rows)
    baseline_case_id = pick_baseline_case_id(manifest_rows)
    baseline_cfg = cfg_map.get(baseline_case_id, {})
    if not baseline_cfg or opt is None:
        return baseline_case_id, cfg_map, {}

    bflat = opt.flatten_leaves(baseline_cfg)
    excluded_exact = set(getattr(opt, "EXCLUDED_EXACT_PATHS", set()))
    excluded_prefixes = tuple(getattr(opt, "EXCLUDED_PATH_PREFIXES", tuple()))
    changes: dict[str, list[tuple[str, Any]]] = {}

    for row in manifest_rows:
        case_id = row.get("case_id", "").strip()
        if not case_id:
            continue
        cfg = cfg_map.get(case_id, {})
        if not cfg:
            continue
        cflat = opt.flatten_leaves(cfg)
        keys = sorted(set(bflat.keys()) | set(cflat.keys()), key=_action_sort_key)
        active_changes: list[tuple[str, Any]] = []
        for path in keys:
            if path in excluded_exact:
                continue
            if any(path.startswith(prefix) for prefix in excluded_prefixes):
                continue
            bval_raw = bflat.get(path, None)
            cval_raw = cflat.get(path, None)
            if opt.normalize_value(bval_raw) == opt.normalize_value(cval_raw):
                continue
            if not opt.is_path_active(cfg, path):
                continue
            active_changes.append((path, cval_raw))
        changes[case_id] = active_changes
    return baseline_case_id, cfg_map, changes


def build_ui_actions_map(
    baseline_case_id: str,
    active_changes: dict[str, list[tuple[str, Any]]],
) -> dict[str, str]:
    result: dict[str, str] = {}
    for case_id, changes in active_changes.items():
        if case_id == baseline_case_id:
            result[case_id] = "Baseline setup: use baseline values exactly."
            continue
        if not changes:
            continue
        actions = [f"{_label_for_path(path)}={_format_ui_value(value)}" for path, value in changes]
        result[case_id] = "Set in TV UI: " + "; ".join(actions)
    return result


def _append_note(notes: list[str], text: str) -> None:
    if text not in notes:
        notes.append(text)


def build_dependency_note_map(
    manifest_rows: list[dict[str, str]],
    baseline_case_id: str,
    cfg_map: dict[str, dict[str, Any]],
    active_changes: dict[str, list[tuple[str, Any]]],
) -> dict[str, str]:
    notes_map: dict[str, str] = {}
    for row in manifest_rows:
        case_id = row.get("case_id", "").strip()
        if not case_id:
            continue
        if case_id == baseline_case_id:
            notes_map[case_id] = "Reference baseline case."
            continue

        cfg = cfg_map.get(case_id, {})
        changed_paths = [p for p, _ in active_changes.get(case_id, [])]
        notes: list[str] = []

        if str(row.get("expected_trade_behavior", "")).strip().upper() == "ZERO_TRADE_EXPECTED":
            _append_note(notes, "Signal Mode=None oldugu icin 0 trade beklenir.")

        if "trade.exit_on_filter_block" in changed_paths or any(p.startswith("exit_filter_block.") for p in changed_paths):
            _append_note(
                notes,
                "Exit-filter testi icin 3 kosul birlikte gerekir: Exit if Selected Filters Block=ON + ilgili Use Filter=ON + ilgili Exit on ... Block=ON.",
            )

        if any(p.startswith("time_stop.") and p != "time_stop.enabled" for p in changed_paths):
            _append_note(notes, "Time-stop alt ayarlari icin Use Time Stop=ON gerekir.")

        if any(p.startswith("stop_loss.") and p != "stop_loss.use_sl" for p in changed_paths):
            _append_note(notes, "Stop Loss alt ayarlari icin Use Stop Loss=ON gerekir.")

        if any(p.startswith("take_profit.") and p != "take_profit.use_tp" for p in changed_paths):
            _append_note(notes, "Take Profit alt ayarlari icin Use TP=ON gerekir.")

        if any(p.startswith("multi_tp.") for p in changed_paths):
            _append_note(notes, "Multi TP alt ayarlari icin Use TP=ON ve Use Multi TP=ON gerekir.")

        if any(p.startswith("break_even.") and p != "break_even.use_break_even" for p in changed_paths):
            _append_note(notes, "Break Even alt ayarlari icin Use Break Even=ON gerekir.")

        if any(p.startswith("trailing.") and p != "trailing.use_trailing" for p in changed_paths):
            _append_note(notes, "Trailing alt ayarlari icin Use Trailing Stop=ON gerekir.")

        if "risk.max_daily_loss_percent" in changed_paths:
            _append_note(notes, "Max Daily Loss % sadece Use Daily Loss Limit=ON iken etkilidir.")

        if "risk.max_trades_per_day" in changed_paths:
            _append_note(notes, "Max Trades Per Day sadece Use Max Trades Per Day=ON iken etkilidir.")

        if any(p.startswith("filters.range_regime_") or p in {"filters.use_range_regime_filter", "exit_filter_block.exit_on_range_block"} for p in changed_paths):
            _append_note(notes, "Range-regime ile ilgili ayarlar icin Use Range Filters=ON gerekir.")

        if any(p.startswith("guards.guard_recovery_") for p in changed_paths) or "guards.use_guard_recovery" in changed_paths:
            _append_note(notes, "Guard Recovery icin en az bir guard ON olmalidir (DD/ConsecLoss/Cooldown/EqCurve/MAE).")

        if any(p.startswith("supertrend.") for p in changed_paths):
            _append_note(notes, "Supertrend ayarlari sadece Signal Mode=Supertrend iken etkilidir.")

        if any(p.startswith("range_filter.") for p in changed_paths):
            _append_note(notes, "Range Filter ayarlari sadece Signal Mode=Range Filter Hybrid iken etkilidir.")

        manifest_note = compose_dependency_note(row)
        if manifest_note:
            _append_note(notes, manifest_note)

        if not notes:
            # Fallback with concise behavior reminder.
            signal_mode = str(cfg.get("signal_mode", "")).strip()
            if signal_mode:
                notes_map[case_id] = f"No extra parent dependency. Signal Mode={signal_mode}."
            else:
                notes_map[case_id] = "No extra parent dependency."
        else:
            notes_map[case_id] = " | ".join(notes)
    return notes_map


def build_rows(
    suite_root: Path,
    manifest_rows: list[dict[str, str]],
    preserved: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    baseline_case_id, cfg_map, active_changes = _collect_case_active_changes(suite_root, manifest_rows)
    ui_actions_map = build_ui_actions_map(baseline_case_id, active_changes)
    dependency_note_map = build_dependency_note_map(manifest_rows, baseline_case_id, cfg_map, active_changes)
    out: list[dict[str, str]] = []
    for row in manifest_rows:
        case_id = row.get("case_id", "").strip()
        if not case_id:
            continue
        run_order = parse_int(row.get("run_order", 0), 0)
        keep = preserved.get(case_id, {})
        guide_row = {
            "run_order": str(run_order) if run_order else "",
            "package": row.get("pack", "").strip(),
            "case_id": case_id,
            "case_folder": make_case_folder(run_order, case_id),
            "tv_preset_name": row.get("tv_preset_name", "").strip(),
            "primary_change": row.get("primary_change", "").strip(),
            "tv_ui_actions": row.get("ui_actions", "").strip() or ui_actions_map.get(case_id, ""),
            "dependency_note": dependency_note_map.get(case_id, compose_dependency_note(row)),
            "expected_trade_behavior": row.get("expected_trade_behavior", "NORMAL").strip() or "NORMAL",
            "tv_xlsx_status": keep.get(
                "tv_xlsx_status",
                "SKIP" if str(row.get("ui_actions", "")).strip().lower().startswith("reuse baseline xlsx from ") else "PENDING",
            ),
            "tv_xlsx_file": keep.get("tv_xlsx_file", ""),
            "tv_download_date": keep.get("tv_download_date", ""),
            "setup_check": keep.get("setup_check", "NOT_CHECKED"),
            "compare_status": keep.get("compare_status", "NOT_RUN"),
            "clip_strict_status": keep.get("clip_strict_status", ""),
            "raw_strict_status": keep.get("raw_strict_status", ""),
            "early_trade_end_candidate": keep.get("early_trade_end_candidate", ""),
            "gap_days": keep.get("gap_days", ""),
            "clip_tv_trades": keep.get("clip_tv_trades", ""),
            "clip_py_trades": keep.get("clip_py_trades", ""),
            "raw_tv_trades": keep.get("raw_tv_trades", ""),
            "raw_py_trades": keep.get("raw_py_trades", ""),
            "notes": keep.get("notes", row.get("notes", "").strip()),
        }
        out.append(guide_row)
    out.sort(key=lambda r: parse_int(r["run_order"], 10**9))
    return out


def write_readme_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    lines = [
        "CASE_SETUP_GUIDE usage",
        "",
        "1) Open 'Cases' sheet and filter by tv_xlsx_status = PENDING.",
        "2) Apply UI settings from 'tv_ui_actions'.",
        "3) Download TV XLSX and place into tv_manual_inputs/raw_tv_exports.",
        "4) Run route_tv_xlsx.py to auto-route files and update this workbook.",
        "5) Use setup_check and compare_status to track parity progression.",
        "6) Inspect dual diagnostics columns for raw-vs-clip mismatch split.",
        "",
        "Status values:",
        "- tv_xlsx_status: PENDING, DOWNLOADED, ROUTED, SETUP_CHECKED, DONE, SKIP",
        "- setup_check: NOT_CHECKED, OK, MISMATCH, NA",
        "- compare_status: NOT_RUN, PASS, MISMATCH, SETUP_MISMATCH, NO_TRADE_EXPECTED_PASS, ERROR",
        "- clip_strict_status/raw_strict_status: PASS or FAIL",
        "- early_trade_end_candidate: yes/no (TV trade-list truncation candidate)",
        "",
        f"Generated at UTC: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 120


def apply_case_sheet_format(ws, max_row: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(GUIDE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 10,
        "B": 10,
        "C": 46,
        "D": 50,
        "E": 34,
        "F": 44,
        "G": 52,
        "H": 42,
        "I": 24,
        "J": 18,
        "K": 52,
        "L": 18,
        "M": 16,
        "N": 24,
        "O": 16,
        "P": 16,
        "Q": 20,
        "R": 12,
        "S": 14,
        "T": 14,
        "U": 14,
        "V": 14,
        "W": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(GUIDE_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(2, max_row)}"

    for row in range(2, max_row + 1):
        for col in range(1, len(GUIDE_HEADERS) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    # Drop-downs
    dv_xlsx = DataValidation(
        type="list",
        formula1='"PENDING,DOWNLOADED,ROUTED,SETUP_CHECKED,DONE,SKIP"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_xlsx)
    dv_xlsx.add(f"J2:J{max(2, max_row)}")

    dv_setup = DataValidation(
        type="list",
        formula1='"NOT_CHECKED,OK,MISMATCH,NA"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_setup)
    dv_setup.add(f"M2:M{max(2, max_row)}")

    dv_compare = DataValidation(
        type="list",
        formula1='"NOT_RUN,PASS,MISMATCH,SETUP_MISMATCH,NO_TRADE_EXPECTED_PASS,ERROR"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_compare)
    dv_compare.add(f"N2:N{max(2, max_row)}")

    # Conditional format (xlsx status)
    xlsx_rules = [
        ("PENDING", "FFFDE9D9"),
        ("DOWNLOADED", "FFFFF2CC"),
        ("ROUTED", "FFD9E1F2"),
        ("SETUP_CHECKED", "FFD9EAD3"),
        ("DONE", "FFC6EFCE"),
        ("SKIP", "FFE7E6E6"),
    ]
    for value, color in xlsx_rules:
        ws.conditional_formatting.add(
            f"J2:J{max(2, max_row)}",
            FormulaRule(formula=[f'$J2="{value}"'], fill=PatternFill(fill_type="solid", fgColor=color)),
        )

    compare_rules = [
        ("PASS", "FFC6EFCE"),
        ("NO_TRADE_EXPECTED_PASS", "FFC6EFCE"),
        ("MISMATCH", "FFFFC7CE"),
        ("SETUP_MISMATCH", "FFFFEB9C"),
        ("ERROR", "FFFFC7CE"),
    ]
    for value, color in compare_rules:
        ws.conditional_formatting.add(
            f"N2:N{max(2, max_row)}",
            FormulaRule(formula=[f'$N2="{value}"'], fill=PatternFill(fill_type="solid", fgColor=color)),
        )

    ws.conditional_formatting.add(
        f"Q2:Q{max(2, max_row)}",
        FormulaRule(formula=['$Q2="yes"'], fill=PatternFill(fill_type="solid", fgColor="FFFFC7CE")),
    )


def build_workbook(out_path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    write_readme_sheet(wb)

    ws = wb.create_sheet("Cases")
    ws.append(GUIDE_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in GUIDE_HEADERS])
    apply_case_sheet_format(ws, ws.max_row)
    wb.save(out_path)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Build user-friendly CASE_SETUP_GUIDE.xlsx.")
    ap.add_argument(
        "--suite-root",
        default="mtc_backtest/parity_suite_350",
        help="Suite root folder",
    )
    ap.add_argument(
        "--manifest",
        default="manifests/cases_manifest_all.csv",
        help="Manifest path (relative to suite root unless absolute)",
    )
    ap.add_argument(
        "--out",
        default="CASE_SETUP_GUIDE.xlsx",
        help="Output workbook path (relative to suite root unless absolute)",
    )
    ap.add_argument(
        "--no-preserve",
        action="store_true",
        help="Do not preserve existing tracker fields from prior workbook.",
    )
    return ap.parse_args()


def resolve_path(base: Path, raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return (base / p).resolve()


def main() -> int:
    args = parse_args()
    suite_root = Path(args.suite_root).resolve()
    suite_root.mkdir(parents=True, exist_ok=True)

    manifest_path = resolve_path(suite_root, args.manifest)
    out_path = resolve_path(suite_root, args.out)

    ensure_manifest(manifest_path)
    rows = read_manifest(manifest_path)
    preserved = {}
    if out_path.exists() and not args.no_preserve:
        preserved = read_existing_tracking(out_path)
    guide_rows = build_rows(suite_root, rows, preserved)

    final_out = out_path
    try:
        build_workbook(out_path, guide_rows)
    except PermissionError:
        fallback = out_path.with_name(f"{out_path.stem}_AUTOSAVE{out_path.suffix}")
        build_workbook(fallback, guide_rows)
        final_out = fallback
    print(f"guide={final_out}")
    print(f"rows={len(guide_rows)}")
    print("status=ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
