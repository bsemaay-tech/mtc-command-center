#!/usr/bin/env python3
"""
Run parity tests for the 24 re-checked FAIL cases that now PASS setup.
Uses subprocess for backtest (same approach as check_and_run_new_cases.py).
"""

import sys, io, csv, json, os, subprocess, tempfile, traceback
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# Setup paths
BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
MTC_BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest")
CASES_DIR = BASE / "cases"
TV_DIR = BASE / "tv_manual_inputs"
DEBUG_DIR = MTC_BASE / "debug" / "parity_suite_350"
COMPARE_RUNS = BASE / "compare_runs"

sys.path.insert(0, str(MTC_BASE))
import openpyxl
from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades, clip_overlap, summarize_report


def run_backtest(case_id):
    """Run Python backtest via subprocess."""
    case_json = CASES_DIR / f"{case_id}.json"
    cmd = ["python", str(MTC_BASE / "scripts" / "run_case.py"), str(case_json)]
    try:
        result = subprocess.run(cmd, cwd=str(MTC_BASE), capture_output=True, text=True, timeout=300)
        if result.returncode != 0:
            # Extract last meaningful error line
            err_lines = [l for l in result.stderr.strip().split("\n") if l.strip()]
            err_msg = err_lines[-1] if err_lines else "returncode != 0"
            return False, err_msg
        return True, ""
    except subprocess.TimeoutExpired:
        return False, "Timeout (300s)"
    except Exception as e:
        return False, str(e)


def compare_parity(ro, case_id):
    """Compare TV vs Python trades."""
    folder_name = f"{ro}_{case_id}"
    case_folder = TV_DIR / folder_name
    if not case_folder.exists():
        # Try glob
        folders = list(TV_DIR.glob(f"{ro}_*"))
        if not folders:
            return "ERROR", f"No TV folder {ro}_*"
        case_folder = folders[0]

    xlsx_files = [f for f in case_folder.glob("*.xlsx") if not f.name.startswith("~$")]
    if not xlsx_files:
        return "ERROR", "No XLSX"

    # Extract TV trades to temp CSV
    tv_csv = None
    try:
        wb = openpyxl.load_workbook(xlsx_files[0], read_only=True, data_only=True)
        sheet_name = None
        for sn in wb.sheetnames:
            if "list of trades" in sn.lower():
                sheet_name = sn
                break
        if not sheet_name:
            wb.close()
            return "SKIP", "No 'List of trades' sheet"

        ws = wb[sheet_name]
        temp_file = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
        writer = csv.writer(temp_file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
        temp_file.close()
        tv_csv = Path(temp_file.name)
        wb.close()
    except Exception as e:
        return "ERROR", f"TV extract error: {e}"

    # Find Python trades
    debug_dir = DEBUG_DIR / case_id
    if not debug_dir.exists():
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()
        return "ERROR", "Debug dir not found"

    py_files = list(debug_dir.glob("debug_python_trades_*.csv"))
    if not py_files:
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()
        return "ERROR", "PY trades not found"

    py_csv = sorted(py_files)[-1]

    try:
        case_json_path = CASES_DIR / f"{case_id}.json"
        tv_tz = "Europe/London"
        if case_json_path.exists():
            with open(case_json_path, "r", encoding="utf-8") as f:
                case_payload = json.load(f)
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
        tv_df = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
        py_df = load_py_trades(py_csv)
        tv_df, py_df = clip_overlap(tv_df, py_df)
        report = build_report(tv_df, py_df)
        summary = summarize_report(tv_df, py_df, report)
        if summary["tv_trades"] == 0 and summary["py_trades"] == 0:
            return "PASS", "No trades"
        if summary["strict_pass"]:
            return "PASS", f"{summary['tv_trades']} trades match"
        return "MISMATCH", (
            f"core={summary['core_match_count']}/{summary['compared']}; "
            f"tv={summary['tv_trades']}; py={summary['py_trades']}; "
            f"extra_tv={summary['extra_tv_trades']}; extra_py={summary['extra_py_trades']}"
        )
    except Exception as e:
        return "ERROR", f"Compare error: {e}"
    finally:
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()


# ─── MAIN ───

# Read cases
cases = []
with open(COMPARE_RUNS / "recheck_fail_results.csv") as f:
    for row in csv.DictReader(f):
        if row["setup_status"] == "PASS":
            cases.append((row["run_order"], row["case_id"]))

print(f"Running parity for {len(cases)} cases...")
print(f"Start: {datetime.now():%H:%M:%S}")
print("=" * 70)

results = []

for idx, (ro, cid) in enumerate(cases):
    print(f"\n[{idx+1}/{len(cases)}] [{ro}] {cid}")

    # Step 1: Backtest
    ok, err = run_backtest(cid)
    if not ok:
        print(f"  BT_FAIL: {err[:120]}")
        results.append((ro, cid, "PASS", "BT_FAIL", err[:200]))
        continue
    print(f"  Backtest OK")

    # Step 2: Parity
    status, notes = compare_parity(ro, cid)
    print(f"  {status}: {notes}")
    results.append((ro, cid, "PASS", status, notes))

# Save
out_path = COMPARE_RUNS / "batch_recheck_results.csv"
with open(out_path, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["run_order", "case_id", "setup_status", "parity_status", "notes"])
    for r in results:
        w.writerow(r)

print("\n" + "=" * 70)
print("SUMMARY")
print("=" * 70)
counts = Counter(r[3] for r in results)
for k, v in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {k:12s} {v:4d}")
print(f"  {'TOTAL':12s} {len(results):4d}")
print(f"\nSaved to {out_path.name}")
print(f"End: {datetime.now():%H:%M:%S}")
