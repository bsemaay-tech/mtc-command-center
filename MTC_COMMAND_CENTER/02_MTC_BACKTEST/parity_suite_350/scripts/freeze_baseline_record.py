#!/usr/bin/env python
"""
Create Phase-0 baseline freeze record from TV XLSX + baseline case JSON.

Outputs:
- manifests/baseline_freeze_record.json
- manifests/baseline_freeze_record.md
- manifests/frozen_baseline_case_<case_id>.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def git_commit_short(workspace_root: Path) -> str:
    try:
        out = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=workspace_root,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        return out.strip()
    except Exception:
        return ""


def parse_pine_version(pine_file: Path) -> str:
    if not pine_file.exists():
        return ""
    text = pine_file.read_text(encoding="utf-8", errors="ignore")
    for line in text.splitlines()[:120]:
        if "MASTER_TEMPLATE_CORE" in line and "v" in line:
            m = re.search(r"v\d+\.\d+\.\d+(?:[-\w\.]+)?", line)
            if m:
                return m.group(0)
    m = re.search(r"v\d+\.\d+\.\d+(?:[-\w\.]+)?", text)
    return m.group(0) if m else ""


def normalize_key(raw: Any) -> str:
    return str(raw).strip().lower() if raw is not None else ""


def first_numeric_cell(row: list[Any], start_idx: int = 1) -> float | int | None:
    for val in row[start_idx:]:
        if isinstance(val, (int, float)):
            return val
    return None


def first_nonempty_cell(row: list[Any], start_idx: int = 1) -> Any:
    for val in row[start_idx:]:
        if val is None:
            continue
        if str(val).strip() == "":
            continue
        return val
    return None


def parse_xlsx_baseline(xlsx_path: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        out: dict[str, Any] = {
            "properties": {},
            "metrics": {},
        }

        # Properties sheet (TradingView export style: col2=name, col3=value).
        prop_sheet = None
        for name in wb.sheetnames:
            if normalize_key(name) == "properties":
                prop_sheet = wb[name]
                break
        if prop_sheet is not None:
            props: dict[str, Any] = {}
            for row in prop_sheet.iter_rows(values_only=True):
                vals = list(row)
                if len(vals) >= 3 and vals[1] is not None:
                    k = str(vals[1]).strip()
                    v = vals[2]
                elif len(vals) >= 2 and vals[0] is not None:
                    k = str(vals[0]).strip()
                    v = vals[1]
                else:
                    continue
                if not k:
                    continue
                props[k] = v
            out["properties"] = props

        def read_label_metric(sheet_name: str, label: str) -> Any:
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            target = label.strip().lower()
            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                if not vals:
                    continue
                first = vals[0]
                if normalize_key(first) == target:
                    # prefer numeric (All USDT / count columns), fallback to first non-empty
                    return first_numeric_cell(vals, 1) or first_nonempty_cell(vals, 1)
            return None

        out["metrics"] = {
            "total_trades": read_label_metric("Trades analysis", "Total trades"),
            "percent_profitable": read_label_metric("Trades analysis", "Percent profitable"),
            "net_profit": read_label_metric("Performance", "Net profit"),
            "max_equity_drawdown": read_label_metric("Performance", "Max equity drawdown"),
            "profit_factor": read_label_metric("Risk-adjusted performance", "Profit factor"),
        }
        return out
    finally:
        wb.close()


def main() -> int:
    ap = argparse.ArgumentParser(description="Freeze baseline record from TV XLSX + baseline JSON.")
    ap.add_argument("--workspace-root", default=".", help="Workspace root (git root).")
    ap.add_argument("--suite-root", default="mtc_backtest/parity_suite_350", help="Parity suite root.")
    ap.add_argument("--baseline-case-id", default="parity_core_005_enable_long_trades_v01")
    ap.add_argument("--baseline-xlsx", required=True, help="Path to baseline TV XLSX export.")
    ap.add_argument("--baseline-case-json", default="", help="Optional baseline case JSON path override.")
    ap.add_argument(
        "--pine-file",
        default="00_MASTER_TEMPLATE/MASTER_TEMPLATE_CORE.pine",
        help="Path to Pine source used for version/tag extraction.",
    )
    args = ap.parse_args()

    workspace_root = Path(args.workspace_root).resolve()
    suite_root = Path(args.suite_root).resolve()
    manifests_dir = suite_root / "manifests"
    manifests_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = Path(args.baseline_xlsx).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"baseline xlsx not found: {xlsx_path}")

    if args.baseline_case_json.strip():
        case_json_path = Path(args.baseline_case_json).resolve()
    else:
        case_json_path = (suite_root / "cases" / f"{args.baseline_case_id}.json").resolve()
    if not case_json_path.exists():
        raise FileNotFoundError(f"baseline case json not found: {case_json_path}")

    pine_file = (workspace_root / args.pine_file).resolve()
    pine_version = parse_pine_version(pine_file)

    parsed = parse_xlsx_baseline(xlsx_path)
    props = parsed.get("properties", {})
    metrics = parsed.get("metrics", {})

    snapshot_name = f"frozen_baseline_case_{args.baseline_case_id}.json"
    snapshot_path = manifests_dir / snapshot_name
    shutil.copy2(case_json_path, snapshot_path)

    record = {
        "generated_at_utc": utc_now(),
        "baseline_case_id": args.baseline_case_id,
        "baseline_xlsx": str(xlsx_path),
        "baseline_xlsx_sha256": sha256_file(xlsx_path),
        "symbol": props.get("Symbol"),
        "timeframe": props.get("Timeframe"),
        "trading_range": props.get("Trading range"),
        "backtesting_range": props.get("Backtesting range"),
        "pine_file": str(pine_file) if pine_file.exists() else "",
        "pine_version_tag": pine_version,
        "python_commit_hash": git_commit_short(workspace_root),
        "baseline_case_json": str(case_json_path),
        "baseline_case_json_sha256": sha256_file(case_json_path),
        "frozen_case_snapshot": str(snapshot_path),
        "frozen_case_snapshot_sha256": sha256_file(snapshot_path),
        "tv_metrics_snapshot": metrics,
        "tv_properties_subset": {
            "Initial capital": props.get("Initial capital"),
            "Order size": props.get("Order size"),
            "Pyramiding": props.get("Pyramiding"),
            "Commission": props.get("Commission"),
            "Slippage": props.get("Slippage"),
            "Recalculate on bar close": props.get("Recalculate on bar close"),
            "Backtesting precision. Use bar magnifier": props.get("Backtesting precision. Use bar magnifier"),
        },
    }

    json_out = manifests_dir / "baseline_freeze_record.json"
    json_out.write_text(json.dumps(record, ensure_ascii=True, indent=2), encoding="utf-8")

    md_lines = [
        "# Baseline Freeze Record",
        "",
        f"- generated_at_utc: `{record['generated_at_utc']}`",
        f"- baseline_case_id: `{record['baseline_case_id']}`",
        f"- baseline_xlsx: `{record['baseline_xlsx']}`",
        f"- symbol: `{record.get('symbol')}`",
        f"- timeframe: `{record.get('timeframe')}`",
        f"- trading_range: `{record.get('trading_range')}`",
        f"- backtesting_range: `{record.get('backtesting_range')}`",
        f"- pine_version_tag: `{record.get('pine_version_tag')}`",
        f"- python_commit_hash: `{record.get('python_commit_hash')}`",
        f"- baseline_case_json: `{record.get('baseline_case_json')}`",
        f"- frozen_case_snapshot: `{record.get('frozen_case_snapshot')}`",
        "",
        "## TV Metrics Snapshot",
        "",
        f"- total_trades: `{metrics.get('total_trades')}`",
        f"- net_profit: `{metrics.get('net_profit')}`",
        f"- max_equity_drawdown: `{metrics.get('max_equity_drawdown')}`",
        f"- percent_profitable: `{metrics.get('percent_profitable')}`",
        f"- profit_factor: `{metrics.get('profit_factor')}`",
        "",
        "## Integrity",
        "",
        f"- baseline_xlsx_sha256: `{record['baseline_xlsx_sha256']}`",
        f"- baseline_case_json_sha256: `{record['baseline_case_json_sha256']}`",
        f"- frozen_case_snapshot_sha256: `{record['frozen_case_snapshot_sha256']}`",
        "",
        f"- json_record: `{json_out}`",
    ]
    md_out = manifests_dir / "baseline_freeze_record.md"
    md_out.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    print(f"baseline_case_id={record['baseline_case_id']}")
    print(f"baseline_xlsx={record['baseline_xlsx']}")
    print(f"symbol={record.get('symbol')}")
    print(f"timeframe={record.get('timeframe')}")
    print(f"python_commit_hash={record.get('python_commit_hash')}")
    print(f"json_record={json_out}")
    print(f"md_record={md_out}")
    print("status=ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
