#!/usr/bin/env python3
"""
Extract TV trades from XLSX and compare against Python trades for Core-1 cases.

Simpler approach: extract trades, run comparison, evaluate results.
"""

import csv
import json
import tempfile
from pathlib import Path
from typing import List, Tuple, Optional
import sys
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades, clip_overlap, summarize_report

def read_pass_cases(csv_path: Path) -> List[Tuple[int, str]]:
    """Read PASS cases from setup check CSV."""
    pass_cases = []
    try:
        with open(csv_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("status") == "PASS":
                    run_order = int(row.get("run_order", 0))
                    case_id = row.get("case_id", "")
                    if case_id:
                        pass_cases.append((run_order, case_id))
    except Exception as e:
        print(f"ERROR reading {csv_path}: {e}")
        return []
    return pass_cases

def find_tv_xlsx(case_folder: Path) -> Optional[Path]:
    """Find TV XLSX in case folder."""
    xlsx_files = list(case_folder.glob("*.xlsx"))
    return xlsx_files[0] if xlsx_files else None

def extract_tv_trades_csv(xlsx_path: Path) -> Optional[Path]:
    """Extract 'List of trades' sheet from TV XLSX to temp CSV."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Try to find "List of trades" sheet
        sheet_name = None
        for sn in wb.sheetnames:
            if "list of trades" in sn.lower():
                sheet_name = sn
                break

        if not sheet_name:
            print(f"  WARNING: no 'List of trades' sheet found")
            wb.close()
            return None

        ws = wb[sheet_name]

        # Write to temp CSV
        temp_file = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
        writer = csv.writer(temp_file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
        temp_file.close()

        wb.close()
        return Path(temp_file.name)
    except Exception as e:
        print(f"  ERROR extracting trades: {e}")
        return None

def compare_case(run_order: int, case_id: str, base_path: Path) -> Tuple[str, str]:
    """Compare a single case. Returns (status, notes)."""
    # Find TV XLSX
    case_folder = base_path / "tv_manual_inputs" / f"{run_order:03d}_{case_id}"
    if not case_folder.exists():
        return "ERROR", "TV folder not found"

    tv_xlsx = find_tv_xlsx(case_folder)
    if not tv_xlsx:
        return "ERROR", "XLSX not found"

    # Extract TV trades to temp CSV
    tv_csv = extract_tv_trades_csv(tv_xlsx)
    if not tv_csv:
        return "SKIP", "Could not extract TV trades"

    # Find Python trades CSV
    debug_case_dir = Path("C:\\LAB\\tradingview-lab\\mtc_backtest\\debug\\parity_suite_350") / case_id
    if not debug_case_dir.exists():
        return "ERROR", "Debug dir not found"

    py_trades_files = list(debug_case_dir.glob("debug_python_trades_*.csv"))
    if not py_trades_files:
        return "ERROR", "PY trades not found"

    py_csv = sorted(py_trades_files)[-1]

    try:
        # Load and compare
        case_json_path = base_path / "cases" / f"{case_id}.json"
        tv_tz = "Europe/London"
        if case_json_path.exists():
            with open(case_json_path, "r", encoding="utf-8") as f:
                case_payload = json.load(f)
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
        tv_df = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
        py_df = load_py_trades(py_csv)

        # Clip to overlapping time
        tv_df, py_df = clip_overlap(tv_df, py_df)

        # Build report
        report = build_report(tv_df, py_df)

        # Evaluate
        summary = summarize_report(tv_df, py_df, report)
        if summary["tv_trades"] == 0 and summary["py_trades"] == 0:
            return "PASS", "No trades to compare"
        if summary["strict_pass"]:
            return "PASS", f"{summary['tv_trades']} trades match"
        return "MISMATCH", (
            f"core={summary['core_match_count']}/{summary['compared']}; "
            f"tv={summary['tv_trades']}; py={summary['py_trades']}; "
            f"extra_tv={summary['extra_tv_trades']}; extra_py={summary['extra_py_trades']}"
        )
    except Exception as e:
        return "ERROR", f"Compare error: {str(e)[:50]}"
    finally:
        # Clean up temp TV CSV
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()

def main():
    base_path = Path("C:\\LAB\\tradingview-lab\\mtc_backtest\\parity_suite_350")
    compare_runs_base = base_path / "compare_runs"
    csv_path = compare_runs_base / "core1_setup_check.csv"

    if not csv_path.exists():
        print(f"ERROR: setup check CSV not found")
        return

    pass_cases = read_pass_cases(csv_path)
    print(f"Comparing {len(pass_cases)} PASS cases...")
    print()

    results = []
    pass_count = 0
    mismatch_count = 0
    skip_count = 0
    error_count = 0

    for run_order, case_id in pass_cases:
        print(f"[{run_order:03d}] {case_id}...", end=" ", flush=True)
        status, notes = compare_case(run_order, case_id, base_path)

        if status == "PASS":
            print("[PASS]")
            pass_count += 1
        elif status == "MISMATCH":
            print("[MISMATCH]")
            mismatch_count += 1
        elif status == "SKIP":
            print("[SKIP]")
            skip_count += 1
        else:
            print(f"[ERROR]")
            error_count += 1

        results.append({
            "run_order": run_order,
            "case_id": case_id,
            "compare_status": status,
            "notes": notes,
        })

    print()
    print(f"Summary: {pass_count} PASS, {mismatch_count} MISMATCH, {skip_count} SKIP, {error_count} ERROR")
    print()

    # Write CSV
    csv_out = compare_runs_base / "core1_parity_results.csv"
    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["run_order", "case_id", "compare_status", "notes"])
        writer.writeheader()
        writer.writerows(results)

    print(f"Results: {csv_out}")

if __name__ == "__main__":
    main()
