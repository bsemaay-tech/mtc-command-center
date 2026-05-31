#!/usr/bin/env python3
"""
Batch execution + parity comparison for all PASS cases.

Phase 2: Run Python backtest for all PASS cases
Phase 3: Compare TV vs Python trades
"""

import subprocess
import csv
import tempfile
import json
from pathlib import Path
from typing import List, Tuple, Optional
import sys
import openpyxl
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades, clip_overlap, summarize_report
from src.parity.dual_status import classify_parity_with_dual_view, compute_gap_days

def read_pass_cases(csv_path: Path) -> List[Tuple[int, str]]:
    """Read PASS cases from setup check CSV."""
    pass_cases = []
    try:
        with open(csv_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("status") == "PASS":
                    run_order = int(row.get("run_order", 0))
                    case_id = row.get("case_id", "")
                    if case_id:
                        pass_cases.append((run_order, case_id))
    except Exception as e:
        print(f"ERROR reading {csv_path}: {e}")
        return []
    return pass_cases

def run_case(case_id: str, mtc_backtest_base: Path) -> bool:
    """Run a single case."""
    case_json = mtc_backtest_base / "parity_suite_350" / "cases" / f"{case_id}.json"
    if not case_json.exists():
        return False

    cmd = ["python", str(mtc_backtest_base / "scripts" / "run_case.py"), str(case_json)]

    try:
        result = subprocess.run(cmd, cwd=str(mtc_backtest_base), capture_output=True, text=True, timeout=300)
        return result.returncode == 0
    except:
        return False

def find_tv_xlsx(case_folder: Path) -> Optional[Path]:
    """Find TV XLSX in case folder."""
    xlsx_files = list(case_folder.glob("*.xlsx"))
    return xlsx_files[0] if xlsx_files else None

def extract_tv_trades_csv(xlsx_path: Path) -> Optional[Path]:
    """Extract 'List of trades' sheet from TV XLSX to temp CSV."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        sheet_name = None
        for sn in wb.sheetnames:
            if "list of trades" in sn.lower():
                sheet_name = sn
                break

        if not sheet_name:
            wb.close()
            return None

        ws = wb[sheet_name]

        temp_file = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
        writer = csv.writer(temp_file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
        temp_file.close()

        wb.close()
        return Path(temp_file.name)
    except:
        return None

def compare_case(run_order: int, case_id: str, base_path: Path) -> Tuple[str, str, dict]:
    """Compare a single case."""
    case_folder = base_path / "tv_manual_inputs" / f"{run_order:03d}_{case_id}"
    if not case_folder.exists():
        return "ERROR", "TV folder not found", {}

    tv_xlsx = find_tv_xlsx(case_folder)
    if not tv_xlsx:
        return "ERROR", "XLSX not found", {}

    tv_csv = extract_tv_trades_csv(tv_xlsx)
    if not tv_csv:
        return "SKIP", "Could not extract TV trades", {}

    debug_case_dir = Path("C:\\LAB\\tradingview-lab\\mtc_backtest\\debug\\parity_suite_350") / case_id
    if not debug_case_dir.exists():
            return "ERROR", "Debug dir not found", {}

    py_trades_files = list(debug_case_dir.glob("debug_python_trades_*.csv"))
    if not py_trades_files:
            return "ERROR", "PY trades not found", {}

    py_csv = sorted(py_trades_files)[-1]

    try:
        case_json_path = base_path / "cases" / f"{case_id}.json"
        tv_tz = "Europe/London"
        case_end_utc = None
        if case_json_path.exists():
            with open(case_json_path, "r", encoding="utf-8") as f:
                case_payload = json.load(f)
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
            end_raw = case_payload.get("end_date")
            if end_raw:
                try:
                    case_end_utc = datetime.fromisoformat(str(end_raw))
                except ValueError:
                    case_end_utc = datetime.strptime(str(end_raw), "%Y-%m-%d")
                if case_end_utc.tzinfo is None:
                    case_end_utc = case_end_utc.replace(tzinfo=timezone.utc)
                else:
                    case_end_utc = case_end_utc.astimezone(timezone.utc)

        tv_raw = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
        py_raw = load_py_trades(py_csv)

        raw_report = build_report(tv_raw, py_raw)
        raw_summary = summarize_report(tv_raw, py_raw, raw_report)

        tv_clip, py_clip = clip_overlap(tv_raw, py_raw)
        clip_report = build_report(tv_clip, py_clip)
        clip_summary = summarize_report(tv_clip, py_clip, clip_report)

        last_tv_exit_utc = None
        if len(tv_raw) > 0:
            try:
                last_tv_exit_utc = tv_raw["exit_time"].max().to_pydatetime()
            except Exception:
                last_tv_exit_utc = None
        gap_days = compute_gap_days(case_end_utc, last_tv_exit_utc)
        return classify_parity_with_dual_view(
            clip_summary=clip_summary,
            raw_summary=raw_summary,
            gap_days=gap_days,
        )
    except Exception as e:
        return "ERROR", f"Compare error", {}
    finally:
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()

def main():
    base_path = Path("C:\\LAB\\tradingview-lab\\mtc_backtest\\parity_suite_350")
    compare_runs_base = base_path / "compare_runs"
    csv_path = compare_runs_base / "all_setup_check.csv"

    if not csv_path.exists():
        print(f"ERROR: setup check CSV not found")
        return

    mtc_backtest_base = base_path.parent
    pass_cases = read_pass_cases(csv_path)

    print(f"Processing {len(pass_cases)} PASS cases...")
    print()

    # PHASE 2: BACKTEST
    print("=" * 60)
    print("PHASE 2: PYTHON BACKTEST")
    print("=" * 60)

    backtest_ok = 0
    backtest_fail = 0

    for i, (run_order, case_id) in enumerate(pass_cases):
        print(f"[{i+1:3d}/{len(pass_cases)}] [{run_order:03d}] {case_id:50s}...", end=" ", flush=True)
        if run_case(case_id, mtc_backtest_base):
            print("[OK]")
            backtest_ok += 1
        else:
            print("[FAILED]")
            backtest_fail += 1

    print()
    print(f"Backtest summary: {backtest_ok} OK, {backtest_fail} FAILED")
    print()

    # PHASE 3: COMPARISON
    print("=" * 60)
    print("PHASE 3: TV VS PYTHON PARITY COMPARISON")
    print("=" * 60)

    results = []
    pass_count = 0
    mismatch_count = 0
    skip_count = 0
    error_count = 0

    for i, (run_order, case_id) in enumerate(pass_cases):
        print(f"[{i+1:3d}/{len(pass_cases)}] [{run_order:03d}] {case_id:50s}...", end=" ", flush=True)
        status, notes, diag = compare_case(run_order, case_id, base_path)

        if status == "PASS":
            print("[PASS]")
            pass_count += 1
        elif status == "MISMATCH":
            print("[MISMATCH]")
            mismatch_count += 1
        elif status == "SKIP":
            print("[SKIP]")
            skip_count += 1
        else:
            print("[ERROR]")
            error_count += 1

        results.append({
            "run_order": run_order,
            "case_id": case_id,
            "compare_status": status,
            "notes": notes,
            "clip_strict_status": diag.get("clip_strict_status", ""),
            "raw_strict_status": diag.get("raw_strict_status", ""),
            "early_trade_end_candidate": diag.get("early_trade_end_candidate", ""),
            "gap_days": diag.get("gap_days", ""),
            "clip_tv_trades": diag.get("clip_tv_trades", ""),
            "clip_py_trades": diag.get("clip_py_trades", ""),
            "raw_tv_trades": diag.get("raw_tv_trades", ""),
            "raw_py_trades": diag.get("raw_py_trades", ""),
        })

    print()
    print(f"Comparison summary: {pass_count} PASS, {mismatch_count} MISMATCH, {skip_count} SKIP, {error_count} ERROR")
    print()

    # Write CSV
    csv_out = compare_runs_base / "all_parity_results.csv"
    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "run_order",
                "case_id",
                "compare_status",
                "notes",
                "clip_strict_status",
                "raw_strict_status",
                "early_trade_end_candidate",
                "gap_days",
                "clip_tv_trades",
                "clip_py_trades",
                "raw_tv_trades",
                "raw_py_trades",
            ],
        )
        writer.writeheader()
        writer.writerows(results)

    print(f"Results: {csv_out}")

    diagnostics_script = base_path / "scripts" / "run_post_compare_diagnostics.py"
    if diagnostics_script.exists():
        print("\nRunning post-compare diagnostics...")
        subprocess.run(
            ["python", str(diagnostics_script)],
            cwd=str(base_path),
            check=False,
        )

    # Summary statistics
    total = len(results)
    print()
    print(f"Final Summary ({total} cases):")
    print(f"  PASS:     {pass_count:3d} ({100*pass_count/total:.1f}%)")
    print(f"  MISMATCH: {mismatch_count:3d} ({100*mismatch_count/total:.1f}%)")
    print(f"  SKIP:     {skip_count:3d} ({100*skip_count/total:.1f}%)")
    print(f"  ERROR:    {error_count:3d} ({100*error_count/total:.1f}%)")

if __name__ == "__main__":
    main()
