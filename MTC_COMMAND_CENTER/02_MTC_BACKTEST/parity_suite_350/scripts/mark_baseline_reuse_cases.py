#!/usr/bin/env python
"""
Mark cases that are UI-identical to baseline defaults as baseline-reuse.

Logic:
- Read defaults from input map (Properties column C by seq).
- For each manifest case, read case json `_tv_case` metadata:
  - target seq/value
  - parent seq/value list
- If target and all parent values equal workbook defaults, case is baseline-identical.
- Keep the earliest run_order as anchor baseline case.
- Rewrite manifests to mark other baseline-identical cases as reusable from anchor.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MANIFEST_FILES = [
    "cases_manifest_all.csv",
    "cases_manifest_core.csv",
    "cases_manifest_boundary.csv",
    "cases_manifest_pairwise.csv",
]


def normalize_text(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "ON" if raw else "OFF"
    txt = str(raw).strip()
    low = txt.lower()
    if low in {"on", "true", "yes", "y", "1"}:
        return "ON"
    if low in {"off", "false", "no", "n", "0"}:
        return "OFF"
    try:
        # Normalize numerics ("1", "1.0", 1)
        f = float(txt.replace(",", "."))
        if f.is_integer():
            return str(int(f))
        return str(f).rstrip("0").rstrip(".")
    except Exception:
        return txt.upper()


def load_defaults_by_seq(input_map_path: Path) -> dict[int, str]:
    wb = load_workbook(input_map_path, data_only=True, read_only=True)
    try:
        ws = wb["Properties"]
        out: dict[int, str] = {}
        for r in range(7, ws.max_row + 1):
            seq_raw = ws.cell(r, 1).value
            if not isinstance(seq_raw, (int, float)):
                continue
            seq = int(seq_raw)
            default_value = ws.cell(r, 3).value
            out[seq] = normalize_text(default_value)
        return out
    finally:
        wb.close()


def load_manifest(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def write_manifest(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def parse_int(raw: Any, default: int = 0) -> int:
    try:
        return int(str(raw).strip())
    except Exception:
        return default


def is_case_baseline_identical(case_json_path: Path, defaults_by_seq: dict[int, str]) -> bool:
    if not case_json_path.exists():
        return False
    try:
        obj = json.loads(case_json_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    tv_case = obj.get("_tv_case", {})
    if not isinstance(tv_case, dict):
        return False

    seq = tv_case.get("seq")
    target_val = tv_case.get("target_value")
    if not isinstance(seq, int):
        return False

    if normalize_text(target_val) != defaults_by_seq.get(seq, "__MISSING__"):
        return False

    parents = tv_case.get("parents", [])
    if not isinstance(parents, list):
        return False
    for p in parents:
        if not isinstance(p, dict):
            return False
        pseq = p.get("seq")
        pval = p.get("value")
        if not isinstance(pseq, int):
            return False
        if normalize_text(pval) != defaults_by_seq.get(pseq, "__MISSING__"):
            return False
    return True


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Mark baseline-identical cases as reusable from one anchor baseline case.")
    ap.add_argument("--suite-root", default="mtc_backtest/parity_suite_350", help="Suite root")
    ap.add_argument("--input-map", default="manifests/input_map_FILLED_v6.xlsx", help="Input map path")
    ap.add_argument("--manifest-dir", default="manifests", help="Manifest dir path")
    ap.add_argument("--cases-dir", default="cases", help="Cases dir path")
    return ap.parse_args()


def resolve(base: Path, raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return (base / p).resolve()


def main() -> int:
    args = parse_args()
    suite_root = Path(args.suite_root).resolve()
    input_map_path = resolve(suite_root, args.input_map)
    manifest_dir = resolve(suite_root, args.manifest_dir)
    cases_dir = resolve(suite_root, args.cases_dir)

    if not input_map_path.exists():
        print(f"ERROR: input map not found: {input_map_path}")
        return 2

    all_manifest = manifest_dir / "cases_manifest_all.csv"
    rows_all = load_manifest(all_manifest)
    if not rows_all:
        print(f"ERROR: empty/all manifest missing: {all_manifest}")
        return 2

    defaults_by_seq = load_defaults_by_seq(input_map_path)

    baseline_identical: list[dict[str, str]] = []
    for row in rows_all:
        case_id = row.get("case_id", "").strip()
        if not case_id:
            continue
        cj = cases_dir / f"{case_id}.json"
        if is_case_baseline_identical(cj, defaults_by_seq):
            baseline_identical.append(row)

    if not baseline_identical:
        print("baseline_identical_count=0")
        print("anchor_case_id=")
        print("status=ok")
        return 0

    baseline_identical.sort(key=lambda r: parse_int(r.get("run_order", 0), 10**9))
    anchor_case_id = baseline_identical[0]["case_id"].strip()
    baseline_identical_ids = {r["case_id"].strip() for r in baseline_identical}

    reuse_map_rows: dict[str, dict[str, Any]] = {}

    def rewrite(rows: list[dict[str, str]]) -> list[dict[str, str]]:
        out: list[dict[str, str]] = []
        for row in rows:
            cid = row.get("case_id", "").strip()
            if cid in baseline_identical_ids and cid != anchor_case_id:
                row = dict(row)
                row["primary_change"] = f"reuse baseline ({anchor_case_id})"
                row["ui_actions"] = f"Reuse baseline XLSX from {anchor_case_id} (same UI as baseline)."
                row["depends_on"] = anchor_case_id
                row["parent_required"] = ""
                notes = row.get("notes", "").strip()
                tag = f"baseline_reuse_from={anchor_case_id}"
                if tag not in notes:
                    row["notes"] = f"{notes}; {tag}".strip("; ").strip()
                reuse_map_rows[cid] = {
                    "reuse_case_id": cid,
                    "reuse_from_case_id": anchor_case_id,
                    "run_order": row.get("run_order", ""),
                }
            out.append(row)
        return out

    # Rewrite all manifests consistently.
    for mf in MANIFEST_FILES:
        p = manifest_dir / mf
        rows = load_manifest(p)
        if not rows:
            continue
        updated = rewrite(rows)
        write_manifest(p, updated)

    # Export reuse map.
    reuse_map_path = manifest_dir / "baseline_reuse_map.csv"
    with reuse_map_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["reuse_case_id", "reuse_from_case_id", "run_order"])
        writer.writeheader()
        writer.writerows(
            sorted(reuse_map_rows.values(), key=lambda r: parse_int(r["run_order"], 10**9))
        )

    print(f"baseline_identical_count={len(baseline_identical)}")
    print(f"reuse_case_count={len(baseline_identical)-1}")
    print(f"anchor_case_id={anchor_case_id}")
    print(f"reuse_map={reuse_map_path}")
    print("status=ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
