#!/usr/bin/env python
"""
Generate a fresh manifest/case set from FILLED_v6 input-map workbook.

Design goals:
- Honor source count in Properties!J (numeric rows), no silent dedupe.
- Build one case per requested test value (K..U) for each input row (A seq).
- Keep router-compatible manifest/ui_actions and case folders.
- Apply known UI->config mappings; keep unknowns in metadata/notes.
"""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MANIFEST_HEADERS = [
    "run_order",
    "pack",
    "case_id",
    "case_json",
    "tv_preset_name",
    "enabled",
    "expected_trade_behavior",
    "primary_change",
    "ui_actions",
    "depends_on",
    "parent_required",
    "canonical_config_hash",
    "semantic_fingerprint",
    "symbol",
    "timeframe",
    "start_date",
    "end_date",
    "notes",
]


UI_TO_CONFIG_PATH: dict[str, str] = {
    "enable long trades": "trade.enable_long",
    "enable short trades": "trade.enable_short",
    "allow position flips (longshort same bar)": "trade.allow_flip",
    "exit on opposite signal": "trade.exit_on_opposite_signal",
    "exit if selected filters block while in position": "trade.exit_on_filter_block",
    "regime lock (prevent re-entry same direction)": "trade.use_regime_lock",
    "entry mode": "trade.entry_mode",
    "exit on ma filter block": "exit_filter_block.exit_on_ma_block",
    "exit on ma slope filter block": "exit_filter_block.exit_on_ma_slope_block",
    "exit on mcginley filter block": "exit_filter_block.exit_on_mcginley_block",
    "exit on htf trend filter block": "exit_filter_block.exit_on_htf_trend_block",
    "exit on volume filter block": "exit_filter_block.exit_on_vol_block",
    "exit on atr vol filter block": "exit_filter_block.exit_on_atr_vol_block",
    "exit on range filter block": "exit_filter_block.exit_on_range_block",
    "use time stop (position duration exit)": "time_stop.enabled",
    "time stop bars": "time_stop.bars",
    "exit at end of day": "time_stop.eod",
    "exit at end of week": "time_stop.eow",
    "time stop condition": "time_stop.condition",
    "signal mode max entries": "trade.signal_mode_max_entries",
    "signal mode cooldown bars": "trade.signal_mode_cooldown_bars",
    "max pyramid positions": "trade.max_pyramid_positions",
    "signal mode": "signal_mode",
    "use ha for supertrend": "supertrend.use_ha",
    "[st] use wicks": "supertrend.use_wicks",
    "[st] atr len": "supertrend.atr_len",
    "[st] factor": "supertrend.factor",
    "[rf] use bb filter (range mode)": "range_filter.use_bb_filter",
    "risk per long (% of equity)": "risk.risk_long_percent",
    "risk per short (% of equity)": "risk.risk_short_percent",
    "fallback position size when sl off (% of equity)": "risk.fallback_qty_pct",
    "use daily loss limit?": "risk.use_daily_loss_limit",
    "max daily loss (% of equity)": "risk.max_daily_loss_percent",
    "use max trades per day?": "risk.use_max_trades_per_day",
    "max trades per day": "risk.max_trades_per_day",
    "max leverage cap (limit)": "risk.max_leverage_cap",
    "risk equity source": "risk.risk_equity_mode",
    "use notional hard assert": "risk.use_notional_hard_assert",
    "sl mode": "stop_loss.mode",
    "swing sl basis": "stop_loss.swing_basis",
    "swing sl lookback": "stop_loss.swing_lookback",
    "sl % (distance)": "stop_loss.percent",
    "tp mode (single tp)": "take_profit.mode",
    "tp % (distance)": "take_profit.percent",
    "tp r multiple (single tp)": "take_profit.rr",
    "use stop loss": "stop_loss.use_sl",
    "use take profit": "take_profit.use_tp",
    "use break-even?": "break_even.use_break_even",
    "use multi tp (2 tps, requires take profit = on)": "multi_tp.use_multi_tp",
    "tp2 at r multiple": "multi_tp.tp2_rr",
    "swing sl atr length": "stop_loss.swing_atr_len",
    "swing sl atr multiplier": "stop_loss.swing_atr_mult",
    "sl atr length": "stop_loss.atr_len",
    "sl atr multiplier": "stop_loss.atr_mult",
    "tp atr length": "take_profit.atr_len",
    "tp atr multiplier": "take_profit.atr_mult",
    "be trigger (r multiple)": "break_even.rr",
    "be buffer (r multiple)": "break_even.buffer_r",
    "tp1 at r multiple": "multi_tp.tp1_rr",
    "tp1 close % of position": "multi_tp.tp1_pct",
    "use trailing stop": "trailing.use_trailing",
    "trailing atr length": "trailing.atr_len",
    "start after (r multiple)": "trailing.start_r",
    "trail distance (r multiple)": "trailing.dist_r",
    "use ma filter": "filters.use_ma_filter",
    "use ma slope filter": "filters.use_ma_slope_filter",
    "use mcginley filter": "filters.use_mcginley_filter",
    "[mcginley] length": "filters.mcginley_len",
    "[mcginley] use higher timeframe": "filters.use_mcginley_htf",
    "[mcginley] htf timeframe": "filters.mcginley_htf_timeframe",
    "use htf trend filter": "filters.use_htf_trend_filter",
    "use volume participation filter": "filters.use_volume_filter",
    "use range filters (entry pause)": "filters.use_range_filters",
    "range aggregation": "filters.range_agg_mode",
    "count mode min pass": "filters.range_min_pass",
    "use atr volatility floor": "filters.use_atr_vol_filter",
    "[htf] trend timeframe": "filters.htf_trend_timeframe",
    "use max drawdown guard": "guards.use_dd_guard",
    "max drawdown %": "guards.dd_guard_pct",
    "use consecutive loss halt": "guards.use_consec_loss_guard",
    "max consecutive losses": "guards.consec_loss_max",
    "use trade cooldown": "guards.use_cooldown_guard",
    "use equity curve filter": "guards.use_eq_curve_guard",
    "use mae guard (in-trade)": "guards.use_mae_guard",
    "use guard recovery (auto-resume after block)": "guards.use_guard_recovery",
    "recovery mode": "guards.guard_recovery_mode",
    "enable macd filter hub": "filters.use_macd_filter",
    "mode": "filters.macd_gate_mode",
    "use confirmation: swing break + momentum": "confirmation.enabled",
    "swing left bars": "confirmation.p_left",
    "swing right bars": "confirmation.p_right",
    "require close beyond swing (else wick)": "confirmation.require_close_beyond",
    "max bars to wait (timeout)": "confirmation.confirm_timeout_bars",
    "min bars to wait after raw": "confirmation.min_wait_bars",
    "apply confirmation only when flat": "confirmation.gate_only_when_flat",
    "update swing level while waiting (dynamic)": "confirmation.dynamic_level_while_waiting",
    "dynamic update mode": "confirmation.dyn_update_mode",
    "if dynamic updates level this bar, defer break check 1 bar": "confirmation.defer_break_on_level_update",
    "if new raw event arrives, refresh level+timer": "confirmation.refresh_on_new_raw_signal",
    "raw event mode": "confirmation.raw_event_mode",
    "if long & short raw same bar": "confirmation.same_bar_tie_rule",
    "bar close only (barstate.isconfirmed)": "confirmation.bar_close_only",
    "require raw still true at confirmation": "confirmation.require_raw_still_true",
    "break buffer (ticks)": "confirmation.break_buffer_ticks",
    "max swing distance % (0 = off)": "confirmation.max_swing_distance_pct",
    "max pivot age bars (0 = off)": "confirmation.max_pivot_age_bars",
    "use momentum filter": "confirmation.use_momentum",
    "momentum mode": "confirmation.momentum_mode",
    "atr len": "confirmation.atr_len",
    "atr body mult": "confirmation.mom_atr_mult",
    "roc(1) min % (0.15 = 0.15%)": "confirmation.roc_min_pct",
    "use session filter": "confirmation.use_session_filter",
    "session": "confirmation.session",
}


def normalize_space(raw: Any) -> str:
    return re.sub(r"\s+", " ", str(raw or "")).strip()


def normalize_name(raw: Any) -> str:
    return normalize_space(raw).lower()


def slugify(raw: str, max_len: int = 42) -> str:
    txt = normalize_name(raw)
    txt = re.sub(r"[^a-z0-9]+", "_", txt).strip("_")
    if not txt:
        txt = "unnamed"
    return txt[:max_len].strip("_")


def parse_count(raw: Any) -> int:
    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        try:
            return int(raw)
        except Exception:
            return 0
    txt = normalize_name(raw)
    if txt in {"p", "skip", "na", "n/a"}:
        return 0
    try:
        return int(float(txt))
    except Exception:
        return 0


def to_display_value(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "On" if raw else "Off"
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and raw.is_integer():
            return str(int(raw))
        return str(raw).rstrip("0").rstrip(".") if isinstance(raw, float) else str(raw)
    return normalize_space(raw)


def parse_generic(raw: Any) -> Any:
    txt = normalize_space(raw)
    low = txt.lower()
    if low in {"on", "true", "yes", "y"}:
        return True
    if low in {"off", "false", "no", "n"}:
        return False
    if re.fullmatch(r"[-+]?\d+", txt):
        try:
            return int(txt)
        except Exception:
            return txt
    if re.fullmatch(r"[-+]?\d+\.\d+", txt):
        try:
            return float(txt)
        except Exception:
            return txt
    return txt


def coerce_like(raw: Any, current: Any) -> Any:
    if isinstance(current, bool):
        v = parse_generic(raw)
        if isinstance(v, bool):
            return v
        return str(raw).strip().lower() in {"1", "on", "true", "yes", "y"}
    if isinstance(current, int) and not isinstance(current, bool):
        v = parse_generic(raw)
        if isinstance(v, (int, float)):
            return int(v)
        return int(float(str(raw).strip()))
    if isinstance(current, float):
        v = parse_generic(raw)
        if isinstance(v, (int, float)):
            return float(v)
        return float(str(raw).strip())
    if isinstance(current, str):
        return to_display_value(raw)
    return parse_generic(raw)


def deep_get(obj: dict[str, Any], path: str) -> Any:
    cur: Any = obj
    for part in path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return None
        cur = cur[part]
    return cur


def deep_set(obj: dict[str, Any], path: str, value: Any) -> None:
    cur = obj
    parts = path.split(".")
    for part in parts[:-1]:
        nxt = cur.get(part)
        if not isinstance(nxt, dict):
            nxt = {}
            cur[part] = nxt
        cur = nxt
    cur[parts[-1]] = value


def apply_path_with_aliases(
    cfg: dict[str, Any],
    base_config: dict[str, Any],
    path: str,
    raw_value: Any,
) -> list[str]:
    current = deep_get(base_config, path)
    coerced = coerce_like(raw_value, current)
    deep_set(cfg, path, coerced)
    mapped = [path]
    # Range regime uses a single TV master toggle. Keep the legacy internal
    # alias synchronized so regenerated manifests cannot silently disable it.
    if path == "filters.use_range_filters":
        deep_set(cfg, "filters.use_range_regime_filter", bool(coerced))
        mapped.append("filters.use_range_regime_filter")
    return mapped


def stable_hash_config(config: dict[str, Any]) -> str:
    payload = json.dumps(config, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_parent_value(raw_parent_val: Any, parent_row: dict[str, Any]) -> Any:
    txt = normalize_space(raw_parent_val)
    if not txt:
        return parent_row.get("default_value")

    upper = txt.upper()
    if " OR " in upper:
        txt = re.split(r"\s+OR\s+", txt, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    if txt.startswith(">"):
        threshold = txt[1:].strip()
        try:
            tval = float(threshold)
            for v in parent_row.get("values", []):
                pv = parse_generic(v)
                if isinstance(pv, (int, float)) and float(pv) > tval:
                    return pv
            return int(tval) + 1
        except Exception:
            return parent_row.get("default_value")
    return parse_generic(txt)


def load_input_rows(input_map_path: Path) -> tuple[dict[int, dict[str, Any]], int]:
    wb = load_workbook(input_map_path, data_only=True, read_only=True)
    try:
        ws = wb["Properties"]
        by_seq: dict[int, dict[str, Any]] = {}
        target_total = 0
        total_cell = ws.cell(263, 10).value
        if isinstance(total_cell, (int, float)):
            target_total = int(total_cell)
        for row_idx in range(7, ws.max_row + 1):
            seq_raw = ws.cell(row_idx, 1).value
            if not isinstance(seq_raw, (int, float)):
                continue
            seq = int(seq_raw)
            name = normalize_space(ws.cell(row_idx, 2).value)
            default_value = ws.cell(row_idx, 3).value

            p1 = ws.cell(row_idx, 4).value
            p1v = ws.cell(row_idx, 5).value
            p2 = ws.cell(row_idx, 6).value
            p2v = ws.cell(row_idx, 7).value
            p3 = ws.cell(row_idx, 8).value
            p3v = ws.cell(row_idx, 9).value
            count_raw = ws.cell(row_idx, 10).value
            count = parse_count(count_raw)
            values: list[Any] = []
            # K..U (11..21)
            for col in range(11, 22):
                val = ws.cell(row_idx, col).value
                if val is None:
                    continue
                txt = normalize_space(val)
                if txt == "":
                    continue
                values.append(val)

            if count > 0:
                if len(values) == 0 and default_value is not None:
                    values = [default_value]
                if len(values) < count:
                    default_val = default_value
                    while len(values) < count:
                        values.append(default_val)
                if len(values) > count:
                    values = values[:count]
            else:
                values = []

            by_seq[seq] = {
                "seq": seq,
                "name": name,
                "norm_name": normalize_name(name),
                "default_value": default_value,
                "count": count,
                "values": values,
                "parent1": int(p1) if isinstance(p1, (int, float)) else None,
                "parent1_val": p1v,
                "parent2": int(p2) if isinstance(p2, (int, float)) else None,
                "parent2_val": p2v,
                "parent3": int(p3) if isinstance(p3, (int, float)) else None,
                "parent3_val": p3v,
            }
        return by_seq, target_total
    finally:
        wb.close()


def ensure_manifest_header(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_HEADERS)
        writer.writeheader()


def write_manifest(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in MANIFEST_HEADERS})


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Generate case set from FILLED_v6 input map.")
    ap.add_argument("--suite-root", default="mtc_backtest/parity_suite_350", help="Suite root path")
    ap.add_argument(
        "--input-map",
        default="manifests/input_map_FILLED_v6.xlsx",
        help="Input map workbook path (relative to suite root unless absolute)",
    )
    ap.add_argument(
        "--baseline-case-json",
        default="manifests/frozen_baseline_case_parity_core_001_baseline_touch.json",
        help="Baseline JSON template path (relative to suite root unless absolute)",
    )
    ap.add_argument("--cases-dir", default="cases", help="Case json output dir (relative to suite root)")
    ap.add_argument("--manifest-dir", default="manifests", help="Manifest output dir (relative to suite root)")
    ap.add_argument("--purge-cases", action="store_true", help="Delete existing case JSON files before writing")
    return ap.parse_args()


def resolve(base: Path, raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return (base / p).resolve()


def main() -> int:
    args = parse_args()
    suite_root = Path(args.suite_root).resolve()
    input_map_path = resolve(suite_root, args.input_map)
    baseline_path = resolve(suite_root, args.baseline_case_json)
    cases_dir = resolve(suite_root, args.cases_dir)
    manifest_dir = resolve(suite_root, args.manifest_dir)

    if not input_map_path.exists():
        print(f"ERROR: input map not found: {input_map_path}")
        return 2
    if not baseline_path.exists():
        print(f"ERROR: baseline template not found: {baseline_path}")
        return 2

    by_seq, target_total = load_input_rows(input_map_path)
    baseline_obj = json.loads(baseline_path.read_text(encoding="utf-8"))
    base_config = baseline_obj.get("config", {})
    if not isinstance(base_config, dict):
        print("ERROR: baseline config missing/invalid")
        return 2

    if args.purge_cases:
        if cases_dir.exists():
            for fp in cases_dir.glob("*.json"):
                fp.unlink()
    cases_dir.mkdir(parents=True, exist_ok=True)
    manifest_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[dict[str, Any]] = []
    core_rows: list[dict[str, Any]] = []
    bnd_rows: list[dict[str, Any]] = []
    pw_rows: list[dict[str, Any]] = []
    used_case_ids: set[str] = set()

    mapped_case_count = 0
    unmapped_case_count = 0
    unknown_inputs: dict[str, int] = {}
    run_order = 0

    seqs = sorted(by_seq.keys())
    for seq in seqs:
        row = by_seq[seq]
        count = int(row.get("count", 0) or 0)
        if count <= 0:
            continue

        input_name = str(row.get("name", "")).strip()
        norm_name = row.get("norm_name", "")
        values = list(row.get("values", []))
        if not values:
            continue

        parent_specs: list[tuple[int, Any]] = []
        for idx in (1, 2, 3):
            pseq = row.get(f"parent{idx}")
            pval = row.get(f"parent{idx}_val")
            if isinstance(pseq, int):
                parent_specs.append((pseq, pval))

        for value_idx, target_value in enumerate(values, start=1):
            run_order += 1
            pack = "core" if value_idx == 1 else "boundary"
            prefix = "parity_core" if pack == "core" else "parity_bnd"
            base_slug = slugify(input_name)
            raw_case_id = f"{prefix}_{seq:03d}_{base_slug}_v{value_idx:02d}"
            case_id = raw_case_id
            dup = 2
            while case_id in used_case_ids:
                case_id = f"{raw_case_id}_{dup:02d}"
                dup += 1
            used_case_ids.add(case_id)

            case_obj = copy.deepcopy(baseline_obj)
            cfg = case_obj.setdefault("config", {})
            if not isinstance(cfg, dict):
                cfg = {}
                case_obj["config"] = cfg

            mapped_paths: list[str] = []
            parent_actions: list[str] = []
            depends_on: list[str] = []
            parent_required: list[str] = []

            for pseq, pval in parent_specs:
                prow = by_seq.get(pseq)
                if not prow:
                    continue
                pname = str(prow.get("name", "")).strip()
                parsed_parent_value = parse_parent_value(pval, prow)
                parent_actions.append(f"{pname}={to_display_value(parsed_parent_value)}")
                depends_on.append(str(pseq))
                parent_required.append(f"{pseq}={to_display_value(parsed_parent_value)}")

                pnorm = normalize_name(pname)
                ppath = UI_TO_CONFIG_PATH.get(pnorm)
                if ppath:
                    mapped_paths.extend(apply_path_with_aliases(cfg, base_config, ppath, parsed_parent_value))

            target_path = UI_TO_CONFIG_PATH.get(norm_name)
            if target_path:
                mapped_paths.extend(apply_path_with_aliases(cfg, base_config, target_path, target_value))
                mapped_case_count += 1
            else:
                unknown_inputs[norm_name] = unknown_inputs.get(norm_name, 0) + 1
                unmapped_case_count += 1

            parity_cfg = cfg.setdefault("parity", {})
            if isinstance(parity_cfg, dict):
                parity_cfg["debug_dir"] = f"debug/parity_suite_350/{case_id}"

            expected_trade_behavior = "NORMAL"
            if norm_name == "signal mode" and normalize_name(target_value) == "none":
                expected_trade_behavior = "ZERO_TRADE_EXPECTED"

            action_parts = parent_actions + [f"{input_name}={to_display_value(target_value)}"]
            ui_actions = "Set in TV UI: " + "; ".join(action_parts)
            primary_change = f"{input_name}={to_display_value(target_value)}"
            config_hash = stable_hash_config(cfg)

            case_obj["_generated"] = {
                "suite": "parity_suite_350",
                "generator": "generate_case_set_from_input_map.py",
                "source_input_map": str(input_map_path),
                "source_seq": seq,
                "source_input_name": input_name,
                "source_value_index": value_idx,
                "source_target_value": to_display_value(target_value),
                "mapped_paths": sorted(set(mapped_paths)),
                "mapped_target_path": target_path or "",
            }
            case_obj["_tv_case"] = {
                "seq": seq,
                "input_name": input_name,
                "target_value": to_display_value(target_value),
                "parents": [
                    {
                        "seq": pseq,
                        "name": str(by_seq.get(pseq, {}).get("name", "")),
                        "value": to_display_value(parse_parent_value(pval, by_seq.get(pseq, {}))),
                    }
                    for pseq, pval in parent_specs
                ],
            }

            case_path = cases_dir / f"{case_id}.json"
            case_path.write_text(json.dumps(case_obj, ensure_ascii=True, indent=2), encoding="utf-8")

            manifest_row: dict[str, Any] = {
                "run_order": run_order,
                "pack": pack,
                "case_id": case_id,
                "case_json": f"cases/{case_id}.json",
                "tv_preset_name": case_id,
                "enabled": 1,
                "expected_trade_behavior": expected_trade_behavior,
                "primary_change": primary_change,
                "ui_actions": ui_actions,
                "depends_on": ",".join(depends_on),
                "parent_required": "; ".join(parent_required),
                "canonical_config_hash": config_hash,
                "semantic_fingerprint": f"{pack}:{seq}:{value_idx}:{config_hash[:16]}",
                "symbol": "BINANCE:BTCUSDT.P",
                "timeframe": "15",
                "start_date": case_obj.get("start_date", "2025-06-30T05:15:00"),
                "end_date": case_obj.get("end_date", "2026-02-01"),
                "notes": (
                    f"source_seq={seq}; source_value_idx={value_idx}; "
                    f"mapped_target={'yes' if bool(target_path) else 'no'}"
                ),
            }
            all_rows.append(manifest_row)
            if pack == "core":
                core_rows.append(manifest_row)
            elif pack == "boundary":
                bnd_rows.append(manifest_row)
            else:
                pw_rows.append(manifest_row)

    all_manifest = manifest_dir / "cases_manifest_all.csv"
    core_manifest = manifest_dir / "cases_manifest_core.csv"
    bnd_manifest = manifest_dir / "cases_manifest_boundary.csv"
    pw_manifest = manifest_dir / "cases_manifest_pairwise.csv"

    write_manifest(all_manifest, all_rows)
    write_manifest(core_manifest, core_rows)
    write_manifest(bnd_manifest, bnd_rows)
    write_manifest(pw_manifest, pw_rows)

    report = {
        "source_input_map": str(input_map_path),
        "target_total_from_workbook": target_total,
        "generated_total_cases": len(all_rows),
        "generated_core_cases": len(core_rows),
        "generated_boundary_cases": len(bnd_rows),
        "generated_pairwise_cases": len(pw_rows),
        "mapped_case_count": mapped_case_count,
        "unmapped_case_count": unmapped_case_count,
        "unmapped_unique_inputs": sorted(unknown_inputs.items(), key=lambda kv: (-kv[1], kv[0])),
        "manifests": {
            "all": str(all_manifest),
            "core": str(core_manifest),
            "boundary": str(bnd_manifest),
            "pairwise": str(pw_manifest),
        },
    }
    report_path = manifest_dir / "input_map_generation_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=True, indent=2), encoding="utf-8")

    print(f"source_target_total={target_total}")
    print(f"generated_total={len(all_rows)}")
    print(f"core={len(core_rows)}")
    print(f"boundary={len(bnd_rows)}")
    print(f"pairwise={len(pw_rows)}")
    print(f"mapped_cases={mapped_case_count}")
    print(f"unmapped_cases={unmapped_case_count}")
    print(f"all_manifest={all_manifest}")
    print(f"report={report_path}")
    print("status=ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
