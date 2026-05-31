#!/usr/bin/env python
"""
Route TradingView XLSX exports from raw_tv_exports into case folders.

Matching priority:
1) case_id extracted from filename (regex from case_rules.json)
2) case_id extracted from Properties sheet configured keys
3) tv_preset_name extracted from Properties sheet

Outputs:
- manifests/tv_collection_status.csv
- manifests/tv_unmatched.csv

Optional:
- update CASE_SETUP_GUIDE.xlsx tracker fields after routing
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MANIFEST_REQUIRED_HEADERS = [
    "run_order",
    "pack",
    "case_id",
    "tv_preset_name",
    "enabled",
]


DEFAULT_STATUS_HEADERS = [
    "processed_at_utc",
    "source_file",
    "action",
    "match_result",
    "match_method",
    "case_id",
    "run_order",
    "pack",
    "destination_file",
    "props_case_id",
    "props_preset",
    "error",
]


DEFAULT_UNMATCHED_HEADERS = [
    "processed_at_utc",
    "source_file",
    "suggested_case_id",
    "props_case_id",
    "props_preset",
    "reason",
]


PROPERTY_NAME_ALIASES = {
    "use volume filter": "use volume participation filter",
    "use atr volatility filter": "use atr volatility floor",
    "exit end of day": "exit at end of day",
    "exit end of week": "exit at end of week",
    "use break even": "use break-even?",
    "use break-even": "use break-even?",
}


APPLY_ONLY_ACTION_OVERRIDES: dict[str, dict[str, str]] = {
    "exit filter block global on": {
        "Exit if Selected Filters Block While In Position": "ON",
    }
}


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def normalize_key(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).strip().lower()


def normalize_spaces(raw: str) -> str:
    return re.sub(r"\s+", " ", raw).strip()


def canonical_property_name(raw: Any) -> str:
    key = normalize_spaces(str(raw or "")).lower()
    if not key:
        return ""
    return PROPERTY_NAME_ALIASES.get(key, key)


def normalize_value(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "ON" if raw else "OFF"
    if isinstance(raw, (int, float)):
        # Keep deterministic numeric text.
        if isinstance(raw, float) and raw.is_integer():
            return str(int(raw))
        return f"{raw}".rstrip("0").rstrip(".") if isinstance(raw, float) else str(raw)

    text = normalize_spaces(str(raw))
    if not text:
        return ""

    low = text.lower()
    if low in {"on", "true", "yes", "y", "1"}:
        return "ON"
    if low in {"off", "false", "no", "n", "0"}:
        return "OFF"

    num = text.replace(",", ".")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", num):
        fval = float(num)
        if fval.is_integer():
            return str(int(fval))
        return f"{fval}".rstrip("0").rstrip(".")

    return text.upper()


def parse_int(raw: Any, default: int = 0) -> int:
    try:
        return int(str(raw).strip())
    except Exception:
        return default


def to_bool(raw: Any, default: bool = True) -> bool:
    if raw is None:
        return default
    val = str(raw).strip().lower()
    if val in {"1", "true", "yes", "y"}:
        return True
    if val in {"0", "false", "no", "n"}:
        return False
    return default


def next_available_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for i in range(1, 1000):
        candidate = path.with_name(f"{stem}_dup{i:02d}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Cannot allocate destination filename for: {path}")


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_manifest(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MANIFEST_REQUIRED_HEADERS)
        writer.writeheader()


def load_manifest(path: Path) -> tuple[list[dict[str, str]], dict[str, list[dict[str, str]]], dict[str, list[dict[str, str]]]]:
    ensure_manifest(path)
    with path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    for row in rows:
        for header in MANIFEST_REQUIRED_HEADERS:
            row.setdefault(header, "")
        row.setdefault("case_folder", "")
        row.setdefault("primary_change", "")
        row.setdefault("ui_actions", "")
        row.setdefault("notes", "")

    enabled_rows = [r for r in rows if to_bool(r.get("enabled", "1"), True)]
    enabled_rows.sort(key=lambda r: parse_int(r.get("run_order", 0), 0))

    by_case_id: dict[str, list[dict[str, str]]] = {}
    by_preset: dict[str, list[dict[str, str]]] = {}
    for row in enabled_rows:
        case_id = row.get("case_id", "").strip()
        preset = row.get("tv_preset_name", "").strip()
        if case_id:
            by_case_id.setdefault(case_id.lower(), []).append(row)
        if preset:
            by_preset.setdefault(preset.lower(), []).append(row)
    return enabled_rows, by_case_id, by_preset


def pick_single(rows: list[dict[str, str]]) -> dict[str, str] | None:
    if not rows:
        return None
    if len(rows) == 1:
        return rows[0]
    ordered = sorted(rows, key=lambda r: parse_int(r.get("run_order", 0), 10**9))
    return ordered[0]


def extract_case_id_from_text(raw: Any, case_id_regex: re.Pattern[str]) -> str:
    if raw is None:
        return ""
    match = case_id_regex.search(str(raw))
    if not match:
        return ""
    return match.group(1).strip().lower()


def resolve_case_folder(row: dict[str, str]) -> str:
    raw = row.get("case_folder", "").strip()
    if raw:
        return raw
    run_order = parse_int(row.get("run_order", 0), 0)
    case_id = row.get("case_id", "").strip()
    if run_order > 0 and case_id:
        return f"{run_order:03d}_{case_id}"
    return case_id


def load_baseline_reuse_map(
    manifest_rows: list[dict[str, str]],
) -> tuple[dict[str, list[dict[str, str]]], dict[str, dict[str, str]]]:
    by_source: dict[str, list[dict[str, str]]] = {}
    by_case_id: dict[str, dict[str, str]] = {}
    for row in manifest_rows:
        cid = str(row.get("case_id", "")).strip().lower()
        if cid:
            by_case_id[cid] = row
        action = str(row.get("ui_actions", "")).strip()
        m = re.match(r"^Reuse baseline XLSX from\s+([a-zA-Z0-9_]+)\s*\(same UI as baseline\)\.?$", action, flags=re.IGNORECASE)
        if not m:
            continue
        src = m.group(1).strip().lower()
        if src:
            by_source.setdefault(src, []).append(row)
    return by_source, by_case_id


def find_latest_strategy_report_xlsx(case_dir: Path) -> Path | None:
    if not case_dir.exists():
        return None
    files = [p for p in case_dir.glob("*_strategy_report.xlsx") if p.is_file()]
    if not files:
        return None
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0]


def read_properties(xlsx_path: Path) -> dict[str, Any]:
    props: dict[str, Any] = {}
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        prop_sheet_name = None
        for name in wb.sheetnames:
            if normalize_key(name) == "properties":
                prop_sheet_name = name
                break
        if prop_sheet_name is None:
            return props

        ws = wb[prop_sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return props

        for row in rows:
            vals = list(row)
            key = None
            value = None

            # TV export "Properties" format: [seq?, name, value, ...]
            if len(vals) >= 3 and vals[1] is not None:
                key = vals[1]
                value = vals[2]
            # Generic fallback: [name, value]
            elif len(vals) >= 2 and vals[0] is not None:
                key = vals[0]
                value = vals[1]

            if key is None:
                continue
            k = str(key).strip()
            if not k:
                continue
            if normalize_key(k) == "name":
                continue
            props[k] = value
        return props
    finally:
        wb.close()


def props_lookup(props: dict[str, Any], aliases: list[str]) -> list[str]:
    if not props:
        return []
    norm = {normalize_key(k): v for k, v in props.items()}
    values: list[str] = []
    for alias in aliases:
        key = normalize_key(alias)
        if key in norm and norm[key] is not None:
            raw = str(norm[key]).strip()
            if raw:
                values.append(raw)
    return values


def parse_ui_actions_to_expectations(action_text: str) -> dict[str, str]:
    text = (action_text or "").strip()
    if not text:
        return {}
    if text.startswith("Baseline setup:"):
        return {}
    if text.startswith("Apply only this case change in TV UI:"):
        token = text.split(":", 1)[1].strip().lower()
        raw = APPLY_ONLY_ACTION_OVERRIDES.get(token, {})
        return {canonical_property_name(k): normalize_value(v) for k, v in raw.items()}
    if not text.startswith("Set in TV UI:"):
        return {}

    body = text.split("Set in TV UI:", 1)[1].strip()
    pairs = [p.strip() for p in body.split(";") if p.strip()]
    out: dict[str, str] = {}
    for part in pairs:
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        ckey = canonical_property_name(key)
        cval = normalize_value(value)
        if ckey:
            out[ckey] = cval
    return out


def load_case_expectations_from_guide(guide_path: Path) -> dict[str, dict[str, Any]]:
    if not guide_path.exists():
        return {}
    try:
        wb = load_workbook(guide_path, data_only=True, read_only=True)
    except Exception:
        return {}
    try:
        if "Cases" not in wb.sheetnames:
            return {}
        ws = wb["Cases"]
        if ws.max_row < 2:
            return {}
        header_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            header_map[normalize_key(ws.cell(row=1, column=c).value)] = c
        cid_col = header_map.get("case_id")
        action_col = header_map.get("tv_ui_actions")
        run_col = header_map.get("run_order")
        if not cid_col or not action_col:
            return {}

        out: dict[str, dict[str, Any]] = {}
        baseline_case_ids: list[str] = []
        for r in range(2, ws.max_row + 1):
            cid_val = ws.cell(row=r, column=cid_col).value
            cid = str(cid_val).strip() if cid_val is not None else ""
            if not cid:
                continue
            action_val = ws.cell(row=r, column=action_col).value
            action_text = str(action_val).strip() if action_val is not None else ""
            run_order = parse_int(ws.cell(row=r, column=run_col).value if run_col else 0, 0)
            if action_text.startswith("Baseline setup:"):
                baseline_case_ids.append(cid.lower())
            out[cid.lower()] = {
                "case_id": cid,
                "run_order": run_order,
                "action_text": action_text,
                "expected": parse_ui_actions_to_expectations(action_text),
            }

        # Baseline case has no explicit assignments in guide text.
        # Build strict baseline expectation from baseline freeze XLSX on keys used by other cases.
        if baseline_case_ids:
            union_keys: set[str] = set()
            for meta in out.values():
                union_keys.update(meta.get("expected", {}).keys())
            if union_keys:
                baseline_record_path = guide_path.parent / "manifests" / "baseline_freeze_record.json"
                if baseline_record_path.exists():
                    try:
                        baseline_record = json.loads(baseline_record_path.read_text(encoding="utf-8"))
                        baseline_xlsx = Path(str(baseline_record.get("baseline_xlsx", "")))
                        if baseline_xlsx.exists():
                            bprops_raw = read_properties(baseline_xlsx)
                            bprops = {
                                canonical_property_name(k): normalize_value(v)
                                for k, v in bprops_raw.items()
                                if canonical_property_name(k)
                            }
                            baseline_expected = {
                                k: bprops[k] for k in sorted(union_keys) if k in bprops
                            }
                            for bcid in baseline_case_ids:
                                if bcid in out:
                                    out[bcid]["expected"] = baseline_expected
                    except Exception:
                        pass
        return out
    finally:
        wb.close()


def load_pending_queue_from_guide(guide_path: Path) -> list[str]:
    if not guide_path.exists():
        return []
    try:
        wb = load_workbook(guide_path, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        if "Cases" not in wb.sheetnames:
            return []
        ws = wb["Cases"]
        if ws.max_row < 2:
            return []
        header_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            header_map[normalize_key(ws.cell(row=1, column=c).value)] = c

        cid_col = header_map.get("case_id")
        run_col = header_map.get("run_order")
        status_col = header_map.get("tv_xlsx_status")
        if not cid_col:
            return []

        rows: list[tuple[int, str]] = []
        for r in range(2, ws.max_row + 1):
            cid_val = ws.cell(row=r, column=cid_col).value
            cid = str(cid_val).strip() if cid_val is not None else ""
            if not cid:
                continue
            status = str(ws.cell(row=r, column=status_col).value).strip().upper() if status_col else ""
            if status in {"DONE", "ROUTED", "SKIP"}:
                continue
            run_order = parse_int(ws.cell(row=r, column=run_col).value if run_col else 0, 0)
            rows.append((run_order, cid.lower()))

        rows.sort(key=lambda x: x[0] if x[0] > 0 else 10**9)
        return [cid for _, cid in rows]
    finally:
        wb.close()


def pop_next_pending_match(
    pending_queue: list[str],
    assigned_case_ids: set[str],
    by_case_id: dict[str, list[dict[str, str]]],
) -> dict[str, str] | None:
    while pending_queue:
        cid = pending_queue.pop(0)
        if cid in assigned_case_ids:
            continue
        row = pick_single(by_case_id.get(cid, []))
        if row is None:
            continue
        assigned_case_ids.add(cid)
        return row
    return None


def match_case_by_expected_properties(
    props: dict[str, Any],
    by_case_id: dict[str, list[dict[str, str]]],
    expectations: dict[str, dict[str, Any]],
) -> tuple[dict[str, str] | None, str]:
    if not props or not expectations:
        return None, "NO_EXPECTATIONS"

    normalized_props: dict[str, str] = {}
    for key, value in props.items():
        ckey = canonical_property_name(key)
        if not ckey:
            continue
        normalized_props[ckey] = normalize_value(value)

    candidates: list[tuple[int, int, dict[str, str]]] = []
    for cid_lower, meta in expectations.items():
        expected: dict[str, str] = meta.get("expected", {})
        if not expected:
            continue
        ok = True
        for ek, ev in expected.items():
            actual = normalized_props.get(ek, "")
            if actual != ev:
                ok = False
                break
        if not ok:
            continue
        row = pick_single(by_case_id.get(cid_lower, []))
        if row is None:
            continue
        score = len(expected)
        run_order = parse_int(row.get("run_order", 0), 0)
        candidates.append((score, -run_order, row))

    if not candidates:
        return None, "NO_PROPERTY_MATCH"

    candidates.sort(reverse=True)
    best_score = candidates[0][0]
    best_rows = [row for score, _, row in candidates if score == best_score]
    if len(best_rows) == 1:
        return best_rows[0], "OK"
    return None, f"AMBIGUOUS_TOP_SCORE_{best_score}"


def match_case(
    source_file: Path,
    props: dict[str, Any],
    by_case_id: dict[str, list[dict[str, str]]],
    by_preset: dict[str, list[dict[str, str]]],
    expectations: dict[str, dict[str, Any]],
    case_id_regex: re.Pattern[str],
    property_case_id_keys: list[str],
    property_strategy_name_keys: list[str],
) -> tuple[dict[str, str] | None, str, str, str, str]:
    # 1) filename case_id
    file_case_id = extract_case_id_from_text(source_file.name, case_id_regex)
    if file_case_id:
        row = pick_single(by_case_id.get(file_case_id, []))
        if row:
            return row, "filename_case_id", file_case_id, "", ""

    # 2) properties case_id
    prop_case_id_candidates: list[str] = []
    for value in props_lookup(props, property_case_id_keys):
        case_id = extract_case_id_from_text(value, case_id_regex)
        if case_id:
            prop_case_id_candidates.append(case_id)
    for cid in prop_case_id_candidates:
        row = pick_single(by_case_id.get(cid, []))
        if row:
            return row, "properties_case_id", cid, "", ""

    # 3) properties preset
    prop_preset_candidates = props_lookup(props, property_strategy_name_keys + property_case_id_keys)
    for preset in prop_preset_candidates:
        row = pick_single(by_preset.get(preset.lower(), []))
        if row:
            return row, "properties_preset", "", preset, ""

    # 4) compare UI assignments from CASE_SETUP_GUIDE against Properties values.
    row, prop_reason = match_case_by_expected_properties(props, by_case_id, expectations)
    if row is not None:
        return row, "properties_ui_actions", "", "", ""

    candidate = file_case_id or (prop_case_id_candidates[0] if prop_case_id_candidates else "")
    preset = prop_preset_candidates[0] if prop_preset_candidates else ""
    return None, "", candidate, preset, prop_reason


def move_or_copy(src: Path, dst: Path, copy_mode: bool, dry_run: bool) -> str:
    dst.parent.mkdir(parents=True, exist_ok=True)
    safe_dst = next_available_path(dst)
    if dry_run:
        return str(safe_dst)
    if copy_mode:
        shutil.copy2(src, safe_dst)
    else:
        shutil.move(str(src), str(safe_dst))
    return str(safe_dst)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in headers})


def prepare_case_folders(tv_root: Path, manifest_rows: list[dict[str, str]], dry_run: bool) -> int:
    prepared = 0
    for row in manifest_rows:
        folder = resolve_case_folder(row)
        if not folder:
            continue
        path = tv_root / folder
        if path.exists():
            continue
        if not dry_run:
            path.mkdir(parents=True, exist_ok=True)
        prepared += 1
    return prepared


def update_case_setup_guide(guide_path: Path, routed_rows: list[dict[str, Any]], dry_run: bool) -> bool:
    if not guide_path.exists():
        return False
    try:
        wb = load_workbook(guide_path)
    except Exception:
        return False
    if "Cases" not in wb.sheetnames:
        return False

    ws = wb["Cases"]
    if ws.max_row < 2:
        return False
    header_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        header_map[normalize_key(ws.cell(row=1, column=c).value)] = c

    case_col = header_map.get("case_id")
    status_col = header_map.get("tv_xlsx_status")
    file_col = header_map.get("tv_xlsx_file")
    date_col = header_map.get("tv_download_date")
    if case_col is None:
        return False

    case_row_map: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        cid_val = ws.cell(row=r, column=case_col).value
        cid = str(cid_val).strip().lower() if cid_val is not None else ""
        if cid:
            case_row_map[cid] = r

    touched = False
    now_local = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row in routed_rows:
        if row.get("match_result") not in {"MATCHED", "REUSED"}:
            continue
        cid = str(row.get("case_id", "")).strip().lower()
        if not cid or cid not in case_row_map:
            continue
        r = case_row_map[cid]
        if status_col:
            ws.cell(row=r, column=status_col, value="ROUTED" if not dry_run else "DOWNLOADED")
        if file_col:
            destination = str(row.get("destination_file", "")).strip()
            ws.cell(row=r, column=file_col, value=Path(destination).name if destination else "")
        if date_col:
            ws.cell(row=r, column=date_col, value=now_local)
        touched = True

    if touched and not dry_run:
        try:
            wb.save(guide_path)
        except PermissionError:
            # Keep routing successful even if guide is open/locked in Excel.
            return False
    return touched


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Route TV XLSX exports into parity suite case folders.")
    ap.add_argument("--suite-root", default="mtc_backtest/parity_suite_350", help="Suite root folder")
    ap.add_argument("--rules", default="scripts/case_rules.json", help="Rules JSON path")
    ap.add_argument("--manifest", default="manifests/cases_manifest_all.csv", help="Case manifest CSV path")
    ap.add_argument("--raw-dir", default="tv_manual_inputs/raw_tv_exports", help="Raw XLSX input dir")
    ap.add_argument("--tv-root", default="tv_manual_inputs", help="TV manual input root")
    ap.add_argument("--status-csv", default="manifests/tv_collection_status.csv", help="Route status CSV output")
    ap.add_argument("--unmatched-csv", default="manifests/tv_unmatched.csv", help="Unmatched CSV output")
    ap.add_argument("--update-guide", default="CASE_SETUP_GUIDE.xlsx", help="Guide workbook path to update")
    ap.add_argument(
        "--sequential-fallback",
        action="store_true",
        help="If no strong match, route to next pending case in CASE_SETUP_GUIDE run_order.",
    )
    ap.add_argument("--copy", action="store_true", help="Copy files instead of move")
    ap.add_argument("--dry-run", action="store_true", help="Analyze only, do not move/copy")
    return ap.parse_args()


def resolve(base: Path, raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return (base / p).resolve()


def main() -> int:
    args = parse_args()
    suite_root = Path(args.suite_root).resolve()
    rules_path = resolve(suite_root, args.rules)
    manifest_path = resolve(suite_root, args.manifest)
    raw_dir = resolve(suite_root, args.raw_dir)
    tv_root = resolve(suite_root, args.tv_root)
    status_csv = resolve(suite_root, args.status_csv)
    unmatched_csv = resolve(suite_root, args.unmatched_csv)
    guide_path = resolve(suite_root, args.update_guide)

    if not rules_path.exists():
        print(f"ERROR: rules not found: {rules_path}")
        return 2

    tv_root.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)
    (tv_root / "unmatched").mkdir(parents=True, exist_ok=True)

    rules = load_json(rules_path)
    routing = rules.get("routing", {})
    case_id_pattern = routing.get("case_id_regex", r"(parity_[a-z0-9_]+)")
    case_id_regex = re.compile(case_id_pattern, flags=re.IGNORECASE)
    property_case_id_keys = routing.get("property_case_id_keys", [])
    property_strategy_name_keys = routing.get("property_strategy_name_keys", [])

    manifest_rows, by_case_id, by_preset = load_manifest(manifest_path)
    reuse_by_source, manifest_by_case_id = load_baseline_reuse_map(manifest_rows)
    prepared_count = prepare_case_folders(tv_root, manifest_rows, args.dry_run)

    source_files = sorted(
        [
            p for p in raw_dir.glob("*.xlsx")
            if p.is_file() and not p.name.startswith("~$")
        ],
        key=lambda p: p.stat().st_mtime,
    )

    if not source_files:
        write_csv(status_csv, DEFAULT_STATUS_HEADERS, [])
        write_csv(unmatched_csv, DEFAULT_UNMATCHED_HEADERS, [])
        print(f"prepared_case_folders={prepared_count}")
        print("processed_files=0")
        print("matched=0")
        print("reused=0")
        print("unmatched=0")
        print("errors=0")
        print(f"status_csv={status_csv}")
        print(f"unmatched_csv={unmatched_csv}")
        print("guide_updated=0")
        return 0

    expectations = load_case_expectations_from_guide(guide_path)
    pending_queue = load_pending_queue_from_guide(guide_path) if args.sequential_fallback else []
    assigned_case_ids: set[str] = set()
    reuse_applied_sources: set[str] = set()

    status_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []

    for src in source_files:
        timestamp = utc_now()
        action = "COPY" if args.copy else "MOVE"
        if args.dry_run:
            action = f"DRY_RUN_{action}"
        record = {
            "processed_at_utc": timestamp,
            "source_file": str(src),
            "action": action,
            "match_result": "",
            "match_method": "",
            "case_id": "",
            "run_order": "",
            "pack": "",
            "destination_file": "",
            "props_case_id": "",
            "props_preset": "",
            "error": "",
        }

        try:
            props = read_properties(src)
            row, match_method, props_case_id, props_preset, prop_reason = match_case(
                source_file=src,
                props=props,
                by_case_id=by_case_id,
                by_preset=by_preset,
                expectations=expectations,
                case_id_regex=case_id_regex,
                property_case_id_keys=property_case_id_keys,
                property_strategy_name_keys=property_strategy_name_keys,
            )
            record["props_case_id"] = props_case_id
            record["props_preset"] = props_preset

            if row is not None:
                cid_candidate = str(row.get("case_id", "")).strip().lower()
                if cid_candidate and cid_candidate in assigned_case_ids and args.sequential_fallback:
                    fallback_row = pop_next_pending_match(pending_queue, assigned_case_ids, by_case_id)
                    if fallback_row is not None:
                        row = fallback_row
                        match_method = "sequential_fallback_duplicate_case"
                        prop_reason = f"fallback_from=duplicate_case:{cid_candidate}"

            if row is None:
                if args.sequential_fallback:
                    fallback_row = pop_next_pending_match(pending_queue, assigned_case_ids, by_case_id)
                    if fallback_row is not None:
                        row = fallback_row
                        match_method = "sequential_fallback"
                        prop_reason = f"fallback_from={prop_reason or 'NO_MATCH'}"

            if row is None:
                unmatched_dst = tv_root / "unmatched" / src.name
                out_path = move_or_copy(src, unmatched_dst, args.copy, args.dry_run)
                record["match_result"] = "UNMATCHED"
                record["destination_file"] = out_path
                record["error"] = "No strong match found." + (f" ({prop_reason})" if prop_reason else "")
                unmatched_rows.append(
                    {
                        "processed_at_utc": timestamp,
                        "source_file": str(src),
                        "suggested_case_id": props_case_id,
                        "props_case_id": props_case_id,
                        "props_preset": props_preset,
                        "reason": "NO_MATCH",
                    }
                )
            else:
                case_id = row.get("case_id", "").strip()
                run_order = parse_int(row.get("run_order", 0), 0)
                pack = row.get("pack", "").strip()
                case_folder = resolve_case_folder(row)
                case_dir = tv_root / case_folder
                canonical_name = f"{case_folder}_strategy_report.xlsx"
                destination = case_dir / canonical_name
                out_path = move_or_copy(src, destination, args.copy, args.dry_run)

                record["match_result"] = "MATCHED"
                record["match_method"] = match_method
                record["case_id"] = case_id
                record["run_order"] = str(run_order)
                record["pack"] = pack
                record["destination_file"] = out_path
                if prop_reason:
                    record["error"] = prop_reason

                if case_id:
                    assigned_case_ids.add(case_id.lower())

                # Auto-propagate baseline-reuse cases if this matched case is reuse source.
                src_cid_lower = case_id.strip().lower()
                if src_cid_lower in reuse_by_source and src_cid_lower not in reuse_applied_sources:
                    reuse_applied_sources.add(src_cid_lower)
                    src_dir = tv_root / resolve_case_folder(manifest_by_case_id.get(src_cid_lower, row))
                    src_xlsx = find_latest_strategy_report_xlsx(src_dir)
                    if src_xlsx is not None:
                        for reuse_row in reuse_by_source[src_cid_lower]:
                            reuse_case_id = reuse_row.get("case_id", "").strip()
                            if not reuse_case_id:
                                continue
                            reuse_case_lower = reuse_case_id.lower()
                            if reuse_case_lower in assigned_case_ids:
                                continue
                            reuse_folder = resolve_case_folder(reuse_row)
                            reuse_dir = tv_root / reuse_folder
                            reuse_filename = f"{reuse_folder}_strategy_report.xlsx"
                            reuse_dst = reuse_dir / reuse_filename
                            reuse_out = move_or_copy(src_xlsx, reuse_dst, copy_mode=True, dry_run=args.dry_run)
                            assigned_case_ids.add(reuse_case_lower)
                            status_rows.append(
                                {
                                    "processed_at_utc": timestamp,
                                    "source_file": str(src_xlsx),
                                    "action": "DRY_RUN_COPY" if args.dry_run else "COPY",
                                    "match_result": "REUSED",
                                    "match_method": "baseline_reuse_auto_copy",
                                    "case_id": reuse_case_id,
                                    "run_order": str(parse_int(reuse_row.get("run_order", 0), 0)),
                                    "pack": reuse_row.get("pack", "").strip(),
                                    "destination_file": reuse_out,
                                    "props_case_id": "",
                                    "props_preset": "",
                                    "error": f"reuse_from={case_id}",
                                }
                            )
        except Exception as exc:
            record["match_result"] = "ERROR"
            record["error"] = str(exc)
        status_rows.append(record)

    write_csv(status_csv, DEFAULT_STATUS_HEADERS, status_rows)
    write_csv(unmatched_csv, DEFAULT_UNMATCHED_HEADERS, unmatched_rows)
    guide_updated = update_case_setup_guide(guide_path, status_rows, args.dry_run)

    matched = sum(1 for r in status_rows if r["match_result"] == "MATCHED")
    reused = sum(1 for r in status_rows if r["match_result"] == "REUSED")
    unmatched = sum(1 for r in status_rows if r["match_result"] == "UNMATCHED")
    errors = sum(1 for r in status_rows if r["match_result"] == "ERROR")
    print(f"prepared_case_folders={prepared_count}")
    print(f"processed_files={len(status_rows)}")
    print(f"matched={matched}")
    print(f"reused={reused}")
    print(f"unmatched={unmatched}")
    print(f"errors={errors}")
    print(f"status_csv={status_csv}")
    print(f"unmatched_csv={unmatched_csv}")
    print(f"guide_updated={int(guide_updated)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
