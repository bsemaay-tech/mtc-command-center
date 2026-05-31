#!/usr/bin/env python3
"""Full reuse analysis: check ALL non-downloaded cases for potential reuse."""

import json, sys, io
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

CASES_DIR = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350\cases")
XLSX_PATH = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350\CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx")

# --- Load XLSX ---
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["Cases"]
headers = {}
for cell in ws[1]:
    if cell.value:
        headers[cell.value] = cell.column

all_cases = []
for row in ws.iter_rows(min_row=2, values_only=False):
    cid = row[headers["case_id"] - 1].value
    if not cid:
        continue
    all_cases.append({
        "run_order": row[headers["run_order"] - 1].value,
        "case_id": cid,
        "tv_xlsx_status": row[headers["tv_xlsx_status"] - 1].value or "",
        "setup_check": row[headers["setup_check"] - 1].value or "",
        "compare_status": row[headers["compare_status"] - 1].value or "",
    })
wb.close()


def get_case_info(case_id):
    p = CASES_DIR / f"{case_id}.json"
    if not p.exists():
        return None
    with open(p) as f:
        data = json.load(f)
    gen = data.get("_generated", {})
    tv = data.get("_tv_case", {})
    parents = tv.get("parents", [])
    pname = gen.get("source_input_name", "")
    pvalue = gen.get("source_target_value", "")
    settings = {p["name"]: str(p["value"]) for p in parents}
    if pname:
        settings[pname] = str(pvalue)
    parent_only = {p["name"]: str(p["value"]) for p in parents}
    return {
        "primary_name": pname,
        "primary_value": str(pvalue),
        "full_settings": settings,
        "parent_settings": parent_only,
    }


# Build cache
cache = {}
for c in all_cases:
    info = get_case_info(c["case_id"])
    if info:
        cache[c["case_id"]] = info

# Build downloaded index: settings_key -> case
downloaded_cases = [c for c in all_cases if c["tv_xlsx_status"] == "DOWNLOADED"]
downloaded_index = {}
for c in downloaded_cases:
    if c["case_id"] in cache:
        key = tuple(sorted(cache[c["case_id"]]["full_settings"].items()))
        if key not in downloaded_index:
            downloaded_index[key] = c

# Baseline
baseline_case = [c for c in all_cases if c["tv_xlsx_status"] == "baseline"]
baseline_key = None
if baseline_case and baseline_case[0]["case_id"] in cache:
    baseline_key = tuple(sorted(cache[baseline_case[0]["case_id"]]["full_settings"].items()))

# Analyze every non-downloaded, non-baseline, non-skip case
results = []
for c in all_cases:
    status = c["tv_xlsx_status"]
    if status in ("DOWNLOADED", "baseline", "SKIP"):
        continue

    cid = c["case_id"]
    if cid not in cache:
        results.append({"case": c, "match": None, "match_type": None, "primary": "?"})
        continue

    info = cache[cid]
    full_key = tuple(sorted(info["full_settings"].items()))
    parent_key = tuple(sorted(info["parent_settings"].items()))

    match = None
    match_type = None

    # 1. Exact match with downloaded
    if full_key in downloaded_index:
        m = downloaded_index[full_key]
        if m["case_id"] != cid:
            match, match_type = m, "exact"

    # 2. Parent-only matches a downloaded (primary = default)
    if not match and parent_key in downloaded_index:
        m = downloaded_index[parent_key]
        if m["case_id"] != cid:
            match, match_type = m, "parent_default"

    # 3. Matches baseline (empty settings or same as baseline)
    if not match:
        if full_key == baseline_key or (len(info["full_settings"]) == 0 and baseline_case):
            match, match_type = baseline_case[0], "baseline"

    # 4. Parent matches baseline (primary = default, no parents)
    if not match:
        if parent_key == baseline_key or (len(info["parent_settings"]) == 0 and baseline_case):
            # This means: no parents, just a primary change that = default
            # Check if the full_key matches any downloaded
            if not match:
                match, match_type = baseline_case[0], "baseline_default"

    # 5. Chain: parent matches another non-downloaded case that matches downloaded
    if not match:
        for oc in all_cases:
            if oc["case_id"] == cid or oc["case_id"] not in cache:
                continue
            oc_full = tuple(sorted(cache[oc["case_id"]]["full_settings"].items()))
            if oc_full == parent_key:
                if oc_full in downloaded_index:
                    match = downloaded_index[oc_full]
                    match_type = "chain"
                    break

    results.append({
        "case": c,
        "match": match,
        "match_type": match_type,
        "primary": f"{info['primary_name']}={info['primary_value']}",
    })

# --- Results ---
can_reuse = [r for r in results if r["match"]]
needs_xlsx = [r for r in results if not r["match"]]

print(f"Non-downloaded cases analyzed: {len(results)}")
print(f"Can reuse another XLSX:        {len(can_reuse)}")
print(f"Needs own XLSX (truly unique): {len(needs_xlsx)}")
print()

# Group can_reuse by status
print("=" * 70)
print(f"CAN REUSE ({len(can_reuse)} cases)")
print("=" * 70)

for status_label in ["baseline;resuse", ""]:
    group = [r for r in can_reuse if r["case"]["tv_xlsx_status"] == status_label]
    if not group:
        continue
    label = status_label if status_label else "(henuz indirilmemis)"
    print(f"\n  -- {label}: {len(group)} cases --")
    for r in sorted(group, key=lambda x: int(x["case"]["run_order"]) if str(x["case"]["run_order"]).isdigit() else 0):
        c = r["case"]
        m = r["match"]
        ro = str(c["run_order"]).zfill(3)
        mro = str(m["run_order"]).zfill(3)
        print(f"  [{ro}] {c['case_id']}")
        print(f"       -> reuses [{mro}] {m['case_id']} ({r['match_type']})")

# Check setup_fail cases that can reuse
fail_reuse = [r for r in can_reuse if r["case"]["setup_check"] == "FAIL"]
if fail_reuse:
    print(f"\n  -- Setup FAIL cases that can reuse: {len(fail_reuse)} --")
    for r in sorted(fail_reuse, key=lambda x: int(x["case"]["run_order"]) if str(x["case"]["run_order"]).isdigit() else 0):
        c = r["case"]
        m = r["match"]
        ro = str(c["run_order"]).zfill(3)
        mro = str(m["run_order"]).zfill(3)
        print(f"  [{ro}] {c['case_id']} (setup={c['setup_check']})")
        print(f"       -> reuses [{mro}] {m['case_id']} ({r['match_type']})")

print()
print("=" * 70)
print(f"NEEDS OWN XLSX ({len(needs_xlsx)} cases)")
print("=" * 70)
for r in sorted(needs_xlsx, key=lambda x: int(x["case"]["run_order"]) if str(x["case"]["run_order"]).isdigit() else 0):
    c = r["case"]
    ro = str(c["run_order"]).zfill(3)
    st = c["tv_xlsx_status"] or "(empty)"
    su = c["setup_check"] or "-"
    print(f"  [{ro}] {c['case_id']}  status={st}  setup={su}  primary={r['primary']}")
