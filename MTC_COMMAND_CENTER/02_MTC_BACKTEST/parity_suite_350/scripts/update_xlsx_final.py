#!/usr/bin/env python3
"""
Final XLSX updater:
1. Compute reuse relationships for ALL cases
2. Merge ALL batch CSV results (setup_check + compare_status)
3. Update XLSX with proper reuse labels + test results
4. Print cumulative status report
"""

import csv, json, sys, io, shutil
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
CASES_DIR = BASE / "cases"
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
COMPARE_RUNS = BASE / "compare_runs"
DIAG_FIELDS = [
    "clip_strict_status",
    "raw_strict_status",
    "early_trade_end_candidate",
    "gap_days",
    "clip_tv_trades",
    "clip_py_trades",
    "raw_tv_trades",
    "raw_py_trades",
]


# ============================================================
# STEP 1: Load all batch results into unified dictionaries
# ============================================================

def load_all_results():
    """Merge all CSV result files into setup_results and parity_results dicts."""
    setup_results = {}   # case_id -> status
    parity_results = {}  # case_id -> {"status","notes",diag...}

    def _upsert_parity(row, status_key, notes_key="notes"):
        cid = row.get("case_id", "")
        status = row.get(status_key, "")
        if not cid or not status:
            return
        payload = {
            "status": status,
            "notes": row.get(notes_key, ""),
        }
        for f in DIAG_FIELDS:
            payload[f] = row.get(f, "")
        parity_results[cid] = payload

    # --- all_parity_results.csv (original 121 cases from VS Code Claude) ---
    csv_path = COMPARE_RUNS / "all_parity_results.csv"
    if csv_path.exists():
        with open(csv_path, "r") as f:
            for row in csv.DictReader(f):
                _upsert_parity(row, "compare_status")

    # --- core1_parity_results.csv ---
    csv_path = COMPARE_RUNS / "core1_parity_results.csv"
    if csv_path.exists():
        with open(csv_path, "r") as f:
            for row in csv.DictReader(f):
                _upsert_parity(row, "compare_status")

    # --- core1_setup_check.csv ---
    csv_path = COMPARE_RUNS / "core1_setup_check.csv"
    if csv_path.exists():
        with open(csv_path, "r") as f:
            for row in csv.DictReader(f):
                cid = row["case_id"]
                if row["status"] != "NO_FILE":
                    setup_results[cid] = row["status"]

    # --- all_setup_check.csv ---
    csv_path = COMPARE_RUNS / "all_setup_check.csv"
    if csv_path.exists():
        with open(csv_path, "r") as f:
            for row in csv.DictReader(f):
                cid = row["case_id"]
                if row["status"] != "NO_FILE":
                    setup_results[cid] = row["status"]

    # --- batch CSV files (batch2 through batch11) ---
    for batch_csv in sorted(COMPARE_RUNS.glob("batch*_results.csv")):
        with open(batch_csv, "r") as f:
            for row in csv.DictReader(f):
                cid = row["case_id"]
                if "setup_status" in row and row["setup_status"]:
                    setup_results[cid] = row["setup_status"]
                if "parity_status" in row and row["parity_status"]:
                    _upsert_parity(row, "parity_status")

    # --- new_cases_results.csv (last batch overwrite) ---
    csv_path = COMPARE_RUNS / "new_cases_results.csv"
    if csv_path.exists():
        with open(csv_path, "r") as f:
            for row in csv.DictReader(f):
                cid = row["case_id"]
                if "setup_status" in row and row["setup_status"]:
                    setup_results[cid] = row["setup_status"]
                if "parity_status" in row and row["parity_status"]:
                    _upsert_parity(row, "parity_status")

    return setup_results, parity_results


# ============================================================
# STEP 2: Compute reuse relationships
# ============================================================

def get_case_info(case_id):
    """Extract settings from case JSON."""
    p = CASES_DIR / f"{case_id}.json"
    if not p.exists():
        return None
    with open(p) as f:
        data = json.load(f)
    gen = data.get("_generated", {})
    tv = data.get("_tv_case", {})
    parents = tv.get("parents", [])
    pname = gen.get("source_input_name", "")
    pvalue = gen.get("source_target_value", "")

    # Full settings = parents + primary change
    settings = {p["name"]: str(p["value"]) for p in parents}
    if pname:
        settings[pname] = str(pvalue)

    parent_only = {p["name"]: str(p["value"]) for p in parents}

    return {
        "primary_name": pname,
        "primary_value": str(pvalue),
        "full_settings": settings,
        "parent_settings": parent_only,
    }


def compute_reuse_map(all_cases, cache):
    """
    For each non-downloaded case, find which downloaded case (or baseline)
    it can reuse. Returns dict: case_id -> (reuse_case_id, reuse_run_order, match_type)
    """
    # Build downloaded index: settings_key -> case
    downloaded_cases = [c for c in all_cases if c["tv_xlsx_status"] == "DOWNLOADED"]
    downloaded_index = {}
    for c in downloaded_cases:
        if c["case_id"] in cache:
            key = tuple(sorted(cache[c["case_id"]]["full_settings"].items()))
            if key not in downloaded_index:
                downloaded_index[key] = c

    # Baseline
    baseline_case = [c for c in all_cases if c["tv_xlsx_status"] == "baseline"]
    baseline_key = None
    baseline = None
    if baseline_case:
        baseline = baseline_case[0]
        if baseline["case_id"] in cache:
            baseline_key = tuple(sorted(cache[baseline["case_id"]]["full_settings"].items()))

    reuse_map = {}  # case_id -> (reuse_case_id, reuse_run_order, match_type)

    for c in all_cases:
        status = c["tv_xlsx_status"]
        if status in ("DOWNLOADED", "baseline", "SKIP"):
            continue

        cid = c["case_id"]
        if cid not in cache:
            continue

        info = cache[cid]
        full_key = tuple(sorted(info["full_settings"].items()))
        parent_key = tuple(sorted(info["parent_settings"].items()))

        match = None
        match_type = None

        # 1. Exact match with downloaded
        if full_key in downloaded_index:
            m = downloaded_index[full_key]
            if m["case_id"] != cid:
                match, match_type = m, "exact"

        # 2. Parent-only matches a downloaded (primary = default)
        if not match and parent_key in downloaded_index:
            m = downloaded_index[parent_key]
            if m["case_id"] != cid:
                match, match_type = m, "parent_default"

        # 3. Matches baseline (empty settings or same as baseline)
        if not match and baseline:
            if full_key == baseline_key or (len(info["full_settings"]) == 0):
                match, match_type = baseline, "baseline"

        # 4. Parent matches baseline (primary = default, no parents)
        if not match and baseline:
            if parent_key == baseline_key or (len(info["parent_settings"]) == 0):
                match, match_type = baseline, "baseline_default"

        # 5. Chain: find any other case (downloaded or reuse-resolvable) with matching settings
        if not match:
            for oc in all_cases:
                if oc["case_id"] == cid or oc["case_id"] not in cache:
                    continue
                oc_full = tuple(sorted(cache[oc["case_id"]]["full_settings"].items()))
                if oc_full == full_key and oc["tv_xlsx_status"] == "DOWNLOADED":
                    match = oc
                    match_type = "chain"
                    break
                # Also check parent_key match with downloaded
                if oc_full == parent_key and oc["tv_xlsx_status"] == "DOWNLOADED":
                    match = oc
                    match_type = "chain_parent"
                    break

        if match:
            reuse_map[cid] = (match["case_id"], match["run_order"], match_type)

    # Second pass: resolve chains (cases that reuse other reuse cases)
    # Find cases NOT in reuse_map that can chain through reuse cases
    unresolved = []
    for c in all_cases:
        cid = c["case_id"]
        if cid in reuse_map or c["tv_xlsx_status"] in ("DOWNLOADED", "baseline", "SKIP"):
            continue
        if cid not in cache:
            continue
        unresolved.append(c)

    # Try to resolve through intermediate reuse cases
    for c in unresolved:
        cid = c["case_id"]
        info = cache[cid]
        full_key = tuple(sorted(info["full_settings"].items()))
        parent_key = tuple(sorted(info["parent_settings"].items()))

        # Check if any case in reuse_map has matching settings
        for other_cid, (reuse_cid, reuse_ro, rtype) in reuse_map.items():
            if other_cid in cache:
                other_full = tuple(sorted(cache[other_cid]["full_settings"].items()))
                if other_full == full_key or other_full == parent_key:
                    # This case can reuse the same source as the other case
                    reuse_map[cid] = (reuse_cid, reuse_ro, f"chain_via_{other_cid}")
                    break

    return reuse_map


# ============================================================
# STEP 3: Update XLSX
# ============================================================

def main():
    print("=" * 70)
    print("FINAL XLSX UPDATE")
    print("=" * 70)
    print()

    # Load all batch results
    setup_results, parity_results = load_all_results()
    print(f"Merged setup results: {len(setup_results)} cases")
    print(f"Merged parity results: {len(parity_results)} cases")
    print()

    # Backup XLSX
    backup = XLSX_PATH.parent / f"CASE_SETUP_GUIDE_BACKUP_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(XLSX_PATH, backup)
    print(f"Backup: {backup.name}")

    # Open XLSX
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Cases"]

    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.value] = cell.column
    for field in DIAG_FIELDS:
        if field not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = field
            headers[field] = new_col

    col_run_order = headers["run_order"]
    col_case_id = headers["case_id"]
    col_tv_xlsx_status = headers["tv_xlsx_status"]
    col_setup_check = headers["setup_check"]
    col_compare_status = headers["compare_status"]
    col_diag = {f: headers.get(f) for f in DIAG_FIELDS}
    col_notes = headers.get("notes")

    # Read all cases from XLSX
    all_cases = []
    for row_idx in range(2, ws.max_row + 1):
        cid = ws.cell(row=row_idx, column=col_case_id).value
        if not cid:
            continue
        all_cases.append({
            "row_idx": row_idx,
            "run_order": ws.cell(row=row_idx, column=col_run_order).value,
            "case_id": cid,
            "tv_xlsx_status": ws.cell(row=row_idx, column=col_tv_xlsx_status).value or "",
            "setup_check": ws.cell(row=row_idx, column=col_setup_check).value or "",
            "compare_status": ws.cell(row=row_idx, column=col_compare_status).value or "",
        })

    # Build case info cache
    cache = {}
    for c in all_cases:
        info = get_case_info(c["case_id"])
        if info:
            cache[c["case_id"]] = info

    # Compute reuse map
    reuse_map = compute_reuse_map(all_cases, cache)
    print(f"Reuse relationships found: {len(reuse_map)} cases")
    print()

    # --- Apply updates ---
    updated_reuse = 0
    updated_setup = 0
    updated_parity = 0
    updated_notes_count = 0

    for c in all_cases:
        cid = c["case_id"]
        row_idx = c["row_idx"]
        current_status = c["tv_xlsx_status"]

        # --- Update reuse labels ---
        if cid in reuse_map:
            reuse_cid, reuse_ro, match_type = reuse_map[cid]
            reuse_label = f"reuses:{str(reuse_ro).zfill(3)}_{reuse_cid}"
            if current_status != reuse_label:
                ws.cell(row=row_idx, column=col_tv_xlsx_status).value = reuse_label
                updated_reuse += 1

            # Propagate parity from reused case
            current_compare = ws.cell(row=row_idx, column=col_compare_status).value or ""
            if not current_compare:
                # Check if the reused case has parity result
                if reuse_cid in parity_results:
                    reuse_payload = parity_results[reuse_cid]
                    reuse_status = reuse_payload.get("status", "")
                    if reuse_status == "PASS":
                        ws.cell(row=row_idx, column=col_compare_status).value = "PASS(reuse)"
                        ws.cell(row=row_idx, column=col_setup_check).value = "PASS(reuse)"
                        for field in DIAG_FIELDS:
                            col = col_diag.get(field)
                            if col:
                                ws.cell(row=row_idx, column=col).value = reuse_payload.get(field, "")
                        if col_notes:
                            ws.cell(row=row_idx, column=col_notes).value = f"reuses {reuse_cid}"
                        updated_parity += 1

        # --- Update setup_check from merged results ---
        if cid in setup_results:
            new_val = setup_results[cid]
            current_setup = ws.cell(row=row_idx, column=col_setup_check).value or ""
            if not current_setup or current_setup != new_val:
                ws.cell(row=row_idx, column=col_setup_check).value = new_val
                updated_setup += 1

        # --- Update compare_status from merged results ---
        if cid in parity_results:
            payload = parity_results[cid]
            new_status = payload.get("status", "")
            new_notes = payload.get("notes", "")
            current_compare = ws.cell(row=row_idx, column=col_compare_status).value or ""
            if not current_compare or current_compare != new_status:
                ws.cell(row=row_idx, column=col_compare_status).value = new_status
                updated_parity += 1
            for field in DIAG_FIELDS:
                col = col_diag.get(field)
                if col:
                    ws.cell(row=row_idx, column=col).value = payload.get(field, "")
            # Update notes
            if col_notes and new_notes:
                existing_notes = ws.cell(row=row_idx, column=col_notes).value or ""
                if new_notes not in existing_notes:
                    note = f"{existing_notes}; {new_notes}" if existing_notes else new_notes
                    ws.cell(row=row_idx, column=col_notes).value = note
                    updated_notes_count += 1

    # --- Save ---
    wb.save(XLSX_PATH)
    wb.close()

    print("=" * 70)
    print("UPDATE RESULTS")
    print("=" * 70)
    print(f"  Reuse labels updated:    {updated_reuse}")
    print(f"  setup_check updated:     {updated_setup}")
    print(f"  compare_status updated:  {updated_parity}")
    print(f"  Notes updated:           {updated_notes_count}")
    print()

    # ============================================================
    # STEP 4: Cumulative Status Report
    # ============================================================

    # Re-read updated XLSX for final counts
    wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws2 = wb2["Cases"]
    headers2 = {}
    for cell in ws2[1]:
        if cell.value:
            headers2[cell.value] = cell.column

    status_counts = Counter()
    setup_counts = Counter()
    compare_counts = Counter()
    reuse_type_counts = Counter()
    total = 0

    for row_idx in range(2, ws2.max_row + 1):
        cid = ws2.cell(row=row_idx, column=headers2["case_id"]).value
        if not cid:
            continue
        total += 1

        tv_status = ws2.cell(row=row_idx, column=headers2["tv_xlsx_status"]).value or "(empty)"
        setup = ws2.cell(row=row_idx, column=headers2["setup_check"]).value or "(empty)"
        compare = ws2.cell(row=row_idx, column=headers2["compare_status"]).value or "(empty)"

        # Categorize tv_xlsx_status
        if tv_status.startswith("reuses:"):
            status_counts["reuses:*"] += 1
            # Extract match type from reuse_map
            if cid in reuse_map:
                _, _, mt = reuse_map[cid]
                base_mt = mt.split("_via_")[0] if "_via_" in mt else mt
                reuse_type_counts[base_mt] += 1
        else:
            status_counts[tv_status] += 1

        setup_counts[setup] += 1
        compare_counts[compare] += 1

    wb2.close()

    print("=" * 70)
    print("CUMULATIVE STATUS (457 cases)")
    print("=" * 70)
    print()

    print("tv_xlsx_status:")
    for k in ["DOWNLOADED", "reuses:*", "baseline", "baseline;resuse", "SKIP", "(empty)"]:
        if k in status_counts:
            print(f"  {k:25s} {status_counts[k]:4d}")
    for k, v in sorted(status_counts.items(), key=lambda x: -x[1]):
        if k not in ["DOWNLOADED", "reuses:*", "baseline", "baseline;resuse", "SKIP", "(empty)"]:
            print(f"  {k:25s} {v:4d}")
    print(f"  {'TOTAL':25s} {total:4d}")
    print()

    if reuse_type_counts:
        print("  Reuse breakdown:")
        for k, v in sorted(reuse_type_counts.items(), key=lambda x: -x[1]):
            print(f"    {k:25s} {v:4d}")
        print()

    print("setup_check:")
    for k, v in sorted(setup_counts.items(), key=lambda x: -x[1]):
        print(f"  {k:25s} {v:4d}")
    print()

    print("compare_status:")
    for k, v in sorted(compare_counts.items(), key=lambda x: -x[1]):
        print(f"  {k:25s} {v:4d}")
    print()

    # --- Cases still needing action ---
    print("=" * 70)
    print("REMAINING ACTIONS")
    print("=" * 70)

    wb3 = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws3 = wb3["Cases"]
    h3 = {}
    for cell in ws3[1]:
        if cell.value:
            h3[cell.value] = cell.column

    needs_download = []
    needs_reexport = []
    bt_fail_cases = []
    error_cases = []
    reuse_no_parity = []

    for row_idx in range(2, ws3.max_row + 1):
        cid = ws3.cell(row=row_idx, column=h3["case_id"]).value
        ro = ws3.cell(row=row_idx, column=h3["run_order"]).value
        if not cid:
            continue

        tv_status = ws3.cell(row=row_idx, column=h3["tv_xlsx_status"]).value or ""
        setup = ws3.cell(row=row_idx, column=h3["setup_check"]).value or ""
        compare = ws3.cell(row=row_idx, column=h3["compare_status"]).value or ""

        ro_str = str(ro).zfill(3)

        if not tv_status or tv_status == "":
            needs_download.append(f"  [{ro_str}] {cid}")
        elif tv_status == "baseline;resuse":
            # Still has old label - couldn't resolve reuse
            needs_download.append(f"  [{ro_str}] {cid}  (was baseline;resuse, no reuse match found)")
        elif setup == "FAIL":
            needs_reexport.append(f"  [{ro_str}] {cid}  setup=FAIL")
        elif compare == "BT_FAIL":
            bt_fail_cases.append(f"  [{ro_str}] {cid}  BT_FAIL (engine limitation)")
        elif compare == "ERROR":
            error_cases.append(f"  [{ro_str}] {cid}  ERROR")
        elif tv_status.startswith("reuses:") and not compare:
            reuse_no_parity.append(f"  [{ro_str}] {cid}  -> {tv_status}")

    if needs_download:
        print(f"\nNeeds download ({len(needs_download)}):")
        for line in needs_download:
            print(line)

    if needs_reexport:
        print(f"\nNeeds re-export (setup FAIL) ({len(needs_reexport)}):")
        for line in needs_reexport:
            print(line)

    if bt_fail_cases:
        print(f"\nEngine limitation - BT_FAIL ({len(bt_fail_cases)}):")
        for line in bt_fail_cases:
            print(line)

    if error_cases:
        print(f"\nErrors ({len(error_cases)}):")
        for line in error_cases:
            print(line)

    if reuse_no_parity:
        print(f"\nReuse cases - parity not yet run ({len(reuse_no_parity)}):")
        for line in reuse_no_parity[:20]:
            print(line)
        if len(reuse_no_parity) > 20:
            print(f"  ... and {len(reuse_no_parity) - 20} more")

    tested = sum(v for k, v in compare_counts.items() if k != "(empty)")
    remaining = total - tested - status_counts.get("SKIP", 0)
    print()
    print("=" * 70)
    print(f"PROGRESS: {tested} tested / {total} total")
    print(f"  Parity PASS:    {compare_counts.get('PASS', 0)}")
    print(f"  SKIP:           {status_counts.get('SKIP', 0)}")
    print(f"  BT_FAIL:        {compare_counts.get('BT_FAIL', 0)}")
    print(f"  MISMATCH:       {compare_counts.get('MISMATCH', 0)}")
    print(f"  ERROR:          {compare_counts.get('ERROR', 0)}")
    print(f"  Remaining:      {remaining} (needs download/reuse parity/re-export)")
    print("=" * 70)

    wb3.close()


if __name__ == "__main__":
    main()
