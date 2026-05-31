#!/usr/bin/env python3
"""Run parity for the 8 previously BT_FAIL cases after engine fixes."""

import sys, io, csv, json, subprocess, tempfile
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
MTC_BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest")
CASES_DIR = BASE / "cases"
TV_DIR = BASE / "tv_manual_inputs"
DEBUG_DIR = MTC_BASE / "debug" / "parity_suite_350"

sys.path.insert(0, str(MTC_BASE))
import openpyxl
from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades, clip_overlap, summarize_report

bt_fail_cases = [
    ("048", "parity_bnd_026_max_pyramid_positions_v02"),
    ("049", "parity_bnd_026_max_pyramid_positions_v03"),
    ("145", "parity_bnd_062_use_stop_loss_v02"),
    ("238", "parity_core_097_htf_trend_timeframe_v01"),
    ("239", "parity_bnd_097_htf_trend_timeframe_v02"),
    ("335", "parity_bnd_134_recovery_mode_v03"),
    ("362", "parity_core_194_mode_v01"),
    ("363", "parity_bnd_194_mode_v02"),
]

print(f"Running {len(bt_fail_cases)} previously BT_FAIL cases...")
print(f"Start: {datetime.now():%H:%M:%S}")
print("=" * 70)

results = []

for idx, (ro, cid) in enumerate(bt_fail_cases):
    print(f"\n[{idx+1}/{len(bt_fail_cases)}] [{ro}] {cid}")
    cj = CASES_DIR / f"{cid}.json"

    # Backtest
    cmd = ["python", str(MTC_BASE / "scripts" / "run_case.py"), str(cj)]
    r = subprocess.run(cmd, cwd=str(MTC_BASE), capture_output=True, text=True, timeout=300)
    if r.returncode != 0:
        err = [l for l in r.stderr.strip().split("\n") if l.strip()]
        msg = err[-1][:200] if err else "unknown"
        print(f"  BT_FAIL: {msg}")
        results.append((ro, cid, "BT_FAIL", msg))
        continue
    print(f"  Backtest OK")

    # Parity
    folders = list(TV_DIR.glob(f"{ro}_*"))
    if not folders:
        print(f"  ERROR: No TV folder")
        results.append((ro, cid, "ERROR", "No TV folder"))
        continue

    xlsx_files = [f for f in folders[0].glob("*.xlsx") if not f.name.startswith("~")]
    if not xlsx_files:
        print(f"  ERROR: No XLSX")
        results.append((ro, cid, "ERROR", "No XLSX"))
        continue

    tv_csv = None
    try:
        wb = openpyxl.load_workbook(xlsx_files[0], read_only=True, data_only=True)
        sheet_name = None
        for sn in wb.sheetnames:
            if "list of trades" in sn.lower():
                sheet_name = sn
                break
        if not sheet_name:
            wb.close()
            results.append((ro, cid, "SKIP", "No List of trades sheet"))
            continue

        ws = wb[sheet_name]
        tf = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
        w = csv.writer(tf)
        for row in ws.iter_rows(values_only=True):
            w.writerow(row)
        tf.close()
        tv_csv = Path(tf.name)
        wb.close()

        debug_dir = DEBUG_DIR / cid
        py_files = sorted(debug_dir.glob("debug_python_trades_*.csv"))
        if not py_files:
            results.append((ro, cid, "ERROR", "No Python trades CSV"))
            continue
        py_csv = py_files[-1]

        case_json_path = CASES_DIR / f"{cid}.json"
        tv_tz = "Europe/London"
        if case_json_path.exists():
            with open(case_json_path, "r", encoding="utf-8") as f:
                case_payload = json.load(f)
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
        tv_df = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
        py_df = load_py_trades(py_csv)
        tv_df, py_df = clip_overlap(tv_df, py_df)
        report = build_report(tv_df, py_df)
        summary = summarize_report(tv_df, py_df, report)

        if summary["tv_trades"] == 0 and summary["py_trades"] == 0:
            print(f"  PASS: No trades")
            results.append((ro, cid, "PASS", "No trades"))
        elif summary["strict_pass"]:
            print(f"  PASS: {summary['tv_trades']} trades match")
            results.append((ro, cid, "PASS", f"{summary['tv_trades']} trades match"))
        else:
            note = (
                f"core={summary['core_match_count']}/{summary['compared']}; "
                f"tv={summary['tv_trades']}; py={summary['py_trades']}; "
                f"extra_tv={summary['extra_tv_trades']}; extra_py={summary['extra_py_trades']}"
            )
            print(f"  MISMATCH: {note}")
            results.append((ro, cid, "MISMATCH", note))
    except Exception as e:
        print(f"  ERROR: {e}")
        results.append((ro, cid, "ERROR", str(e)[:200]))
    finally:
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()

# Save
out_path = BASE / "compare_runs" / "batch_btfail_fix_results.csv"
with open(out_path, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["run_order", "case_id", "parity_status", "notes"])
    for r in results:
        w.writerow(r)

print("\n" + "=" * 70)
print("SUMMARY")
print("=" * 70)
counts = Counter(r[2] for r in results)
for k, v in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {k:12s} {v:4d}")
print(f"  {'TOTAL':12s} {len(results):4d}")
print(f"\nEnd: {datetime.now():%H:%M:%S}")
