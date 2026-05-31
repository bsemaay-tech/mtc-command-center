#!/usr/bin/env python
from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook
from zoneinfo import ZoneInfo


PROPERTY_NAME_ALIASES = {
    "use volume filter": "use volume participation filter",
    "use atr volatility filter": "use atr volatility floor",
    "exit end of day": "exit at end of day",
    "exit end of week": "exit at end of week",
    "use break even": "use break-even?",
    "use break-even": "use break-even?",
}


def canonical_property_name(raw: Any) -> str:
    key = re.sub(r"\s+", " ", str(raw or "")).strip().lower()
    if not key:
        return ""
    return PROPERTY_NAME_ALIASES.get(key, key)


def normalize_value(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "ON" if raw else "OFF"
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and raw.is_integer():
            return str(int(raw))
        return f"{raw}".rstrip("0").rstrip(".") if isinstance(raw, float) else str(raw)
    text = re.sub(r"\s+", " ", str(raw)).strip()
    if not text:
        return ""
    low = text.lower()
    if low in {"on", "true", "yes", "y", "1"}:
        return "ON"
    if low in {"off", "false", "no", "n", "0"}:
        return "OFF"
    num = text.replace(",", ".")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", num):
        fval = float(num)
        if fval.is_integer():
            return str(int(fval))
        return f"{fval}".rstrip("0").rstrip(".")
    return text.upper()


def parse_ui_actions_to_expectations(action_text: str) -> dict[str, str]:
    text = (action_text or "").strip()
    if not text.startswith("Set in TV UI:"):
        return {}
    body = text.split("Set in TV UI:", 1)[1].strip()
    out: dict[str, str] = {}
    for part in [p.strip() for p in body.split(";") if p.strip()]:
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        ckey = canonical_property_name(key)
        cval = normalize_value(value)
        if ckey:
            out[ckey] = cval
    return out


def read_properties(xlsx_path: Path) -> tuple[dict[str, str], dict[str, Any]]:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        prop_sheet_name = None
        for name in wb.sheetnames:
            if str(name).strip().lower() == "properties":
                prop_sheet_name = name
                break
        if prop_sheet_name is None:
            return {}, {}
        ws = wb[prop_sheet_name]
        props_norm: dict[str, str] = {}
        props_raw: dict[str, Any] = {}
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            key = None
            value = None
            if len(vals) >= 3 and vals[1] is not None:
                key = vals[1]
                value = vals[2]
            elif len(vals) >= 2 and vals[0] is not None:
                key = vals[0]
                value = vals[1]
            if key is None:
                continue
            k = canonical_property_name(key)
            if not k or k == "name":
                continue
            props_norm[k] = normalize_value(value)
            props_raw[k] = value
        return props_norm, props_raw
    finally:
        wb.close()


def parse_case_datetime(raw: str) -> datetime:
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        dt = datetime.strptime(raw, "%Y-%m-%d")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def parse_range_to_utc(raw: Any, local_tz: str) -> tuple[datetime | None, datetime | None]:
    text = re.sub(r"\s+", " ", str(raw or "")).strip()
    if not text:
        return None, None
    parts = [p.strip() for p in re.split(r"\s*[—-]\s*", text, maxsplit=1) if p.strip()]
    if len(parts) != 2:
        return None, None
    try:
        start_local = pd.to_datetime(parts[0], format="%b %d, %Y, %H:%M", errors="raise")
        end_local = pd.to_datetime(parts[1], format="%b %d, %Y, %H:%M", errors="raise")
        tz = ZoneInfo(local_tz)
        start_dt = start_local.to_pydatetime().replace(tzinfo=tz).astimezone(timezone.utc)
        end_dt = end_local.to_pydatetime().replace(tzinfo=tz).astimezone(timezone.utc)
        return start_dt, end_dt
    except Exception:
        return None, None


def find_latest_xlsx(case_dir: Path) -> Path | None:
    files = [p for p in case_dir.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    if not files:
        return None
    files.sort(key=lambda p: (p.stat().st_mtime, p.name), reverse=True)
    return files[0]


def parse_margin_call_and_last_time(xlsx_path: Path) -> tuple[int, str]:
    try:
        xl = pd.ExcelFile(xlsx_path)
    except Exception:
        return 0, ""
    if "List of trades" not in xl.sheet_names:
        return 0, ""
    try:
        df = xl.parse("List of trades")
    except Exception:
        return 0, ""
    if df.empty:
        return 0, ""
    margin_rows = 0
    s = df.astype(str).apply(lambda col: col.str.contains("margin call", case=False, na=False))
    if not s.empty:
        margin_rows = int(s.any(axis=1).sum())

    time_col = None
    for cand in ("Date and time", "Date/Time", "Time", "Entry Time", "Exit Time"):
        if cand in df.columns:
            time_col = cand
            break
    if time_col is None:
        return margin_rows, ""
    dt = pd.to_datetime(df[time_col], errors="coerce")
    dt = dt.dropna()
    if dt.empty:
        return margin_rows, ""
    return margin_rows, dt.max().isoformat()


def load_run_case_module(script_path: Path):
    spec = importlib.util.spec_from_file_location("run_case_module", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load module: {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def extract_tv_trades_csv(xlsx_path: Path, out_csv: Path) -> int:
    xl = pd.ExcelFile(xlsx_path)
    if "List of trades" not in xl.sheet_names:
        raise RuntimeError(f"'List of trades' sheet not found: {xlsx_path}")
    trades = xl.parse("List of trades")
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    trades.to_csv(out_csv, index=False)
    if "Trade #" in trades.columns:
        return int(pd.to_numeric(trades["Trade #"], errors="coerce").nunique())
    return int(len(trades))


def is_reuse_case(ui_actions: str) -> bool:
    return bool(re.match(r"^Reuse baseline XLSX from\s+.+", (ui_actions or "").strip(), flags=re.IGNORECASE))


def main() -> int:
    ap = argparse.ArgumentParser(description="Audit setup + parity for manual case folder range.")
    ap.add_argument("--suite-root", default="mtc_backtest/parity_suite_350")
    ap.add_argument("--manifest", default="manifests/cases_manifest_all.csv")
    ap.add_argument("--run-order-start", type=int, required=True)
    ap.add_argument("--run-order-end", type=int, required=True)
    ap.add_argument("--baseline-xlsx", default="")
    ap.add_argument("--output-prefix", default="cases_001_039_recheck")
    ap.add_argument(
        "--use-tv-trading-range",
        action="store_true",
        help="Run Python with per-case Trading range (from XLSX Properties) instead of case JSON date window.",
    )
    args = ap.parse_args()

    suite_root = Path(args.suite_root).resolve()
    workspace_root = suite_root.parent.parent
    manifest_path = (suite_root / args.manifest).resolve()
    compare_root = (suite_root / "compare_runs").resolve()
    compare_root.mkdir(parents=True, exist_ok=True)
    tmp_case_root = compare_root / "_tmp_case_overrides"
    tmp_case_root.mkdir(parents=True, exist_ok=True)

    run_case_module = load_run_case_module(workspace_root / "mtc_backtest" / "scripts" / "run_case.py")
    import sys

    sys.path.insert(0, str(workspace_root / "mtc_backtest"))
    from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades, summarize_report  # type: ignore

    with manifest_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    rows = [
        r
        for r in rows
        if args.run_order_start <= int(str(r.get("run_order", "0")).strip() or 0) <= args.run_order_end
    ]
    rows.sort(key=lambda r: int(str(r.get("run_order", "0")).strip() or 0))

    baseline_props: dict[str, str] = {}
    baseline_xlsx = Path(args.baseline_xlsx).resolve() if args.baseline_xlsx else None
    if baseline_xlsx and baseline_xlsx.exists():
        baseline_props, _ = read_properties(baseline_xlsx)

    out_rows: list[dict[str, Any]] = []

    for row in rows:
        run_order = int(str(row.get("run_order", "0")).strip() or 0)
        case_id = str(row.get("case_id", "")).strip()
        folder = suite_root / "tv_manual_inputs" / f"{run_order:03d}_{case_id}"
        xlsx = find_latest_xlsx(folder)
        ui_actions = str(row.get("ui_actions", "")).strip()
        expected = parse_ui_actions_to_expectations(ui_actions)

        base = {
            "run_order": run_order,
            "case_id": case_id,
            "folder": str(folder),
            "xlsx": str(xlsx) if xlsx else "",
            "setup_status": "",
            "setup_issues": "",
            "trading_range_raw": "",
            "trading_range_start_utc": "",
            "trading_range_end_utc": "",
            "backtesting_range_raw": "",
            "backtesting_range_start_utc": "",
            "backtesting_range_end_utc": "",
            "date_range_note": "",
            "entry_mode": "",
            "margin_call_rows": "",
            "last_trade_time": "",
            "tv_trades": "",
            "py_trades": "",
            "core_match": "",
            "entry_price_match": "",
            "exit_price_match": "",
            "qty_match": "",
            "all_price_qty_match": "",
            "parity_status": "NOT_RUN",
            "action": "",
        }

        if xlsx is None:
            if is_reuse_case(ui_actions):
                base["setup_status"] = "SKIP_BASELINE_REUSE_EMPTY_OK"
                base["action"] = "no_action"
            else:
                base["setup_status"] = "MISSING_XLSX"
                base["action"] = "download_xlsx"
            out_rows.append(base)
            continue

        props_norm, props_raw = read_properties(xlsx)
        base["trading_range_raw"] = str(props_raw.get("trading range", "") or "")
        base["backtesting_range_raw"] = str(props_raw.get("backtesting range", "") or "")
        base["entry_mode"] = props_norm.get("entry mode", "")
        margin_rows, last_trade_time = parse_margin_call_and_last_time(xlsx)
        base["margin_call_rows"] = margin_rows
        base["last_trade_time"] = last_trade_time

        issues: list[str] = []
        for key, exp in expected.items():
            got = props_norm.get(key, "")
            if got != exp:
                issues.append(f"{key}:exp={exp}|got={got}")

        tv_tz = "Europe/London"
        case_json_path = (suite_root / str(row.get("case_json", ""))).resolve()
        try:
            case_payload = json.loads(case_json_path.read_text(encoding="utf-8"))
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
            exp_start = parse_case_datetime(str(case_payload.get("start_date")))
            exp_end = parse_case_datetime(str(case_payload.get("end_date")))
        except Exception:
            exp_start = None
            exp_end = None

        got_start, got_end = parse_range_to_utc(props_raw.get("trading range", ""), tv_tz)
        bck_start, bck_end = parse_range_to_utc(props_raw.get("backtesting range", ""), tv_tz)
        base["trading_range_start_utc"] = got_start.isoformat() if got_start else ""
        base["trading_range_end_utc"] = got_end.isoformat() if got_end else ""
        base["backtesting_range_start_utc"] = bck_start.isoformat() if bck_start else ""
        base["backtesting_range_end_utc"] = bck_end.isoformat() if bck_end else ""

        # NOTE:
        # "Trading range" on TV exports can drift based on strategy behavior (first/last active trade window),
        # so keep it as diagnostic metadata and do not fail setup on this field.
        if exp_start and exp_end and got_start and got_end:
            if got_start != exp_start or got_end != exp_end:
                base["date_range_note"] = (
                    f"trading_range_diff(exp_start={exp_start.isoformat()},got_start={got_start.isoformat()},"
                    f"exp_end={exp_end.isoformat()},got_end={got_end.isoformat()})"
                )
        elif exp_start and exp_end:
            base["date_range_note"] = "trading_range_unparsed"

        if baseline_props:
            for anchor_key in ("entry mode", "global heikin ashi", "signal mode"):
                if anchor_key in expected:
                    continue
                got = props_norm.get(anchor_key, "")
                exp = baseline_props.get(anchor_key, "")
                if exp and got != exp:
                    issues.append(f"baseline_drift_{anchor_key}:exp={exp}|got={got}")

        base["setup_issues"] = " || ".join(issues)
        base["setup_status"] = "SETUP_OK" if not issues else "SETUP_FAIL"

        if issues:
            base["action"] = "fix_tv_setup_and_reexport"
            out_rows.append(base)
            continue

        try:
            tv_csv = xlsx.with_name(f"{xlsx.stem}_trades.csv")
            tv_trade_count = extract_tv_trades_csv(xlsx, tv_csv)
            base["tv_trades"] = tv_trade_count

            run_case_path = case_json_path
            if args.use_tv_trading_range and got_start is not None and got_end is not None:
                with case_json_path.open("r", encoding="utf-8") as f_case:
                    case_payload = json.load(f_case)
                case_payload["start_date"] = got_start.replace(tzinfo=None).isoformat()
                case_payload["end_date"] = got_end.replace(tzinfo=None).isoformat()
                # Limit preroll to the TV backtesting window lead-in to avoid
                # over-seeding filters (e.g., MA/HTF) from unrelated older history.
                if bck_start is not None and got_start > bck_start:
                    lead_sec = (got_start - bck_start).total_seconds()
                    lead_days = max(1, int(math.ceil(lead_sec / 86400.0)))
                    original_preroll = int(case_payload.get("preroll_days", 0) or 0)
                    # Preserve the case's native preroll when it is already larger.
                    # Some parity cases rely on deeper indicator seeding than the
                    # visible TV backtesting lead-in alone.
                    case_payload["preroll_days"] = max(original_preroll, lead_days)
                # Keep ATR seed aligned to TV backtesting start so early-window
                # ATR-dependent SL/TP sizing is not under-seeded.
                if bck_start is not None:
                    cfg = case_payload.setdefault("config", {})
                    parity_cfg = cfg.setdefault("parity", {})
                    parity_cfg["atr_seed_start_utc"] = bck_start.isoformat()
                tmp_case_path = tmp_case_root / f"{run_order:03d}_{case_id}_{uuid4().hex[:8]}.json"
                with tmp_case_path.open("w", encoding="utf-8") as f_tmp:
                    json.dump(case_payload, f_tmp, ensure_ascii=False, indent=2)
                run_case_path = tmp_case_path

            results = run_case_module.run(str(run_case_path))
            debug_exports = results.get("debug_exports", {})
            py_csv_rel = str(debug_exports.get("debug_python_trades", "")).strip()
            py_csv = (workspace_root / py_csv_rel).resolve() if py_csv_rel else Path()
            if not py_csv.exists():
                raise RuntimeError("python trades csv missing")

            tv_df = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
            py_df = load_py_trades(py_csv, py_shift_min=0)
            rep = build_report(tv_df, py_df)
            summary = summarize_report(tv_df, py_df, rep)

            compared = int(summary["compared"])
            core_match = int(summary["core_match_count"])
            entry_price_match = int(summary["entry_price_match_count"])
            exit_price_match = int(summary["exit_price_match_count"])
            qty_match = int(summary["qty_match_count"])
            all_pq_match = int(summary["all_price_qty_match_count"])

            base["tv_trades"] = int(summary["tv_trades"])
            base["py_trades"] = int(summary["py_trades"])
            base["core_match"] = f"{core_match}/{compared}"
            base["entry_price_match"] = f"{entry_price_match}/{compared}"
            base["exit_price_match"] = f"{exit_price_match}/{compared}"
            base["qty_match"] = f"{qty_match}/{compared}"
            base["all_price_qty_match"] = f"{all_pq_match}/{compared}"
            base["parity_status"] = "PASS" if bool(summary["strict_pass"]) else "MISMATCH"
            base["action"] = "none" if base["parity_status"] == "PASS" else "analyze_mismatch"
        except Exception as exc:
            base["parity_status"] = "ERROR"
            base["action"] = f"error:{exc}"

        out_rows.append(base)

    out_csv = compare_root / f"{args.output_prefix}.csv"
    pd.DataFrame(out_rows).to_csv(out_csv, index=False)

    status_counts = pd.Series([r["setup_status"] for r in out_rows]).value_counts().to_dict()
    parity_counts = pd.Series([r["parity_status"] for r in out_rows]).value_counts().to_dict()

    out_md = compare_root / f"{args.output_prefix}.md"
    md = [
        "# Case Range Audit",
        "",
        f"- range: {args.run_order_start}-{args.run_order_end}",
        f"- rows: {len(out_rows)}",
        f"- baseline_xlsx: {str(baseline_xlsx) if baseline_xlsx else ''}",
        "",
        "## Setup Status",
    ]
    for k, v in sorted(status_counts.items()):
        md.append(f"- {k}: {v}")
    md.append("")
    md.append("## Parity Status")
    for k, v in sorted(parity_counts.items()):
        md.append(f"- {k}: {v}")
    md.append("")
    md.append("## Margin Call Rows")
    margin_cases = [r for r in out_rows if str(r.get("margin_call_rows", "")).strip() not in {"", "0"}]
    if not margin_cases:
        md.append("- none")
    else:
        for r in margin_cases:
            md.append(f"- {r['run_order']} {r['case_id']}: margin_call_rows={r['margin_call_rows']}")
    out_md.write_text("\n".join(md), encoding="utf-8")

    print(f"output_csv={out_csv}")
    print(f"output_md={out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
