#!/usr/bin/env python3
"""
Setup check + backtest + parity comparison for newly downloaded cases.

Phase 1: Validate TV XLSX Properties vs case JSON expectations
Phase 2: Run Python backtest for PASS cases
Phase 3: Compare TV vs Python trades for PASS cases
Phase 4: Update XLSX tracker with results
"""

import csv
import json
import shutil
import subprocess
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Dict, Tuple, Optional, List
import openpyxl
from datetime import timezone


BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
CASES_DIR = BASE / "cases"
TV_INPUTS = BASE / "tv_manual_inputs"
COMPARE_RUNS = BASE / "compare_runs"
PENDING_CSV = COMPARE_RUNS / "new_downloads_pending.csv"
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
MTC_BASE = BASE.parent  # mtc_backtest/
DIAG_FIELDS = [
    "clip_strict_status",
    "raw_strict_status",
    "early_trade_end_candidate",
    "gap_days",
    "clip_tv_trades",
    "clip_py_trades",
    "raw_tv_trades",
    "raw_py_trades",
]

import sys
sys.path.insert(0, str(MTC_BASE))
from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades, clip_overlap, summarize_report
from src.parity.dual_status import classify_parity_with_dual_view, compute_gap_days


# ─── SETUP CHECK FUNCTIONS ───

def read_tv_properties(xlsx_path: Path) -> Dict[str, str]:
    """Read Properties sheet from TV XLSX."""
    props = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb["Properties"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                props[str(row[0]).strip()] = str(row[1]).strip()
        wb.close()
    except Exception as e:
        print(f"    ERROR reading Properties: {e}")
    return props


def normalize_value(value: str) -> str:
    if not value:
        return ""
    value = value.strip().lower()
    if value in ("on", "yes", "true", "enabled"):
        return "On"
    if value in ("off", "no", "false", "disabled"):
        return "Off"
    return value.strip()


def compare_values(tv_value: str, expected_value: str) -> bool:
    tv_norm = normalize_value(tv_value)
    exp_norm = normalize_value(expected_value)
    if tv_norm == exp_norm:
        return True
    try:
        return abs(float(tv_value) - float(expected_value)) < 0.001
    except (ValueError, TypeError):
        return False


def check_case_setup(run_order: int, case_id: str) -> Tuple[str, str]:
    """Check TV XLSX Properties vs case JSON. Returns (status, notes)."""
    # Find TV XLSX
    folder_name = f"{run_order:03d}_{case_id}"
    case_folder = TV_INPUTS / folder_name
    if not case_folder.exists():
        return "NO_FILE", f"Folder not found: {folder_name}"

    xlsx_files = [f for f in case_folder.glob("*.xlsx") if not f.name.startswith("~$")]
    if not xlsx_files:
        return "NO_FILE", f"No XLSX in {folder_name}"

    tv_xlsx = xlsx_files[0]

    # Load case JSON
    case_json_path = CASES_DIR / f"{case_id}.json"
    if not case_json_path.exists():
        return "NO_FILE", f"Case JSON not found"

    with open(case_json_path, "r") as f:
        case_data = json.load(f)

    # Read TV properties
    tv_props = read_tv_properties(tv_xlsx)
    if not tv_props:
        return "FAIL", "Cannot read Properties sheet"

    # Extract expected settings
    generated = case_data.get("_generated", {})
    primary_name = generated.get("source_input_name", "")
    primary_value = generated.get("source_target_value", "")

    tv_case = case_data.get("_tv_case", {})
    parents = tv_case.get("parents", [])

    failures = []

    # Check primary setting
    if primary_name:
        if primary_name in tv_props:
            if not compare_values(tv_props[primary_name], primary_value):
                failures.append(f"PRIMARY {primary_name}: expected '{primary_value}', got '{tv_props[primary_name]}'")
        else:
            failures.append(f"PRIMARY '{primary_name}' not found in Properties")

    # Check parent conditions
    for parent in parents:
        pname = parent.get("name", "")
        pvalue = parent.get("value", "")
        if pname in tv_props:
            if not compare_values(tv_props[pname], pvalue):
                failures.append(f"PARENT {pname}: expected '{pvalue}', got '{tv_props[pname]}'")
        else:
            failures.append(f"PARENT '{pname}' not found")

    if failures:
        return "FAIL", "; ".join(failures)
    return "PASS", ""


# ─── BACKTEST FUNCTION ───

def run_backtest(case_id: str) -> bool:
    """Run Python backtest for a single case."""
    case_json = CASES_DIR / f"{case_id}.json"
    cmd = ["python", str(MTC_BASE / "scripts" / "run_case.py"), str(case_json)]
    try:
        result = subprocess.run(cmd, cwd=str(MTC_BASE), capture_output=True, text=True, timeout=300)
        return result.returncode == 0
    except Exception as e:
        print(f"    Backtest error: {e}")
        return False


# ─── PARITY COMPARISON ───

def compare_parity(run_order: int, case_id: str) -> Tuple[str, str, Dict[str, str]]:
    """Compare TV vs Python trades. Returns (status, notes, diagnostics)."""
    # Find TV XLSX
    folder_name = f"{run_order:03d}_{case_id}"
    case_folder = TV_INPUTS / folder_name
    xlsx_files = [f for f in case_folder.glob("*.xlsx") if not f.name.startswith("~$")]
    if not xlsx_files:
        return "ERROR", "No XLSX", {}

    # Extract TV trades to temp CSV
    tv_csv = None
    try:
        wb = openpyxl.load_workbook(xlsx_files[0], read_only=True, data_only=True)
        sheet_name = None
        for sn in wb.sheetnames:
            if "list of trades" in sn.lower():
                sheet_name = sn
                break
        if not sheet_name:
            wb.close()
            return "SKIP", "No 'List of trades' sheet", {}

        ws = wb[sheet_name]
        temp_file = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
        writer = csv.writer(temp_file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
        temp_file.close()
        tv_csv = Path(temp_file.name)
        wb.close()
    except Exception as e:
        return "ERROR", f"TV extract error: {e}", {}

    # Find Python trades
    debug_dir = MTC_BASE / "debug" / "parity_suite_350" / case_id
    if not debug_dir.exists():
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()
        return "ERROR", "Debug dir not found", {}

    py_files = list(debug_dir.glob("debug_python_trades_*.csv"))
    if not py_files:
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()
        return "ERROR", "PY trades not found", {}

    py_csv = sorted(py_files)[-1]

    try:
        case_json_path = CASES_DIR / f"{case_id}.json"
        tv_tz = "Europe/London"
        case_end_utc = None
        if case_json_path.exists():
            with open(case_json_path, "r", encoding="utf-8") as f:
                case_payload = json.load(f)
            tv_tz = str(case_payload.get("tv_tz", tv_tz))
            end_raw = case_payload.get("end_date")
            if end_raw:
                try:
                    case_end_utc = datetime.fromisoformat(str(end_raw))
                except ValueError:
                    case_end_utc = datetime.strptime(str(end_raw), "%Y-%m-%d")
                if case_end_utc.tzinfo is None:
                    case_end_utc = case_end_utc.replace(tzinfo=timezone.utc)
                else:
                    case_end_utc = case_end_utc.astimezone(timezone.utc)

        tv_raw = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
        py_raw = load_py_trades(py_csv)

        raw_report = build_report(tv_raw, py_raw)
        raw_summary = summarize_report(tv_raw, py_raw, raw_report)

        tv_clip, py_clip = clip_overlap(tv_raw, py_raw)
        clip_report = build_report(tv_clip, py_clip)
        clip_summary = summarize_report(tv_clip, py_clip, clip_report)

        last_tv_exit_utc = None
        if len(tv_raw) > 0:
            try:
                last_tv_exit_utc = tv_raw["exit_time"].max().to_pydatetime()
            except Exception:
                last_tv_exit_utc = None
        gap_days = compute_gap_days(case_end_utc, last_tv_exit_utc)
        return classify_parity_with_dual_view(
            clip_summary=clip_summary,
            raw_summary=raw_summary,
            gap_days=gap_days,
        )
    except Exception as e:
        return "ERROR", f"Compare error: {e}", {}
    finally:
        if tv_csv and tv_csv.exists():
            tv_csv.unlink()


# ─── XLSX UPDATER ───

def update_xlsx(results: List[dict]):
    """Update XLSX tracker with new results."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Cases"]

    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.value] = cell.column
    for field in DIAG_FIELDS:
        if field not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = field
            headers[field] = new_col

    col_case_id = headers["case_id"]
    col_setup = headers["setup_check"]
    col_compare = headers["compare_status"]
    col_clip = headers.get("clip_strict_status")
    col_raw = headers.get("raw_strict_status")
    col_early = headers.get("early_trade_end_candidate")
    col_gap = headers.get("gap_days")
    col_clip_tv = headers.get("clip_tv_trades")
    col_clip_py = headers.get("clip_py_trades")
    col_raw_tv = headers.get("raw_tv_trades")
    col_raw_py = headers.get("raw_py_trades")
    col_notes = headers.get("notes")

    # Build lookup
    result_map = {r["case_id"]: r for r in results}

    for row_idx in range(2, ws.max_row + 1):
        case_id = ws.cell(row=row_idx, column=col_case_id).value
        if case_id and case_id in result_map:
            r = result_map[case_id]
            ws.cell(row=row_idx, column=col_setup).value = r["setup_status"]
            if r.get("parity_status"):
                ws.cell(row=row_idx, column=col_compare).value = r["parity_status"]
            if col_clip:
                ws.cell(row=row_idx, column=col_clip).value = r.get("clip_strict_status", "")
            if col_raw:
                ws.cell(row=row_idx, column=col_raw).value = r.get("raw_strict_status", "")
            if col_early:
                ws.cell(row=row_idx, column=col_early).value = r.get("early_trade_end_candidate", "")
            if col_gap:
                ws.cell(row=row_idx, column=col_gap).value = r.get("gap_days", "")
            if col_clip_tv:
                ws.cell(row=row_idx, column=col_clip_tv).value = r.get("clip_tv_trades", "")
            if col_clip_py:
                ws.cell(row=row_idx, column=col_clip_py).value = r.get("clip_py_trades", "")
            if col_raw_tv:
                ws.cell(row=row_idx, column=col_raw_tv).value = r.get("raw_tv_trades", "")
            if col_raw_py:
                ws.cell(row=row_idx, column=col_raw_py).value = r.get("raw_py_trades", "")
            if col_notes and r.get("notes"):
                existing = ws.cell(row=row_idx, column=col_notes).value or ""
                note = r["notes"]
                if note and note not in existing:
                    ws.cell(row=row_idx, column=col_notes).value = (
                        f"{existing}; {note}" if existing else note
                    )

    wb.save(XLSX_PATH)
    wb.close()


# ─── MAIN ───

def main():
    # Load pending cases
    if not PENDING_CSV.exists():
        print("ERROR: new_downloads_pending.csv not found")
        return

    cases = []
    with open(PENDING_CSV, "r") as f:
        for row in csv.DictReader(f):
            cases.append((int(row["run_order"]), row["case_id"]))

    print(f"Processing {len(cases)} new downloads...")
    print()

    # ─── PHASE 1: SETUP CHECK ───
    print("=" * 70)
    print("PHASE 1: SETUP CHECK")
    print("=" * 70)

    all_results = []
    pass_cases = []

    for run_order, case_id in cases:
        print(f"  [{run_order:03d}] {case_id}...", end=" ")
        status, notes = check_case_setup(run_order, case_id)

        result = {
            "run_order": run_order,
            "case_id": case_id,
            "setup_status": status,
            "setup_notes": notes,
            "parity_status": None,
            "notes": notes,
            "clip_strict_status": "",
            "raw_strict_status": "",
            "early_trade_end_candidate": "",
            "gap_days": "",
            "clip_tv_trades": "",
            "clip_py_trades": "",
            "raw_tv_trades": "",
            "raw_py_trades": "",
        }
        all_results.append(result)

        if status == "PASS":
            print("[PASS]")
            pass_cases.append((run_order, case_id))
        else:
            print(f"[{status}] {notes}")

    setup_pass = sum(1 for r in all_results if r["setup_status"] == "PASS")
    setup_fail = sum(1 for r in all_results if r["setup_status"] == "FAIL")
    print(f"\nSetup: {setup_pass} PASS, {setup_fail} FAIL, {len(cases) - setup_pass - setup_fail} OTHER")
    print()

    if not pass_cases:
        print("No PASS cases to continue with.")
        update_xlsx(all_results)
        return

    # ─── PHASE 2: BACKTEST ───
    print("=" * 70)
    print(f"PHASE 2: PYTHON BACKTEST ({len(pass_cases)} cases)")
    print("=" * 70)

    backtest_pass = []
    for i, (run_order, case_id) in enumerate(pass_cases):
        print(f"  [{i+1:2d}/{len(pass_cases)}] [{run_order:03d}] {case_id}...", end=" ", flush=True)
        if run_backtest(case_id):
            print("[OK]")
            backtest_pass.append((run_order, case_id))
        else:
            print("[FAILED]")
            for r in all_results:
                if r["case_id"] == case_id:
                    r["parity_status"] = "BT_FAIL"
                    r["notes"] = "Backtest failed"

    print(f"\nBacktest: {len(backtest_pass)}/{len(pass_cases)} OK")
    print()

    # ─── PHASE 3: PARITY COMPARISON ───
    print("=" * 70)
    print(f"PHASE 3: PARITY COMPARISON ({len(backtest_pass)} cases)")
    print("=" * 70)

    parity_pass = 0
    parity_mismatch = 0
    parity_error = 0

    for i, (run_order, case_id) in enumerate(backtest_pass):
        print(f"  [{i+1:2d}/{len(backtest_pass)}] [{run_order:03d}] {case_id}...", end=" ", flush=True)
        status, notes, diag = compare_parity(run_order, case_id)

        for r in all_results:
            if r["case_id"] == case_id:
                r["parity_status"] = status
                if notes:
                    r["notes"] = notes
                for k, v in diag.items():
                    r[k] = v

        if status == "PASS":
            print(f"[PASS] {notes}")
            parity_pass += 1
        elif status == "MISMATCH":
            print(f"[MISMATCH] {notes}")
            parity_mismatch += 1
        else:
            print(f"[{status}] {notes}")
            parity_error += 1

    print(f"\nParity: {parity_pass} PASS, {parity_mismatch} MISMATCH, {parity_error} ERROR/SKIP")
    print()

    # ─── PHASE 4: UPDATE XLSX ───
    print("=" * 70)
    print("PHASE 4: UPDATE XLSX TRACKER")
    print("=" * 70)

    update_xlsx(all_results)
    print("XLSX updated.")

    # ─── SAVE DETAILED RESULTS CSV ───
    csv_out = COMPARE_RUNS / "new_cases_results.csv"
    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "run_order", "case_id", "setup_status", "parity_status", "notes",
            "clip_strict_status", "raw_strict_status",
            "early_trade_end_candidate", "gap_days",
            "clip_tv_trades", "clip_py_trades",
            "raw_tv_trades", "raw_py_trades"
        ])
        writer.writeheader()
        for r in all_results:
            writer.writerow({
                "run_order": r["run_order"],
                "case_id": r["case_id"],
                "setup_status": r["setup_status"],
                "parity_status": r["parity_status"] or "",
                "notes": r["notes"],
                "clip_strict_status": r.get("clip_strict_status", ""),
                "raw_strict_status": r.get("raw_strict_status", ""),
                "early_trade_end_candidate": r.get("early_trade_end_candidate", ""),
                "gap_days": r.get("gap_days", ""),
                "clip_tv_trades": r.get("clip_tv_trades", ""),
                "clip_py_trades": r.get("clip_py_trades", ""),
                "raw_tv_trades": r.get("raw_tv_trades", ""),
                "raw_py_trades": r.get("raw_py_trades", ""),
            })
    print(f"Detailed results: {csv_out}")

    diagnostics_script = BASE / "scripts" / "run_post_compare_diagnostics.py"
    if diagnostics_script.exists():
        print("\nRunning post-compare diagnostics...")
        subprocess.run(
            ["python", str(diagnostics_script)],
            cwd=str(BASE),
            check=False,
        )

    # ─── FINAL SUMMARY ───
    print()
    print("=" * 70)
    print("FINAL SUMMARY")
    print("=" * 70)
    print(f"  Total cases processed:  {len(cases)}")
    print(f"  Setup PASS:             {setup_pass}")
    print(f"  Setup FAIL:             {setup_fail}")
    print(f"  Backtest OK:            {len(backtest_pass)}")
    print(f"  Parity PASS:            {parity_pass}")
    print(f"  Parity MISMATCH:        {parity_mismatch}")
    print(f"  Parity ERROR/SKIP:      {parity_error}")

    # Show FAIL details
    fail_cases = [r for r in all_results if r["setup_status"] == "FAIL"]
    if fail_cases:
        print()
        print("SETUP FAIL DETAILS:")
        for r in fail_cases:
            print(f"  [{r['run_order']:03d}] {r['case_id']}")
            print(f"       {r['setup_notes']}")

    mismatch_cases = [r for r in all_results if r.get("parity_status") == "MISMATCH"]
    if mismatch_cases:
        print()
        print("PARITY MISMATCH DETAILS:")
        for r in mismatch_cases:
            print(f"  [{r['run_order']:03d}] {r['case_id']}")
            print(f"       {r['notes']}")


if __name__ == "__main__":
    main()
