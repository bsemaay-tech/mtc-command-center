#!/usr/bin/env python3
"""
Create a 2nd sheet 'Retest Candidates' in the XLSX with same format as 'Cases'.
Lists cases where primary_change produced NO trade count difference,
with suggested alternative values that should produce different trades.
"""

import sys, io, json, re, copy
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
CASES_DIR = BASE / "cases"
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"

# =====================================================================
# KNOWN DEFAULTS (from defaults.py)
# =====================================================================
DEFAULTS = {
    # Range Filter
    "rsi_len": 14, "bb_len": 20, "use_bb_filter": True,
    "adx_trend_threshold": 25, "adx_range_threshold": 20,
    "chop_trend_threshold": 50, "chop_range_threshold": 62,
    "rsi_oversold": 30, "rsi_overbought": 70, "bb_mult": 2.0,
    # MACD
    "macd_fast_len": 12, "macd_slow_len": 26, "macd_signal_len": 9,
    "macd_distance_pct": 0.0,
    # Confirmation
    "swing_lookback": 5, "momentum_len": 10,
}

# =====================================================================
# RECOMMENDATIONS: (case_ro, current_value, suggested_value, reason)
# =====================================================================

recommendations = {}

# --- RANGE FILTER GROUP (28 cases, all 137 trades) ---
# These sub-parameters don't affect the filtered result
# Need extreme values to force different filtering
rf_suggestions = {
    # RSI Length: default=14, tested 7,14,21 -> all same
    "066": ("14", "40", "Extreme RSI length should produce very different smoothing"),
    "067": ("21", "5", "Minimum RSI produces very reactive signals"),
    "068": ("7", "50", "Maximum RSI length - very smooth, fewer signals"),
    # BB Filter On/Off: tested On,Off -> same
    "069": ("On", "On", "BB filter On combined with extreme BB params needed - see BB cases"),
    "070": ("Off", "Off", "Expected: if RF doesn't use BB, this is normal"),
    # Use Strategy TF Source: tested On,Off -> same
    "071": ("Off", "Off", "May not affect if data is same TF"),
    "072": ("On", "On", "May not affect if data is same TF"),
    # ADX Trend: default=25, tested 25 -> same as baseline
    "073": ("25", "45", "Very high threshold - only strong trends pass"),
    # ADX Range: default=20, tested 20,25,15 -> same
    "076": ("20", "10", "Minimum range threshold - very permissive"),
    "077": ("25", "30", "Maximum range threshold - very restrictive"),
    "078": ("15", "10", "At minimum - check if field ge=10"),
    # Chop Trend: default=50, tested 50,60 -> same
    "079": ("50", "35", "Near minimum (ge=30) - allows very choppy as trend"),
    "080": ("60", "70", "Maximum - only clear trends pass"),
    # Chop Range: default=62, tested 62,55,70 -> same
    "082": ("62", "80", "Maximum - almost nothing qualifies as range"),
    "083": ("55", "50", "Minimum - very permissive range detection"),
    "084": ("70", "50", "Minimum ge=50 - maximally permissive"),
    # RSI Oversold: default=30, tested 30,35,20 -> same
    "085": ("30", "10", "Extreme low - almost never oversold"),
    "086": ("35", "40", "Maximum - very broad oversold zone"),
    "087": ("20", "10", "Minimum ge=10"),
    # RSI Overbought: default=70, tested 70,80 -> same
    "088": ("70", "60", "Minimum ge=60 - wide overbought zone"),
    "089": ("80", "90", "Maximum - only extreme overbought"),
    # BB Length: default=20, tested 20,25,10 -> same
    "091": ("20", "50", "Maximum - very smooth BB"),
    "092": ("25", "10", "Minimum - very reactive BB"),
    "093": ("10", "50", "Maximum BB length"),
    # BB Multiplier: default=2, tested 2,3,1 -> same
    "094": ("2", "1.0", "Tightest BB - most signals filtered"),
    "095": ("3", "1.0", "Tightest BB for comparison"),
    "096": ("1", "3.0", "Widest BB - least filtered"),
}

# --- MACD HUB GROUP ---
macd_suggestions = {
    # Signal: default=9, tested 5,15 -> both 146 trades
    "373": ("5", "1", "Minimum signal - fastest MACD crossover"),
    "374": ("15", "30", "Very slow signal - delays crossovers significantly"),
    # Min distance: tested 0.1, 0.5 -> both 146
    "385": ("0.1", "5", "Large distance gate - should filter many crosses"),
    "386": ("0.5", "10", "Very large distance - filters most crosses"),
    # Fast: default=12, tested 8,20 -> both 152
    "367": ("8", "3", "Fastest MACD - very reactive"),
    "368": ("20", "50", "Very slow fast-line - huge lag"),
    # Slow: default=26, tested 34 -> 152
    "371": ("34", "100", "Very slow MACD baseline - extreme divergence from fast"),
    # Regime/CrossState/Histogram/Distance gates: tested On -> 146
    "375": ("On", "On", "Regime filter On - check with extreme MACD params"),
    "384": ("On", "On", "Zero-line distance gate - needs higher min_distance"),
    # HTF Bias: tested Off -> 170
    "389": ("Off", "Off", "HTF bias Off - expected if not in HTF mode"),
    "394": ("Off", "Off", "Apply HTF regime sign Off"),
    # CrossState/Histogram individual
    "378": ("On", "On", "CrossState On - check with extreme Signal length"),
    "380": ("On", "On", "Histogram gate On - check with extreme params"),
}

# --- CONFIRMATION/SWING GROUP (225 trades, 14 cases) ---
swing_suggestions = {
    # Swing Left Bars: tested 7 (default=5?) -> 225
    "397": ("7", "20", "Large lookback - finds wider/rarer swing levels"),
    # Max bars to wait: tested 50 -> 225
    "407": ("50", "3", "Very short timeout - many entries will timeout"),
    # Apply confirmation only when flat: tested On -> 225
    "412": ("On", "On", "On is expected behavior if already flat most entries"),
    # If new RAW arrives refresh: tested Off -> 225
    "420": ("Off", "On", "Enable refresh - dynamic behavior difference expected"),
    # If LONG & SHORT raw same bar: tested SHORT_WINS, IGNORE -> 225
    "424": ("SHORT_WINS", "CANCEL_BOTH", "Cancel both should reduce entries"),
    "425": ("IGNORE", "CANCEL_BOTH", "Cancel both - different handling"),
    # Bar close only: tested Off -> 225
    "427": ("Off", "On", "Bar close only should change timing significantly"),
    # Require RAW still true: tested On -> 225
    "429": ("On", "Off", "Off = more permissive, might allow more entries"),
    # Break buffer: tested 0,1,5 -> all 225
    "430": ("1", "50", "Large buffer - price must break far beyond swing"),
    "431": ("5", "100", "Very large buffer - significant price movement needed"),
    "432": ("0", "50", "Buffer 50 ticks should definitely filter some"),
    # Max swing distance: tested 2%, 5% -> both 225
    "433": ("2", "0.5", "Very tight distance - only closest swings valid"),
    "434": ("5", "0.1", "Extremely tight - almost no swing qualifies"),
}

# --- BASELINE MATCH: SUSPICIOUS (241 trades) ---
baseline_suspicious = {
    # Global Heikin Ashi = On -> still 241
    "055": ("On", "On", "HA candles should change Supertrend signals - INVESTIGATE if HA is actually applied to ST calc"),
    # Use Range Filters (Entry Pause) = On -> still 241
    "252": ("On", "On", "Range Filter enabled but same trades - RF may not filter in Supertrend mode - check if RF only active in RF signal modes"),
    # Chop Off Threshold = 57 -> still 241
    "278": ("57", "38", "Lower off-threshold: more restrictive, trades exit chop range sooner. But check if chop filter is enabled"),
    # Use MAE Guard = On, Max MAE% = 1-4 -> still 241
    "325": ("On", "On", "MAE guard on but no trades exceeded MAE% - try 0.1% to guarantee some exits"),
    "326": ("2", "0.1", "0.1% MAE - almost every trade will be force-closed"),
    "327": ("1", "0.1", "Minimum MAE - should force close nearly all trades early"),
    "328": ("4", "0.1", "Minimum MAE for maximum impact"),
}

# =====================================================================
# BUILD RETEST SHEET
# =====================================================================

# Merge all recommendations
all_suggestions = {}
all_suggestions.update(rf_suggestions)
all_suggestions.update(macd_suggestions)
all_suggestions.update(swing_suggestions)
all_suggestions.update(baseline_suspicious)

# Load XLSX
wb = openpyxl.load_workbook(XLSX_PATH)
ws_cases = wb["Cases"]

# Get headers from Cases sheet
case_headers = {}
for cell in ws_cases[1]:
    if cell.value:
        case_headers[cell.value] = cell.column

# Create or clear 'Retest Candidates' sheet
sheet_name = "Retest Candidates"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws_retest = wb.create_sheet(sheet_name)

# Copy headers from Cases sheet + add new columns
orig_headers = []
for cell in ws_cases[1]:
    if cell.value:
        orig_headers.append(cell.value)

# Add extra columns
extra_headers = [
    "issue_group",
    "current_trade_count",
    "suggested_value",
    "suggestion_reason",
    "retest_status",
]

all_headers = orig_headers + extra_headers

# Write headers with formatting
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_idx, h in enumerate(all_headers, 1):
    cell = ws_retest.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font

# Collect matching rows from Cases sheet
retest_rows = []
for row_idx in range(2, ws_cases.max_row + 1):
    ro = ws_cases.cell(row=row_idx, column=case_headers["run_order"]).value
    if ro is None:
        continue
    ro_str = str(ro).zfill(3)

    if ro_str in all_suggestions:
        current_val, suggested_val, reason = all_suggestions[ro_str]

        # Determine issue group
        if ro_str in rf_suggestions:
            group = "RANGE_FILTER_NO_EFFECT"
        elif ro_str in macd_suggestions:
            group = "MACD_HUB_NO_EFFECT"
        elif ro_str in swing_suggestions:
            group = "SWING_CONFIRM_NO_EFFECT"
        elif ro_str in baseline_suspicious:
            group = "BASELINE_MATCH_SUSPICIOUS"
        else:
            group = "OTHER"

        # Get trade count from notes
        notes = ws_cases.cell(row=row_idx, column=case_headers.get("notes", 15)).value or ""
        tc_match = re.search(r"(\d+)\s+trades", notes)
        trade_count = int(tc_match.group(1)) if tc_match else None

        # Copy all original columns
        row_data = []
        for h in orig_headers:
            val = ws_cases.cell(row=row_idx, column=case_headers[h]).value
            row_data.append(val)

        # Add extra columns
        row_data.append(group)
        row_data.append(trade_count)
        row_data.append(suggested_val)
        row_data.append(reason)
        row_data.append("")  # retest_status - empty for now

        retest_rows.append(row_data)

# Sort by issue_group then run_order
group_order = {
    "RANGE_FILTER_NO_EFFECT": 1,
    "MACD_HUB_NO_EFFECT": 2,
    "SWING_CONFIRM_NO_EFFECT": 3,
    "BASELINE_MATCH_SUSPICIOUS": 4,
}
retest_rows.sort(key=lambda r: (group_order.get(r[len(orig_headers)], 99), r[0] or 0))

# Write rows
group_fills = {
    "RANGE_FILTER_NO_EFFECT": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "MACD_HUB_NO_EFFECT": PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
    "SWING_CONFIRM_NO_EFFECT": PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid"),
    "BASELINE_MATCH_SUSPICIOUS": PatternFill(start_color="F0E0FF", end_color="F0E0FF", fill_type="solid"),
}

for r_idx, row_data in enumerate(retest_rows, 2):
    group = row_data[len(orig_headers)]
    fill = group_fills.get(group)
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_retest.cell(row=r_idx, column=c_idx, value=val)
        if fill:
            cell.fill = fill

# Auto-fit column widths (approximate)
col_widths = {
    "run_order": 10, "package": 10, "case_id": 55, "case_folder": 55,
    "tv_preset_name": 55, "primary_change": 40, "tv_ui_actions": 60,
    "dependency_note": 50, "expected_trade_behavior": 12, "tv_xlsx_status": 20,
    "tv_xlsx_file": 45, "tv_download_date": 15, "setup_check": 12,
    "compare_status": 14, "notes": 25, "exec_conflict_tag": 15,
    "exec_conflict_ready": 15, "issue_group": 28, "current_trade_count": 18,
    "suggested_value": 15, "suggestion_reason": 65, "retest_status": 14,
}
for col_idx, h in enumerate(all_headers, 1):
    ws_retest.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(h, 15)

# Save
wb.save(XLSX_PATH)
wb.close()

print(f"Created '{sheet_name}' sheet with {len(retest_rows)} cases")
print()
from collections import Counter
groups = Counter(r[len(orig_headers)] for r in retest_rows)
for g, c in sorted(groups.items(), key=lambda x: group_order.get(x[0], 99)):
    print(f"  {g:35s} {c:3d} cases")
print(f"  {'TOTAL':35s} {sum(groups.values()):3d} cases")
