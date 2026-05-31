#!/usr/bin/env python
"""
Run + compare pipeline for routed parity_suite_350 cases.

Flow:
1) Read manifests/tv_collection_status.csv and take MATCHED rows.
2) For each matched case:
   - Extract List of trades from routed *_strategy_report.xlsx to canonical *_trades.csv.
   - Run Python backtest for the case JSON.
   - Compare TV trades vs Python trades via src.parity.compare_tv_trades.
3) Write per-case compare summary + mismatch reports under compare_runs/.
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def resolve(base: Path, raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return (base / p).resolve()


def parse_int(raw: Any, default: int = 0) -> int:
    try:
        return int(str(raw).strip())
    except Exception:
        return default


PROPERTY_NAME_ALIASES = {
    "use volume filter": "use volume participation filter",
    "use atr volatility filter": "use atr volatility floor",
    "exit end of day": "exit at end of day",
    "exit end of week": "exit at end of week",
    "use break even": "use break-even?",
    "use break-even": "use break-even?",
}

BASELINE_ANCHOR_KEYS = [
    "trading range",
    "signal mode",
    "global heikin ashi",
    "use ha for supertrend",
    "[st] use wicks",
    "[st] atr len",
    "[st] factor",
    "use stop loss",
    "sl mode",
    "sl atr length",
    "sl atr multiplier",
    "sl % (distance)",
    "swing sl basis",
    "swing sl lookback",
    "use take profit",
    "tp mode (single tp)",
    "tp atr length",
    "tp atr multiplier",
    "tp % (distance)",
    "tp r multiple (single tp)",
    "use multi tp (2 tps, requires take profit = on)",
    "use break-even?",
    "use trailing stop",
    "trailing atr length",
    "start after (r multiple)",
    "trail distance (r multiple)",
    "time stop bars",
    "exit at end of day",
    "exit at end of week",
    "time stop condition",
]


def canonical_property_name(raw: Any) -> str:
    key = re.sub(r"\s+", " ", str(raw or "")).strip().lower()
    if not key:
        return ""
    return PROPERTY_NAME_ALIASES.get(key, key)


def normalize_value(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "ON" if raw else "OFF"
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and raw.is_integer():
            return str(int(raw))
        return f"{raw}".rstrip("0").rstrip(".") if isinstance(raw, float) else str(raw)

    text = re.sub(r"\s+", " ", str(raw)).strip()
    if not text:
        return ""
    low = text.lower()
    if low in {"on", "true", "yes", "y", "1"}:
        return "ON"
    if low in {"off", "false", "no", "n", "0"}:
        return "OFF"
    num = text.replace(",", ".")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", num):
        fval = float(num)
        if fval.is_integer():
            return str(int(fval))
        return f"{fval}".rstrip("0").rstrip(".")
    return text.upper()


def parse_ui_actions_to_expectations(action_text: str) -> dict[str, str]:
    text = (action_text or "").strip()
    if not text.startswith("Set in TV UI:"):
        return {}
    body = text.split("Set in TV UI:", 1)[1].strip()
    out: dict[str, str] = {}
    for part in [p.strip() for p in body.split(";") if p.strip()]:
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        ckey = canonical_property_name(key)
        cval = normalize_value(value)
        if ckey:
            out[ckey] = cval
    return out


def read_properties(xlsx_path: Path) -> dict[str, str]:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        prop_sheet_name = None
        for name in wb.sheetnames:
            if str(name).strip().lower() == "properties":
                prop_sheet_name = name
                break
        if prop_sheet_name is None:
            return {}
        ws = wb[prop_sheet_name]
        props: dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            key = None
            value = None
            if len(vals) >= 3 and vals[1] is not None:
                key = vals[1]
                value = vals[2]
            elif len(vals) >= 2 and vals[0] is not None:
                key = vals[0]
                value = vals[1]
            if key is None:
                continue
            k = canonical_property_name(key)
            if not k or k == "name":
                continue
            props[k] = normalize_value(value)
        return props
    finally:
        wb.close()


def find_baseline_xlsx(suite_root: Path, override: str) -> Path | None:
    if override:
        p = resolve(suite_root, override)
        return p if p.exists() else None
    candidate = (suite_root / "manifests" / "baseline_sources" / "baseline_tv_export_FILLED_v6.xlsx").resolve()
    if candidate.exists():
        return candidate
    return None


def compute_baseline_drift(
    *,
    baseline_props: dict[str, str] | None,
    actual_props: dict[str, str],
    setup_expected: dict[str, str],
) -> list[str]:
    if not baseline_props:
        return []
    drift_keys: list[str] = []
    for key in BASELINE_ANCHOR_KEYS:
        # Case-intended UI changes are allowed drift on anchor keys.
        if key in setup_expected:
            continue
        baseline_val = baseline_props.get(key, "")
        actual_val = actual_props.get(key, "")
        if baseline_val != actual_val:
            drift_keys.append(f"{key}:base={baseline_val}|got={actual_val}")
    return drift_keys


def load_manifest_rows(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        cid = str(row.get("case_id", "")).strip()
        if cid:
            out[cid] = row
    return out


def load_matched_status_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = [r for r in csv.DictReader(f) if str(r.get("match_result", "")).strip() == "MATCHED"]
    rows.sort(key=lambda r: parse_int(r.get("run_order", 0), 10**9))
    return rows


def load_run_case_module(script_path: Path):
    spec = importlib.util.spec_from_file_location("run_case_module", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load run_case module from {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def extract_tv_trades_csv(xlsx_path: Path, out_csv: Path) -> int:
    xl = pd.ExcelFile(xlsx_path)
    if "List of trades" not in xl.sheet_names:
        raise RuntimeError(f"'List of trades' sheet not found in {xlsx_path}")
    trades = xl.parse("List of trades")
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    trades.to_csv(out_csv, index=False)
    if "Trade #" in trades.columns:
        return int(pd.to_numeric(trades["Trade #"], errors="coerce").nunique())
    return int(len(trades))


def parse_tv_metrics(xlsx_path: Path) -> dict[str, Any]:
    out: dict[str, Any] = {
        "tv_total_trades": None,
        "tv_net_profit": None,
        "tv_percent_profitable": None,
        "tv_profit_factor": None,
    }
    xl = pd.ExcelFile(xlsx_path)
    if "Performance" in xl.sheet_names:
        perf = xl.parse("Performance")
        if not perf.empty:
            names = perf.iloc[:, 0].astype(str).str.strip().str.lower()
            for key, target in [
                ("net profit", "tv_net_profit"),
                ("profit factor", "tv_profit_factor"),
            ]:
                idx = names[names == key]
                if len(idx):
                    row = perf.loc[idx.index[0]]
                    val = row.iloc[1]
                    out[target] = float(val) if pd.notna(val) else None
    if "Trades analysis" in xl.sheet_names:
        ta = xl.parse("Trades analysis")
        if not ta.empty:
            names = ta.iloc[:, 0].astype(str).str.strip().str.lower()
            idx = names[names == "total trades"]
            if len(idx):
                v = ta.loc[idx.index[0]].iloc[1]
                out["tv_total_trades"] = int(float(v)) if pd.notna(v) else None
            idx = names[names == "percent profitable"]
            if len(idx):
                v = ta.loc[idx.index[0]].iloc[2]
                out["tv_percent_profitable"] = float(v) if pd.notna(v) else None
            idx = names[names == "profit factor"]
            if len(idx):
                v = ta.loc[idx.index[0]].iloc[1]
                out["tv_profit_factor"] = float(v) if pd.notna(v) else out["tv_profit_factor"]
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Process routed MATCHED cases for parity_suite_350.")
    ap.add_argument("--suite-root", default="mtc_backtest/parity_suite_350", help="Suite root")
    ap.add_argument("--manifest", default="manifests/cases_manifest_all.csv", help="Manifest CSV path")
    ap.add_argument("--status-csv", default="manifests/tv_collection_status.csv", help="Routing status CSV path")
    ap.add_argument("--compare-root", default="compare_runs", help="Compare output root")
    ap.add_argument("--max-cases", type=int, default=0, help="Optional limit for MATCHED rows")
    ap.add_argument(
        "--baseline-xlsx",
        default="",
        help="Optional baseline strategy_report xlsx (Properties anchor source).",
    )
    args = ap.parse_args()

    suite_root = Path(args.suite_root).resolve()
    manifest_path = resolve(suite_root, args.manifest)
    status_path = resolve(suite_root, args.status_csv)
    compare_root = resolve(suite_root, args.compare_root)
    compare_dir = compare_root / f"compare_{utc_stamp()}"
    reports_dir = compare_dir / "case_reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    if not manifest_path.exists():
        print(f"ERROR: manifest not found: {manifest_path}")
        return 2
    if not status_path.exists():
        print(f"ERROR: status csv not found: {status_path}")
        return 2

    manifest_by_case = load_manifest_rows(manifest_path)
    matched_rows = load_matched_status_rows(status_path)
    baseline_xlsx_path = find_baseline_xlsx(suite_root, args.baseline_xlsx)
    baseline_props = read_properties(baseline_xlsx_path) if baseline_xlsx_path else {}
    if args.max_cases > 0:
        matched_rows = matched_rows[: args.max_cases]

    if not matched_rows:
        print("matched_cases=0")
        print("status=ok")
        return 0

    workspace_root = suite_root.parent.parent
    run_case_script = workspace_root / "mtc_backtest" / "scripts" / "run_case.py"
    run_case_module = load_run_case_module(run_case_script)

    import sys

    sys.path.insert(0, str(workspace_root / "mtc_backtest"))
    from src.parity.compare_tv_trades import build_report, load_py_trades, load_tv_trades  # type: ignore

    summary_rows: list[dict[str, Any]] = []

    for idx, r in enumerate(matched_rows, start=1):
        case_id = str(r.get("case_id", "")).strip()
        manifest_row = manifest_by_case.get(case_id)
        if manifest_row is None:
            summary_rows.append(
                {
                    "case_id": case_id,
                    "run_order": r.get("run_order", ""),
                    "pack": r.get("pack", ""),
                    "status": "ERROR",
                    "error": "case_id not found in manifest",
                }
            )
            continue

        case_json = resolve(suite_root, manifest_row.get("case_json", ""))
        destination_file = Path(str(r.get("destination_file", "")))
        xlsx_path = destination_file if destination_file.exists() else None
        if xlsx_path is None:
            summary_rows.append(
                {
                    "case_id": case_id,
                    "run_order": r.get("run_order", ""),
                    "pack": r.get("pack", ""),
                    "status": "ERROR",
                    "error": "routed strategy_report xlsx missing",
                }
            )
            continue

        print(f"[{idx}/{len(matched_rows)}] case={case_id}")
        try:
            tv_csv = xlsx_path.with_name(f"{xlsx_path.stem.replace('_strategy_report', '')}_trades.csv")
            tv_trade_count = extract_tv_trades_csv(xlsx_path, tv_csv)
            tv_metrics = parse_tv_metrics(xlsx_path)
            setup_expected = parse_ui_actions_to_expectations(manifest_row.get("ui_actions", ""))
            setup_actual = read_properties(xlsx_path)
            baseline_drift_keys = compute_baseline_drift(
                baseline_props=baseline_props,
                actual_props=setup_actual,
                setup_expected=setup_expected,
            )
            setup_mismatch_keys: list[str] = []
            setup_matched_count = 0
            for key, expected_val in setup_expected.items():
                actual_val = setup_actual.get(key, "")
                if actual_val == expected_val:
                    setup_matched_count += 1
                else:
                    setup_mismatch_keys.append(f"{key}:exp={expected_val}|got={actual_val}")

            if setup_expected and setup_mismatch_keys:
                summary_rows.append(
                    {
                        "case_id": case_id,
                        "run_order": parse_int(r.get("run_order", 0), 0),
                        "pack": r.get("pack", ""),
                        "status": "SETUP_MISMATCH",
                        "setup_status": "SETUP_MISMATCH",
                        "setup_expected_count": len(setup_expected),
                        "setup_matched_count": setup_matched_count,
                        "setup_mismatch_keys": " || ".join(setup_mismatch_keys),
                        "baseline_anchor_status": "SKIPPED_DUE_TO_SETUP_MISMATCH",
                        "baseline_drift_count": len(baseline_drift_keys),
                        "baseline_drift_keys": " || ".join(baseline_drift_keys),
                        "tv_trades": int(tv_trade_count),
                        "py_trades": "",
                        "trade_delta": "",
                        "compared_rows": "",
                        "full_core_match": "",
                        "mismatches": "",
                        "mismatch_ratio": "",
                        "tv_trade_count_raw": tv_trade_count,
                        "tv_net_profit": tv_metrics.get("tv_net_profit"),
                        "py_net_profit": "",
                        "tv_percent_profitable": tv_metrics.get("tv_percent_profitable"),
                        "py_win_rate": "",
                        "tv_profit_factor": tv_metrics.get("tv_profit_factor"),
                        "py_profit_factor": "",
                        "tv_xlsx": str(xlsx_path),
                        "tv_csv": str(tv_csv),
                        "py_csv": "",
                        "report_csv": "",
                        "error": "",
                    }
                )
                continue

            if baseline_drift_keys:
                summary_rows.append(
                    {
                        "case_id": case_id,
                        "run_order": parse_int(r.get("run_order", 0), 0),
                        "pack": r.get("pack", ""),
                        "status": "BASELINE_DRIFT",
                        "setup_status": "SETUP_OK" if setup_expected else "SETUP_UNCHECKED",
                        "setup_expected_count": len(setup_expected),
                        "setup_matched_count": setup_matched_count,
                        "setup_mismatch_keys": "",
                        "baseline_anchor_status": "BASELINE_DRIFT",
                        "baseline_drift_count": len(baseline_drift_keys),
                        "baseline_drift_keys": " || ".join(baseline_drift_keys),
                        "tv_trades": int(tv_trade_count),
                        "py_trades": "",
                        "trade_delta": "",
                        "compared_rows": "",
                        "full_core_match": "",
                        "mismatches": "",
                        "mismatch_ratio": "",
                        "tv_trade_count_raw": tv_trade_count,
                        "tv_net_profit": tv_metrics.get("tv_net_profit"),
                        "py_net_profit": "",
                        "tv_percent_profitable": tv_metrics.get("tv_percent_profitable"),
                        "py_win_rate": "",
                        "tv_profit_factor": tv_metrics.get("tv_profit_factor"),
                        "py_profit_factor": "",
                        "tv_xlsx": str(xlsx_path),
                        "tv_csv": str(tv_csv),
                        "py_csv": "",
                        "report_csv": "",
                        "error": "",
                    }
                )
                continue

            results = run_case_module.run(str(case_json))
            py_metrics = results.get("metrics", {})
            debug_exports = results.get("debug_exports", {})
            py_csv_rel = str(debug_exports.get("debug_python_trades", "")).strip()
            py_csv = (workspace_root / py_csv_rel).resolve() if py_csv_rel else Path()
            if not py_csv.exists():
                raise RuntimeError(f"python trades csv missing: {py_csv}")

            case_payload = json.loads(case_json.read_text(encoding="utf-8"))
            tv_tz = str(case_payload.get("tv_tz", "Europe/London"))

            tv_df = load_tv_trades(tv_csv, tv_tz=tv_tz, tv_shift_min=0)
            py_df = load_py_trades(py_csv, py_shift_min=0)
            report = build_report(tv_df, py_df)

            compared = int(len(report))
            full_match = int(report["all_core_match"].sum()) if compared > 0 else 0
            mismatches = compared - full_match
            mismatch_ratio = float(mismatches / compared) if compared > 0 else 0.0
            trade_delta = int(len(py_df) - len(tv_df))
            status = "PASS" if trade_delta == 0 and mismatches == 0 else "MISMATCH"

            report_csv = reports_dir / f"{case_id}_report.csv"
            report.to_csv(report_csv, index=False)

            summary_rows.append(
                {
                    "case_id": case_id,
                    "run_order": parse_int(r.get("run_order", 0), 0),
                    "pack": r.get("pack", ""),
                    "status": status,
                    "setup_status": "SETUP_OK" if setup_expected else "SETUP_UNCHECKED",
                    "setup_expected_count": len(setup_expected),
                    "setup_matched_count": setup_matched_count,
                    "setup_mismatch_keys": "",
                    "baseline_anchor_status": "BASELINE_OK",
                    "baseline_drift_count": 0,
                    "baseline_drift_keys": "",
                    "tv_trades": int(len(tv_df)),
                    "py_trades": int(len(py_df)),
                    "trade_delta": trade_delta,
                    "compared_rows": compared,
                    "full_core_match": full_match,
                    "mismatches": mismatches,
                    "mismatch_ratio": mismatch_ratio,
                    "tv_trade_count_raw": tv_trade_count,
                    "tv_net_profit": tv_metrics.get("tv_net_profit"),
                    "py_net_profit": py_metrics.get("net_profit"),
                    "tv_percent_profitable": tv_metrics.get("tv_percent_profitable"),
                    "py_win_rate": py_metrics.get("win_rate"),
                    "tv_profit_factor": tv_metrics.get("tv_profit_factor"),
                    "py_profit_factor": py_metrics.get("profit_factor"),
                    "tv_xlsx": str(xlsx_path),
                    "tv_csv": str(tv_csv),
                    "py_csv": str(py_csv),
                    "report_csv": str(report_csv),
                    "error": "",
                }
            )
        except Exception as exc:
            summary_rows.append(
                {
                    "case_id": case_id,
                    "run_order": parse_int(r.get("run_order", 0), 0),
                    "pack": r.get("pack", ""),
                    "status": "ERROR",
                    "setup_status": "",
                    "setup_expected_count": "",
                    "setup_matched_count": "",
                    "setup_mismatch_keys": "",
                    "baseline_anchor_status": "",
                    "baseline_drift_count": "",
                    "baseline_drift_keys": "",
                    "error": str(exc),
                }
            )

    summary_df = pd.DataFrame(summary_rows)
    summary_csv = compare_dir / "matched_batch_compare_summary.csv"
    summary_df.to_csv(summary_csv, index=False)

    stats = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "suite_root": str(suite_root),
        "matched_input_count": len(matched_rows),
        "processed_count": int(len(summary_df)),
        "pass": int((summary_df["status"] == "PASS").sum()) if "status" in summary_df.columns else 0,
        "baseline_drift": int((summary_df["status"] == "BASELINE_DRIFT").sum()) if "status" in summary_df.columns else 0,
        "mismatch": int((summary_df["status"] == "MISMATCH").sum()) if "status" in summary_df.columns else 0,
        "error": int((summary_df["status"] == "ERROR").sum()) if "status" in summary_df.columns else 0,
        "baseline_xlsx_used": str(baseline_xlsx_path) if baseline_xlsx_path else "",
        "summary_csv": str(summary_csv),
        "reports_dir": str(reports_dir),
    }
    (compare_dir / "matched_batch_compare_summary.json").write_text(
        json.dumps(stats, ensure_ascii=True, indent=2),
        encoding="utf-8",
    )

    print(f"matched_input_count={stats['matched_input_count']}")
    print(f"processed_count={stats['processed_count']}")
    print(f"pass={stats['pass']}")
    print(f"baseline_drift={stats['baseline_drift']}")
    print(f"mismatch={stats['mismatch']}")
    print(f"error={stats['error']}")
    print(f"summary_csv={summary_csv}")
    print(f"compare_dir={compare_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
