#!/usr/bin/env python3
"""
Rebuild 'Retest Candidates' sheet from true TV trade counts.

Source of truth:
  compare_runs/true_no_effect_cases.csv

Behavior:
  - Keeps the 'Cases' sheet untouched
  - Recreates 'Retest Candidates' with the same base columns
  - Reuses prior suggestion text where the run_order still exists
  - Adds guidance for newly discovered no-effect cases
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\LAB\tradingview-lab\mtc_backtest\parity_suite_350")
XLSX_PATH = BASE / "CASE_SETUP_GUIDE_tagged_v6_conflict_candidates.xlsx"
CSV_PATH = BASE / "compare_runs" / "true_no_effect_cases.csv"


EXTRA_HEADERS = [
    "issue_group",
    "current_trade_count",
    "suggested_value",
    "suggestion_reason",
    "retest_status",
]


GROUP_FILLS = {
    "RANGE_FILTER_NO_EFFECT": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "MACD_HUB_NO_EFFECT": PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
    "SWING_CONFIRM_NO_EFFECT": PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid"),
    "DEPENDENCY_NO_EFFECT": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
    "METRICS_ONLY_EXPECTED": PatternFill(start_color="E7F6E7", end_color="E7F6E7", fill_type="solid"),
    "LOW_SAMPLE_NO_EFFECT": PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
    "MAE_GUARD_NO_EFFECT": PatternFill(start_color="F4E1FF", end_color="F4E1FF", fill_type="solid"),
    "MCGINLEY_NO_EFFECT": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
}


GROUP_ORDER = {
    "RANGE_FILTER_NO_EFFECT": 1,
    "MACD_HUB_NO_EFFECT": 2,
    "SWING_CONFIRM_NO_EFFECT": 3,
    "DEPENDENCY_NO_EFFECT": 4,
    "MAE_GUARD_NO_EFFECT": 5,
    "MCGINLEY_NO_EFFECT": 6,
    "LOW_SAMPLE_NO_EFFECT": 7,
    "METRICS_ONLY_EXPECTED": 8,
    "BASELINE_MATCH_SUSPICIOUS": 9,
}


FORCE_RECLASSIFY = {"326", "327", "328"}


def load_true_no_effect_rows() -> list[dict]:
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        row["run_order"] = str(row["run_order"]).zfill(3)
        row["tv_trade_count"] = int(row["tv_trade_count"])
    return rows


def load_existing_retest_map(wb: openpyxl.Workbook) -> Dict[str, Tuple[str, str, str]]:
    if "Retest Candidates" not in wb.sheetnames:
        return {}
    ws = wb["Retest Candidates"]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    old: Dict[str, Tuple[str, str, str]] = {}
    for r in range(2, ws.max_row + 1):
        ro = ws.cell(r, idx["run_order"]).value
        if ro is None:
            continue
        ro = str(ro).zfill(3)
        old[ro] = (
            ws.cell(r, idx["issue_group"]).value or "",
            ws.cell(r, idx["suggested_value"]).value or "",
            ws.cell(r, idx["suggestion_reason"]).value or "",
        )
    return old


def classify_and_suggest(run_order: str, primary_name: str, primary_value: str) -> Tuple[str, str, str]:
    if primary_name.startswith("[RF]"):
        if primary_name == "[RF] RSI Overbought" and primary_value == "60":
            return (
                "RANGE_FILTER_NO_EFFECT",
                "90",
                "Push overbought threshold to an extreme. 60/70/80 all held 144 TV trades; 90 should separate behavior if this input is live.",
            )
        return ("RANGE_FILTER_NO_EFFECT", "", "Retest with a more extreme threshold/length than the current RF sweep.")

    if primary_name == "Range Aggregation":
        return (
            "DEPENDENCY_NO_EFFECT",
            "COUNT + ADX+CHOP ON",
            "AND vs COUNT often stays identical when only one range sub-filter is effectively active. Retest after enabling both ADX and Choppiness filters together.",
        )

    if primary_name == "Max Pyramid Positions":
        return (
            "DEPENDENCY_NO_EFFECT",
            "Pair with Max Entries=3",
            "Pyramiding slots alone may stay inert when Signal Mode Max Entries remains 1. Retest together with a higher signal-mode entry allowance.",
        )

    if primary_name in {
        "Risk per Long (% of equity)",
        "Risk per Short (% of equity)",
        "Fallback Position Size When SL OFF (% of equity)",
        "Max Leverage Cap (Limit)",
        "TP1 Close % of position",
    }:
        return (
            "METRICS_ONLY_EXPECTED",
            "P&L / qty check",
            "This input primarily changes position sizing or partial-close distribution. Same trade count is expected; validate PnL, qty, and drawdown instead.",
        )

    if primary_name == "Max MAE %":
        return (
            "MAE_GUARD_NO_EFFECT",
            "0.1",
            "Current MAE sweep still matches baseline on TV. Retest with an extreme threshold to force visible guard blocking if the path is active.",
        )

    if primary_name in {
        "SL Mode",
        "Swing SL Basis",
        "Swing SL ATR Length",
        "Swing SL ATR Multiplier",
        "TP ATR Length",
        "Trailing ATR Length",
    }:
        return (
            "LOW_SAMPLE_NO_EFFECT",
            "More extreme + longer window",
            "This branch produced too few trades to separate variants reliably. Expand the date window or use a much more extreme value.",
        )

    if primary_name in {"[McGinley] Length", "[McGinley] HTF Timeframe"}:
        return (
            "MCGINLEY_NO_EFFECT",
            "Extreme TF / length",
            "McGinley settings stayed on the same TV count. Retest with a much larger length or a clearly different HTF timeframe.",
        )

    if primary_name in {"Mode", "Signal", "Min distance (0=no effect)"}:
        return (
            "MACD_HUB_NO_EFFECT",
            "Use a different MACD family",
            "These specific MACD variants stayed identical in TV. Retest against a materially different mode or a much larger distance gate.",
        )

    if primary_name in {
        "If LONG & SHORT raw same bar",
        "Break buffer (ticks)",
        "Max swing distance % (0 = off)",
    }:
        return (
            "SWING_CONFIRM_NO_EFFECT",
            "More extreme confirmation value",
            "Current confirmation sweep did not move TV trade count. Use a materially tighter/looser value to force divergence.",
        )

    return (
        "METRICS_ONLY_EXPECTED",
        "Manual review",
        "No automatic suggestion was mapped for this input; review whether it should change trade count or only PnL/qty.",
    )


def main() -> int:
    rows = load_true_no_effect_rows()
    wb = openpyxl.load_workbook(XLSX_PATH)
    existing_retest = load_existing_retest_map(wb)
    ws_cases = wb["Cases"]

    case_headers = [c.value for c in ws_cases[1] if c.value]
    case_header_idx = {h: i + 1 for i, h in enumerate([c.value for c in ws_cases[1]]) if h}

    case_rows: Dict[str, list] = {}
    for r in range(2, ws_cases.max_row + 1):
        ro = ws_cases.cell(r, case_header_idx["run_order"]).value
        if ro is None:
            continue
        ro = str(ro).zfill(3)
        case_rows[ro] = [ws_cases.cell(r, case_header_idx[h]).value for h in case_headers]

    if "Retest Candidates" in wb.sheetnames:
        del wb["Retest Candidates"]
    ws = wb.create_sheet("Retest Candidates")

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    all_headers = case_headers + EXTRA_HEADERS
    for c, header in enumerate(all_headers, 1):
        cell = ws.cell(1, c, header)
        cell.fill = header_fill
        cell.font = header_font

    output_rows = []
    for row in rows:
        ro = row["run_order"]
        base_row = case_rows.get(ro)
        if not base_row:
            continue
        if ro in existing_retest and ro not in FORCE_RECLASSIFY:
            issue_group, suggested_value, suggestion_reason = existing_retest[ro]
        else:
            issue_group, suggested_value, suggestion_reason = classify_and_suggest(
                ro, row["primary_name"], row["primary_value"]
            )
        output_rows.append(
            (
                ro,
                base_row
                + [
                    issue_group,
                    row["tv_trade_count"],
                    suggested_value,
                    suggestion_reason,
                    "",
                ],
            )
        )

    output_rows.sort(key=lambda x: (GROUP_ORDER.get(x[1][len(case_headers)], 99), int(x[0])))

    for r_idx, (_, row_data) in enumerate(output_rows, 2):
        group = row_data[len(case_headers)]
        fill = GROUP_FILLS.get(group)
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx, value)
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{ws.max_row}"

    widths = {
        "run_order": 10,
        "package": 10,
        "case_id": 55,
        "case_folder": 55,
        "tv_preset_name": 55,
        "primary_change": 40,
        "tv_ui_actions": 65,
        "dependency_note": 58,
        "expected_trade_behavior": 16,
        "tv_xlsx_status": 20,
        "tv_xlsx_file": 45,
        "tv_download_date": 15,
        "setup_check": 12,
        "compare_status": 14,
        "notes": 28,
        "exec_conflict_tag": 16,
        "exec_conflict_ready": 16,
        "issue_group": 28,
        "current_trade_count": 18,
        "suggested_value": 26,
        "suggestion_reason": 78,
        "retest_status": 14,
    }
    for c, header in enumerate(all_headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(header, 15)

    wb.save(XLSX_PATH)
    wb.close()

    print(f"Rebuilt 'Retest Candidates' with {len(output_rows)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
