from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
TRACKER_PATH = REPO_ROOT / "01_MASTER TEMPLATE_V2" / "05_PARITY" / "MTC_V2_TW_EXPORT_CASE_TRACKER.xlsx"


@dataclass(slots=True)
class TrackerCase:
    case_no: str
    phase: str
    title: str
    description: str
    folder: str
    status: str
    export_file: str
    row_index: int
    raw: dict[str, Any]


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def read_tracker(path: Path = TRACKER_PATH) -> list[TrackerCase]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [_norm(c.value) for c in ws[1]]
    cases: list[TrackerCase] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        case_no = _norm(data.get("case_no"))
        if not case_no:
            continue
        cases.append(
            TrackerCase(
                case_no=case_no,
                phase=_norm(data.get("phase")),
                title=_norm(data.get("description") or data.get("title")),
                description=_norm(data.get("description")),
                folder=_norm(data.get("folder")),
                status=_norm(data.get("status")),
                export_file=_norm(data.get("export_file")),
                row_index=row_idx,
                raw=data,
            )
        )
    return cases


def main() -> int:
    cases = read_tracker()
    print(json.dumps([asdict(c) for c in cases[:5]], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
