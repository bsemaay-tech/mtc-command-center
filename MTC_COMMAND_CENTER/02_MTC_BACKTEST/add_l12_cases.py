"""
add_l12_cases.py
----------------
Adds L12 MA filter + MA slope filter case rows to the parity tracker.
- Adds new columns: use_ma_filter, ma_type, ma_length,
                    use_ma_slope_filter, ma_slope_len, ma_slope_min_pct
- Adds AUTO_017 ... AUTO_024 case rows
- Updates both MTC_V2_PARITY_CASES.csv and MTC_V2_PARITY_CASE_TRACKER.xlsx

Run once:  python add_l12_cases.py
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKTEST_ROOT = REPO_ROOT / "MTC_COMMAND_CENTER/02_MTC_BACKTEST"
TRACKER_CSV = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASES.csv"
TRACKER_XLSX = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASE_TRACKER.xlsx"
REPORTS_DIR = BACKTEST_ROOT / "reports/tracker_cases"

# New columns to add (inserted after use_notional_assert)
NEW_COLS = [
    "use_ma_filter",
    "ma_type",
    "ma_length",
    "use_ma_slope_filter",
    "ma_slope_len",
    "ma_slope_min_pct",
]

# Pine defaults for the new columns
MA_DEFAULTS = {
    "use_ma_filter": "False",
    "ma_type": "EMA",
    "ma_length": "200",
    "use_ma_slope_filter": "False",
    "ma_slope_len": "200",
    "ma_slope_min_pct": "0.005",
}

# ── Case definitions ──────────────────────────────────────────────────────────

# Common base: same as AUTO_001 defaults
BASE = dict(
    layer="L12",
    case_package="FILTER_L12_MA",
    case_type="AUTO_ACTIVE",
    status="NOT_RUN",
    depends_on="AUTO_001",
    symbol="BTCUSDT",
    timeframe="1h",
    bars="1000",
    data_source="Binance via mtc_bridge.mjs / PineTS exact bars",
    enable_long="True",
    enable_short="True",
    allow_flip="True",
    regime_lock="False",
    max_entries="1",
    cooldown_bars="0",
    signal_mode="Supertrend",
    st_atr_len="21",
    st_factor="4.0",
    st_use_wicks="False",
    st_use_ha="False",
    equity_source="Realized",
    use_notional_assert="False",
    use_ma_filter="False",
    ma_type="EMA",
    ma_length="200",
    use_ma_slope_filter="False",
    ma_slope_len="200",
    ma_slope_min_pct="0.005",
    risk_per_long_pct="1.0",
    risk_per_short_pct="1.0",
    fallback_size_pct="10.0",
    max_leverage_cap="1.0",
    sl_mode="ATR",
    sl_atr_len="14",
    sl_atr_mult="2.0",
    sl_pct="",
    tp_mode="None",
    tp_atr_len="",
    tp_atr_mult="",
    llm_auto_status="",
    llm_last_run_utc="",
    llm_agent="",
    pine_trades="",
    pine_win_rate_pct="",
    pine_net_pnl_pct="",
    pine_max_dd_pct="",
    python_trades="",
    python_win_rate_pct="",
    python_net_pnl_pct="",
    python_max_dd_pct="",
    parity_verdict="NOT_RUN",
    artifacts="",
    manual_tw_status="",
    manual_tw_date="",
    manual_tw_xlsx="",
    manual_tw_trade_count="",
    manual_tw_win_rate_pct="",
    manual_tw_net_pnl_pct="",
    manual_tw_max_dd_pct="",
    manual_tw_verdict="",
)

def _case(run_order, case_id, preset, overrides, primary_change, expected, notes):
    c = dict(BASE)
    c.update(overrides)
    c.update(dict(
        run_order=str(run_order),
        case_id=case_id,
        pine_preset_name=preset,
        primary_change=primary_change,
        expected_trade_behavior=expected,
        notes=notes,
    ))
    return c


L12_CASES = [
    _case(30, "AUTO_017", "auto_017_ma_filter_ema200",
          {"use_ma_filter": "True", "ma_type": "EMA", "ma_length": "200"},
          "MA filter ON (EMA 200)",
          "Fewer trades than baseline — only entries where price is on correct side of EMA200",
          "L12 MA filter toggle. Gate: price > EMA200 for longs, < EMA200 for shorts."),
    _case(31, "AUTO_018", "auto_018_ma_filter_sma200",
          {"use_ma_filter": "True", "ma_type": "SMA", "ma_length": "200"},
          "MA filter ON (SMA 200)",
          "Different trade count vs AUTO_017 — SMA reacts slower than EMA",
          "L12 MA type variant: SMA200 vs EMA200."),
    _case(32, "AUTO_019", "auto_019_ma_filter_wma200",
          {"use_ma_filter": "True", "ma_type": "WMA", "ma_length": "200"},
          "MA filter ON (WMA 200)",
          "Different trade count vs AUTO_017/018 — WMA weights recent bars more",
          "L12 MA type variant: WMA200."),
    _case(33, "AUTO_020", "auto_020_ma_filter_rma200",
          {"use_ma_filter": "True", "ma_type": "RMA", "ma_length": "200"},
          "MA filter ON (RMA 200)",
          "Smooth MA — laggier than EMA, different gate timing",
          "L12 MA type variant: RMA200 (Wilder's MA)."),
    _case(34, "AUTO_021", "auto_021_ma_filter_ema50",
          {"use_ma_filter": "True", "ma_type": "EMA", "ma_length": "50"},
          "MA filter ON (EMA 50)",
          "More trades than EMA200 — shorter MA crosses price more often",
          "L12 MA length variant: EMA50 vs EMA200."),
    _case(35, "AUTO_022", "auto_022_slope_filter_default",
          {"use_ma_slope_filter": "True", "ma_slope_len": "200", "ma_slope_min_pct": "0.005"},
          "MA slope filter ON (default threshold 0.005)",
          "Entries blocked when MA200 slope is flat — fewer trades than baseline",
          "L12 slope filter toggle with default threshold."),
    _case(36, "AUTO_023", "auto_023_slope_filter_loose",
          {"use_ma_slope_filter": "True", "ma_slope_len": "200", "ma_slope_min_pct": "0.001"},
          "MA slope filter ON (loose threshold 0.001)",
          "More trades than AUTO_022 — lower slope requirement lets more entries through",
          "L12 slope threshold variant: 0.001 vs default 0.005."),
    _case(37, "AUTO_024", "auto_024_ma_and_slope_combined",
          {"use_ma_filter": "True", "ma_type": "EMA", "ma_length": "200",
           "use_ma_slope_filter": "True", "ma_slope_len": "200", "ma_slope_min_pct": "0.005"},
          "MA filter + slope filter combined",
          "Fewest trades — both gates must pass simultaneously",
          "L12 combined filter. Price side gate AND slope gate both active."),
]


# ── Override JSON files ──────────────────────────────────────────────────────

# Keys to include in pine_input_overrides.json (only the changed values vs Pine defaults)
MA_OVERRIDE_KEYS = {
    "use_ma_filter": ("bool", False),
    "ma_type": ("str", "EMA"),
    "ma_length": ("int", 200),
    "use_ma_slope_filter": ("bool", False),
    "ma_slope_len": ("int", 200),
    "ma_slope_min_pct": ("float", 0.005),
}


def _coerce(raw: str, kind: str) -> object:
    if kind == "bool":
        return raw.strip().lower() == "true"
    if kind == "int":
        return int(raw)
    if kind == "float":
        return float(raw)
    return raw.strip()


def create_override_json(case: dict) -> dict:
    """Only include fields that differ from their Pine default."""
    overrides = {}
    for key, (kind, default) in MA_OVERRIDE_KEYS.items():
        raw = case.get(key, "")
        if raw == "":
            continue
        value = _coerce(raw, kind)
        if value != default:
            overrides[key] = value
    return overrides


# ── CSV update ───────────────────────────────────────────────────────────────

def update_csv(new_cases: list[dict]) -> None:
    with TRACKER_CSV.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        original_fieldnames = list(reader.fieldnames or [])
        existing_rows = list(reader)

    # Guard: skip if already added
    existing_ids = {r["case_id"] for r in existing_rows}
    if "AUTO_017" in existing_ids:
        print("  CSV: AUTO_017 already exists — skipping CSV update")
        return

    # Insert new columns after use_notional_assert
    new_fieldnames = list(original_fieldnames)
    if NEW_COLS[0] not in new_fieldnames:
        insert_after = "use_notional_assert"
        try:
            idx = new_fieldnames.index(insert_after)
            for i, col in enumerate(NEW_COLS):
                new_fieldnames.insert(idx + 1 + i, col)
        except ValueError:
            new_fieldnames.extend(NEW_COLS)

    # Fill existing rows with MA defaults
    for row in existing_rows:
        for col in NEW_COLS:
            if col not in row:
                row[col] = MA_DEFAULTS.get(col, "")

    # Build new rows aligned to fieldnames
    rows_to_add = []
    for c in new_cases:
        row = {fn: c.get(fn, "") for fn in new_fieldnames}
        rows_to_add.append(row)

    all_rows = existing_rows + rows_to_add

    with TRACKER_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=new_fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"  CSV: added {len(rows_to_add)} rows, {len(NEW_COLS)} new columns")


# ── XLSX update ──────────────────────────────────────────────────────────────

FILL_NOT_RUN = PatternFill("solid", start_color="FFF2CC")   # yellow tint
FILL_PARAM   = PatternFill("solid", start_color="D9E8FB")   # blue tint
DATA_FONT    = Font(name="Arial", size=9)
CENTER       = Alignment(horizontal="center")
LEFT         = Alignment(horizontal="left")


def update_xlsx(new_cases: list[dict]) -> None:
    wb = load_workbook(TRACKER_XLSX)
    if "Cases" not in wb.sheetnames:
        print("  XLSX: 'Cases' sheet not found — skipping XLSX update")
        return

    ws = wb["Cases"]
    headers = [cell.value for cell in ws[1]]

    # Guard: skip if already added
    case_id_col = headers.index("case_id") + 1 if "case_id" in headers else None
    if case_id_col:
        for row_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row_idx, case_id_col).value or "") == "AUTO_017":
                print("  XLSX: AUTO_017 already exists — skipping XLSX update")
                return

    # Add new columns to header row if missing
    new_col_positions: dict[str, int] = {}
    for col_name in NEW_COLS:
        if col_name not in headers:
            next_col = len(headers) + 1
            # Insert after use_notional_assert
            if "use_notional_assert" in headers:
                ins_after = headers.index("use_notional_assert") + 1  # 0-based
                # shift existing columns right
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    pass  # we'll just append at the end for simplicity
                # Append at end (simpler and avoids shifting)
            ws.cell(row=1, column=next_col, value=col_name)
            ws.cell(row=1, column=next_col).font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            ws.cell(row=1, column=next_col).fill = PatternFill("solid", start_color="1F4E79")
            ws.cell(row=1, column=next_col).alignment = CENTER
            ws.column_dimensions[get_column_letter(next_col)].width = 14
            headers.append(col_name)
            new_col_positions[col_name] = next_col

    col_map = {str(h): i + 1 for i, h in enumerate(headers) if h}

    # Append new case rows
    added = 0
    for c in new_cases:
        row_idx = ws.max_row + 1
        for col_name, col_idx in col_map.items():
            value = c.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = CENTER if col_name not in ("primary_change", "expected_trade_behavior", "notes", "case_id", "case_package") else LEFT
            # Highlight NOT_RUN status
            if col_name == "status":
                cell.fill = FILL_NOT_RUN
            elif col_name in NEW_COLS and value not in ("", "False", "EMA", "200", "0.005"):
                cell.fill = FILL_PARAM
        ws.row_dimensions[row_idx].height = 14
        added += 1

    wb.save(TRACKER_XLSX)
    print(f"  XLSX: added {added} rows, {len(new_col_positions)} new columns")


# ── Override JSON files ───────────────────────────────────────────────────────

def create_override_files(new_cases: list[dict]) -> None:
    for case in new_cases:
        case_id = case["case_id"]
        overrides = create_override_json(case)
        out_dir = REPORTS_DIR / case_id
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / "pine_input_overrides.json"
        if out_file.exists():
            print(f"  JSON: {case_id}/pine_input_overrides.json already exists — skipping")
            continue
        with out_file.open("w", encoding="utf-8") as fh:
            json.dump(overrides, fh, indent=2)
        print(f"  JSON: {case_id}/pine_input_overrides.json  = {overrides}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=== add_l12_cases.py ===\n")

    print(f"Adding {len(L12_CASES)} L12 cases: AUTO_017 to AUTO_024\n")

    print("1. Updating CSV...")
    update_csv(L12_CASES)

    print("\n2. Updating XLSX...")
    update_xlsx(L12_CASES)

    print("\n3. Creating override JSON files...")
    create_override_files(L12_CASES)

    print("\nDone.")
    print("\nTo run a case:")
    print("  python parity_compare.py --fetch-fresh \\")
    print("    --case reports/tracker_cases/AUTO_017/pine_input_overrides.json \\")
    print("    --tracker-case AUTO_017")
    print("\nOr run all 8 sequentially:")
    for c in L12_CASES:
        print(f"  python parity_compare.py --fetch-fresh --case reports/tracker_cases/{c['case_id']}/pine_input_overrides.json --tracker-case {c['case_id']}")


if __name__ == "__main__":
    main()
