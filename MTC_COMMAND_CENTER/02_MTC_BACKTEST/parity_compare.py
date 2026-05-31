"""
parity_compare.py - PineTS vs MTC_V2 Python parity check for the current V2 dev loop.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
BACKTEST_ROOT = REPO_ROOT / "MTC_COMMAND_CENTER/02_MTC_BACKTEST"
V2_PYTHON_ROOT = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/00_PYTHON"
PINE_SOURCE = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/01_PINE/MTC_V2.pine"
TRACKER_WORKBOOK_PATH = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASE_TRACKER.xlsx"
TRACKER_WORKBOOK_FALLBACK_PATH = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASE_TRACKER_close_only_wave1_pending.xlsx"
TRACKER_CSV_PATH = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASES.csv"
REPORTS_DIR = BACKTEST_ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

sys.path.insert(0, str(V2_PYTHON_ROOT))

from mtc_v2.core.config import resolve_config  # noqa: E402
from mtc_v2.core.runner import Runner  # noqa: E402
from mtc_v2.core.types import Bar  # noqa: E402


METRIC_TOLERANCES: dict[str, float] = {
    "total_trades": 0.0,
    "win_rate_pct": 0.25,
    "profit_factor": 0.02,
    "net_pnl_pct": 0.05,
    "max_drawdown_pct": 0.05,
}

PINE_REASON_MAP = {
    "sl_atr_hit": "SL",
    "sl_percent_hit": "SL",
    "sl_swing_atr_hit": "SL",
    "tp_atr_hit": "TP",
    "tp_percent_hit": "TP",
    "tp_r_hit": "TP",
    "tp1_hit": "TP1",
    "tp2_hit": "TP2",
    "be_hit": "BE",
    "trail_hit": "TRAIL",
    "opp_signal": "OPP_SIGNAL",
    "terminal_close": "MANUAL",
    "price_exit_ambiguous": "SL",
    "filter_block": "FILTER_BLOCK",
    "time_stop": "TIME_STOP",
    "eod_exit": "EOD_EXIT",
    "eow_exit": "EOW_EXIT",
}

PYTHON_REASON_MAP = {
    "sl_atr_hit": "SL",
    "sl_percent_hit": "SL",
    "sl_swing_atr_hit": "SL",
    "tp_atr_hit": "TP",
    "tp_percent_hit": "TP",
    "tp_r_hit": "TP",
    "tp1_hit": "TP1",
    "tp2_hit": "TP2",
    "be_hit": "BE",
    "trail_hit": "TRAIL",
    "opp_signal": "OPP_SIGNAL",
    "terminal_close": "MANUAL",
    "filter_block": "FILTER_BLOCK",
    "time_stop": "TIME_STOP",
    "eod_exit": "EOD_EXIT",
    "eow_exit": "EOW_EXIT",
    "eod": "EOD_EXIT",
    "eow": "EOW_EXIT",
}

SUPPORTED_PINE_INPUTS: dict[str, str] = {
    "enable_long": "bool",
    "enable_short": "bool",
    "allow_flip": "bool",
    "regime_lock": "bool",
    "max_entries": "int",
    "cooldown_bars": "int",
    "signal_mode": "string",
    "st_atr_len": "int",
    "st_factor": "float",
    "st_use_wicks": "bool",
    "st_use_ha": "bool",
    "risk_per_long_pct": "float",
    "risk_per_short_pct": "float",
    "fallback_size_pct": "float",
    "max_leverage_cap": "float",
    "use_ma_filter": "bool",
    "ma_type": "string",
    "ma_length": "int",
    "use_ma_mtf": "bool",
    "ma_htf_timeframe": "string",
    "use_ma_slope_filter": "bool",
    "ma_slope_len": "int",
    "ma_slope_min_pct": "float",
    "sl_mode": "string",
    "sl_atr_len": "int",
    "sl_atr_mult": "float",
    "sl_pct": "float",
    "sl_swing_basis": "string",
    "sl_swing_lookback": "int",
    "sl_swing_atr_len": "int",
    "sl_swing_atr_mult": "float",
    "tp_mode": "string",
    "tp_atr_len": "int",
    "tp_atr_mult": "float",
    "tp_percent": "float",
    "tp_r_multiple": "float",
    "tp1_r_multiple": "float",
    "tp1_close_pct": "float",
    "tp2_r_multiple": "float",
    "use_break_even": "bool",
    "be_trigger_r": "float",
    "be_buffer_r": "float",
    "use_trailing": "bool",
    "trail_atr_len": "int",
    "trail_start_r": "float",
    "trail_distance_atr_mult": "float",
    # McGinley filter
    "use_mcginley_filter": "bool",
    "mcginley_length": "int",
    "mcginley_use_higher_timeframe": "bool",
    "mcginley_htf_timeframe": "string",
    # Volume filter
    "use_volume_filter": "bool",
    "vol_sma_length": "int",
    "vol_sma_mult": "float",
    # ADX filter
    "use_adx_filter": "bool",
    "adx_length": "int",
    "adx_threshold": "float",
    "adx_use_higher_timeframe": "bool",
    "adx_htf_timeframe": "string",
    # Chop filter
    "use_chop_filter": "bool",
    "chop_length": "int",
    "chop_threshold": "float",
    "chop_use_higher_timeframe": "bool",
    "chop_htf_timeframe": "string",
    # ATR Vol Floor filter
    "use_atr_vol_floor": "bool",
    "atr_vol_floor_fast_len": "int",
    "atr_vol_floor_baseline_len": "int",
    "atr_vol_floor_mult": "float",
    # MACD filters
    "use_macd_regime_filter": "bool",
    "use_macd_cross_filter": "bool",
    "use_macd_hist_filter": "bool",
    "macd_hist_mode": "string",
    "use_macd_zero_dist_filter": "bool",
    "macd_zero_dist_min": "float",
    "macd_fast_len": "int",
    "macd_slow_len": "int",
    "macd_sig_len": "int",
    # L13 - opposite signal exit
    "exit_on_opposite_signal": "bool",
    # L14 - per-filter exits
    "exit_on_ma_block": "bool",
    "exit_on_ma_slope_block": "bool",
    "exit_on_mcginley_block": "bool",
    "exit_on_htf_trend_block": "bool",
    "exit_on_vol_block": "bool",
    "exit_on_atr_vol_block": "bool",
    "exit_on_range_block": "bool",
    "exit_on_candle_pattern_block": "bool",
    "exit_on_level_prox_block": "bool",
    # L15 - time-based exits
    "use_time_stop": "bool",
    "time_stop_bars": "int",
    "time_stop_condition": "string",
    "time_stop_eod": "bool",
    "time_stop_eow": "bool",
    # L16 - guards
    "use_daily_loss_limit": "bool",
    "max_daily_loss_pct": "float",
    "use_max_trades_per_day": "bool",
    "max_trades_per_day": "int",
    "use_max_drawdown_guard": "bool",
    "max_drawdown_pct": "float",
    "use_consecutive_loss_halt": "bool",
    "max_consecutive_losses": "int",
    "use_equity_curve_filter": "bool",
    "equity_ma_length": "int",
    "use_mae_guard": "bool",
    "max_mae_pct": "float",
    "use_guard_recovery": "bool",
    "guard_recovery_mode": "string",
    "guard_recovery_bars": "int",
    "guard_recovery_signals": "int",
    # L16 - Trade Cooldown After Exit
    "use_trade_cooldown": "bool",
    "cooldown_bars_after_exit": "int",
    # L18 - confirmation transform
    "use_confirm_transform": "bool",
    "confirm_bars": "int",
    "confirm_close_crosses": "bool",
    # L18b - advanced confirmation
    "require_raw_still_true": "bool",
    "refresh_on_new_raw": "bool",
    # L20 - level proximity gate
    "use_level_proximity_gate": "bool",
    "level_proximity_threshold_pct": "float",
    "level_proximity_lookback": "int",
    # L21 - level retest transform
    "use_level_retest": "bool",
    "retest_timeout_bars": "int",
    "retest_buffer_pct": "float",
    # L22 - candle pattern gate
    "use_candle_pattern_gate": "bool",
    "candle_pattern_lookback": "int",
    # L12 - HTF Trend Filter
    "use_htf_trend_filter": "bool",
    "htf_trend_timeframe": "string",
    "htf_trend_ma_type": "string",
    "htf_trend_ma_len": "int",
    "htf_trend_buffer_pct": "float",
    # L12 - MA Filter HTF path
    "use_ma_mtf": "bool",
    "ma_htf_timeframe": "string",
    # L12 - MACD HTF Bias + MACD Source
    "use_macd_htf_bias": "bool",
    "macd_htf_timeframe": "string",
    "macd_source": "string",
    # L12 - Momentum Filter
    "use_momentum_filter": "bool",
    "momentum_mode": "string",
    "momentum_atr_len": "int",
    "momentum_atr_mult": "float",
    "momentum_roc_min_pct": "float",
    # L12 - Session Filter
    "use_session_filter": "bool",
    "session_name": "string",
    "session_custom_string": "string",
}


def _extract_strategy_initial_capital(pine_text: str) -> float:
    match = re.search(r"strategy\([^)]*initial_capital\s*=\s*([0-9.]+)", pine_text, re.MULTILINE)
    if not match:
        raise ValueError("Could not find strategy initial_capital in MTC_V2.pine")
    return float(match.group(1))


def _extract_input_default(pine_text: str, var_name: str, input_type: str) -> Any:
    pattern = rf"^{re.escape(var_name)}\s*=\s*input\.{input_type}\(\s*(?P<value>(?:\"[^\"]*\"|[^,\n]+))"
    match = re.search(pattern, pine_text, re.MULTILINE)
    if not match:
        raise ValueError(f"Could not find Pine input default for '{var_name}'")

    raw = match.group("value").strip()
    if input_type == "bool":
        if raw == "true":
            return True
        if raw == "false":
            return False
        raise ValueError(f"Unexpected bool literal for {var_name}: {raw}")
    if input_type == "int":
        return int(raw)
    if input_type == "float":
        return float(raw)
    if input_type == "string":
        if not (raw.startswith('"') and raw.endswith('"')):
            raise ValueError(f"Unexpected string literal for {var_name}: {raw}")
        return raw[1:-1]
    raise ValueError(f"Unsupported Pine input type: {input_type}")


def extract_supported_pine_defaults(pine_path: Path) -> dict[str, Any]:
    pine_text = pine_path.read_text(encoding="utf-8")
    defaults: dict[str, Any] = {
        "initial_capital": _extract_strategy_initial_capital(pine_text),
    }
    for name, input_type in SUPPORTED_PINE_INPUTS.items():
        defaults[name] = _extract_input_default(pine_text, name, input_type)
    return defaults


def load_tracker_rows(csv_path: Path = TRACKER_CSV_PATH) -> list[dict[str, str]]:
    if not csv_path.exists():
        raise FileNotFoundError(f"Tracker CSV missing: {csv_path}")
    with csv_path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def load_tracker_case_row(case_id: str, csv_path: Path = TRACKER_CSV_PATH) -> dict[str, str]:
    target = str(case_id).strip()
    for row in load_tracker_rows(csv_path):
        if str(row.get("case_id", "")).strip() == target:
            return row
    raise ValueError(f"Tracker case not found: {case_id}")


def _parse_tracker_override(raw: str, input_type: str) -> Any:
    value = str(raw).strip()
    if value == "":
        raise ValueError("Empty tracker override")
    if input_type == "bool":
        lowered = value.lower()
        if lowered == "true":
            return True
        if lowered == "false":
            return False
        raise ValueError(f"Invalid bool tracker value: {value}")
    if input_type == "int":
        return int(float(value))
    if input_type == "float":
        return float(value)
    if input_type == "string":
        return value
    raise ValueError(f"Unsupported tracker input type: {input_type}")


def tracker_row_to_pine_overrides(row: dict[str, str]) -> dict[str, Any]:
    overrides: dict[str, Any] = {}
    for name, input_type in SUPPORTED_PINE_INPUTS.items():
        raw = row.get(name)
        if raw is None or str(raw).strip() == "":
            continue
        overrides[name] = _parse_tracker_override(str(raw), input_type)

    sl_mode = str(overrides.get("sl_mode", ""))
    if sl_mode != "Percent":
        overrides.pop("sl_pct", None)
    if sl_mode != "Swing+ATR":
        overrides.pop("sl_swing_basis", None)
        overrides.pop("sl_swing_lookback", None)
        overrides.pop("sl_swing_atr_len", None)
        overrides.pop("sl_swing_atr_mult", None)

    tp_mode = str(overrides.get("tp_mode", ""))
    if tp_mode != "ATR":
        overrides.pop("tp_atr_len", None)
        overrides.pop("tp_atr_mult", None)
    if tp_mode != "Percent":
        overrides.pop("tp_percent", None)
    if tp_mode != "R":
        overrides.pop("tp_r_multiple", None)
    if tp_mode != "Multi-TP":
        overrides.pop("tp1_r_multiple", None)
        overrides.pop("tp1_close_pct", None)
        overrides.pop("tp2_r_multiple", None)

    return overrides


def overlay_pine_overrides(
    pine_defaults: dict[str, Any],
    overrides: dict[str, Any],
) -> dict[str, Any]:
    merged = dict(pine_defaults)
    merged.update(overrides)
    return merged


def load_signals_json(path: Path) -> tuple[dict[str, Any], pd.DataFrame]:
    with path.open(encoding="utf-8") as handle:
        payload = json.load(handle)

    signals = pd.DataFrame(payload["signals"])
    if signals.empty:
        raise ValueError(f"No rows found in {path}")

    signals["timestamp_dt"] = pd.to_datetime(signals["timestamp"], unit="ms", utc=True)
    return payload["meta"], signals


def write_exact_ohlcv_dataset(signals_df: pd.DataFrame, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    dataset = signals_df.loc[:, ["timestamp_dt", "open", "high", "low", "close", "volume"]].copy()
    dataset = dataset.rename(columns={"timestamp_dt": "timestamp"})
    dataset["timestamp"] = dataset["timestamp"].astype(str)
    dataset.to_csv(out_path, index=False, encoding="utf-8")
    return out_path


def _float_or_default(value: Any, default: float) -> float:
    if value is None:
        return float(default)
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def build_runner_config_from_pine_defaults(
    pine_defaults: dict[str, Any],
    meta: dict[str, Any],
) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []

    signal_mode = str(pine_defaults["signal_mode"])
    if signal_mode != "Supertrend":
        raise ValueError(f"Only signal_mode='Supertrend' is supported in 00_PYTHON right now, got {signal_mode!r}")

    sl_mode = str(pine_defaults["sl_mode"])
    if sl_mode not in {"None", "ATR", "Percent", "Swing+ATR"}:
        raise ValueError(f"Unsupported sl_mode in Pine defaults: {sl_mode!r}")

    tp_mode = str(pine_defaults["tp_mode"])
    if tp_mode not in {"None", "ATR", "Percent", "R", "Multi-TP"}:
        raise ValueError(f"Unsupported tp_mode in Pine defaults: {tp_mode!r}")

    syminfo = meta.get("syminfo_patched", {})
    price_tick = max(_float_or_default(syminfo.get("mintick"), 0.01), 1e-9)
    qty_step = max(_float_or_default(syminfo.get("mincontract"), 0.001), 1e-9)
    point_value = _float_or_default(syminfo.get("pointvalue"), 1.0)

    config = {
        "enable_long": bool(pine_defaults["enable_long"]),
        "enable_short": bool(pine_defaults["enable_short"]),
        "allow_flip": bool(pine_defaults["allow_flip"]),
        "regime_lock": bool(pine_defaults["regime_lock"]),
        "max_entries": int(pine_defaults["max_entries"]),
        "cooldown_bars": int(pine_defaults["cooldown_bars"]),
        "signal_mode": signal_mode,
        "st_atr_len": int(pine_defaults["st_atr_len"]),
        "st_factor": float(pine_defaults["st_factor"]),
        "st_use_wicks": bool(pine_defaults["st_use_wicks"]),
        "st_use_ha": bool(pine_defaults["st_use_ha"]),
        "instrument_symbol": str(meta.get("symbol") or "UNKNOWN"),
        "instrument_point_value": point_value,
        "instrument_price_tick": price_tick,
        "instrument_qty_step": qty_step,
        "instrument_min_qty": qty_step,
        "instrument_min_notional": 0.0,
        "instrument_contract_multiplier": point_value,
        "initial_capital": float(pine_defaults["initial_capital"]),
        "margin_long_pct": 100.0,
        "margin_short_pct": 100.0,
        "risk_per_long_pct": float(pine_defaults["risk_per_long_pct"]),
        "risk_per_short_pct": float(pine_defaults["risk_per_short_pct"]),
        "fallback_size_pct": float(pine_defaults["fallback_size_pct"]),
        "max_leverage_cap": float(pine_defaults["max_leverage_cap"]),
        "equity_source": "Realized",
        "use_notional_assert": False,
        "execution_profile_id": "close_only_deterministic_v2",
        "tw_audit_semantics_mode": "off",
        "tw_reversal_reentry_mode": "local",
        "tw_reversal_reentry_delay_bars": 0,
        "tw_margin_call_mode": "off",
        "tw_margin_call_split_entries": False,
        "tw_be_semantics_mode": "local",
        "tw_trailing_semantics_mode": "local",
        "use_ma_filter": bool(pine_defaults["use_ma_filter"]),
        "ma_type": str(pine_defaults["ma_type"]),
        "ma_length": int(pine_defaults["ma_length"]),
        "use_ma_mtf": bool(pine_defaults["use_ma_mtf"]),
        "ma_htf_timeframe": str(pine_defaults["ma_htf_timeframe"]),
        "use_ma_slope_filter": bool(pine_defaults["use_ma_slope_filter"]),
        "ma_slope_len": int(pine_defaults["ma_slope_len"]),
        "ma_slope_min_pct": float(pine_defaults["ma_slope_min_pct"]),
        "use_sl": sl_mode != "None",
        "use_sl_atr": sl_mode == "ATR",
        "use_sl_percent": sl_mode == "Percent",
        "use_sl_swing_atr": sl_mode == "Swing+ATR",
        "sl_atr_len": int(pine_defaults["sl_atr_len"]),
        "sl_atr_mult": float(pine_defaults["sl_atr_mult"]),
        "sl_percent": float(pine_defaults["sl_pct"]),
        "sl_swing_basis": str(pine_defaults["sl_swing_basis"]),
        "sl_swing_lookback": int(pine_defaults["sl_swing_lookback"]),
        "sl_swing_atr_len": int(pine_defaults["sl_swing_atr_len"]),
        "sl_swing_atr_mult": float(pine_defaults["sl_swing_atr_mult"]),
        "tp_mode": tp_mode,
        "tp_atr_len": int(pine_defaults["tp_atr_len"]),
        "tp_atr_mult": float(pine_defaults["tp_atr_mult"]),
        "tp_percent": float(pine_defaults["tp_percent"]),
        "tp_r_multiple": float(pine_defaults["tp_r_multiple"]),
        "tp1_r_multiple": float(pine_defaults["tp1_r_multiple"]),
        "tp1_close_pct": float(pine_defaults["tp1_close_pct"]),
        "tp2_r_multiple": float(pine_defaults["tp2_r_multiple"]),
        "use_break_even": bool(pine_defaults["use_break_even"]),
        "be_trigger_r": float(pine_defaults["be_trigger_r"]),
        "be_buffer_r": float(pine_defaults["be_buffer_r"]),
        "use_trailing": bool(pine_defaults["use_trailing"]),
        "trail_atr_len": int(pine_defaults["trail_atr_len"]),
        "trail_start_r": float(pine_defaults["trail_start_r"]),
        "trail_distance_atr_mult": float(pine_defaults["trail_distance_atr_mult"]),
        # McGinley filter
        "use_mcginley_filter": bool(pine_defaults["use_mcginley_filter"]),
        "mcginley_length": int(pine_defaults["mcginley_length"]),
        # Volume filter
        "use_volume_filter": bool(pine_defaults["use_volume_filter"]),
        "vol_sma_length": int(pine_defaults["vol_sma_length"]),
        "vol_sma_mult": float(pine_defaults["vol_sma_mult"]),
        # ADX filter
        "use_adx_filter": bool(pine_defaults["use_adx_filter"]),
        "adx_length": int(pine_defaults["adx_length"]),
        "adx_threshold": float(pine_defaults["adx_threshold"]),
        # Chop filter
        "use_chop_filter": bool(pine_defaults["use_chop_filter"]),
        "chop_length": int(pine_defaults["chop_length"]),
        "chop_threshold": float(pine_defaults["chop_threshold"]),
        # ATR Vol Floor
        "use_atr_vol_floor": bool(pine_defaults["use_atr_vol_floor"]),
        "atr_vol_floor_fast_len": int(pine_defaults["atr_vol_floor_fast_len"]),
        "atr_vol_floor_baseline_len": int(pine_defaults["atr_vol_floor_baseline_len"]),
        "atr_vol_floor_mult": float(pine_defaults["atr_vol_floor_mult"]),
        # MACD filters
        "use_macd_regime_filter": bool(pine_defaults["use_macd_regime_filter"]),
        "use_macd_cross_filter": bool(pine_defaults["use_macd_cross_filter"]),
        "use_macd_hist_filter": bool(pine_defaults["use_macd_hist_filter"]),
        "macd_hist_mode": str(pine_defaults["macd_hist_mode"]),
        "use_macd_zero_dist_filter": bool(pine_defaults["use_macd_zero_dist_filter"]),
        "macd_zero_dist_min": float(pine_defaults["macd_zero_dist_min"]),
        "macd_fast_len": int(pine_defaults["macd_fast_len"]),
        "macd_slow_len": int(pine_defaults["macd_slow_len"]),
        "macd_sig_len": int(pine_defaults["macd_sig_len"]),
        # L13 opposite signal exit
        "exit_on_opposite_signal": bool(pine_defaults["exit_on_opposite_signal"]),
        # L14 per-filter exits
        "exit_on_ma_block": bool(pine_defaults["exit_on_ma_block"]),
        "exit_on_ma_slope_block": bool(pine_defaults["exit_on_ma_slope_block"]),
        "exit_on_mcginley_block": bool(pine_defaults["exit_on_mcginley_block"]),
        "exit_on_htf_trend_block": bool(pine_defaults["exit_on_htf_trend_block"]),
        "exit_on_vol_block": bool(pine_defaults["exit_on_vol_block"]),
        "exit_on_atr_vol_block": bool(pine_defaults["exit_on_atr_vol_block"]),
        "exit_on_range_block": bool(pine_defaults["exit_on_range_block"]),
        "exit_on_candle_pattern_block": bool(pine_defaults["exit_on_candle_pattern_block"]),
        "exit_on_level_prox_block": bool(pine_defaults["exit_on_level_prox_block"]),
        # L15 time exits
        "use_time_stop": bool(pine_defaults["use_time_stop"]),
        "time_stop_bars": int(pine_defaults["time_stop_bars"]),
        "time_stop_condition": str(pine_defaults["time_stop_condition"]),
        "time_stop_eod": bool(pine_defaults["time_stop_eod"]),
        "time_stop_eow": bool(pine_defaults["time_stop_eow"]),
        # L16 guards
        "use_daily_loss_limit": bool(pine_defaults["use_daily_loss_limit"]),
        "max_daily_loss_pct": float(pine_defaults["max_daily_loss_pct"]),
        "use_max_trades_per_day": bool(pine_defaults["use_max_trades_per_day"]),
        "max_trades_per_day": int(pine_defaults["max_trades_per_day"]),
        "use_max_drawdown_guard": bool(pine_defaults["use_max_drawdown_guard"]),
        "max_drawdown_pct": float(pine_defaults["max_drawdown_pct"]),
        "use_consecutive_loss_halt": bool(pine_defaults["use_consecutive_loss_halt"]),
        "max_consecutive_losses": int(pine_defaults["max_consecutive_losses"]),
        "use_equity_curve_filter": bool(pine_defaults["use_equity_curve_filter"]),
        "equity_ma_length": int(pine_defaults["equity_ma_length"]),
        "use_mae_guard": bool(pine_defaults["use_mae_guard"]),
        "max_mae_pct": float(pine_defaults["max_mae_pct"]),
        "use_guard_recovery": bool(pine_defaults["use_guard_recovery"]),
        "guard_recovery_mode": str(pine_defaults["guard_recovery_mode"]),
        "guard_recovery_bars": int(pine_defaults["guard_recovery_bars"]),
        "guard_recovery_signals": int(pine_defaults["guard_recovery_signals"]),
        # L16 trade cooldown
        "use_trade_cooldown": bool(pine_defaults["use_trade_cooldown"]),
        "cooldown_bars_after_exit": int(pine_defaults["cooldown_bars_after_exit"]),
        # L18 confirmation transform
        "use_confirm_transform": bool(pine_defaults["use_confirm_transform"]),
        "confirm_bars": int(pine_defaults["confirm_bars"]),
        "confirm_close_crosses": bool(pine_defaults["confirm_close_crosses"]),
        # L18b advanced confirmation
        "require_raw_still_true": bool(pine_defaults["require_raw_still_true"]),
        "refresh_on_new_raw": bool(pine_defaults["refresh_on_new_raw"]),
        # L20 level proximity gate
        "use_level_proximity_gate": bool(pine_defaults["use_level_proximity_gate"]),
        "level_proximity_threshold_pct": float(pine_defaults["level_proximity_threshold_pct"]),
        "level_proximity_lookback": int(pine_defaults["level_proximity_lookback"]),
        # L21 level retest
        "use_level_retest": bool(pine_defaults["use_level_retest"]),
        "retest_timeout_bars": int(pine_defaults["retest_timeout_bars"]),
        "retest_buffer_pct": float(pine_defaults["retest_buffer_pct"]),
        # L22 candle pattern gate
        "use_candle_pattern_gate": bool(pine_defaults["use_candle_pattern_gate"]),
        # L12 HTF Trend Filter
        "use_htf_trend_filter": bool(pine_defaults["use_htf_trend_filter"]),
        "htf_trend_timeframe": str(pine_defaults["htf_trend_timeframe"]),
        "htf_trend_ma_type": str(pine_defaults["htf_trend_ma_type"]),
        "htf_trend_ma_len": int(pine_defaults["htf_trend_ma_len"]),
        "htf_trend_buffer_pct": float(pine_defaults["htf_trend_buffer_pct"]),
        # L12 MA Filter HTF path
        "use_ma_mtf": bool(pine_defaults["use_ma_mtf"]),
        "ma_htf_timeframe": str(pine_defaults["ma_htf_timeframe"]),
        # L12 MACD HTF Bias + MACD Source
        "use_macd_htf_bias": bool(pine_defaults["use_macd_htf_bias"]),
        "macd_htf_timeframe": str(pine_defaults["macd_htf_timeframe"]),
        "macd_source": str(pine_defaults["macd_source"]),
        # L12 Momentum Filter
        "use_momentum_filter": bool(pine_defaults["use_momentum_filter"]),
        "momentum_mode": str(pine_defaults["momentum_mode"]),
        "momentum_atr_len": int(pine_defaults["momentum_atr_len"]),
        "momentum_atr_mult": float(pine_defaults["momentum_atr_mult"]),
        "momentum_roc_min_pct": float(pine_defaults["momentum_roc_min_pct"]),
        # L12 Session Filter
        "use_session_filter": bool(pine_defaults["use_session_filter"]),
        "session_name": str(pine_defaults["session_name"]),
        "session_custom_string": str(pine_defaults["session_custom_string"]),
        # L12 McGinley HTF
        "mcginley_use_higher_timeframe": bool(pine_defaults["mcginley_use_higher_timeframe"]),
        "mcginley_htf_timeframe": str(pine_defaults["mcginley_htf_timeframe"]),
        # L12 ADX HTF
        "adx_use_higher_timeframe": bool(pine_defaults["adx_use_higher_timeframe"]),
        "adx_htf_timeframe": str(pine_defaults["adx_htf_timeframe"]),
        # L12 Chop HTF
        "chop_use_higher_timeframe": bool(pine_defaults["chop_use_higher_timeframe"]),
        "chop_htf_timeframe": str(pine_defaults["chop_htf_timeframe"]),
    }

    if tp_mode in {"R", "Multi-TP"} and sl_mode == "None":
        warnings.append("R and Multi-TP require SL. Current Pine defaults disable SL, so parity run will fail validation.")
    return config, warnings


def load_runner_overrides(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        payload = json.load(handle)

    if not isinstance(payload, dict):
        raise ValueError(f"Override config must be a JSON object: {path}")

    if "config" in payload and isinstance(payload["config"], dict):
        payload = payload["config"]

    nested_keys = sorted(key for key, value in payload.items() if isinstance(value, dict))
    if nested_keys:
        raise ValueError(
            "Legacy nested case schema is not supported in parity_compare.py anymore. "
            f"Use flat 00_PYTHON config overrides only. Nested keys: {nested_keys}"
        )

    return payload


def normalize_runner_overrides(payload: dict[str, Any]) -> dict[str, Any]:
    overrides = dict(payload)
    sl_mode = str(overrides.pop("sl_mode", "") or "").strip()
    if "sl_pct" in overrides and "sl_percent" not in overrides:
        overrides["sl_percent"] = overrides.pop("sl_pct")
    if sl_mode:
        overrides["use_sl"] = sl_mode != "None"
        overrides["use_sl_atr"] = sl_mode == "ATR"
        overrides["use_sl_percent"] = sl_mode == "Percent"
        overrides["use_sl_swing_atr"] = sl_mode == "Swing+ATR"
    return overrides


def bars_from_signals(signals_df: pd.DataFrame) -> list[Bar]:
    bars: list[Bar] = []
    for idx, row in signals_df.iterrows():
        bar_index = int(row.get("bar_index", idx))
        bars.append(
            Bar(
                timestamp=row["timestamp_dt"].to_pydatetime(),
                open=float(row["open"]),
                high=float(row["high"]),
                low=float(row["low"]),
                close=float(row["close"]),
                volume=float(row["volume"]),
                bar_index=bar_index,
            )
        )
    return bars


def _snapshot_open_trade(current: dict[str, Any] | None, position: Any, bar: Bar, sizing_equity: float) -> dict[str, Any]:
    direction = "long" if str(position.side).lower() == "long" else "short"
    base = dict(current) if current is not None else {}
    base.update(
        {
            "lifecycle_id": int(position.lifecycle_id),
            "direction": direction,
            "entry_bar": int(position.entry_bar),
            "entry_time": base.get("entry_time", str(bar.timestamp)),
            "entry_price": float(position.entry_price),
            "avg_entry_price": float(position.avg_entry_price),
            "qty": float(position.qty),
            "initial_stop": base.get("initial_stop", position.active_stop_price),
            "initial_tp": base.get("initial_tp", position.active_tp_price),
            "active_stop_price": position.active_stop_price,
            "active_tp_price": position.active_tp_price,
            "sizing_equity": float(sizing_equity),
        }
    )
    return base


def _close_trade(
    open_trade: dict[str, Any],
    *,
    bar: Bar,
    exit_price: float,
    exit_reason: str,
    exit_qty: float,
    realized_pnl: float,
    initial_capital: float,
) -> dict[str, Any]:
    trade = dict(open_trade)
    trade["qty"] = float(exit_qty)
    trade["exit_bar"] = int(bar.bar_index)
    trade["exit_time"] = str(bar.timestamp)
    trade["exit_price"] = float(exit_price)
    trade["exit_reason"] = str(exit_reason)
    trade["gross_pnl"] = round(float(realized_pnl), 6)
    trade["pnl_pct"] = round((float(realized_pnl) / float(initial_capital)) * 100.0, 4)
    trade["win"] = float(realized_pnl) > 0.0
    trade["duration_bars"] = int(bar.bar_index) - int(trade["entry_bar"])
    return trade


def _terminal_close_trade(
    open_trade: dict[str, Any],
    *,
    bar: Bar,
    initial_capital: float,
) -> dict[str, Any]:
    entry_price = float(open_trade.get("avg_entry_price") or open_trade.get("entry_price") or 0.0)
    exit_price = float(bar.close)
    qty = float(open_trade.get("qty") or 0.0)
    side = 1.0 if str(open_trade.get("direction")) == "long" else -1.0
    gross_pnl = side * (exit_price - entry_price) * qty
    return _close_trade(
        open_trade,
        bar=bar,
        exit_price=exit_price,
        exit_reason="terminal_close",
        exit_qty=qty,
        realized_pnl=gross_pnl,
        initial_capital=initial_capital,
    )


def run_v2_runner(
    bars: list[Bar],
    effective_config: dict[str, Any],
    signals_df: "pd.DataFrame | None" = None,
    chart_timezone: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    runner_input = {
        key: value
        for key, value in effective_config.items()
        if key
        not in {
            "use_tp",
            "use_tp_single_atr",
            "use_tp_single_pct",
            "use_tp_single_r",
            "use_tp_multi",
        }
    }
    runner = Runner(runner_input)
    # Optional: inject exact PineTS HTF-derived series for parity-critical gates.
    if signals_df is not None:
        def _series_map(column: str) -> dict[int, float] | None:
            if column not in signals_df.columns:
                return None
            values = signals_df[column].tolist()
            mapping = {i: float(v) for i, v in enumerate(values) if pd.notna(v)}
            return mapping or None

        runner.set_gate_overrides(
            adx_map=(
                _series_map("adx")
                if bool(effective_config.get("use_adx_filter")) and bool(effective_config.get("adx_use_higher_timeframe"))
                else None
            ),
            chop_map=(
                _series_map("chop")
                if bool(effective_config.get("use_chop_filter")) and bool(effective_config.get("chop_use_higher_timeframe"))
                else None
            ),
            macd_htf_line_map=(
                _series_map("macd_htf_line")
                if bool(effective_config.get("use_macd_htf_bias"))
                else None
            ),
            htf_trend_line_map=(
                _series_map("htf_trend_line")
                if bool(effective_config.get("use_htf_trend_filter"))
                else None
            ),
            ma_filter_line_map=(
                _series_map("ma_filter_line")
                if bool(effective_config.get("use_ma_filter")) and bool(effective_config.get("use_ma_mtf"))
                else None
            ),
            mcginley_line_map=(
                _series_map("mcginley_line")
                if bool(effective_config.get("use_mcginley_filter")) and bool(effective_config.get("mcginley_use_higher_timeframe"))
                else None
            ),
        )
    open_trades: dict[int, dict[str, Any]] = {}
    closed_trades: list[dict[str, Any]] = []
    initial_capital = float(effective_config["initial_capital"])

    # Build HTF lookup if any HTF feature is enabled
    htf_data: "dict | None" = None
    _needs_htf = (
        bool(effective_config.get("use_htf_trend_filter"))
        or (bool(effective_config.get("use_ma_filter")) and bool(effective_config.get("use_ma_mtf")))
        or bool(effective_config.get("use_macd_htf_bias"))
        or (bool(effective_config.get("use_mcginley_filter")) and bool(effective_config.get("mcginley_use_higher_timeframe")))
        or (bool(effective_config.get("use_adx_filter")) and bool(effective_config.get("adx_use_higher_timeframe")))
        or (bool(effective_config.get("use_chop_filter")) and bool(effective_config.get("chop_use_higher_timeframe")))
    )
    if _needs_htf and signals_df is not None:
        from mtc_v2.core.htf import build_htf_lookup
        df_ltf = signals_df[["timestamp_dt", "open", "high", "low", "close", "volume"]].copy()
        df_ltf = df_ltf.rename(columns={"timestamp_dt": "timestamp"}).set_index("timestamp")
        df_ltf.index = pd.to_datetime(df_ltf.index, utc=True)
        # Determine which HTF timeframe to use for the lookup:
        # Priority: htf_trend_filter > macd_htf_bias > ma_mtf > mcginley/adx/chop HTF
        if bool(effective_config.get("use_htf_trend_filter")):
            htf_tf = str(effective_config["htf_trend_timeframe"])
        elif bool(effective_config.get("use_macd_htf_bias")):
            htf_tf = str(effective_config.get("macd_htf_timeframe", "240"))
        elif bool(effective_config.get("use_ma_filter")) and bool(effective_config.get("use_ma_mtf")):
            htf_tf = str(effective_config["ma_htf_timeframe"])
        elif bool(effective_config.get("use_mcginley_filter")) and bool(effective_config.get("mcginley_use_higher_timeframe")):
            htf_tf = str(effective_config.get("mcginley_htf_timeframe", "240"))
        elif bool(effective_config.get("use_adx_filter")) and bool(effective_config.get("adx_use_higher_timeframe")):
            htf_tf = str(effective_config.get("adx_htf_timeframe", "240"))
        elif bool(effective_config.get("use_chop_filter")) and bool(effective_config.get("chop_use_higher_timeframe")):
            htf_tf = str(effective_config.get("chop_htf_timeframe", "240"))
        else:
            htf_tf = str(effective_config.get("ma_htf_timeframe", "240"))
        raw_htf = build_htf_lookup(df_ltf, htf_tf, chart_timezone=chart_timezone)
        # Convert pd.Timestamp keys â†’ Python datetime (matches bar.timestamp in runner)
        htf_data = {ts.to_pydatetime(): v for ts, v in raw_htf.items()}

    _first_bar = True
    for bar in bars:
        pre_position = runner.state.position
        pre_lifecycle = int(pre_position.lifecycle_id) if pre_position is not None else None
        pre_open_trade = open_trades.get(pre_lifecycle) if pre_lifecycle is not None else None

        runner.run([bar], htf_data=htf_data if _first_bar else None)
        _first_bar = False
        state = runner.state

        exit_events = list(getattr(state, "exit_events_this_bar", []) or [])
        if not exit_events and state.closed_this_bar_reason is not None:
            exit_events = [
                {
                    "exit_price": float(state.last_exit_price if state.last_exit_price is not None else bar.close),
                    "exit_reason": str(state.closed_this_bar_reason),
                    "exit_qty": float(state.last_exit_qty if state.last_exit_qty > 0.0 else (pre_position.qty if pre_position is not None else 0.0)),
                    "realized_pnl": float(state.last_realized_pnl),
                    "was_partial": bool(state.last_exit_was_partial),
                }
            ]

        if exit_events and pre_position is not None and pre_lifecycle is not None:
            if pre_open_trade is None:
                pre_open_trade = _snapshot_open_trade(None, pre_position, bar, state.last_sizing_equity_snapshot)
            for exit_event in exit_events:
                closed_trades.append(
                    _close_trade(
                        pre_open_trade,
                        bar=bar,
                        exit_price=float(
                            exit_event.exit_price if hasattr(exit_event, "exit_price") else exit_event.get("exit_price", bar.close)
                        ),
                        exit_reason=str(
                            exit_event.exit_reason if hasattr(exit_event, "exit_reason") else exit_event.get("exit_reason", "")
                        ),
                        exit_qty=float(
                            exit_event.exit_qty if hasattr(exit_event, "exit_qty") else exit_event.get("exit_qty", pre_position.qty)
                        ),
                        realized_pnl=float(
                            exit_event.realized_pnl if hasattr(exit_event, "realized_pnl") else exit_event.get("realized_pnl", 0.0)
                        ),
                        initial_capital=initial_capital,
                    )
                )
                was_partial = bool(
                    exit_event.was_partial if hasattr(exit_event, "was_partial") else exit_event.get("was_partial", False)
                )
                if not was_partial:
                    open_trades.pop(pre_lifecycle, None)

        if state.position is not None:
            lifecycle_id = int(state.position.lifecycle_id)
            open_trades[lifecycle_id] = _snapshot_open_trade(
                open_trades.get(lifecycle_id),
                state.position,
                bar,
                state.last_sizing_equity_snapshot,
            )

    if bars and runner.state.position is not None:
        lifecycle_id = int(runner.state.position.lifecycle_id)
        open_trade = open_trades.pop(lifecycle_id, None)
        if open_trade is not None:
            closed_trades.append(_terminal_close_trade(open_trade, bar=bars[-1], initial_capital=initial_capital))

    return closed_trades, {
        "warmup_bars": runner.state.warmup_bars,
        "final_equity": runner.state.equity,
        "realized_equity": runner.state.realized_equity,
        "total_entries": runner.state.total_entries,
        "total_exits": runner.state.total_exits,
    }


def compute_metrics(trades: list[dict[str, Any]], initial_capital: float) -> dict[str, Any]:
    if not trades:
        return {
            "total_trades": 0,
            "closed_trades": 0,
            "win_rate_pct": 0.0,
            "net_pnl": 0.0,
            "net_pnl_pct": 0.0,
            "profit_factor": 0.0,
            "max_drawdown_pct": 0.0,
            "gross_wins": 0.0,
            "gross_losses": 0.0,
            "avg_trade_pnl": 0.0,
            "long_trades": 0,
            "short_trades": 0,
            "exit_reasons": {},
        }

    closed = [trade for trade in trades if trade.get("gross_pnl") is not None]
    wins = [trade for trade in closed if trade.get("win")]
    losses = [trade for trade in closed if not trade.get("win") and trade.get("gross_pnl", 0.0) != 0.0]

    gross_wins = sum(float(trade["gross_pnl"]) for trade in wins)
    gross_losses = abs(sum(float(trade["gross_pnl"]) for trade in losses))
    net_pnl = sum(float(trade["gross_pnl"]) for trade in closed)
    net_pnl_pct = (net_pnl / float(initial_capital)) * 100.0
    profit_factor = (gross_wins / gross_losses) if gross_losses > 0 else float("inf")
    win_rate = (len(wins) / len(closed) * 100.0) if closed else 0.0

    equity = float(initial_capital)
    peak = float(initial_capital)
    max_dd = 0.0
    for trade in closed:
        equity += float(trade["gross_pnl"])
        if equity > peak:
            peak = equity
        dd = ((peak - equity) / peak * 100.0) if peak > 0 else 0.0
        if dd > max_dd:
            max_dd = dd

    longs = [trade for trade in trades if str(trade.get("direction")) == "long"]
    shorts = [trade for trade in trades if str(trade.get("direction")) == "short"]

    exit_reasons: dict[str, int] = {}
    for trade in trades:
        reason = str(trade.get("exit_reason", "unknown"))
        exit_reasons[reason] = exit_reasons.get(reason, 0) + 1

    return {
        "total_trades": len(trades),
        "closed_trades": len(closed),
        "win_rate_pct": round(win_rate, 2),
        "net_pnl": round(net_pnl, 4),
        "net_pnl_pct": round(net_pnl_pct, 4),
        "profit_factor": round(profit_factor, 4) if profit_factor != float("inf") else "inf",
        "max_drawdown_pct": round(max_dd, 2),
        "gross_wins": round(gross_wins, 4),
        "gross_losses": round(gross_losses, 4),
        "avg_trade_pnl": round(net_pnl / len(closed), 4) if closed else 0.0,
        "long_trades": len(longs),
        "short_trades": len(shorts),
        "exit_reasons": exit_reasons,
    }


def load_pine_payload(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def load_pine_metrics(pine_payload: dict[str, Any], initial_capital: float | None = None) -> dict[str, Any]:
    metrics = pine_payload["metrics"]
    meta = pine_payload["meta"]
    if initial_capital is not None:
        metrics = compute_metrics(list(pine_payload.get("trades", [])), float(initial_capital))
    return {
        "total_trades": int(metrics.get("total_trades", 0)),
        "win_rate_pct": float(metrics.get("win_rate_pct", 0.0)),
        "net_pnl": float(metrics.get("net_pnl", 0.0)),
        "net_pnl_pct": float(metrics.get("net_pnl_pct", 0.0)),
        "profit_factor": metrics.get("profit_factor", 0.0),
        "max_drawdown_pct": float(metrics.get("max_drawdown_pct", 0.0)),
        "long_trades": int(metrics.get("long_trades", 0)),
        "short_trades": int(metrics.get("short_trades", 0)),
        "exit_reasons": metrics.get("exit_reasons", {}),
        "symbol": meta.get("symbol"),
        "timeframe": meta.get("timeframe"),
        "bars": meta.get("bars"),
    }


def normalize_pine_reason(reason: str) -> str:
    return PINE_REASON_MAP.get(str(reason or "").strip(), str(reason or "").strip().upper())


def normalize_python_reason(reason: str) -> str:
    return PYTHON_REASON_MAP.get(str(reason or "").strip(), str(reason or "").strip().upper())


def build_metric_rows(pine: dict[str, Any], python: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    def _rel_diff(a: float, b: float) -> float:
        base = max(abs(a), abs(b), 1e-9)
        return abs(a - b) / base * 100.0

    specs = [
        ("total_trades", "Trade Count", "{:.0f}"),
        ("win_rate_pct", "Win Rate %", "{:.2f}%"),
        ("profit_factor", "Profit Factor", "{:.4f}"),
        ("net_pnl_pct", "Net PnL %", "{:.4f}%"),
        ("max_drawdown_pct", "Max Drawdown %", "{:.4f}%"),
    ]

    for key, label, fmt in specs:
        pine_value = pine.get(key)
        py_value = python.get(key)
        if pine_value is None or py_value is None:
            rows.append(
                {
                    "metric": label,
                    "key": key,
                    "pine": pine_value,
                    "python": py_value,
                    "abs_diff": None,
                    "rel_diff_pct": None,
                    "status": "N/A",
                    "fmt": fmt,
                }
            )
            continue

        if key == "profit_factor" and (pine_value == "inf" or py_value == "inf"):
            abs_diff = 0.0
            rel_diff = 0.0
            status = "PASS"
        else:
            pine_num = float(pine_value)
            py_num = float(py_value)
            abs_diff = abs(pine_num - py_num)
            rel_diff = _rel_diff(pine_num, py_num)
            status = "PASS" if abs_diff <= METRIC_TOLERANCES[key] else "MISMATCH"

        rows.append(
            {
                "metric": label,
                "key": key,
                "pine": pine_value,
                "python": py_value,
                "abs_diff": round(abs_diff, 6) if abs_diff is not None else None,
                "rel_diff_pct": round(rel_diff, 2) if rel_diff is not None else None,
                "status": status,
                "fmt": fmt,
            }
        )

    return rows


def build_pine_trade_frame(pine_payload: dict[str, Any]) -> pd.DataFrame:
    rows = []
    for idx, trade in enumerate(pine_payload.get("trades", []), start=1):
        rows.append(
            {
                "seq": idx,
                "side": str(trade.get("direction", "")).upper(),
                "entry_time": pd.to_datetime(trade.get("entry_time"), utc=True, errors="coerce"),
                "exit_time": pd.to_datetime(trade.get("exit_time"), utc=True, errors="coerce"),
                "entry_price": float(trade.get("avg_entry_price") or trade.get("entry_price") or 0.0),
                "exit_price": float(trade.get("exit_price") or 0.0),
                "qty": float(trade.get("qty") or 0.0),
                "reason": normalize_pine_reason(str(trade.get("exit_reason", ""))),
                "pine_source_reason": str(trade.get("exit_reason", "")),
            }
        )
    return pd.DataFrame(rows)


def build_python_trade_frame(trades: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for idx, trade in enumerate(trades, start=1):
        rows.append(
            {
                "seq": idx,
                "side": str(trade.get("direction", "")).upper(),
                "entry_time": pd.to_datetime(trade.get("entry_time"), utc=True, errors="coerce"),
                "exit_time": pd.to_datetime(trade.get("exit_time"), utc=True, errors="coerce"),
                "entry_price": float(trade.get("avg_entry_price") or trade.get("entry_price") or 0.0),
                "exit_price": float(trade.get("exit_price") or 0.0),
                "qty": float(trade.get("qty") or 0.0),
                "reason": normalize_python_reason(str(trade.get("exit_reason", ""))),
                "python_source_reason": str(trade.get("exit_reason", "")),
            }
        )
    return pd.DataFrame(rows)


def summarize_trade_parity(
    pine_df: pd.DataFrame,
    python_df: pd.DataFrame,
    *,
    price_tol: float,
    qty_tol: float,
) -> tuple[dict[str, Any], pd.DataFrame]:
    compared = min(len(pine_df), len(python_df))
    if compared == 0:
        return {
            "pine_trades": int(len(pine_df)),
            "python_trades": int(len(python_df)),
            "compared": 0,
            "core_match_count": 0,
            "entry_price_match_count": 0,
            "exit_price_match_count": 0,
            "qty_match_count": 0,
            "extra_pine_trades": int(len(pine_df)),
            "extra_python_trades": int(len(python_df)),
            "trade_delta": int(len(python_df) - len(pine_df)),
            "strict_pass": len(pine_df) == len(python_df),
            "price_tol": price_tol,
            "qty_tol": qty_tol,
            "first_mismatch": None,
        }, pd.DataFrame()

    pine_n = pine_df.iloc[:compared].copy().add_prefix("pine_")
    python_n = python_df.iloc[:compared].copy().add_prefix("python_")
    report = pd.concat([pine_n.reset_index(drop=True), python_n.reset_index(drop=True)], axis=1)
    report["side_match"] = report["pine_side"] == report["python_side"]
    report["reason_match"] = report["pine_reason"] == report["python_reason"]
    report["entry_time_diff_min"] = (report["python_entry_time"] - report["pine_entry_time"]).dt.total_seconds() / 60.0
    report["exit_time_diff_min"] = (report["python_exit_time"] - report["pine_exit_time"]).dt.total_seconds() / 60.0
    report["entry_price_diff"] = report["python_entry_price"] - report["pine_entry_price"]
    report["exit_price_diff"] = report["python_exit_price"] - report["pine_exit_price"]
    report["qty_diff"] = report["python_qty"] - report["pine_qty"]
    report["entry_time_match"] = report["entry_time_diff_min"].abs() < 1e-9
    report["exit_time_match"] = report["exit_time_diff_min"].abs() < 1e-9
    report["entry_price_match"] = report["entry_price_diff"].abs() <= float(price_tol)
    report["exit_price_match"] = report["exit_price_diff"].abs() <= float(price_tol)
    report["qty_match"] = report["qty_diff"].abs() <= float(qty_tol)
    report["all_core_match"] = (
        report["side_match"]
        & report["reason_match"]
        & report["entry_time_match"]
        & report["exit_time_match"]
        & report["entry_price_match"]
        & report["exit_price_match"]
        & report["qty_match"]
    )

    first_mismatch: dict[str, Any] | None = None
    mismatches = report[~report["all_core_match"]]
    if not mismatches.empty:
        first = mismatches.iloc[0]
        first_mismatch = {
            "pine_side": str(first["pine_side"]),
            "python_side": str(first["python_side"]),
            "pine_entry_time": str(first["pine_entry_time"]),
            "python_entry_time": str(first["python_entry_time"]),
            "pine_exit_time": str(first["pine_exit_time"]),
            "python_exit_time": str(first["python_exit_time"]),
            "pine_reason": str(first["pine_reason"]),
            "python_reason": str(first["python_reason"]),
            "entry_time_diff_min": float(first["entry_time_diff_min"]),
            "exit_time_diff_min": float(first["exit_time_diff_min"]),
            "entry_price_diff": float(first["entry_price_diff"]),
            "exit_price_diff": float(first["exit_price_diff"]),
            "qty_diff": float(first["qty_diff"]),
        }

    summary = {
        "pine_trades": int(len(pine_df)),
        "python_trades": int(len(python_df)),
        "compared": int(compared),
        "core_match_count": int(report["all_core_match"].sum()),
        "entry_price_match_count": int(report["entry_price_match"].sum()),
        "exit_price_match_count": int(report["exit_price_match"].sum()),
        "qty_match_count": int(report["qty_match"].sum()),
        "extra_pine_trades": int(max(0, len(pine_df) - compared)),
        "extra_python_trades": int(max(0, len(python_df) - compared)),
        "trade_delta": int(len(python_df) - len(pine_df)),
        "strict_pass": bool(len(pine_df) == len(python_df) and report["all_core_match"].all()),
        "price_tol": price_tol,
        "qty_tol": qty_tol,
        "first_mismatch": first_mismatch,
    }
    return summary, report


def print_metric_table(rows: list[dict[str, Any]], pine: dict[str, Any], python_label: str) -> None:
    width = 78
    print("=" * width)
    print("  PARITY COMPARISON")
    print("=" * width)
    print(
        f"  Pine   : PineTS / {pine.get('symbol', '?')} {pine.get('timeframe', '?')} "
        f"({pine.get('bars', '?')} bars)"
    )
    print(f"  Python : {python_label}")
    print("-" * width)
    print(f"  {'Metric':<20} {'Pine':>12} {'Python':>12} {'AbsDiff':>12} {'Status':>10}")
    print("-" * width)

    for row in rows:
        fmt = row["fmt"]
        pine_value = row["pine"]
        py_value = row["python"]
        abs_diff = row["abs_diff"]
        status = row["status"]
        icon = "OK" if status == "PASS" else status

        try:
            pine_str = fmt.format(float(pine_value)) if pine_value is not None else "-"
        except Exception:
            pine_str = str(pine_value)
        try:
            py_str = fmt.format(float(py_value)) if py_value is not None else "-"
        except Exception:
            py_str = str(py_value)

        diff_str = "-" if abs_diff is None else f"{abs_diff:.4f}"
        print(f"  {row['metric']:<20} {pine_str:>12} {py_str:>12} {diff_str:>12} {icon:>10}")

    print("-" * width)


def print_trade_summary(summary: dict[str, Any], pine_metrics: dict[str, Any], python_metrics: dict[str, Any]) -> None:
    print("  Trade Parity")
    print(f"    Pine trades      : {summary['pine_trades']}")
    print(f"    Python trades    : {summary['python_trades']}")
    print(f"    Compared         : {summary['compared']}")
    print(f"    Core match count : {summary['core_match_count']}")
    print(f"    Entry px match   : {summary['entry_price_match_count']}")
    print(f"    Exit px match    : {summary['exit_price_match_count']}")
    print(f"    Qty match        : {summary['qty_match_count']}")
    print(f"    Extra Pine       : {summary['extra_pine_trades']}")
    print(f"    Extra Python     : {summary['extra_python_trades']}")
    print(f"    Strict trade PASS: {'YES' if summary['strict_pass'] else 'NO'}")
    print()
    print("  Exit Reasons")
    print(f"    Pine   : {pine_metrics.get('exit_reasons', {})}")
    print(f"    Python : {python_metrics.get('exit_reasons', {})}")
    if summary.get("first_mismatch"):
        print()
        print("  First Trade Mismatch")
        for key, value in summary["first_mismatch"].items():
            print(f"    {key}: {value}")


def save_report(
    out_path: Path,
    *,
    verdict: str,
    pine_metrics: dict[str, Any],
    python_metrics: dict[str, Any],
    metric_rows: list[dict[str, Any]],
    trade_summary: dict[str, Any],
    trade_report_csv: Path,
    python_trades_csv: Path,
    dataset_csv: Path,
    effective_config_json: Path,
    python_config_mode: str,
    warnings: list[str],
    override_source: str | None,
    runner_summary: dict[str, Any],
) -> None:
    payload = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "verdict": verdict,
        "python_config_mode": python_config_mode,
        "override_source": override_source,
        "pine_meta": {
            "symbol": pine_metrics.get("symbol"),
            "timeframe": pine_metrics.get("timeframe"),
            "bars": pine_metrics.get("bars"),
        },
        "warnings": warnings,
        "runner_summary": runner_summary,
        "metrics": metric_rows,
        "trade_parity": trade_summary,
        "pine_exit_reasons": pine_metrics.get("exit_reasons", {}),
        "python_exit_reasons": python_metrics.get("exit_reasons", {}),
        "artifacts": {
            "dataset_csv": str(dataset_csv),
            "trade_report_csv": str(trade_report_csv),
            "python_trades_csv": str(python_trades_csv),
            "effective_config_json": str(effective_config_json),
        },
    }
    with out_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, default=str)


def _relative_str(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def _merge_auto_note(existing: str, auto_note: str) -> str:
    prefix = str(existing or "").split(" | auto_run:", 1)[0].strip()
    if prefix:
        return f"{prefix} | auto_run:{auto_note}"
    return f"auto_run:{auto_note}"


def _next_tracker_status(existing_status: str, verdict: str) -> str:
    if existing_status in {"BLOCKED_BY_LAYER_ORDER", "PLANNED_LAYER_GATE"}:
        return existing_status
    if verdict == "PASS":
        return "DONE"
    return "IN_PROGRESS"


def _tracker_update_payload(
    *,
    tracker_case: str,
    tracker_agent: str,
    verdict: str,
    override_source: str | None,
    effective_config: dict[str, Any],
    signals_meta: dict[str, Any],
    pine_metrics: dict[str, Any],
    python_metrics: dict[str, Any],
    warnings: list[str],
    trade_report_csv: Path,
    python_trades_csv: Path,
    effective_config_json: Path,
    report_json: Path,
) -> dict[str, str]:
    preset_name = Path(override_source).stem if override_source else "current_defaults"
    note_bits = [f"case={tracker_case}", f"verdict={verdict}"]
    if warnings:
        note_bits.append("warnings=" + " || ".join(warnings))
    artifacts = "; ".join(
        [
            _relative_str(report_json),
            _relative_str(trade_report_csv),
            _relative_str(python_trades_csv),
            _relative_str(effective_config_json),
        ]
    )
    tp_mode = str(effective_config.get("tp_mode", "None"))
    if not bool(effective_config.get("use_sl", False)):
        sl_mode = "None"
    elif bool(effective_config.get("use_sl_percent", False)):
        sl_mode = "Percent"
    elif bool(effective_config.get("use_sl_swing_atr", False)):
        sl_mode = "Swing+ATR"
    else:
        sl_mode = "ATR"
    return {
        "symbol": str(signals_meta.get("symbol", "")),
        "timeframe": str(signals_meta.get("timeframe", "")),
        "bars": str(signals_meta.get("bars", pine_metrics.get("bars", ""))),
        "data_source": "Binance via mtc_bridge.mjs / PineTS exact bars",
        "pine_preset_name": preset_name,
        "enable_long": str(bool(effective_config.get("enable_long", True))),
        "enable_short": str(bool(effective_config.get("enable_short", True))),
        "allow_flip": str(bool(effective_config.get("allow_flip", True))),
        "regime_lock": str(bool(effective_config.get("regime_lock", False))),
        "max_entries": str(int(effective_config.get("max_entries", 1))),
        "cooldown_bars": str(int(effective_config.get("cooldown_bars", 0))),
        "signal_mode": str(effective_config.get("signal_mode", "")),
        "st_atr_len": str(int(effective_config.get("st_atr_len", 0))),
        "st_factor": str(float(effective_config.get("st_factor", 0.0))),
        "st_use_wicks": str(bool(effective_config.get("st_use_wicks", False))),
        "st_use_ha": str(bool(effective_config.get("st_use_ha", False))),
        "equity_source": str(effective_config.get("equity_source", "")),
        "risk_per_long_pct": str(float(effective_config.get("risk_per_long_pct", 0.0))),
        "risk_per_short_pct": str(float(effective_config.get("risk_per_short_pct", 0.0))),
        "fallback_size_pct": str(float(effective_config.get("fallback_size_pct", 0.0))),
        "max_leverage_cap": str(float(effective_config.get("max_leverage_cap", 0.0))),
        "sl_mode": sl_mode,
        "sl_atr_len": str(int(effective_config.get("sl_atr_len", 0))) if sl_mode == "ATR" else "",
        "sl_atr_mult": str(float(effective_config.get("sl_atr_mult", 0.0))) if sl_mode == "ATR" else "",
        "sl_pct": str(float(effective_config.get("sl_percent", 0.0))) if sl_mode == "Percent" else "",
        "tp_mode": tp_mode,
        "tp_atr_len": str(int(effective_config.get("tp_atr_len", 0))) if tp_mode != "None" else "",
        "tp_atr_mult": str(float(effective_config.get("tp_atr_mult", 0.0))) if tp_mode != "None" else "",
        "llm_auto_status": verdict,
        "llm_last_run_utc": datetime.now(timezone.utc).isoformat(),
        "llm_agent": tracker_agent,
        "pine_trades": str(int(pine_metrics.get("total_trades", 0))),
        "pine_win_rate_pct": str(float(pine_metrics.get("win_rate_pct", 0.0))),
        "pine_net_pnl_pct": str(float(pine_metrics.get("net_pnl_pct", 0.0))),
        "pine_max_dd_pct": str(float(pine_metrics.get("max_drawdown_pct", 0.0))),
        "python_trades": str(int(python_metrics.get("total_trades", 0))),
        "python_win_rate_pct": str(float(python_metrics.get("win_rate_pct", 0.0))),
        "python_net_pnl_pct": str(float(python_metrics.get("net_pnl_pct", 0.0))),
        "python_max_dd_pct": str(float(python_metrics.get("max_drawdown_pct", 0.0))),
        "parity_verdict": verdict,
        "artifacts": artifacts,
        "_auto_note": " | ".join(note_bits),
    }


def _update_tracker_csv(
    *,
    tracker_case: str,
    update_values: dict[str, str],
    csv_path: Path,
) -> bool:
    if not csv_path.exists():
        print(f"Tracker CSV missing: {csv_path}")
        return False

    with csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
        fieldnames = list(rows[0].keys()) if rows else []

    if not fieldnames:
        print(f"Tracker CSV has no headers: {csv_path}")
        return False

    found = False
    for row in rows:
        if row.get("case_id") != tracker_case:
            continue
        found = True
        row["status"] = _next_tracker_status(row.get("status", ""), update_values["parity_verdict"])
        for key, value in update_values.items():
            if key == "_auto_note" or key not in row:
                continue
            row[key] = value
        row["notes"] = _merge_auto_note(row.get("notes", ""), update_values["_auto_note"])
        break

    if not found:
        print(f"Tracker case not found in CSV: {tracker_case}")
        return False

    # Add "notes" field if it was added but not in original fieldnames
    if "notes" not in fieldnames:
        fieldnames.append("notes")

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return True


def _update_tracker_workbook(
    *,
    tracker_case: str,
    update_values: dict[str, str],
    workbook_path: Path,
    emit_missing: bool = True,
    emit_locked: bool = True,
) -> bool:
    if not workbook_path.exists():
        print(f"Tracker workbook missing: {workbook_path}")
        return False

    workbook = load_workbook(workbook_path)
    if "Cases" not in workbook.sheetnames:
        print(f"Tracker workbook missing Cases sheet: {workbook_path}")
        return False

    sheet = workbook["Cases"]
    headers = [cell.value for cell in sheet[1]]
    col_map = {str(name): idx + 1 for idx, name in enumerate(headers) if name}
    if "case_id" not in col_map:
        print(f"Tracker workbook missing case_id column: {workbook_path}")
        return False

    found_row = None
    for row_index in range(2, sheet.max_row + 1):
        if str(sheet.cell(row_index, col_map["case_id"]).value or "") == tracker_case:
            found_row = row_index
            break

    if found_row is None:
        if emit_missing:
            print(f"Tracker case not found in workbook: {tracker_case}")
        return False

    status_col = col_map.get("status")
    notes_col = col_map.get("notes")
    if status_col is not None:
        existing_status = str(sheet.cell(found_row, status_col).value or "")
        sheet.cell(found_row, status_col).value = _next_tracker_status(existing_status, update_values["parity_verdict"])

    for key, value in update_values.items():
        if key == "_auto_note":
            continue
        col_idx = col_map.get(key)
        if col_idx is None:
            continue
        sheet.cell(found_row, col_idx).value = value

    if notes_col is not None:
        existing_notes = str(sheet.cell(found_row, notes_col).value or "")
        sheet.cell(found_row, notes_col).value = _merge_auto_note(existing_notes, update_values["_auto_note"])

    try:
        workbook.save(workbook_path)
    except PermissionError:
        if emit_locked:
            print(f"Tracker workbook is locked for write: {workbook_path}")
        return False
    return True


def _candidate_tracker_workbooks(tracker_case: str) -> list[Path]:
    candidates: list[Path] = [TRACKER_WORKBOOK_PATH]
    # Close-only canonical rows live in the pending workbook until the original
    # workbook becomes writable again. Prefer checking the fallback as well.
    if tracker_case.startswith("CO_") and TRACKER_WORKBOOK_FALLBACK_PATH not in candidates:
        candidates.append(TRACKER_WORKBOOK_FALLBACK_PATH)
    return candidates


def _update_tracker_workbooks(
    *,
    tracker_case: str,
    update_values: dict[str, str],
) -> bool:
    candidates = _candidate_tracker_workbooks(tracker_case)
    updated_any = False
    for i, workbook_path in enumerate(candidates):
        result = _update_tracker_workbook(
            tracker_case=tracker_case,
            update_values=update_values,
            workbook_path=workbook_path,
            emit_missing=(i == len(candidates) - 1),
            emit_locked=(i == len(candidates) - 1),
        )
        updated_any = updated_any or result
    return updated_any


def update_tracker_artifacts(
    *,
    tracker_case: str,
    tracker_agent: str,
    verdict: str,
    override_source: str | None,
    effective_config: dict[str, Any],
    signals_meta: dict[str, Any],
    pine_metrics: dict[str, Any],
    python_metrics: dict[str, Any],
    warnings: list[str],
    trade_report_csv: Path,
    python_trades_csv: Path,
    effective_config_json: Path,
    report_json: Path,
) -> None:
    update_values = _tracker_update_payload(
        tracker_case=tracker_case,
        tracker_agent=tracker_agent,
        verdict=verdict,
        override_source=override_source,
        effective_config=effective_config,
        signals_meta=signals_meta,
        pine_metrics=pine_metrics,
        python_metrics=python_metrics,
        warnings=warnings,
        trade_report_csv=trade_report_csv,
        python_trades_csv=python_trades_csv,
        effective_config_json=effective_config_json,
        report_json=report_json,
    )
    csv_ok = _update_tracker_csv(
        tracker_case=tracker_case,
        update_values=update_values,
        csv_path=TRACKER_CSV_PATH,
    )
    wb_ok = _update_tracker_workbooks(
        tracker_case=tracker_case,
        update_values=update_values,
    )
    if csv_ok and wb_ok:
        print(f"Updated tracker case: {tracker_case}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--chart-timezone", default="UTC+3")
    parser.add_argument(
        "--case",
        default="",
        help="Optional flat 00_PYTHON config overrides JSON. Legacy nested backtest cases are not supported.",
    )
    parser.add_argument(
        "--python-case",
        default="",
        help="Optional extra flat 00_PYTHON config overrides JSON applied to Python only after --case.",
    )
    parser.add_argument("--pine-json", default="data/pine_trades.json")
    parser.add_argument("--signals-json", default="data/mtc_signals.json")
    parser.add_argument("--fetch-fresh", action="store_true", help="Run mtc_bridge.mjs and pine_trades.py before comparing.")
    parser.add_argument("--bars", type=int, default=1000, help="Bars for PineTS fetch.")
    parser.add_argument("--symbol", default="BTCUSDT", help="Symbol for mtc_bridge.mjs when --fetch-fresh is used.")
    parser.add_argument("--tf", default="1h", help="Timeframe for mtc_bridge.mjs when --fetch-fresh is used.")
    parser.add_argument("--start", default="", help="Optional UTC date/date-time start for mtc_bridge.mjs. Example: 2025-01-01")
    parser.add_argument("--end", default="", help="Optional UTC date/date-time end for mtc_bridge.mjs. Example: 2025-12-31")
    parser.add_argument("--out", default="reports/parity_compare.json")
    parser.add_argument("--trade-report-csv", default="reports/parity_trade_report.csv")
    parser.add_argument("--python-trades-csv", default="reports/python_trades_v2.csv")
    parser.add_argument("--effective-config-json", default="reports/parity_effective_config.json")
    parser.add_argument(
        "--tracker-case",
        default="",
        help="Optional tracker case_id to auto-update. If omitted and --case is not used, baseline AUTO_001 is updated.",
    )
    parser.add_argument("--tracker-agent", default="Codex", help="Agent label written into the parity tracker.")
    parser.add_argument(
        "--assume-pine-case-aligned",
        action="store_true",
        help="Allow tracker updates for custom --case runs only if PineTS was already run with the same case inputs.",
    )
    parser.add_argument(
        "--use-case-for-pine",
        action="store_true",
        help="Apply flat --case overrides to Pine inputs too during --fetch-fresh when keys are Pine-supported.",
    )
    args = parser.parse_args()

    warnings: list[str] = []
    default_out = "reports/parity_compare.json"
    default_trade_report = "reports/parity_trade_report.csv"
    default_python_trades = "reports/python_trades_v2.csv"
    default_effective_config = "reports/parity_effective_config.json"

    tracker_case = str(args.tracker_case or "").strip()
    if not args.case and not tracker_case:
        tracker_case = "AUTO_001"

    tracker_row: dict[str, str] | None = None
    tracker_pine_overrides: dict[str, Any] = {}
    fetch_symbol = args.symbol
    fetch_tf = args.tf
    fetch_bars = int(args.bars)
    if tracker_case:
        tracker_row = load_tracker_case_row(tracker_case)
        tracker_pine_overrides = tracker_row_to_pine_overrides(tracker_row)
        fetch_symbol = str(tracker_row.get("symbol") or fetch_symbol)
        fetch_tf = str(tracker_row.get("timeframe") or fetch_tf)
        if str(tracker_row.get("bars") or "").strip():
            fetch_bars = int(float(str(tracker_row["bars"]).strip()))

        if tracker_case != "AUTO_001":
            case_dir = REPORTS_DIR / "tracker_cases" / tracker_case
            case_dir.mkdir(parents=True, exist_ok=True)
            if args.out == default_out:
                args.out = str(case_dir / "parity_compare.json")
            if args.trade_report_csv == default_trade_report:
                args.trade_report_csv = str(case_dir / "parity_trade_report.csv")
            if args.python_trades_csv == default_python_trades:
                args.python_trades_csv = str(case_dir / "python_trades_v2.csv")
            if args.effective_config_json == default_effective_config:
                args.effective_config_json = str(case_dir / "parity_effective_config.json")

    pine_fetch_overrides: dict[str, Any] = {}
    if tracker_pine_overrides:
        pine_fetch_overrides.update(tracker_pine_overrides)
    if args.case and args.use_case_for_pine:
        raw_case_overrides = load_runner_overrides(Path(args.case))
        pine_case_overrides = {k: v for k, v in raw_case_overrides.items() if k in SUPPORTED_PINE_INPUTS}
        pine_fetch_overrides.update(pine_case_overrides)

    if args.fetch_fresh:
        bridge_args = [
            "node",
            "mtc_bridge.mjs",
            "--bars",
            str(fetch_bars),
            "--symbol",
            fetch_symbol,
            "--tf",
            fetch_tf,
        ]
        if args.start:
            bridge_args.extend(["--start", args.start])
        if args.end:
            bridge_args.extend(["--end", args.end])
        if pine_fetch_overrides:
            overrides_path = Path(args.effective_config_json).with_name("pine_input_overrides.json")
            overrides_path.parent.mkdir(parents=True, exist_ok=True)
            with overrides_path.open("w", encoding="utf-8") as handle:
                json.dump(pine_fetch_overrides, handle, indent=2, ensure_ascii=False)
            bridge_args.extend(["--overrides-json", str(overrides_path)])

        print(">>> Running mtc_bridge.mjs...")
        bridge = subprocess.run(
            bridge_args,
            cwd=BACKTEST_ROOT,
            capture_output=False,
        )
        if bridge.returncode != 0:
            print("mtc_bridge.mjs FAILED")
            return 1

        print("\n>>> Running pine_trades.py...")
        extractor = subprocess.run(
            [sys.executable, "pine_trades.py"],
            cwd=BACKTEST_ROOT,
            capture_output=False,
        )
        if extractor.returncode != 0:
            print("pine_trades.py FAILED")
            return 1
        print()

    pine_json = BACKTEST_ROOT / args.pine_json
    signals_json = BACKTEST_ROOT / args.signals_json
    if not pine_json.exists():
        print(f"ERROR: {pine_json} not found. Use --fetch-fresh.")
        return 1
    if not signals_json.exists():
        print(f"ERROR: {signals_json} not found. Use --fetch-fresh.")
        return 1

    pine_payload = load_pine_payload(pine_json)
    signals_meta, signals_df = load_signals_json(signals_json)

    dataset_csv = BACKTEST_ROOT / "data" / "parity_input_from_pinets.csv"
    write_exact_ohlcv_dataset(signals_df, dataset_csv)

    pine_defaults = extract_supported_pine_defaults(PINE_SOURCE)
    if pine_fetch_overrides:
        pine_defaults = overlay_pine_overrides(pine_defaults, pine_fetch_overrides)
    runner_config, config_warnings = build_runner_config_from_pine_defaults(pine_defaults, signals_meta)
    warnings.extend(config_warnings)

    python_config_mode = "generated_from_current_pine_defaults"
    override_source: str | None = None
    if tracker_row is not None:
        python_config_mode = "tracker_case_aligned_pine_python"
        override_source = str(tracker_row.get("pine_preset_name") or tracker_case)
    if args.case:
        overrides = normalize_runner_overrides(load_runner_overrides(Path(args.case)))
        runner_config.update(overrides)
        python_config_mode = "pine_defaults_plus_v2_overrides"
        override_source = str(Path(args.case))
        warnings.append("Applied flat 00_PYTHON config overrides on top of current Pine defaults.")
    if args.python_case:
        python_only_overrides = normalize_runner_overrides(load_runner_overrides(Path(args.python_case)))
        runner_config.update(python_only_overrides)
        python_config_mode = "pine_defaults_plus_v2_overrides_plus_python_only_research"
        if override_source:
            override_source = f"{override_source} + {Path(args.python_case)}"
        else:
            override_source = str(Path(args.python_case))
        warnings.append("Applied extra Python-only overrides after Pine-aligned config.")

    if args.case and tracker_case and not args.assume_pine_case_aligned:
        warnings.append(
            "Custom --case only changes Python config. Tracker auto-update is disabled unless Pine was run with the same case inputs and --assume-pine-case-aligned is passed."
        )

    try:
        effective_config = resolve_config(runner_config)
    except Exception as exc:
        print(f"ERROR: invalid 00_PYTHON config for parity run: {exc}")
        return 1

    pine_metrics = load_pine_metrics(pine_payload, float(effective_config["initial_capital"]))

    effective_config_json = REPO_ROOT / args.effective_config_json
    effective_config_json.parent.mkdir(parents=True, exist_ok=True)
    with effective_config_json.open("w", encoding="utf-8") as handle:
        json.dump(effective_config, handle, indent=2, ensure_ascii=False)

    print(f"Pine metrics loaded: {pine_json.name}")
    print(f"Exact OHLCV dataset: {dataset_csv}")
    print(f"Python config mode : {python_config_mode}")
    if override_source:
        print(f"Override source    : {override_source}")

    bars = bars_from_signals(signals_df)

    print("\nRunning canonical V2 Python runner on exact PineTS bars...")
    try:
        python_trades, runner_summary = run_v2_runner(bars, effective_config, signals_df=signals_df, chart_timezone=args.chart_timezone)
    except Exception as exc:
        print(f"Python runner FAILED: {exc}")
        import traceback

        traceback.print_exc()
        return 1

    python_metrics = compute_metrics(python_trades, float(effective_config["initial_capital"]))
    metric_rows = build_metric_rows(pine_metrics, python_metrics)

    mincontract = max(_float_or_default(signals_meta.get("syminfo_patched", {}).get("mincontract"), 0.001), 1e-9)
    mintick = max(_float_or_default(signals_meta.get("syminfo_patched", {}).get("mintick"), 0.01), 1e-9)
    pine_trades = build_pine_trade_frame(pine_payload)
    python_trade_frame = build_python_trade_frame(python_trades)
    trade_summary, trade_report = summarize_trade_parity(
        pine_trades,
        python_trade_frame,
        price_tol=max(mintick * 5.0, 0.05),
        qty_tol=max(mincontract, 1e-6),
    )

    first_mismatch = trade_summary.get("first_mismatch") or {}
    eval_start_iso = str(signals_df["timestamp_dt"].min())
    if trade_summary["extra_python_trades"] > 0 and str(first_mismatch.get("python_entry_time", "")) == eval_start_iso:
        warnings.append(
            "Python opened on the first eval bar while Pine stayed flat. This now points to a real V2 signal/entry semantics mismatch."
        )
    if not trade_summary["strict_pass"] and trade_summary["core_match_count"] == 0 and trade_summary["compared"] > 0:
        warnings.append(
            "Shared-prefix trade parity is fully mismatched. Inspect parity_trade_report.csv and python_trades_v2.csv before changing Pine."
        )

    trade_report_csv = REPO_ROOT / args.trade_report_csv
    trade_report_csv.parent.mkdir(parents=True, exist_ok=True)
    trade_report.to_csv(trade_report_csv, index=False, encoding="utf-8")

    python_trades_csv = REPO_ROOT / args.python_trades_csv
    python_trades_csv.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(python_trades).to_csv(python_trades_csv, index=False, encoding="utf-8")

    metric_mismatches = sum(1 for row in metric_rows if row["status"] == "MISMATCH")
    verdict = "PASS" if (trade_summary["strict_pass"] and metric_mismatches == 0) else "MISMATCH"

    print()
    print_metric_table(metric_rows, pine_metrics, "MTC_V2 00_PYTHON Runner")
    print_trade_summary(trade_summary, pine_metrics, python_metrics)
    print("=" * 78)
    print(f"  FINAL VERDICT: {verdict}")
    print("=" * 78)

    out_path = REPO_ROOT / args.out
    save_report(
        out_path,
        verdict=verdict,
        pine_metrics=pine_metrics,
        python_metrics=python_metrics,
        metric_rows=metric_rows,
        trade_summary=trade_summary,
        trade_report_csv=trade_report_csv,
        python_trades_csv=python_trades_csv,
        dataset_csv=dataset_csv,
        effective_config_json=effective_config_json,
        python_config_mode=python_config_mode,
        warnings=warnings,
        override_source=override_source,
        runner_summary=runner_summary,
    )
    if tracker_case and (not args.case or args.assume_pine_case_aligned):
        update_tracker_artifacts(
            tracker_case=tracker_case,
            tracker_agent=str(args.tracker_agent),
            verdict=verdict,
            override_source=override_source,
            effective_config=effective_config,
            signals_meta=signals_meta,
            pine_metrics=pine_metrics,
            python_metrics=python_metrics,
            warnings=warnings,
            trade_report_csv=trade_report_csv,
            python_trades_csv=python_trades_csv,
            effective_config_json=effective_config_json,
            report_json=out_path,
        )
    elif tracker_case and args.case and not args.assume_pine_case_aligned:
        print(f"Skipped tracker update for {tracker_case}: Pine case inputs were not explicitly aligned.")
    print(f"\nSaved report: {out_path}")
    print(f"Saved trade report: {trade_report_csv}")
    print(f"Saved python trades: {python_trades_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())






