"""
run_2025_audit.py
-----------------
Python-only backtest for BTCUSDT 1h, 2025-01-01 to 2025-12-31.
Runs every DONE AUTO_* case from the parity tracker and writes
results to a new "2025_AUDIT" sheet in MTC_V2_PARITY_CASE_TRACKER.xlsx.

Usage:
    python run_2025_audit.py
"""

from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
V2_PYTHON_ROOT = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/00_PYTHON"
PINE_SOURCE = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/01_PINE/MTC_V2.pine"
TRACKER_WORKBOOK = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASE_TRACKER.xlsx"
TRACKER_CSV = REPO_ROOT / "MTC_COMMAND_CENTER/01_MTC_PROJECT/05_PARITY/MTC_V2_PARITY_CASES.csv"
DATA_FILE = REPO_ROOT / "MTC_COMMAND_CENTER/02_MTC_BACKTEST/data/BTCUSDT_1h_20180701_20260308.parquet"

START_DATE = datetime(2025, 1, 1, tzinfo=timezone.utc)
END_DATE = datetime(2025, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
SHEET_NAME = "2025_AUDIT"

sys.path.insert(0, str(V2_PYTHON_ROOT))
from mtc_v2.core.runner import Runner  # noqa: E402
from mtc_v2.core.types import Bar  # noqa: E402

# Reuse helpers from parity_compare
sys.path.insert(0, str(SCRIPT_DIR))
from parity_compare import (  # noqa: E402
    extract_supported_pine_defaults,
    load_tracker_rows,
    tracker_row_to_pine_overrides,
    overlay_pine_overrides,
    build_runner_config_from_pine_defaults,
    run_v2_runner,
    compute_metrics,
)

# BTCUSDT Binance spot instrument spec
BTCUSDT_META = {
    "symbol": "BTCUSDT",
    "timeframe": "60",
    "bars": None,
    "syminfo_patched": {
        "mintick": 0.01,
        "mincontract": 0.001,
        "pointvalue": 1.0,
    },
}


def load_2025_bars() -> list[Bar]:
    df = pd.read_parquet(DATA_FILE)

    # timestamp is a column, not the index
    ts_col = pd.to_datetime(df["timestamp"], utc=True)
    mask = (ts_col >= START_DATE) & (ts_col <= END_DATE)
    df = df[mask].copy()
    df = df.sort_values("timestamp").reset_index(drop=True)
    ts_col = pd.to_datetime(df["timestamp"], utc=True)

    print(f"  Loaded {len(df)} bars: {ts_col.iloc[0]} → {ts_col.iloc[-1]}")

    bars: list[Bar] = []
    for i, row in df.iterrows():
        bars.append(
            Bar(
                timestamp=pd.to_datetime(row["timestamp"], utc=True).to_pydatetime(),
                open=float(row["open"]),
                high=float(row["high"]),
                low=float(row["low"]),
                close=float(row["close"]),
                volume=float(row.get("volume", 0.0)),
                bar_index=int(i),
            )
        )
    return bars


def run_case(
    case_row: dict[str, str],
    pine_defaults: dict[str, Any],
    bars: list[Bar],
) -> dict[str, Any]:
    overrides = tracker_row_to_pine_overrides(case_row)
    merged = overlay_pine_overrides(pine_defaults, overrides)
    config, warnings = build_runner_config_from_pine_defaults(merged, BTCUSDT_META)

    trades, _ = run_v2_runner(bars, config)
    metrics = compute_metrics(trades, float(config["initial_capital"]))

    if warnings:
        print(f"    WARNINGS: {warnings}")

    return metrics


def _pf_str(pf: Any) -> str:
    if pf == "inf" or pf == float("inf"):
        return "∞"
    try:
        return f"{float(pf):.4f}"
    except Exception:
        return str(pf)


def write_audit_sheet(results: list[dict[str, Any]]) -> None:
    wb = load_workbook(TRACKER_WORKBOOK)

    # Remove old sheet if exists
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    section_fills = {
        "id": PatternFill("solid", start_color="D6E4F0"),
        "config": PatternFill("solid", start_color="EBF5FB"),
        "python": PatternFill("solid", start_color="E8F8E8"),
    }

    columns = [
        ("Case ID", "case_id", "id", 12),
        ("Layer", "layer", "id", 8),
        ("Description", "primary_change", "id", 28),
        ("enable_long", "enable_long", "config", 11),
        ("enable_short", "enable_short", "config", 11),
        ("allow_flip", "allow_flip", "config", 10),
        ("SL mode", "sl_mode", "config", 10),
        ("sl_atr_mult", "sl_atr_mult", "config", 10),
        ("risk_pct", "risk_per_long_pct", "config", 9),
        ("lev_cap", "max_leverage_cap", "config", 8),
        ("equity_src", "equity_source", "config", 10),
        ("Trades", "total_trades", "python", 8),
        ("Long", "long_trades", "python", 7),
        ("Short", "short_trades", "python", 7),
        ("Win %", "win_rate_pct", "python", 8),
        ("Net PnL %", "net_pnl_pct", "python", 10),
        ("Profit Factor", "profit_factor", "python", 12),
        ("Max DD %", "max_drawdown_pct", "python", 10),
    ]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"MTC_V2 — Python Backtest Audit  |  BTCUSDT 1h  |  2025-01-01 → 2025-12-31"
    title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="0D3349")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header row
    for col_idx, (header, _, section, width) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, result in enumerate(results, start=3):
        for col_idx, (_, key, section, _) in enumerate(columns, start=1):
            raw = result.get(key, "")
            if key == "profit_factor":
                raw = _pf_str(raw)
            elif isinstance(raw, float):
                raw = round(raw, 4)
            cell = ws.cell(row=row_idx, column=col_idx, value=raw)
            cell.font = data_font
            cell.fill = section_fills.get(section, PatternFill())
            cell.alignment = center if section in ("python", "config") else left

        ws.row_dimensions[row_idx].height = 16

    # Freeze panes
    ws.freeze_panes = "A3"

    # Timestamp note
    note_row = len(results) + 4
    ws.cell(row=note_row, column=1).value = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Python-only run. Pine TW comparison is manual."
    ws.cell(row=note_row, column=1).font = Font(name="Arial", italic=True, size=9, color="888888")

    wb.save(TRACKER_WORKBOOK)
    print(f"\n  Sheet '{SHEET_NAME}' written to {TRACKER_WORKBOOK.name}")


def main() -> None:
    print("=== MTC_V2 2025 Annual Audit ===\n")
    print(f"  Data  : {DATA_FILE.name}")
    print(f"  Period: 2025-01-01 → 2025-12-31  (BTCUSDT 1h)\n")

    bars = load_2025_bars()
    pine_defaults = extract_supported_pine_defaults(PINE_SOURCE)
    tracker_rows = load_tracker_rows(TRACKER_CSV)

    # All AUTO_* cases that are not future layer gates
    done_rows = [
        r for r in tracker_rows
        if r["case_id"].startswith("AUTO_")
        and r["status"] not in ("BLOCKED_BY_LAYER_ORDER", "PLANNED_LAYER_GATE")
    ]
    print(f"  Running {len(done_rows)} AUTO_* cases (DONE + NOT_RUN)...\n")

    results: list[dict[str, Any]] = []
    for row in done_rows:
        case_id = row["case_id"]
        desc = row.get("primary_change", "")
        print(f"  [{case_id}] {desc}")
        try:
            metrics = run_case(row, pine_defaults, bars)
            entry = {
                "case_id": case_id,
                "layer": row.get("layer", ""),
                "primary_change": desc,
                "enable_long": row.get("enable_long", ""),
                "enable_short": row.get("enable_short", ""),
                "allow_flip": row.get("allow_flip", ""),
                "sl_mode": row.get("sl_mode", ""),
                "sl_atr_mult": row.get("sl_atr_mult", ""),
                "risk_per_long_pct": row.get("risk_per_long_pct", ""),
                "max_leverage_cap": row.get("max_leverage_cap", ""),
                "equity_source": row.get("equity_source", ""),
                **metrics,
            }
            results.append(entry)
            print(
                f"    → trades={metrics['total_trades']}  win%={metrics['win_rate_pct']}  "
                f"pnl%={metrics['net_pnl_pct']}  dd%={metrics['max_drawdown_pct']}  "
                f"pf={_pf_str(metrics['profit_factor'])}"
            )
        except Exception as exc:
            print(f"    ERROR: {exc}")
            results.append({
                "case_id": case_id,
                "layer": row.get("layer", ""),
                "primary_change": desc,
                "enable_long": row.get("enable_long", ""),
                "enable_short": row.get("enable_short", ""),
                "allow_flip": row.get("allow_flip", ""),
                "sl_mode": row.get("sl_mode", ""),
                "sl_atr_mult": row.get("sl_atr_mult", ""),
                "risk_per_long_pct": row.get("risk_per_long_pct", ""),
                "max_leverage_cap": row.get("max_leverage_cap", ""),
                "equity_source": row.get("equity_source", ""),
                "total_trades": "ERROR",
                "long_trades": "",
                "short_trades": "",
                "win_rate_pct": str(exc)[:60],
                "net_pnl_pct": "",
                "profit_factor": "",
                "max_drawdown_pct": "",
            })

    write_audit_sheet(results)

    print("\n=== SUMMARY ===")
    print(f"{'Case':<12} {'Trades':>7} {'Win%':>7} {'PnL%':>9} {'MaxDD%':>8} {'PF':>8}")
    print("-" * 58)
    for r in results:
        if r.get("total_trades") == "ERROR":
            print(f"{r['case_id']:<12}  ERROR")
        else:
            print(
                f"{r['case_id']:<12} {r['total_trades']:>7} {r['win_rate_pct']:>7.2f} "
                f"{r['net_pnl_pct']:>9.4f} {r['max_drawdown_pct']:>8.2f} "
                f"{_pf_str(r['profit_factor']):>8}"
            )
    print()


if __name__ == "__main__":
    main()
