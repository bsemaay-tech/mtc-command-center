from __future__ import annotations

import argparse
import csv
import glob
import sys
from dataclasses import dataclass, replace
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "CASE_SETUP_GUIDE_L4_120_baseline.xlsx"
CSV_PATH = ROOT / "01_TW_CHART_DATA" / "BINANCE_BTCUSDT.P, 60_consolidated_stable.csv"
PYTHON_ROOT = ROOT.parent / "00_PYTHON"
RANGE_SEPARATOR = chr(8212)
QTY_STEP = 1e-6
QTY_TOL = 1e-5
PRICE_TOL = 0.11
TIME_TOL_SECONDS = 3600
SPRING_SHIFT_CUTOFF = datetime(2025, 3, 30, 0, 0, tzinfo=timezone.utc)
FALL_SHIFT_CUTOFF = datetime(2025, 10, 26, 0, 0, tzinfo=timezone.utc)
SHIFT_LABEL = "eu_dst_2025_spring_window_minus_1h"
SHIFT_DELTA = timedelta(hours=-1)

sys.path.insert(0, str(PYTHON_ROOT))

from mtc_v2.core.runner import Runner  # noqa: E402
from mtc_v2.core.types import Bar  # noqa: E402


@dataclass(frozen=True, slots=True)
class Trade:
    side: str
    entry_time: datetime
    exit_time: datetime
    entry_price: float
    exit_price: float
    qty: float
    pnl: float
    reason: str | None = None


@dataclass(frozen=True, slots=True)
class CompareSummary:
    status: str
    detail: str
    variant: str
    margin_calls: int


def parse_range(text: str) -> tuple[datetime, datetime]:
    parts = [part.strip() for part in str(text).split(RANGE_SEPARATOR)]
    if len(parts) != 2:
        raise ValueError(f"Unexpected range format: {text!r}")
    fmt = "%b %d, %Y, %H:%M"
    start = datetime.strptime(parts[0], fmt).replace(tzinfo=timezone.utc)
    end = datetime.strptime(parts[1], fmt).replace(tzinfo=timezone.utc)
    return start, end


def load_all_rows() -> list[dict[str, str]]:
    with CSV_PATH.open(newline="") as handle:
        return list(csv.DictReader(handle))


def build_bars(rows: list[dict[str, str]], start: datetime, end: datetime) -> list[Bar]:
    bars: list[Bar] = []
    for row in rows:
        timestamp = datetime.fromtimestamp(int(float(row["time"])), tz=timezone.utc)
        if start <= timestamp <= end:
            bars.append(
                Bar(
                    timestamp=timestamp,
                    open=float(row["open"]),
                    high=float(row["high"]),
                    low=float(row["low"]),
                    close=float(row["close"]),
                    volume=0.0,
                    bar_index=len(bars),
                )
            )
    return bars


def case_export_file(case_dir: Path) -> Path:
    matches = glob.glob(str(case_dir / "*.xlsx"))
    if len(matches) != 1:
        raise FileNotFoundError(f"{case_dir.name}: expected exactly 1 xlsx, found {len(matches)}")
    return Path(matches[0])


def workbook_case_ids() -> list[str]:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Cases"]
    headers = [cell.value for cell in sheet[1]]
    case_col = headers.index("case_id") + 1
    case_ids: list[str] = []
    for row_index in range(2, sheet.max_row + 1):
        case_id = sheet.cell(row_index, case_col).value
        if isinstance(case_id, str) and case_id.startswith("case_"):
            case_ids.append(case_id)
    return case_ids


def export_props(workbook: openpyxl.Workbook) -> dict[str, object]:
    return {
        key: value
        for key, value in workbook["Properties"].iter_rows(min_row=2, values_only=True)
        if key
    }


def export_trades(workbook: openpyxl.Workbook) -> tuple[list[Trade], int]:
    trade_map: dict[int, dict[str, object]] = {}
    margin_calls = 0
    for row in workbook["List of trades"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        trade_no = int(row[0])
        event_type = str(row[1])
        timestamp = row[2].replace(tzinfo=timezone.utc)
        price = float(row[4])
        qty = float(row[5])
        pnl = float(row[7])
        reason = str(row[3])
        if reason == "Margin call":
            margin_calls += 1
        record = trade_map.setdefault(trade_no, {})
        if event_type.startswith("Entry "):
            record["side"] = event_type.split()[1].lower()
            record["entry_time"] = timestamp
            record["entry_price"] = price
            record["qty"] = qty
        elif event_type.startswith("Exit "):
            record["exit_time"] = timestamp
            record["exit_price"] = price
            record["pnl"] = pnl
            record["reason"] = reason

    trades: list[Trade] = []
    for trade_no in sorted(trade_map):
        record = trade_map[trade_no]
        trades.append(
            Trade(
                side=str(record["side"]),
                entry_time=record["entry_time"],
                exit_time=record["exit_time"],
                entry_price=float(record["entry_price"]),
                exit_price=float(record["exit_price"]),
                qty=float(record["qty"]),
                pnl=float(record["pnl"]),
                reason=str(record.get("reason") or ""),
            )
        )
    return trades, margin_calls


def build_config(props: dict[str, object]) -> dict[str, object]:
    sl_mode = str(props["SL Mode"])
    tp_mode = str(props["TP Mode"])
    return {
        "enable_long": props["Enable Long"] == "On",
        "enable_short": props["Enable Short"] == "On",
        "allow_flip": props["Allow Flip"] == "On",
        "regime_lock": props["Regime Lock"] == "On",
        "max_entries": int(props["Max Entries"]),
        # Current Python contract still rejects nonzero cooldown while Pine keeps
        # this input inert.  Force the replay into the active current code path.
        "cooldown_bars": 0,
        "signal_mode": str(props["Signal Mode"]),
        "st_atr_len": int(props["ATR Len"]),
        "st_factor": float(props["Factor"]),
        "st_use_wicks": props["Use Wicks"] == "On",
        "st_use_ha": props["Use HA"] == "On",
        "instrument_symbol": str(props["Symbol"]),
        "instrument_point_value": float(props["Point value"]),
        "instrument_price_tick": float(props["Tick size"]),
        # TradingView exports fractional qty; match that instead of the stale
        # legacy fixed-qty assumptions in older parity scripts.
        "instrument_qty_step": QTY_STEP,
        "instrument_min_qty": QTY_STEP,
        "instrument_min_notional": 0.0,
        "instrument_contract_multiplier": float(props["Point value"]),
        "initial_capital": float(props["Initial capital"]),
        "margin_long_pct": float(str(props["Margin for long positions"]).rstrip("%")),
        "margin_short_pct": float(str(props["Margin for short positions"]).rstrip("%")),
        "fixed_qty": 1.0,
        "risk_per_long_pct": float(props["Risk % Long"]),
        "risk_per_short_pct": float(props["Short"]),
        "fallback_size_pct": float(props["Fallback %"]),
        "max_leverage_cap": float(props["Max Lev"]),
        "execution_profile_id": "close_only_deterministic_v2",
        "equity_source": "Realized",
        "use_notional_assert": False,
        "tw_audit_semantics_mode": "off",
        "tw_reversal_reentry_mode": "local",
        "tw_reversal_reentry_delay_bars": 0,
        "tw_margin_call_mode": "off",
        "tw_margin_call_split_entries": False,
        "tw_be_semantics_mode": "local",
        "tw_trailing_semantics_mode": "local",
        "use_sl": sl_mode == "ATR",
        "use_sl_atr": sl_mode == "ATR",
        "sl_atr_len": int(props["ATR Length"]),
        "sl_atr_mult": float(props["ATR Mult"]),
        "tp_mode": tp_mode,
        "tp_atr_len": int(props["TP ATR Length"]),
        "tp_atr_mult": float(props["TP ATR Mult"]),
    }


def replay_export_truth(props: dict[str, object], rows: list[dict[str, str]]) -> list[Trade]:
    start, end = parse_range(str(props["Backtesting range"]))
    bars = build_bars(rows, start, end)
    runner = Runner(build_config(props))

    trades: list[Trade] = []
    open_trade: Trade | None = None
    for bar in bars:
        prev_position = runner.state.position
        runner.run([bar])
        if runner.state.closed_this_bar_reason is not None:
            if prev_position is None or open_trade is None:
                raise RuntimeError("close event without an open trade snapshot")
            trades.append(
                Trade(
                    side=prev_position.side,
                    entry_time=open_trade.entry_time,
                    exit_time=bar.timestamp,
                    entry_price=open_trade.entry_price,
                    exit_price=float(runner.state.last_exit_price),
                    qty=float(prev_position.qty),
                    pnl=float(runner.state.last_realized_pnl),
                    reason=runner.state.closed_this_bar_reason,
                )
            )
            open_trade = None
        if runner.state.opened_this_bar_reason is not None:
            position = runner.state.position
            if position is None:
                raise RuntimeError("open event without position state")
            open_trade = Trade(
                side=position.side,
                entry_time=bar.timestamp,
                exit_time=bar.timestamp,
                entry_price=float(position.entry_price),
                exit_price=float(position.entry_price),
                qty=float(position.qty),
                pnl=0.0,
                reason=runner.state.opened_this_bar_reason,
            )
    return trades


def shift_trade_timestamps(trades: list[Trade]) -> list[Trade]:
    shifted: list[Trade] = []
    for trade in trades:
        entry_time = (
            trade.entry_time + SHIFT_DELTA
            if SPRING_SHIFT_CUTOFF <= trade.entry_time < FALL_SHIFT_CUTOFF
            else trade.entry_time
        )
        exit_time = (
            trade.exit_time + SHIFT_DELTA
            if SPRING_SHIFT_CUTOFF <= trade.exit_time < FALL_SHIFT_CUTOFF
            else trade.exit_time
        )
        shifted.append(replace(trade, entry_time=entry_time, exit_time=exit_time))
    return shifted


def strip_terminal_open_tail(trades: list[Trade]) -> tuple[list[Trade], bool]:
    if trades and (trades[-1].reason or "") == "Open":
        return trades[:-1], True
    return trades, False


def exact_match(left: Trade, right: Trade) -> bool:
    return (
        left.side == right.side
        and left.entry_time == right.entry_time
        and left.exit_time == right.exit_time
        and abs(left.qty - right.qty) <= QTY_TOL
        and abs(left.entry_price - right.entry_price) <= PRICE_TOL
        and abs(left.exit_price - right.exit_price) <= PRICE_TOL
    )


def tolerant_match(left: Trade, right: Trade) -> bool:
    return (
        left.side == right.side
        and abs((left.entry_time - right.entry_time).total_seconds()) <= TIME_TOL_SECONDS
        and abs((left.exit_time - right.exit_time).total_seconds()) <= TIME_TOL_SECONDS
        and abs(left.qty - right.qty) <= QTY_TOL
        and abs(left.entry_price - right.entry_price) <= PRICE_TOL
        and abs(left.exit_price - right.exit_price) <= PRICE_TOL
    )


def lcs_match_count(left: list[Trade], right: list[Trade], matcher: callable) -> int:
    @lru_cache(maxsize=None)
    def dp(i: int, j: int) -> int:
        if i == len(left) or j == len(right):
            return 0
        if matcher(left[i], right[j]):
            return 1 + dp(i + 1, j + 1)
        return max(dp(i + 1, j), dp(i, j + 1))

    return dp(0, 0)


def best_compare_variant(tv_trades: list[Trade], py_trades: list[Trade]) -> tuple[str, list[Trade], int, int]:
    raw_exact = lcs_match_count(tv_trades, py_trades, exact_match)
    raw_tol = lcs_match_count(tv_trades, py_trades, tolerant_match)
    shifted = shift_trade_timestamps(tv_trades)
    shifted_exact = lcs_match_count(shifted, py_trades, exact_match)
    shifted_tol = lcs_match_count(shifted, py_trades, tolerant_match)

    raw_score = (raw_tol, raw_exact)
    shifted_score = (shifted_tol, shifted_exact)
    if shifted_score > raw_score:
        return SHIFT_LABEL, shifted, shifted_exact, shifted_tol
    return "raw", tv_trades, raw_exact, raw_tol


def first_zipped_exact_diff(left: list[Trade], right: list[Trade]) -> str:
    for index, (tv_trade, py_trade) in enumerate(zip(left, right), start=1):
        if not exact_match(tv_trade, py_trade):
            return f"T{index:03d}"
    if len(left) != len(right):
        return f"LEN:{len(left)}!={len(right)}"
    return "OK"


def compare_trades(tv_trades: list[Trade], py_trades: list[Trade], *, margin_calls: int) -> CompareSummary:
    variant, candidate_tv, exact_lcs, tol_lcs = best_compare_variant(tv_trades, py_trades)
    candidate_core, tail_open = strip_terminal_open_tail(candidate_tv)

    if tail_open:
        exact_lcs = lcs_match_count(candidate_core, py_trades, exact_match)
        tol_lcs = lcs_match_count(candidate_core, py_trades, tolerant_match)

    first_diff = first_zipped_exact_diff(candidate_core, py_trades)
    min_len = min(len(candidate_core), len(py_trades))

    if margin_calls > 0:
        status = "BLOCKED_MARGIN"
    elif len(candidate_core) == len(py_trades) and exact_lcs == len(candidate_core):
        status = "PASS_TAIL_OPEN" if tail_open else "PASS"
    elif len(candidate_core) == len(py_trades) and tol_lcs == len(candidate_core):
        status = "NEAR_TOL"
    elif abs(len(candidate_core) - len(py_trades)) <= 1 and tol_lcs == min_len:
        status = "CLOSE_ALIGN"
    elif len(candidate_tv) == len(py_trades) and exact_lcs == len(candidate_tv):
        status = "PASS"
    elif len(candidate_tv) == len(py_trades) and tol_lcs == len(candidate_tv):
        status = "NEAR_TOL"
    elif abs(len(candidate_tv) - len(py_trades)) <= 1 and tol_lcs == min_len:
        status = "CLOSE_ALIGN"
    else:
        status = "FAIL"

    detail = (
        f"variant={variant} exact_lcs={exact_lcs}/{len(candidate_core)} "
        f"tol_lcs={tol_lcs}/{len(candidate_core)} py={len(py_trades)} "
        f"first={first_diff} margin_calls={margin_calls} tail_open={int(tail_open)}"
    )
    return CompareSummary(status=status, detail=detail, variant=variant, margin_calls=margin_calls)


def update_workbook(results: dict[str, CompareSummary]) -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet = workbook["Cases"]
    headers = [cell.value for cell in sheet[1]]
    col = {name: index + 1 for index, name in enumerate(headers)}

    for row_index in range(2, sheet.max_row + 1):
        case_id = sheet.cell(row_index, col["case_id"]).value
        if case_id not in results:
            continue
        summary = results[case_id]
        sheet.cell(row_index, col["compare_status"]).value = summary.status
        notes = sheet.cell(row_index, col["notes"]).value or ""
        base_note = notes.split(" | export_compare:", 1)[0]
        sheet.cell(row_index, col["notes"]).value = f"{base_note} | export_compare:{summary.detail}"

    workbook.save(WORKBOOK_PATH)


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate TradingView exports against current export-truth replay.")
    parser.add_argument("--write-workbook", action="store_true", help="Write compare_status back into the workbook.")
    args = parser.parse_args()

    rows = load_all_rows()
    print("Validator notes:")
    print(f"- qty_step replay override = {QTY_STEP}")
    print("- cooldown_bars replay override = 0 (current Python contract path)")
    print(f"- timestamp alt variant = {SHIFT_LABEL}")

    results: dict[str, CompareSummary] = {}
    for case_id in workbook_case_ids():
        export_file = case_export_file(ROOT / case_id)
        workbook = openpyxl.load_workbook(export_file, data_only=True)
        props = export_props(workbook)
        tv_trades, margin_calls = export_trades(workbook)
        py_trades = replay_export_truth(props, rows)
        summary = compare_trades(tv_trades, py_trades, margin_calls=margin_calls)
        results[case_id] = summary
        print(f"{case_id} {summary.status} {summary.detail}")

    if args.write_workbook:
        update_workbook(results)


if __name__ == "__main__":
    main()
