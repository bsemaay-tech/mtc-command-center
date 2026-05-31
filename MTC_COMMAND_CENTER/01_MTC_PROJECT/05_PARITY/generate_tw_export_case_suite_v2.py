from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
PARITY_ROOT = REPO_ROOT / "01_MASTER TEMPLATE_V2/05_PARITY"
SUITE_ROOT = PARITY_ROOT / "TW_EXPORT_CASES_V2"
WORKBOOK_PATH = PARITY_ROOT / "MTC_V2_TW_EXPORT_CASE_TRACKER_V2.xlsx"
MANIFEST_JSON = PARITY_ROOT / "MTC_V2_TW_EXPORT_CASE_SUITE_V2.json"
MANIFEST_CSV = PARITY_ROOT / "MTC_V2_TW_EXPORT_CASE_SUITE_V2.csv"
AUDIT_SCRIPT = PARITY_ROOT / "manual_tw_futures_audit.py"
REPORTS_DIR = REPO_ROOT / "reports"

HTF_TFS = ["30", "60", "120", "240", "D", "W"]
MA_TYPES = ["EMA", "SMA", "WMA", "RMA"]
SESSION_NAMES = ["New York", "London", "Asia", "Sydney", "Custom"]

CASE_CUSTOMIZATIONS = {
    "case_005": {
        "title": "regime_lock",
        "description": "regime_lock: FALSE -> TRUE; sl_mode: None -> ATR; risk_per_long/short_pct: 1.0 -> 0.1",
        "overrides": {
            "allow_flip": False,
            "regime_lock": True,
            "sl_mode": "ATR",
            "risk_per_long_pct": 0.1,
            "risk_per_short_pct": 0.1,
        },
    },
    "case_007": {
        "title": "cooldown_bars",
        "description": "cooldown_bars: 0 -> 2 (paired with max_entries=2 for stateful entry-spacing check)",
        "overrides": {
            "max_entries": 2,
            "cooldown_bars": 2,
        },
        "setting_effect_status_hint": "DEPENDENT_ON_L18C",
        "setting_effect_note_hint": (
            "Cooldown spacing test depends on future L18c / trend_rearm_add. "
            "Current Supertrend pulse model does not emit same-side rearm entries, "
            "so cooldown_bars cannot change trade count until L18c exists."
        ),
    },
}

ORDERED_KEYS = [
    "enable_long",
    "enable_short",
    "allow_flip",
    "regime_lock",
    "max_entries",
    "cooldown_bars",
    "risk_per_long_pct",
    "risk_per_short_pct",
    "fallback_size_pct",
    "max_leverage_cap",
    "signal_mode",
    "st_atr_len",
    "st_factor",
    "st_use_wicks",
    "st_use_ha",
    "use_ma_filter",
    "ma_type",
    "ma_length",
    "use_ma_mtf",
    "ma_htf_timeframe",
    "use_htf_trend_filter",
    "htf_trend_timeframe",
    "htf_trend_ma_type",
    "htf_trend_ma_len",
    "htf_trend_buffer_pct",
    "use_ma_slope_filter",
    "ma_slope_len",
    "ma_slope_min_pct",
    "use_mcginley_filter",
    "mcginley_length",
    "mcginley_use_higher_timeframe",
    "mcginley_htf_timeframe",
    "use_volume_filter",
    "vol_sma_length",
    "vol_sma_mult",
    "use_adx_filter",
    "adx_length",
    "adx_threshold",
    "adx_use_higher_timeframe",
    "adx_htf_timeframe",
    "use_chop_filter",
    "chop_length",
    "chop_threshold",
    "chop_use_higher_timeframe",
    "chop_htf_timeframe",
    "use_atr_vol_floor",
    "atr_vol_floor_fast_len",
    "atr_vol_floor_baseline_len",
    "atr_vol_floor_mult",
    "use_macd_regime_filter",
    "use_macd_cross_filter",
    "use_macd_hist_filter",
    "macd_hist_mode",
    "use_macd_zero_dist_filter",
    "macd_zero_dist_min",
    "macd_fast_len",
    "macd_slow_len",
    "macd_sig_len",
    "macd_source",
    "use_macd_htf_bias",
    "macd_htf_timeframe",
    "use_momentum_filter",
    "momentum_mode",
    "momentum_atr_len",
    "momentum_atr_mult",
    "momentum_roc_min_pct",
    "use_session_filter",
    "session_name",
    "session_custom_string",
    "sl_mode",
    "sl_atr_len",
    "sl_atr_mult",
    "sl_pct",
    "sl_swing_basis",
    "sl_swing_lookback",
    "sl_swing_atr_len",
    "sl_swing_atr_mult",
    "tp_mode",
    "tp_atr_len",
    "tp_atr_mult",
    "tp_percent",
    "tp_r_multiple",
    "tp1_r_multiple",
    "tp1_close_pct",
    "tp2_r_multiple",
    "use_break_even",
    "be_trigger_r",
    "be_buffer_r",
    "use_trailing",
    "trail_atr_len",
    "trail_start_r",
    "trail_distance_atr_mult",
    "exit_on_opposite_signal",
    "exit_on_ma_block",
    "exit_on_ma_slope_block",
    "exit_on_mcginley_block",
    "exit_on_htf_trend_block",
    "exit_on_vol_block",
    "exit_on_atr_vol_block",
    "exit_on_range_block",
    "exit_on_candle_pattern_block",
    "exit_on_level_prox_block",
    "use_candle_pattern_gate",
    "use_level_proximity_gate",
    "level_proximity_threshold_pct",
    "level_proximity_lookback",
    "use_time_stop",
    "time_stop_bars",
    "time_stop_condition",
    "time_stop_eod",
    "time_stop_eow",
    "use_daily_loss_limit",
    "max_daily_loss_pct",
    "use_max_trades_per_day",
    "max_trades_per_day",
    "use_max_drawdown_guard",
    "max_drawdown_pct",
    "use_consecutive_loss_halt",
    "max_consecutive_losses",
    "use_equity_curve_filter",
    "equity_ma_length",
    "use_mae_guard",
    "max_mae_pct",
    "use_guard_recovery",
    "guard_recovery_mode",
    "guard_recovery_bars",
    "guard_recovery_signals",
    "use_trade_cooldown",
    "cooldown_bars_after_exit",
    "use_confirm_transform",
    "confirm_bars",
    "confirm_close_crosses",
    "require_raw_still_true",
    "refresh_on_new_raw",
    "use_level_retest",
    "retest_timeout_bars",
    "retest_buffer_pct",
]

PAIRWISE_CASES = [
    ("pair_sl_atr_be", "Critical pairs", "SL=ATR + Break-even", {"sl_mode": "ATR", "use_break_even": True}),
    ("pair_sl_atr_trail", "Critical pairs", "SL=ATR + Trailing", {"sl_mode": "ATR", "use_trailing": True}),
    ("pair_sl_atr_tp_atr", "Critical pairs", "SL=ATR + TP=ATR", {"sl_mode": "ATR", "tp_mode": "ATR"}),
    ("pair_sl_atr_tp_r", "Critical pairs", "SL=ATR + TP=R", {"sl_mode": "ATR", "tp_mode": "R"}),
    ("pair_sl_atr_tp_multi", "Critical pairs", "SL=ATR + TP=Multi-TP", {"sl_mode": "ATR", "tp_mode": "Multi-TP"}),
    ("pair_sl_percent_tp_percent", "Critical pairs", "SL=Percent + TP=Percent", {"sl_mode": "Percent", "tp_mode": "Percent"}),
    ("pair_ma_exit", "Critical pairs", "MA filter + exit_on_ma_block", {"use_ma_filter": True, "exit_on_ma_block": True}),
    ("pair_ma_slope_exit", "Critical pairs", "MA slope + exit_on_ma_slope_block", {"use_ma_slope_filter": True, "exit_on_ma_slope_block": True}),
    ("pair_mcginley_exit", "Critical pairs", "McGinley + exit_on_mcginley_block", {"use_mcginley_filter": True, "exit_on_mcginley_block": True}),
    ("pair_htf_trend_exit", "Critical pairs", "HTF trend + exit_on_htf_trend_block", {"use_htf_trend_filter": True, "exit_on_htf_trend_block": True}),
    ("pair_volume_exit", "Critical pairs", "Volume filter + exit_on_vol_block", {"use_volume_filter": True, "exit_on_vol_block": True}),
    ("pair_atr_floor_exit", "Critical pairs", "ATR vol floor + exit_on_atr_vol_block", {"use_atr_vol_floor": True, "exit_on_atr_vol_block": True}),
    ("pair_candle_exit", "Critical pairs", "Candle gate + exit_on_candle_pattern_block", {"use_candle_pattern_gate": True, "exit_on_candle_pattern_block": True}),
    ("pair_level_exit", "Critical pairs", "Level proximity + exit_on_level_prox_block", {"use_level_proximity_gate": True, "exit_on_level_prox_block": True}),
    ("pair_confirm_require_raw", "Critical pairs", "Confirm transform + require_raw_still_true", {"use_confirm_transform": True, "require_raw_still_true": True}),
    ("pair_confirm_refresh", "Critical pairs", "Confirm transform + refresh_on_new_raw", {"use_confirm_transform": True, "refresh_on_new_raw": True}),
    ("pair_level_retest_confirm", "Critical pairs", "Level retest + confirm transform", {"use_level_retest": True, "use_confirm_transform": True}),
    ("pair_macd_htf_regime", "Critical pairs", "MACD regime + HTF bias", {"use_macd_regime_filter": True, "use_macd_htf_bias": True}),
    ("pair_macd_hist_zero", "Critical pairs", "MACD hist + zero distance", {"use_macd_hist_filter": True, "use_macd_zero_dist_filter": True}),
    ("pair_momentum_session", "Critical pairs", "Momentum ROC + session filter", {"use_momentum_filter": True, "momentum_mode": "ROC", "use_session_filter": True, "session_name": "London"}),
    ("pair_trade_cooldown_max_trades", "Critical pairs", "Trade cooldown + max trades/day", {"use_trade_cooldown": True, "use_max_trades_per_day": True}),
    ("pair_daily_loss_recovery", "Critical pairs", "Daily loss limit + guard recovery", {"use_daily_loss_limit": True, "use_guard_recovery": True}),
    ("pair_max_dd_recovery_signals", "Critical pairs", "Max DD guard + recovery signals", {"use_max_drawdown_guard": True, "use_guard_recovery": True, "guard_recovery_mode": "Signals"}),
    ("pair_range_filters", "Critical pairs", "ADX HTF + Chop HTF + range block", {"use_adx_filter": True, "adx_use_higher_timeframe": True, "use_chop_filter": True, "chop_use_higher_timeframe": True, "exit_on_range_block": True}),
    ("pair_ma_htf_htf_trend", "Critical pairs", "MA HTF + HTF trend", {"use_ma_filter": True, "use_ma_mtf": True, "use_htf_trend_filter": True}),
]

LEGACY_FIXED_KEYS = {
    "execution_profile_id",
    "equity_source",
    "use_notional_assert",
    "tw_audit_semantics_mode",
    "tw_reversal_reentry_mode",
    "tw_reversal_reentry_delay_bars",
    "tw_margin_call_mode",
    "tw_margin_call_split_entries",
    "tw_be_semantics_mode",
    "tw_trailing_semantics_mode",
}

SUMMARY_COLUMNS = [
    "case_no",
    "prev_case_diff",
    "phase",
    "description",
    "folder",
    "status",
    "export_file",
    "setting_effect_status",
    "setting_effect_note",
    "tw_observed_max_open_same_side",
    "tw_stack_entry_events",
    "tw_trade_count",
    "tw_win_rate_pct",
    "tw_net_pnl_pct",
    "tw_profit_factor",
    "pinets_trade_count",
    "pinets_win_rate_pct",
    "pinets_net_pnl_pct",
    "pinets_profit_factor",
    "python_trade_count",
    "python_win_rate_pct",
    "python_net_pnl_pct",
    "python_profit_factor",
    "tw_vs_pinets_strict",
    "tw_vs_python_strict",
    "pinets_vs_python_strict",
    "tw_vs_pinets_outcome",
    "tw_vs_python_outcome",
    "pinets_vs_python_outcome",
]


@dataclass
class CaseDef:
    case_no: int
    case_name: str
    phase: str
    title: str
    description: str
    overrides: dict[str, Any]
    prev_case_diff: str = ""
    folder: str = ""
    export_file: str = ""
    status: str = "PLANNED"
    metrics: dict[str, Any] | None = None
    setting_effect_status_hint: str = ""
    setting_effect_note_hint: str = ""


def _load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load module from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_audit_module():
    return _load_module("tw_export_suite_audit", AUDIT_SCRIPT)


def load_config_module():
    python_root = REPO_ROOT / "01_MASTER TEMPLATE_V2/00_PYTHON"
    if str(python_root) not in sys.path:
        sys.path.insert(0, str(python_root))
    from mtc_v2.core import config as config_module

    return config_module


def stringify(value: Any) -> str:
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value)


def next_option(current: Any, options: list[Any]) -> Any:
    for option in options:
        if option != current:
            return option
    return current


def alt_value(key: str, current: Any) -> Any | None:
    enum_map: dict[str, list[Any]] = {
        "ma_type": MA_TYPES,
        "htf_trend_ma_type": MA_TYPES,
        "session_name": SESSION_NAMES,
        "momentum_mode": ["ATR_BODY", "ROC"],
        "macd_hist_mode": ["POSITIVE", "RISING", "RISING_POSITIVE"],
        "sl_mode": ["None", "ATR", "Percent", "Swing+ATR"],
        "tp_mode": ["None", "ATR", "Percent", "R", "Multi-TP"],
        "sl_swing_basis": ["Wick", "Close"],
        "time_stop_condition": ["Always", "Profit Only", "Loss Only"],
        "guard_recovery_mode": ["Bars", "Signals"],
        "macd_source": ["close", "open", "hl2", "hlc3", "ohlc4"],
        "session_custom_string": ["0930-1600:America/New_York", "0800-1700:UTC"],
        "ma_htf_timeframe": HTF_TFS,
        "htf_trend_timeframe": HTF_TFS,
        "mcginley_htf_timeframe": HTF_TFS,
        "adx_htf_timeframe": HTF_TFS,
        "chop_htf_timeframe": HTF_TFS,
        "macd_htf_timeframe": HTF_TFS,
    }
    float_map: dict[str, Any] = {
        "st_factor": 5.0,
        "ma_slope_min_pct": 0.001,
        "risk_per_long_pct": 2.0,
        "risk_per_short_pct": 2.0,
        "fallback_size_pct": 15.0,
        "max_leverage_cap": 2.0,
        "vol_sma_mult": 1.5,
        "adx_threshold": 20.0,
        "chop_threshold": 55.0,
        "atr_vol_floor_mult": 0.8,
        "macd_zero_dist_min": 0.1,
        "momentum_atr_mult": 0.5,
        "momentum_roc_min_pct": 0.3,
        "level_proximity_threshold_pct": 0.25,
        "sl_atr_mult": 2.5,
        "sl_pct": 1.5,
        "sl_swing_atr_mult": 0.8,
        "tp_atr_mult": 2.5,
        "tp_percent": 2.5,
        "tp_r_multiple": 1.5,
        "tp1_r_multiple": 2.0,
        "tp1_close_pct": 40.0,
        "tp2_r_multiple": 4.0,
        "be_trigger_r": 1.5,
        "be_buffer_r": 0.2,
        "trail_distance_atr_mult": 1.5,
        "trail_start_r": 1.5,
        "max_daily_loss_pct": 1.0,
        "max_drawdown_pct": 5.0,
        "max_mae_pct": 1.0,
        "retest_buffer_pct": 0.2,
        "htf_trend_buffer_pct": 0.2,
    }
    int_map: dict[str, int] = {
        "max_entries": 2,
        "cooldown_bars": 2,
        "st_atr_len": 14,
        "ma_length": 100,
        "htf_trend_ma_len": 50,
        "ma_slope_len": 100,
        "mcginley_length": 100,
        "vol_sma_length": 30,
        "adx_length": 10,
        "chop_length": 10,
        "atr_vol_floor_fast_len": 10,
        "atr_vol_floor_baseline_len": 50,
        "macd_fast_len": 8,
        "macd_slow_len": 21,
        "macd_sig_len": 5,
        "momentum_atr_len": 10,
        "level_proximity_lookback": 30,
        "sl_atr_len": 10,
        "sl_swing_lookback": 10,
        "sl_swing_atr_len": 10,
        "tp_atr_len": 10,
        "trail_atr_len": 10,
        "time_stop_bars": 20,
        "max_trades_per_day": 2,
        "max_consecutive_losses": 2,
        "equity_ma_length": 10,
        "guard_recovery_bars": 10,
        "guard_recovery_signals": 1,
        "cooldown_bars_after_exit": 3,
        "confirm_bars": 2,
        "retest_timeout_bars": 20,
    }

    if key == "signal_mode":
        return None
    if key in enum_map:
        return next_option(current, enum_map[key])
    if key in int_map:
        return int_map[key] if int_map[key] != current else current + 1
    if key in float_map:
        return float_map[key] if float_map[key] != current else float(current) + 0.1
    if isinstance(current, bool):
        return not current
    if isinstance(current, int):
        return current + 1 if current > 0 else 1
    if isinstance(current, float):
        return current + 1.0 if current > 0 else 0.1
    if isinstance(current, str):
        return f"{current}_ALT"
    return None


def apply_dependencies(key: str, overrides: dict[str, Any]) -> None:
    if key in {"use_break_even", "be_trigger_r", "be_buffer_r"}:
        overrides.setdefault("sl_mode", "ATR")
        if key != "use_break_even":
            overrides.setdefault("use_break_even", True)
    if key in {"use_trailing", "trail_atr_len", "trail_start_r", "trail_distance_atr_mult"}:
        overrides.setdefault("sl_mode", "ATR")
        if key != "use_trailing":
            overrides.setdefault("use_trailing", True)
    if key in {"sl_atr_len", "sl_atr_mult"}:
        overrides.setdefault("sl_mode", "ATR")
    if key == "sl_pct":
        overrides.setdefault("sl_mode", "Percent")
    if key in {"sl_swing_basis", "sl_swing_lookback", "sl_swing_atr_len", "sl_swing_atr_mult"}:
        overrides.setdefault("sl_mode", "Swing+ATR")
    if key in {"tp_atr_len", "tp_atr_mult"}:
        overrides.setdefault("tp_mode", "ATR")
    if key == "tp_percent":
        overrides.setdefault("tp_mode", "Percent")
        overrides.setdefault("sl_mode", "ATR")
    if key == "tp_r_multiple":
        overrides.setdefault("tp_mode", "R")
        overrides.setdefault("sl_mode", "ATR")
    if key in {"tp1_r_multiple", "tp1_close_pct", "tp2_r_multiple"}:
        overrides.setdefault("tp_mode", "Multi-TP")
        overrides.setdefault("sl_mode", "ATR")
    if key == "use_ma_mtf":
        overrides.setdefault("use_ma_filter", True)
    if key == "ma_htf_timeframe":
        overrides.setdefault("use_ma_filter", True)
        overrides.setdefault("use_ma_mtf", True)
    if key in {"ma_type", "ma_length", "exit_on_ma_block"}:
        overrides.setdefault("use_ma_filter", True)
    if key in {"use_htf_trend_filter", "htf_trend_timeframe", "htf_trend_ma_type", "htf_trend_ma_len", "htf_trend_buffer_pct", "exit_on_htf_trend_block"}:
        if key != "use_htf_trend_filter":
            overrides.setdefault("use_htf_trend_filter", True)
    if key in {"use_ma_slope_filter", "ma_slope_len", "ma_slope_min_pct", "exit_on_ma_slope_block"}:
        if key != "use_ma_slope_filter":
            overrides.setdefault("use_ma_slope_filter", True)
    if key in {"use_mcginley_filter", "mcginley_length", "mcginley_use_higher_timeframe", "mcginley_htf_timeframe", "exit_on_mcginley_block"}:
        if key != "use_mcginley_filter":
            overrides.setdefault("use_mcginley_filter", True)
    if key == "mcginley_htf_timeframe":
        overrides.setdefault("mcginley_use_higher_timeframe", True)
    if key in {"use_volume_filter", "vol_sma_length", "vol_sma_mult", "exit_on_vol_block"}:
        if key != "use_volume_filter":
            overrides.setdefault("use_volume_filter", True)
    if key in {"use_adx_filter", "adx_length", "adx_threshold", "adx_use_higher_timeframe", "adx_htf_timeframe"}:
        if key != "use_adx_filter":
            overrides.setdefault("use_adx_filter", True)
    if key == "adx_htf_timeframe":
        overrides.setdefault("adx_use_higher_timeframe", True)
    if key in {"use_chop_filter", "chop_length", "chop_threshold", "chop_use_higher_timeframe", "chop_htf_timeframe"}:
        if key != "use_chop_filter":
            overrides.setdefault("use_chop_filter", True)
    if key == "chop_htf_timeframe":
        overrides.setdefault("chop_use_higher_timeframe", True)
    if key == "exit_on_range_block":
        overrides.setdefault("use_adx_filter", True)
    if key in {"use_atr_vol_floor", "atr_vol_floor_fast_len", "atr_vol_floor_baseline_len", "atr_vol_floor_mult", "exit_on_atr_vol_block"}:
        if key != "use_atr_vol_floor":
            overrides.setdefault("use_atr_vol_floor", True)
    if key in {"use_macd_regime_filter", "use_macd_cross_filter", "use_macd_hist_filter", "macd_hist_mode", "use_macd_zero_dist_filter", "macd_zero_dist_min", "macd_fast_len", "macd_slow_len", "macd_sig_len", "macd_source", "use_macd_htf_bias", "macd_htf_timeframe"}:
        if key == "macd_hist_mode":
            overrides.setdefault("use_macd_hist_filter", True)
        if key == "macd_zero_dist_min":
            overrides.setdefault("use_macd_zero_dist_filter", True)
        if key == "macd_htf_timeframe":
            overrides.setdefault("use_macd_htf_bias", True)
    if key in {"use_momentum_filter", "momentum_mode", "momentum_atr_len", "momentum_atr_mult", "momentum_roc_min_pct"}:
        if key != "use_momentum_filter":
            overrides.setdefault("use_momentum_filter", True)
        if key == "momentum_roc_min_pct":
            overrides.setdefault("momentum_mode", "ROC")
    if key in {"use_session_filter", "session_name", "session_custom_string"}:
        if key != "use_session_filter":
            overrides.setdefault("use_session_filter", True)
        if key == "session_custom_string":
            overrides.setdefault("session_name", "Custom")
    if key == "allow_flip":
        overrides.setdefault("exit_on_opposite_signal", True)
    if key in {"use_candle_pattern_gate", "exit_on_candle_pattern_block"}:
        if key != "use_candle_pattern_gate":
            overrides.setdefault("use_candle_pattern_gate", True)
    if key in {"use_level_proximity_gate", "level_proximity_threshold_pct", "level_proximity_lookback", "exit_on_level_prox_block"}:
        if key != "use_level_proximity_gate":
            overrides.setdefault("use_level_proximity_gate", True)
    if key in {"use_time_stop", "time_stop_bars", "time_stop_condition", "time_stop_eod", "time_stop_eow"}:
        if key != "use_time_stop":
            overrides.setdefault("use_time_stop", True)
    if key in {"use_daily_loss_limit", "max_daily_loss_pct"}:
        if key != "use_daily_loss_limit":
            overrides.setdefault("use_daily_loss_limit", True)
    if key in {"use_max_trades_per_day", "max_trades_per_day"}:
        if key != "use_max_trades_per_day":
            overrides.setdefault("use_max_trades_per_day", True)
    if key in {"use_max_drawdown_guard", "max_drawdown_pct"}:
        if key != "use_max_drawdown_guard":
            overrides.setdefault("use_max_drawdown_guard", True)
    if key in {"use_consecutive_loss_halt", "max_consecutive_losses"}:
        if key != "use_consecutive_loss_halt":
            overrides.setdefault("use_consecutive_loss_halt", True)
    if key in {"use_equity_curve_filter", "equity_ma_length"}:
        if key != "use_equity_curve_filter":
            overrides.setdefault("use_equity_curve_filter", True)
    if key in {"use_mae_guard", "max_mae_pct"}:
        if key != "use_mae_guard":
            overrides.setdefault("use_mae_guard", True)
    if key in {"use_guard_recovery", "guard_recovery_mode", "guard_recovery_bars", "guard_recovery_signals"}:
        if key != "use_guard_recovery":
            overrides.setdefault("use_guard_recovery", True)
    if key in {"use_trade_cooldown", "cooldown_bars_after_exit"}:
        if key != "use_trade_cooldown":
            overrides.setdefault("use_trade_cooldown", True)
    if key in {"use_confirm_transform", "confirm_bars", "confirm_close_crosses", "require_raw_still_true", "refresh_on_new_raw"}:
        if key != "use_confirm_transform":
            overrides.setdefault("use_confirm_transform", True)
    if key in {"use_level_retest", "retest_timeout_bars", "retest_buffer_pct"}:
        if key != "use_level_retest":
            overrides.setdefault("use_level_retest", True)
def build_single_case_overrides(key: str, baseline: dict[str, Any]) -> tuple[dict[str, Any], str] | None:
    current = baseline.get(key)
    candidate = alt_value(key, current)
    if candidate is None or candidate == current:
        return None
    overrides = {key: candidate}
    apply_dependencies(key, overrides)
    description = f"{key}: {stringify(current)} -> {stringify(candidate)}"
    return overrides, description


def compute_prev_case_diff(current: dict[str, Any], previous: dict[str, Any] | None, previous_case_name: str | None) -> str:
    if previous is None or previous_case_name is None:
        return "Baseline"
    changed = [key for key in current if current.get(key) != previous.get(key)]
    if not changed:
        return f"vs {previous_case_name}: no config change"
    short = changed[:8]
    parts = [f"{key}: {stringify(previous.get(key))} -> {stringify(current.get(key))}" for key in short]
    suffix = " ..." if len(changed) > len(short) else ""
    return f"vs {previous_case_name}: " + "; ".join(parts) + suffix


def load_baseline_metrics(xlsx_name: str) -> dict[str, Any] | None:
    files = sorted(REPORTS_DIR.glob("manual_tw_futures_audit_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in files:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        rows = payload if isinstance(payload, list) else payload.get("cases", [])
        if not isinstance(rows, list):
            continue
        for row in rows:
            if isinstance(row, dict) and row.get("xlsx") == xlsx_name:
                data = dict(row)
                data["audit_report_json"] = str(path.relative_to(REPO_ROOT))
                return data
    return None


def load_case_metrics(case_name: str, xlsx_name: str) -> dict[str, Any] | None:
    files = sorted(REPORTS_DIR.glob("manual_tw_futures_audit_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in files:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        rows = payload if isinstance(payload, list) else payload.get("cases", [])
        if not isinstance(rows, list):
            continue
        for row in rows:
            if isinstance(row, dict) and row.get("case") == case_name and row.get("xlsx") == xlsx_name:
                data = dict(row)
                data["audit_report_json"] = str(path.relative_to(REPO_ROOT))
                return data
    return None


def latest_suite_export(case_name: str) -> Path | None:
    case_dir = SUITE_ROOT / case_name
    files = sorted(
        [p for p in case_dir.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")],
        key=lambda p: (p.stat().st_mtime, p.name),
        reverse=True,
    )
    return files[0] if files else None


def normalize_compare_value(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 9)
    return value


def overrides_match(planned: dict[str, Any], exported: dict[str, Any]) -> bool:
    if planned.keys() != exported.keys():
        return False
    for key in planned:
        if normalize_compare_value(planned.get(key)) != normalize_compare_value(exported.get(key)):
            return False
    return True


def autosize(ws, wide_columns: set[int] | None = None) -> None:
    wide_columns = wide_columns or set()
    for idx, column in enumerate(ws.columns, start=1):
        values = ["" if cell.value is None else str(cell.value) for cell in column[:200]]
        width = max((len(v) for v in values), default=10)
        if idx in wide_columns:
            width = max(width, 36)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 52)


def write_headers(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_suite(baseline_xlsx: Path) -> tuple[list[CaseDef], list[str], list[str]]:
    audit = load_audit_module()
    chart_tz = audit.parse_chart_timezone("UTC+3")
    xlsx_info = audit.read_tv_workbook(baseline_xlsx, chart_tz)
    baseline = audit.build_case_overrides(xlsx_info)
    for key in LEGACY_FIXED_KEYS:
        baseline.pop(key, None)
    baseline_metrics = load_baseline_metrics(baseline_xlsx.name)

    ordered = [key for key in ORDERED_KEYS if key in baseline]
    ordered.extend(sorted(key for key in baseline if key not in ordered and key not in LEGACY_FIXED_KEYS))

    cases: list[CaseDef] = [
        CaseDef(
            case_no=1,
            case_name="case_001",
            phase="Baseline",
            title="Baseline",
            description="Baseline export-faithful case",
            overrides=dict(baseline),
            export_file=baseline_xlsx.name,
            status="READY",
            metrics=baseline_metrics,
        )
    ]
    skipped_keys: list[str] = []

    for key in ordered:
        built = build_single_case_overrides(key, baseline)
        if built is None:
            skipped_keys.append(key)
            continue
        overrides, description = built
        case_no = len(cases) + 1
        cases.append(
            CaseDef(
                case_no=case_no,
                case_name=f"case_{case_no:03d}",
                phase="Single-setting coverage",
                title=key,
                description=description,
                overrides={**baseline, **overrides},
            )
        )

    for slug, phase, title, override_patch in PAIRWISE_CASES:
        full = dict(baseline)
        full.update(override_patch)
        case_no = len(cases) + 1
        cases.append(
            CaseDef(
                case_no=case_no,
                case_name=f"case_{case_no:03d}",
                phase=phase,
                title=slug,
                description=title,
                overrides=full,
            )
        )

    for case in cases:
        custom = CASE_CUSTOMIZATIONS.get(case.case_name)
        if not custom:
            continue
        case.title = str(custom["title"])
        case.description = str(custom["description"])
        full = dict(baseline)
        full.update(custom["overrides"])
        case.overrides = full
        case.setting_effect_status_hint = str(custom.get("setting_effect_status_hint", ""))
        case.setting_effect_note_hint = str(custom.get("setting_effect_note_hint", ""))

    previous_overrides: dict[str, Any] | None = None
    previous_case_name: str | None = None
    for case in cases:
        case.prev_case_diff = compute_prev_case_diff(case.overrides, previous_overrides, previous_case_name)
        case.folder = str((SUITE_ROOT / case.case_name).relative_to(REPO_ROOT))
        export_path = latest_suite_export(case.case_name)
        if export_path is not None:
            case.export_file = export_path.name
            xlsx_info = audit.read_tv_workbook(export_path, chart_tz)
            exported_overrides = audit.build_case_overrides(xlsx_info)
            if overrides_match(case.overrides, exported_overrides):
                case.status = "READY"
                case.metrics = load_case_metrics(case.case_name, export_path.name) or case.metrics
            else:
                case.status = "STALE_EXPORT"
                case.metrics = None
        previous_overrides = case.overrides
        previous_case_name = case.case_name

    return cases, ordered, skipped_keys


def write_suite_files(
    cases: list[CaseDef],
    ordered_keys: list[str],
    skipped_keys: list[str],
    baseline_xlsx: Path,
    output_workbook: Path,
) -> None:
    SUITE_ROOT.mkdir(parents=True, exist_ok=True)

    for case in cases:
        folder = SUITE_ROOT / case.case_name
        folder.mkdir(parents=True, exist_ok=True)
        plan_path = folder / "case_plan.json"
        plan_payload = {
            "case_no": case.case_no,
            "case_name": case.case_name,
            "phase": case.phase,
            "title": case.title,
            "description": case.description,
            "prev_case_diff": case.prev_case_diff,
            "status": case.status,
            "baseline_reference_xlsx": str(baseline_xlsx.relative_to(REPO_ROOT)),
            "planned_overrides": case.overrides,
            "setting_effect_status_hint": case.setting_effect_status_hint,
            "setting_effect_note_hint": case.setting_effect_note_hint,
        }
        plan_path.write_text(json.dumps(plan_payload, indent=2, ensure_ascii=False), encoding="utf-8")

        note_path = folder / "PLACE_EXPORT_XLSX_HERE.txt"
        if case.case_no == 1:
            baseline_copy = folder / baseline_xlsx.name
            if not baseline_copy.exists():
                shutil.copy2(baseline_xlsx, baseline_copy)
            note_path.write_text(
                "Baseline export copied into this folder. Future reruns should replace it only with a newer baseline export if intended.\n",
                encoding="utf-8",
            )
        else:
            note_path.write_text(
                "Place the TradingView xlsx export for this planned case into this folder.\n",
                encoding="utf-8",
            )

    manifest = {
        "baseline_xlsx": str(baseline_xlsx.relative_to(REPO_ROOT)),
        "suite_root": str(SUITE_ROOT.relative_to(REPO_ROOT)),
        "case_count": len(cases),
        "single_setting_case_count": len([c for c in cases if c.phase == "Single-setting coverage"]),
        "pairwise_case_count": len([c for c in cases if c.phase == "Critical pairs"]),
        "baseline_only_keys": skipped_keys,
        "cases": [
            {
                "case_no": case.case_no,
                "case_name": case.case_name,
                "phase": case.phase,
                "title": case.title,
                "description": case.description,
                "prev_case_diff": case.prev_case_diff,
                "folder": case.folder,
                "status": case.status,
                "export_file": case.export_file,
                "setting_effect_status_hint": case.setting_effect_status_hint,
                "setting_effect_note_hint": case.setting_effect_note_hint,
                "planned_overrides": case.overrides,
            }
            for case in cases
        ],
    }
    MANIFEST_JSON.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    csv_headers = ["case_no", "case_name", "phase", "title", "description", "prev_case_diff", "folder", "status", "export_file"] + ordered_keys
    with MANIFEST_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=csv_headers)
        writer.writeheader()
        for case in cases:
            row = {
                "case_no": case.case_name,
                "case_name": case.case_name,
                "phase": case.phase,
                "title": case.title,
                "description": case.description,
                "prev_case_diff": case.prev_case_diff,
                "folder": case.folder,
                "status": case.status,
                "export_file": case.export_file,
            }
            row.update(case.overrides)
            writer.writerow(row)

    wb = Workbook()
    guide = wb.active
    guide.title = "Guide"
    guide["A1"] = "TW Export Case Suite V2"
    guide["A1"].font = Font(bold=True, size=14)
    guide["A3"] = "Baseline xlsx"
    guide["B3"] = str(baseline_xlsx.relative_to(REPO_ROOT))
    guide["A4"] = "Suite root"
    guide["B4"] = str(SUITE_ROOT.relative_to(REPO_ROOT))
    guide["A5"] = "Case count"
    guide["B5"] = len(cases)
    guide["A6"] = "Phases"
    guide["B6"] = "Baseline + Single-setting coverage + Critical pairs"
    guide["A7"] = "First two columns"
    guide["B7"] = "case_no and prev_case_diff are kept first on purpose."
    guide["A8"] = "Baseline-only keys"
    guide["B8"] = ", ".join(skipped_keys) if skipped_keys else "-"
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 110

    summary = wb.create_sheet("Summary")
    write_headers(summary, SUMMARY_COLUMNS)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}1"

    settings_headers = ["case_no", "prev_case_diff", "phase", "title", "description", "folder", "status", "export_file"] + ordered_keys
    settings = wb.create_sheet("All_Settings")
    write_headers(settings, settings_headers)
    settings.freeze_panes = "A2"
    settings.auto_filter.ref = f"A1:{get_column_letter(len(settings_headers))}1"

    for row_idx, case in enumerate(cases, start=2):
        metrics = case.metrics or {}
        summary_row = {
            "case_no": case.case_name,
            "prev_case_diff": case.prev_case_diff,
            "phase": case.phase,
            "description": case.description,
            "folder": case.folder,
            "status": case.status,
            "export_file": case.export_file,
            "setting_effect_status": metrics.get("setting_effect_status", case.setting_effect_status_hint),
            "setting_effect_note": metrics.get("setting_effect_note", case.setting_effect_note_hint),
            "tw_observed_max_open_same_side": metrics.get("tw_observed_max_open_same_side", ""),
            "tw_stack_entry_events": metrics.get("tw_stack_entry_events", ""),
            "tw_trade_count": metrics.get("tw_trades", ""),
            "tw_win_rate_pct": metrics.get("tw_win_rate_pct", ""),
            "tw_net_pnl_pct": metrics.get("tw_net_pnl_pct", ""),
            "tw_profit_factor": metrics.get("tw_profit_factor", ""),
            "pinets_trade_count": metrics.get("pine_trades", ""),
            "pinets_win_rate_pct": metrics.get("pine_win_rate_pct", ""),
            "pinets_net_pnl_pct": metrics.get("pine_net_pnl_pct", ""),
            "pinets_profit_factor": metrics.get("pine_profit_factor", ""),
            "python_trade_count": metrics.get("python_trades", ""),
            "python_win_rate_pct": metrics.get("python_win_rate_pct", ""),
            "python_net_pnl_pct": metrics.get("python_net_pnl_pct", ""),
            "python_profit_factor": metrics.get("python_profit_factor", ""),
            "tw_vs_pinets_strict": metrics.get("tw_vs_pine_strict", ""),
            "tw_vs_python_strict": metrics.get("tw_vs_python_strict", ""),
            "pinets_vs_python_strict": metrics.get("pine_vs_python_strict", ""),
            "tw_vs_pinets_outcome": metrics.get("tw_vs_pine_outcome_pass", ""),
            "tw_vs_python_outcome": metrics.get("tw_vs_python_outcome_pass", ""),
            "pinets_vs_python_outcome": metrics.get("pine_vs_python_outcome_pass", ""),
        }
        for col_idx, header in enumerate(SUMMARY_COLUMNS, start=1):
            cell = summary.cell(row=row_idx, column=col_idx, value=summary_row.get(header, ""))
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header in {"prev_case_diff", "description", "folder", "setting_effect_note"},
            )

        settings_row = {
            "case_no": case.case_name,
            "prev_case_diff": case.prev_case_diff,
            "phase": case.phase,
            "title": case.title,
            "description": case.description,
            "folder": case.folder,
            "status": case.status,
            "export_file": case.export_file,
        }
        settings_row.update(case.overrides)
        for col_idx, header in enumerate(settings_headers, start=1):
            cell = settings.cell(row=row_idx, column=col_idx, value=stringify(settings_row.get(header, "")))
            cell.alignment = Alignment(vertical="top", wrap_text=header in {"prev_case_diff", "description"})

    autosize(summary, wide_columns={2, 4, 5})
    autosize(settings, wide_columns={2, 5, 6})
    wb.save(output_workbook)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--baseline-xlsx",
        default=str(PARITY_ROOT / "case_001" / "MTC_V2_BINANCE_BTCUSDT.P_2026-04-03_827c5.xlsx"),
    )
    parser.add_argument("--out", default=str(WORKBOOK_PATH))
    args = parser.parse_args()

    baseline_xlsx = Path(args.baseline_xlsx).resolve()
    output_workbook = Path(args.out).resolve()
    cases, ordered_keys, skipped_keys = build_suite(baseline_xlsx)
    write_suite_files(cases, ordered_keys, skipped_keys, baseline_xlsx, output_workbook)
    print(f"Saved {output_workbook}")
    print(f"Saved {MANIFEST_JSON}")
    print(f"Saved {MANIFEST_CSV}")
    print(f"Created {len(cases)} case folders under {SUITE_ROOT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
