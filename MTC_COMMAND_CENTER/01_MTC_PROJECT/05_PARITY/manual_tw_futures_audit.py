from __future__ import annotations

import argparse
import csv
import json
import math
import subprocess
import sys
from dataclasses import dataclass
from datetime import timedelta, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
PARITY_ROOT = REPO_ROOT / "01_MASTER TEMPLATE_V2/05_PARITY"
REPORTS_DIR = REPO_ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

STRICT_OUTCOME_TOLERANCES: dict[str, float] = {
    "value_tol": 0.1,
    "pnl_tol": 0.1,
    "pnl_pct_tol": 0.02,
    "excursion_tol": 0.1,
    "excursion_pct_tol": 0.02,
}

STRICT_TRADE_TOLERANCES: dict[str, float] = {
    "price_tol": 0.05,
    "qty_tol": 0.001,
}

SOFT_TRADE_TOLERANCES: dict[str, float] = {
    "price_tol": 0.05,
    "qty_tol": 0.005,
}

SOFT_OUTCOME_TOLERANCES: dict[str, float] = {
    "value_tol": 50.0,
    "pnl_tol": 5.0,
    "pnl_pct_tol": 0.05,
    "excursion_tol": 250.0,
    "excursion_pct_tol": 0.25,
}

PROPERTY_ROW_TO_OVERRIDE: dict[tuple[str, int], tuple[str, str]] = {
    ("Enable Long", 1): ("enable_long", "bool"),
    ("Enable Short", 1): ("enable_short", "bool"),
    ("Allow Flip", 1): ("allow_flip", "bool"),
    ("Regime Lock", 1): ("regime_lock", "bool"),
    ("Max Entries", 1): ("max_entries", "int"),
    ("Cooldown Bars", 1): ("cooldown_bars", "int"),
    ("Signal Mode", 1): ("signal_mode", "string"),
    ("ATR Len", 1): ("st_atr_len", "int"),
    ("Factor", 1): ("st_factor", "float"),
    ("Use Wicks", 1): ("st_use_wicks", "bool"),
    ("Use HA", 1): ("st_use_ha", "bool"),
    ("Use HA [Not Supported]", 1): ("st_use_ha", "bool"),
    ("Custom Session", 1): ("session_custom_string", "string"),
    ("Use MA Filter", 1): ("use_ma_filter", "bool"),
    ("Type", 1): ("ma_type", "string"),
    ("Length", 1): ("ma_length", "int"),
    ("Use HTF", 1): ("use_ma_mtf", "bool"),
    ("HTF TF", 1): ("ma_htf_timeframe", "string"),
    ("Use MA Slope", 1): ("use_ma_slope_filter", "bool"),
    ("Slope Len", 1): ("ma_slope_len", "int"),
    ("Min Ratio", 1): ("ma_slope_min_pct", "float"),
    ("Use McGinley", 1): ("use_mcginley_filter", "bool"),
    ("McGinley Len", 1): ("mcginley_length", "int"),
    ("MCG Use HTF", 1): ("mcginley_use_higher_timeframe", "bool"),
    ("MCG HTF TF", 1): ("mcginley_htf_timeframe", "string"),
    ("Use Volume", 1): ("use_volume_filter", "bool"),
    ("Vol SMA Len", 1): ("vol_sma_length", "int"),
    ("Vol Mult", 1): ("vol_sma_mult", "float"),
    ("Use ADX", 1): ("use_adx_filter", "bool"),
    ("ADX Len", 1): ("adx_length", "int"),
    ("ADX Min", 1): ("adx_threshold", "float"),
    ("ADX Use HTF", 1): ("adx_use_higher_timeframe", "bool"),
    ("ADX HTF TF", 1): ("adx_htf_timeframe", "string"),
    ("Use Chop", 1): ("use_chop_filter", "bool"),
    ("Chop Len", 1): ("chop_length", "int"),
    ("Chop Max", 1): ("chop_threshold", "float"),
    ("Chop Use HTF", 1): ("chop_use_higher_timeframe", "bool"),
    ("Chop HTF TF", 1): ("chop_htf_timeframe", "string"),
    ("Use ATR Floor", 1): ("use_atr_vol_floor", "bool"),
    ("ATR Fast", 1): ("atr_vol_floor_fast_len", "int"),
    ("ATR Baseline", 1): ("atr_vol_floor_baseline_len", "int"),
    ("ATR Floor Mult", 1): ("atr_vol_floor_mult", "float"),
    ("Use MACD Regime", 1): ("use_macd_regime_filter", "bool"),
    ("Use MACD Cross", 1): ("use_macd_cross_filter", "bool"),
    ("Use MACD Hist", 1): ("use_macd_hist_filter", "bool"),
    ("Hist Mode", 1): ("macd_hist_mode", "string"),
    ("Use MACD Zero Dist", 1): ("use_macd_zero_dist_filter", "bool"),
    ("Zero Dist Min", 1): ("macd_zero_dist_min", "float"),
    ("MACD Fast", 1): ("macd_fast_len", "int"),
    ("MACD Slow", 1): ("macd_slow_len", "int"),
    ("MACD Sig", 1): ("macd_sig_len", "int"),
    ("MACD Src", 1): ("macd_source", "string"),
    ("MACD HTF Bias", 1): ("use_macd_htf_bias", "bool"),
    ("MACD HTF TF", 1): ("macd_htf_timeframe", "string"),
    ("Use Momentum", 1): ("use_momentum_filter", "bool"),
    ("Mode", 1): ("momentum_mode", "string"),
    ("ATR Len", 2): ("momentum_atr_len", "int"),
    ("ATR Mult", 1): ("momentum_atr_mult", "float"),
    ("ROC Min %", 1): ("momentum_roc_min_pct", "float"),
    ("Use Session", 1): ("use_session_filter", "bool"),
    ("Session", 1): ("session_name", "string"),
    ("Candle Pattern Gate", 1): ("use_candle_pattern_gate", "bool"),
    ("Level Lookback", 1): ("level_proximity_lookback", "int"),
    ("Level Proximity Gate", 1): ("use_level_proximity_gate", "bool"),
    ("Threshold %", 1): ("level_proximity_threshold_pct", "float"),
    ("Use HTF Trend", 1): ("use_htf_trend_filter", "bool"),
    ("HTF TF", 2): ("htf_trend_timeframe", "string"),
    ("MA Type", 1): ("htf_trend_ma_type", "string"),
    ("MA Len", 1): ("htf_trend_ma_len", "int"),
    ("Buffer %", 1): ("htf_trend_buffer_pct", "float"),
    ("Risk % Long", 1): ("risk_per_long_pct", "float"),
    ("Short", 1): ("risk_per_short_pct", "float"),
    ("Fallback %", 1): ("fallback_size_pct", "float"),
    ("Max Lev", 1): ("max_leverage_cap", "float"),
    ("Margin for long positions", 1): ("margin_long_pct", "pct_text"),
    ("Margin for short positions", 1): ("margin_short_pct", "pct_text"),
    ("SL Mode", 1): ("sl_mode", "string"),
    ("SL %", 1): ("sl_pct", "float"),
    ("TP Mode", 1): ("tp_mode", "string"),
    ("TP %", 1): ("tp_percent", "float"),
    ("TP R", 1): ("tp_r_multiple", "float"),
    ("ATR Length", 1): ("sl_atr_len", "int"),
    ("ATR Mult", 2): ("sl_atr_mult", "float"),
    ("Swing Basis", 1): ("sl_swing_basis", "string"),
    ("Lookback", 1): ("sl_swing_lookback", "int"),
    ("Swing ATR Len", 1): ("sl_swing_atr_len", "int"),
    ("Swing ATR Mult", 1): ("sl_swing_atr_mult", "float"),
    ("TP ATR Length", 1): ("tp_atr_len", "int"),
    ("TP ATR Mult", 1): ("tp_atr_mult", "float"),
    ("TP1 R", 1): ("tp1_r_multiple", "float"),
    ("TP1 Close %", 1): ("tp1_close_pct", "float"),
    ("TP2 R", 1): ("tp2_r_multiple", "float"),
    ("Use Break Even", 1): ("use_break_even", "bool"),
    ("Trigger R", 1): ("be_trigger_r", "float"),
    ("Buffer R", 1): ("be_buffer_r", "float"),
    ("Use Trailing", 1): ("use_trailing", "bool"),
    ("ATR Len", 3): ("trail_atr_len", "int"),
    ("Start R", 1): ("trail_start_r", "float"),
    ("Dist ATR", 1): ("trail_distance_atr_mult", "float"),
    ("Exit on Opposite Signal", 1): ("exit_on_opposite_signal", "bool"),
    ("Exit on MA Block", 1): ("exit_on_ma_block", "bool"),
    ("Exit on MA Slope Block", 1): ("exit_on_ma_slope_block", "bool"),
    ("Exit on McGinley Block", 1): ("exit_on_mcginley_block", "bool"),
    ("Exit on HTF Trend Block", 1): ("exit_on_htf_trend_block", "bool"),
    ("Exit on Vol Block", 1): ("exit_on_vol_block", "bool"),
    ("Exit on ATR Vol Block", 1): ("exit_on_atr_vol_block", "bool"),
    ("Exit on Range Block", 1): ("exit_on_range_block", "bool"),
    ("Exit on Candle Ptn Block", 1): ("exit_on_candle_pattern_block", "bool"),
    ("Exit on Level Prox Block", 1): ("exit_on_level_prox_block", "bool"),
    ("Time Stop", 1): ("use_time_stop", "bool"),
    ("Bars", 1): ("time_stop_bars", "int"),
    ("Condition", 1): ("time_stop_condition", "string"),
    ("EOD Exit", 1): ("time_stop_eod", "bool"),
    ("EOW Exit", 1): ("time_stop_eow", "bool"),
    ("Daily Loss Limit", 1): ("use_daily_loss_limit", "bool"),
    ("Limit %", 1): ("max_daily_loss_pct", "float"),
    ("Max Trades/Day", 1): ("use_max_trades_per_day", "bool"),
    ("Max", 1): ("max_trades_per_day", "int"),
    ("Max DD Guard", 1): ("use_max_drawdown_guard", "bool"),
    ("Max DD %", 1): ("max_drawdown_pct", "float"),
    ("Consec Loss Guard", 1): ("use_consecutive_loss_halt", "bool"),
    ("Max Losses", 1): ("max_consecutive_losses", "int"),
    ("Equity Curve Guard", 1): ("use_equity_curve_filter", "bool"),
    ("EC MA Len", 1): ("equity_ma_length", "int"),
    ("MAE Guard", 1): ("use_mae_guard", "bool"),
    ("Max MAE %", 1): ("max_mae_pct", "float"),
    ("Guard Recovery", 1): ("use_guard_recovery", "bool"),
    ("Mode", 2): ("guard_recovery_mode", "string"),
    ("Recovery Bars", 1): ("guard_recovery_bars", "int"),
    ("Recovery Signals", 1): ("guard_recovery_signals", "int"),
    ("Trade Cooldown After Exit", 1): ("use_trade_cooldown", "bool"),
    ("Bars", 2): ("cooldown_bars_after_exit", "int"),
    ("Confirm Transform", 1): ("use_confirm_transform", "bool"),
    ("Confirm Bars", 1): ("confirm_bars", "int"),
    ("Close Must Cross", 1): ("confirm_close_crosses", "bool"),
    ("Req Raw Still True", 1): ("require_raw_still_true", "bool"),
    ("Refresh New Raw", 1): ("refresh_on_new_raw", "bool"),
    ("Level Retest", 1): ("use_level_retest", "bool"),
    ("Timeout", 1): ("retest_timeout_bars", "int"),
    ("Buffer %", 2): ("retest_buffer_pct", "float"),
}
TRACKED_PROPERTY_LABELS = {label for label, _ in PROPERTY_ROW_TO_OVERRIDE}
MAX_PROPERTY_OCCURRENCE = {
    label: max(occurrence for mapped_label, occurrence in PROPERTY_ROW_TO_OVERRIDE if mapped_label == label)
    for label in TRACKED_PROPERTY_LABELS
}
REQUIRED_EXPORT_OVERRIDE_KEYS = {override_key for override_key, _ in PROPERTY_ROW_TO_OVERRIDE.values()}


@dataclass
class Trade:
    side: str
    entry_time: pd.Timestamp
    exit_time: pd.Timestamp
    entry_price: float
    exit_price: float
    qty: float
    reason: str
    entry_bar: int | None = None
    exit_bar: int | None = None
    position_value: float | None = None
    net_pnl: float | None = None
    net_pnl_pct: float | None = None
    favorable_excursion: float | None = None
    favorable_excursion_pct: float | None = None
    adverse_excursion: float | None = None
    adverse_excursion_pct: float | None = None
    cumulative_pnl: float | None = None
    cumulative_pnl_pct: float | None = None


def parse_chart_timezone(raw: str) -> timezone:
    text = str(raw or "").strip().upper()
    if text in {"UTC+3", "UTC+03", "UTC+03:00", "+03:00"}:
        return timezone(timedelta(hours=3))
    if text in {"UTC+2", "UTC+02", "UTC+02:00", "+02:00"}:
        return timezone(timedelta(hours=2))
    raise ValueError(f"Unsupported fixed chart timezone: {raw}")


def normalize_override_value(value: Any, value_type: str) -> Any:
    text = str(value).strip()
    if value_type == "bool":
        return text.lower() == "on"
    if value_type == "int":
        return int(float(text))
    if value_type == "float":
        return float(text)
    if value_type == "pct_text":
        return float(text.replace("%", "").strip())
    return text


def _coerce_optional_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return float(text.replace("%", "").strip())
    except ValueError:
        return None


def evaluate_broker_leverage_alignment(
    *,
    props: dict[str, Any],
    overrides: dict[str, Any],
) -> dict[str, Any]:
    max_leverage_cap = float(overrides.get("max_leverage_cap", 1.0))
    expected_margin_pct = 100.0 / max_leverage_cap
    tv_leverage = _coerce_optional_float(props.get("Leverage"))
    tv_margin_long_pct = _coerce_optional_float(props.get("Margin for long positions"))
    tv_margin_short_pct = _coerce_optional_float(props.get("Margin for short positions"))

    status = "UNKNOWN"
    note = ""
    if tv_margin_long_pct is not None and tv_margin_short_pct is not None:
        margins_match = (
            abs(tv_margin_long_pct - expected_margin_pct) <= 1e-9
            and abs(tv_margin_short_pct - expected_margin_pct) <= 1e-9
        )
        leverage_matches = tv_leverage is None or abs(tv_leverage - max_leverage_cap) <= 1e-9
        status = "MATCH" if margins_match and leverage_matches else "MISMATCH"
        note = (
            f"Max Lev input={max_leverage_cap:g}; "
            f"TV Leverage={'' if tv_leverage is None else f'{tv_leverage:g}'}; "
            f"TV Margin long/short={tv_margin_long_pct:g}%/{tv_margin_short_pct:g}%; "
            f"expected margin={expected_margin_pct:g}%."
        )
    return {
        "tv_broker_leverage": tv_leverage,
        "tv_margin_long_pct": tv_margin_long_pct,
        "tv_margin_short_pct": tv_margin_short_pct,
        "expected_margin_pct_from_max_lev": expected_margin_pct,
        "broker_leverage_alignment_status": status,
        "broker_leverage_alignment_note": note,
    }


def parse_tv_range(raw: Any, chart_tz: timezone) -> tuple[str, str]:
    text = " ".join(str(raw or "").split())
    # Normalize common separators: em dash, en dash, hyphen, 'to'
    seps = ["—", "–", "-", "to"]
    sep_found = None
    for sep in seps:
        token = f" {sep} "
        if token in text:
            sep_found = token
            break
    if sep_found is None:
        # Fallback for mojibake 'â€”'
        if "â€”" in text:
            sep_found = " â€” " if " â€” " in text else "â€”"
        else:
            return "", ""
    left, right = [part.strip() for part in text.split(sep_found, 1)]
    start = pd.Timestamp(left).tz_localize(chart_tz).tz_convert("UTC").isoformat()
    end = pd.Timestamp(right).tz_localize(chart_tz).tz_convert("UTC").isoformat()
    return start, end


def latest_xlsx(case_dir: Path) -> Path:
    files = sorted(
        [p for p in case_dir.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")],
        key=lambda p: (p.stat().st_mtime, p.name),
        reverse=True,
    )
    if not files:
        raise FileNotFoundError(f"No xlsx found in {case_dir}")
    return files[0]


def normalize_reason(raw: str) -> str:
    text = str(raw or "").strip().upper()
    return {
        "OPEN": "TERMINAL_CLOSE",
        "SL_ATR_HIT": "SL",
        "SL_PERCENT_HIT": "SL",
        "SL_SWING_ATR_HIT": "SL",
        "TP_ATR_HIT": "TP",
        "TP_PERCENT_HIT": "TP",
        "TP_R_HIT": "TP",
        "OPP_SIGNAL": "OPP_SIGNAL",
        "L4-LONG": "OPP_SIGNAL",
        "L4-SHORT": "OPP_SIGNAL",
        "L4-LONG-SL": "SL",
        "L4-SHORT-SL": "SL",
        "MARGIN CALL": "MARGIN_CALL",
    }.get(text, text)


def read_tv_workbook(xlsx_path: Path, chart_tz: timezone) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        props_ws = wb["Properties"]
        props: dict[str, Any] = {}
        prop_rows: list[tuple[str, Any]] = []
        for row in props_ws.iter_rows(values_only=True):
            vals = [v for v in row if v is not None]
            if len(vals) < 2:
                continue
            label = str(vals[0]).strip()
            props[label] = vals[1]
            prop_rows.append((label, vals[1]))

        df = pd.read_excel(xlsx_path, sheet_name="List of trades")
        qty_col = "Position size (qty)" if "Position size (qty)" in df.columns else "Size (qty)"
        value_col = "Position size (value)" if "Position size (value)" in df.columns else "Size (value)"
        trades: list[Trade] = []
        for _, group in df.groupby("Trade #", sort=True):
            if len(group) != 2:
                continue
            entry_row = group[group["Type"].astype(str).str.startswith("Entry")].iloc[0]
            exit_row = group[group["Type"].astype(str).str.startswith("Exit")].iloc[0]
            entry_type = str(entry_row["Type"]).lower()
            side = "LONG" if "long" in entry_type else "SHORT"
            trades.append(
                Trade(
                    side=side,
                    entry_time=pd.Timestamp(entry_row["Date and time"]).tz_localize(chart_tz).tz_convert("UTC"),
                    exit_time=pd.Timestamp(exit_row["Date and time"]).tz_localize(chart_tz).tz_convert("UTC"),
                    entry_price=float(entry_row["Price USDT"]),
                    exit_price=float(exit_row["Price USDT"]),
                    qty=float(entry_row[qty_col]),
                    reason=normalize_reason(str(exit_row["Signal"])),
                    position_value=float(entry_row[value_col]),
                    net_pnl=float(exit_row["Net P&L USDT"]),
                    net_pnl_pct=float(exit_row["Net P&L %"]),
                    favorable_excursion=float(exit_row["Favorable excursion USDT"]),
                    favorable_excursion_pct=float(exit_row["Favorable excursion %"]),
                    adverse_excursion=float(exit_row["Adverse excursion USDT"]),
                    adverse_excursion_pct=float(exit_row["Adverse excursion %"]),
                    cumulative_pnl=float(exit_row["Cumulative P&L USDT"]),
                    cumulative_pnl_pct=float(exit_row["Cumulative P&L %"]),
                )
            )

        backtesting_start, backtesting_end = parse_tv_range(props.get("Backtesting range"), chart_tz)
        trading_start, trading_end = parse_tv_range(props.get("Trading range"), chart_tz)
        return {
            "props": props,
            "prop_rows": prop_rows,
            "trades": trades,
            "trade_list_df": df,
            "backtesting_start": backtesting_start,
            "backtesting_end": backtesting_end,
            "trading_start": trading_start,
            "trading_end": trading_end,
        }
    finally:
        wb.close()


def build_case_overrides(xlsx_info: dict[str, Any]) -> dict[str, Any]:
    overrides: dict[str, Any] = {}
    occurrences: dict[str, int] = {}
    for label, value in xlsx_info["prop_rows"]:
        if label in TRACKED_PROPERTY_LABELS:
            occurrence = occurrences.get(label, 0) + 1
            occurrences[label] = occurrence
            if occurrence > MAX_PROPERTY_OCCURRENCE[label]:
                raise ValueError(
                    f"Unexpected duplicate property label in workbook: {label!r} occurrence {occurrence}. "
                    "Update manual_tw_futures_audit.py mapping before using this export."
                )
        else:
            occurrences[label] = occurrences.get(label, 0) + 1
            continue

        mapped = PROPERTY_ROW_TO_OVERRIDE.get((label, occurrence))
        if mapped is None:
            continue
        override_key, value_type = mapped
        overrides[override_key] = normalize_override_value(value, value_type)

    missing = sorted(key for key in REQUIRED_EXPORT_OVERRIDE_KEYS if key not in overrides)
    if missing:
        raise ValueError(
            "Workbook Properties sheet is missing mapped config values required for export-faithful parity: "
            f"{missing}"
        )
    return overrides


def write_case_override(case_name: str, overrides: dict[str, Any], report_prefix: str) -> Path:
    path = REPORTS_DIR / f"{report_prefix}_{case_name}_overrides.json"
    path.write_text(json.dumps(overrides, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def build_python_research_overrides(case_name: str) -> dict[str, Any]:
    if case_name in {"case_010", "case_023"}:
        return {
            "tw_audit_semantics_mode": "research",
            "tw_margin_call_mode": "tradingview",
            "tw_margin_call_split_entries": True,
        }
    return {}


def tv_symbol_to_bridge_symbol(raw: Any) -> str:
    text = str(raw or "").strip()
    if ":" in text:
        text = text.split(":", 1)[1]
    return text.strip()


def tv_timeframe_to_bridge_tf(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    mapping = {
        "1 minute": "1m",
        "3 minutes": "3m",
        "5 minutes": "5m",
        "15 minutes": "15m",
        "30 minutes": "30m",
        "45 minutes": "45m",
        "1 hour": "1h",
        "2 hours": "2h",
        "4 hours": "4h",
        "1 day": "1d",
        "1 week": "1w",
    }
    return mapping.get(text, str(raw or "").strip())


def run_case_parity(
    case_name: str,
    xlsx_info: dict[str, Any],
    chart_tz_label: str,
    override_json: Path,
    python_only_override_json: Path | None = None,
) -> dict[str, Path]:
    report_prefix = REPORTS_DIR / f"manual_tw_futures_{case_name}"
    bridge_symbol = tv_symbol_to_bridge_symbol(xlsx_info["props"].get("Symbol", "BTCUSDT.P"))
    bridge_tf = tv_timeframe_to_bridge_tf(xlsx_info["props"].get("Timeframe", "1h"))
    cmd = [
        sys.executable,
        "parity_compare.py",
        "--fetch-fresh",
        "--symbol",
        bridge_symbol,
        "--tf",
        bridge_tf,
        "--start",
        xlsx_info["backtesting_start"],
        "--end",
        xlsx_info["backtesting_end"],
        "--case",
        str(override_json.relative_to(REPO_ROOT)),
        "--use-case-for-pine",
        "--out",
        str((report_prefix.with_suffix(".json")).relative_to(REPO_ROOT)),
        "--trade-report-csv",
        str((REPORTS_DIR / f"{report_prefix.name}_trade_report.csv").relative_to(REPO_ROOT)),
        "--python-trades-csv",
        str((REPORTS_DIR / f"{report_prefix.name}_python_trades.csv").relative_to(REPO_ROOT)),
        "--effective-config-json",
        str((REPORTS_DIR / f"{report_prefix.name}_effective_config.json").relative_to(REPO_ROOT)),
    ]
    cmd.extend(["--chart-timezone", chart_tz_label])
    if python_only_override_json is not None:
        cmd.extend(["--python-case", str(python_only_override_json.relative_to(REPO_ROOT))])
    subprocess.run(cmd, cwd=REPO_ROOT, check=True)
    return {
        "parity_report_json": REPORTS_DIR / f"{report_prefix.name}.json",
        "python_trades_csv": REPORTS_DIR / f"{report_prefix.name}_python_trades.csv",
        "pine_trades_csv": REPO_ROOT / "data/pine_trades.csv",
    }


def load_parity_report_summary(report_json_path: Path) -> dict[str, Any]:
    data = json.loads(report_json_path.read_text(encoding="utf-8"))
    trade_parity = data.get("trade_parity", {}) if isinstance(data, dict) else {}
    return {
        "pine_trades": int(trade_parity.get("pine_trades", 0) or 0),
        "python_trades": int(trade_parity.get("python_trades", 0) or 0),
    }


def load_trades(csv_path: Path) -> list[Trade]:
    try:
        df = pd.read_csv(csv_path)
    except pd.errors.EmptyDataError:
        return []
    out: list[Trade] = []
    for _, row in df.iterrows():
        out.append(
            Trade(
                side=str(row["direction"]).upper(),
                entry_time=pd.Timestamp(row["entry_time"]),
                exit_time=pd.Timestamp(row["exit_time"]),
                entry_price=float(row["entry_price"]),
                exit_price=float(row["exit_price"]),
                qty=float(row["qty"]),
                reason=normalize_reason(str(row["exit_reason"])),
                entry_bar=int(row["entry_bar"]) if "entry_bar" in row and not pd.isna(row["entry_bar"]) else None,
                exit_bar=int(row["exit_bar"]) if "exit_bar" in row and not pd.isna(row["exit_bar"]) else None,
                net_pnl=float(row["gross_pnl"]) if "gross_pnl" in row and not pd.isna(row["gross_pnl"]) else None,
                net_pnl_pct=float(row["pnl_pct"]) if "pnl_pct" in row and not pd.isna(row["pnl_pct"]) else None,
            )
        )
    return out


def load_bars_dataset(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    return df


def _calc_excursions(
    *,
    bars_df: pd.DataFrame,
    entry_time: pd.Timestamp,
    exit_time: pd.Timestamp,
    entry_price: float,
    side: str,
) -> tuple[float, float, float, float]:
    window = bars_df[(bars_df["timestamp"] > entry_time) & (bars_df["timestamp"] <= exit_time)]
    if window.empty or entry_price <= 0:
        return 0.0, 0.0, 0.0, 0.0

    max_high = float(window["high"].max())
    min_low = float(window["low"].min())
    is_long = side.upper() == "LONG"

    if is_long:
        favorable = max(0.0, max_high - entry_price)
        adverse = -max(0.0, entry_price - min_low)
    else:
        favorable = max(0.0, entry_price - min_low)
        adverse = -max(0.0, max_high - entry_price)

    favorable_pct = favorable / entry_price * 100.0
    adverse_pct = adverse / entry_price * 100.0
    return favorable, favorable_pct, adverse, adverse_pct


def enrich_local_trades(
    trades: list[Trade],
    *,
    bars_df: pd.DataFrame,
    initial_capital: float,
    point_value: float,
) -> None:
    cumulative_pnl = 0.0
    for trade in trades:
        trade.position_value = abs(trade.entry_price * trade.qty * point_value)
        if trade.net_pnl is None:
            direction = 1.0 if trade.side == "LONG" else -1.0
            trade.net_pnl = direction * (trade.exit_price - trade.entry_price) * trade.qty * point_value
        trade.net_pnl_pct = (trade.net_pnl / trade.position_value * 100.0) if trade.position_value > 0 else 0.0
        cumulative_pnl += trade.net_pnl
        trade.cumulative_pnl = cumulative_pnl
        trade.cumulative_pnl_pct = (cumulative_pnl / initial_capital * 100.0) if initial_capital > 0 else 0.0
        (
            trade.favorable_excursion,
            trade.favorable_excursion_pct,
            trade.adverse_excursion,
            trade.adverse_excursion_pct,
        ) = _calc_excursions(
            bars_df=bars_df,
            entry_time=trade.entry_time,
            exit_time=trade.exit_time,
            entry_price=trade.entry_price,
            side=trade.side,
        )


def compute_trade_metrics(trades: list[Trade], *, initial_capital: float) -> dict[str, Any]:
    closed = [trade for trade in trades if trade.net_pnl is not None]
    wins = [trade for trade in closed if float(trade.net_pnl or 0.0) > 0.0]
    losses = [trade for trade in closed if float(trade.net_pnl or 0.0) < 0.0]
    gross_wins = sum(float(trade.net_pnl or 0.0) for trade in wins)
    gross_losses = abs(sum(float(trade.net_pnl or 0.0) for trade in losses))
    net_pnl = sum(float(trade.net_pnl or 0.0) for trade in closed)
    win_rate = (len(wins) / len(closed) * 100.0) if closed else 0.0
    profit_factor = float("inf") if gross_losses == 0.0 and gross_wins > 0.0 else (
        gross_wins / gross_losses if gross_losses > 0.0 else 0.0
    )

    equity = float(initial_capital)
    peak = float(initial_capital)
    max_drawdown_pct = 0.0
    for trade in closed:
        equity += float(trade.net_pnl or 0.0)
        peak = max(peak, equity)
        if peak > 0.0:
            max_drawdown_pct = max(max_drawdown_pct, ((peak - equity) / peak) * 100.0)

    return {
        "trade_count": len(trades),
        "win_rate_pct": round(win_rate, 4),
        "net_pnl": round(net_pnl, 6),
        "net_pnl_pct": round((net_pnl / initial_capital * 100.0), 6) if initial_capital > 0 else 0.0,
        "profit_factor": round(profit_factor, 6) if math.isfinite(profit_factor) else "inf",
        "max_drawdown_pct": round(max_drawdown_pct, 6),
    }


def analyze_trade_list_behavior(df: pd.DataFrame, chart_tz: timezone) -> dict[str, Any]:
    events: list[dict[str, Any]] = []
    for row_index, row in df.iterrows():
        trade_type = str(row.get("Type", "")).strip().lower()
        if not trade_type.startswith(("entry", "exit")):
            continue
        if "long" in trade_type:
            side = "LONG"
        elif "short" in trade_type:
            side = "SHORT"
        else:
            continue
        events.append(
            {
                "row_index": int(row_index),
                "trade_no": int(float(row["Trade #"])),
                "time": pd.Timestamp(row["Date and time"]).tz_localize(chart_tz).tz_convert("UTC"),
                "action": "ENTRY" if trade_type.startswith("entry") else "EXIT",
                "side": side,
                "signal": str(row.get("Signal", "")),
                "qty": float(row.get("Position size (qty)", 0.0)),
            }
        )

    events.sort(key=lambda item: (item["time"], item["row_index"]))
    open_counts = {"LONG": 0, "SHORT": 0}
    max_open_counts = {"LONG": 0, "SHORT": 0}
    stack_entry_events = 0
    first_stack_event: dict[str, Any] | None = None

    for event in events:
        side = event["side"]
        if event["action"] == "ENTRY":
            if open_counts[side] > 0:
                stack_entry_events += 1
                if first_stack_event is None:
                    first_stack_event = {
                        "trade_no": event["trade_no"],
                        "side": side,
                        "time": str(event["time"]),
                        "signal": event["signal"],
                        "prior_open_count_same_side": open_counts[side],
                        "qty": event["qty"],
                    }
            open_counts[side] += 1
            max_open_counts[side] = max(max_open_counts[side], open_counts[side])
        else:
            open_counts[side] = max(0, open_counts[side] - 1)

    observed_max_same_side = max(max_open_counts["LONG"], max_open_counts["SHORT"])
    return {
        "event_count": len(events),
        "observed_max_open_long": max_open_counts["LONG"],
        "observed_max_open_short": max_open_counts["SHORT"],
        "observed_max_open_same_side": observed_max_same_side,
        "stack_entry_events": stack_entry_events,
        "pyramiding_observed": observed_max_same_side > 1,
        "first_stack_event": first_stack_event,
    }


def evaluate_setting_effect(overrides: dict[str, Any], trade_behavior: dict[str, Any]) -> dict[str, Any]:
    planned_max_entries = int(overrides.get("max_entries", 1))
    observed_max = int(trade_behavior["observed_max_open_same_side"])
    if planned_max_entries > 1:
        if observed_max > 1:
            return {
                "setting_effect_status": "EFFECT_OBSERVED",
                "setting_effect_note": (
                    f"TW export realized stacking. max_entries={planned_max_entries}, "
                    f"observed_max_open_same_side={observed_max}, "
                    f"stack_entry_events={trade_behavior['stack_entry_events']}."
                ),
            }
        if str(overrides.get("signal_mode", "")) == "Supertrend":
            return {
                "setting_effect_status": "BLOCKED_BY_SIGNAL_MODEL",
                "setting_effect_note": (
                    f"Supertrend producer is pulse-based and did not emit a same-side rearm pulse. "
                    f"max_entries={planned_max_entries}, observed_max_open_same_side={observed_max}, "
                    f"stack_entry_events={trade_behavior['stack_entry_events']}. "
                    "Future stateful add behavior depends on backlog feature L18c / trend_rearm_add."
                ),
            }
        return {
            "setting_effect_status": "NO_EFFECT_OBSERVED",
            "setting_effect_note": (
                f"TW export did not realize stacking. max_entries={planned_max_entries}, "
                f"observed_max_open_same_side={observed_max}, "
                f"stack_entry_events={trade_behavior['stack_entry_events']}."
            ),
        }
    return {
        "setting_effect_status": "N/A",
        "setting_effect_note": "",
    }


def compare_trades(
    ref: list[Trade],
    other: list[Trade],
    price_tol: float = STRICT_TRADE_TOLERANCES["price_tol"],
    qty_tol: float = STRICT_TRADE_TOLERANCES["qty_tol"],
) -> dict[str, Any]:
    compared = min(len(ref), len(other))
    core_match = 0
    entry_price_match = 0
    exit_price_match = 0
    qty_match = 0
    first_mismatch: dict[str, Any] | None = None

    for idx in range(compared):
        a = ref[idx]
        b = other[idx]
        core_ok = (
            a.side == b.side
            and a.entry_time == b.entry_time
            and a.exit_time == b.exit_time
            and normalize_reason(a.reason) == normalize_reason(b.reason)
        )
        entry_ok = math.isclose(a.entry_price, b.entry_price, abs_tol=price_tol)
        exit_ok = math.isclose(a.exit_price, b.exit_price, abs_tol=price_tol)
        qty_ok = math.isclose(a.qty, b.qty, abs_tol=qty_tol)
        if core_ok:
            core_match += 1
        if entry_ok:
            entry_price_match += 1
        if exit_ok:
            exit_price_match += 1
        if qty_ok:
            qty_match += 1
        if first_mismatch is None and not (core_ok and entry_ok and exit_ok and qty_ok):
            first_mismatch = {
                "ref_side": a.side,
                "other_side": b.side,
                "ref_entry_time": str(a.entry_time),
                "other_entry_time": str(b.entry_time),
                "ref_exit_time": str(a.exit_time),
                "other_exit_time": str(b.exit_time),
                "ref_reason": normalize_reason(a.reason),
                "other_reason": normalize_reason(b.reason),
                "entry_time_diff_min": (b.entry_time - a.entry_time).total_seconds() / 60.0,
                "exit_time_diff_min": (b.exit_time - a.exit_time).total_seconds() / 60.0,
                "entry_price_diff": b.entry_price - a.entry_price,
                "exit_price_diff": b.exit_price - a.exit_price,
                "qty_diff": b.qty - a.qty,
            }

    strict_pass = (
        len(ref) == len(other)
        and core_match == compared
        and entry_price_match == compared
        and exit_price_match == compared
        and qty_match == compared
    )
    return {
        "strict_pass": strict_pass,
        "core_match_count": core_match,
        "entry_price_match_count": entry_price_match,
        "exit_price_match_count": exit_price_match,
        "qty_match_count": qty_match,
        "first_mismatch": first_mismatch,
    }


def compare_trade_outcomes(
    ref: list[Trade],
    other: list[Trade],
    *,
    value_tol: float = STRICT_OUTCOME_TOLERANCES["value_tol"],
    pnl_tol: float = STRICT_OUTCOME_TOLERANCES["pnl_tol"],
    pnl_pct_tol: float = STRICT_OUTCOME_TOLERANCES["pnl_pct_tol"],
    excursion_tol: float = STRICT_OUTCOME_TOLERANCES["excursion_tol"],
    excursion_pct_tol: float = STRICT_OUTCOME_TOLERANCES["excursion_pct_tol"],
) -> dict[str, Any]:
    compared = min(len(ref), len(other))
    position_value_match = 0
    net_pnl_match = 0
    net_pnl_pct_match = 0
    favorable_excursion_match = 0
    favorable_excursion_pct_match = 0
    adverse_excursion_match = 0
    adverse_excursion_pct_match = 0
    cumulative_pnl_match = 0
    cumulative_pnl_pct_match = 0
    first_outcome_mismatch: dict[str, Any] | None = None

    for idx in range(compared):
        a = ref[idx]
        b = other[idx]
        pos_value_ok = math.isclose(float(a.position_value or 0.0), float(b.position_value or 0.0), abs_tol=value_tol)
        net_pnl_ok = math.isclose(float(a.net_pnl or 0.0), float(b.net_pnl or 0.0), abs_tol=pnl_tol)
        net_pnl_pct_ok = math.isclose(float(a.net_pnl_pct or 0.0), float(b.net_pnl_pct or 0.0), abs_tol=pnl_pct_tol)
        mfe_ok = math.isclose(float(a.favorable_excursion or 0.0), float(b.favorable_excursion or 0.0), abs_tol=excursion_tol)
        mfe_pct_ok = math.isclose(float(a.favorable_excursion_pct or 0.0), float(b.favorable_excursion_pct or 0.0), abs_tol=excursion_pct_tol)
        mae_ok = math.isclose(float(a.adverse_excursion or 0.0), float(b.adverse_excursion or 0.0), abs_tol=excursion_tol)
        mae_pct_ok = math.isclose(float(a.adverse_excursion_pct or 0.0), float(b.adverse_excursion_pct or 0.0), abs_tol=excursion_pct_tol)
        cum_ok = math.isclose(float(a.cumulative_pnl or 0.0), float(b.cumulative_pnl or 0.0), abs_tol=pnl_tol)
        cum_pct_ok = math.isclose(float(a.cumulative_pnl_pct or 0.0), float(b.cumulative_pnl_pct or 0.0), abs_tol=pnl_pct_tol)

        position_value_match += int(pos_value_ok)
        net_pnl_match += int(net_pnl_ok)
        net_pnl_pct_match += int(net_pnl_pct_ok)
        favorable_excursion_match += int(mfe_ok)
        favorable_excursion_pct_match += int(mfe_pct_ok)
        adverse_excursion_match += int(mae_ok)
        adverse_excursion_pct_match += int(mae_pct_ok)
        cumulative_pnl_match += int(cum_ok)
        cumulative_pnl_pct_match += int(cum_pct_ok)

        if first_outcome_mismatch is None and not (
            pos_value_ok and net_pnl_ok and net_pnl_pct_ok and mfe_ok and mfe_pct_ok and mae_ok and mae_pct_ok and cum_ok and cum_pct_ok
        ):
            first_outcome_mismatch = {
                "ref_position_value": a.position_value,
                "other_position_value": b.position_value,
                "position_value_diff": float(b.position_value or 0.0) - float(a.position_value or 0.0),
                "ref_net_pnl": a.net_pnl,
                "other_net_pnl": b.net_pnl,
                "net_pnl_diff": float(b.net_pnl or 0.0) - float(a.net_pnl or 0.0),
                "ref_net_pnl_pct": a.net_pnl_pct,
                "other_net_pnl_pct": b.net_pnl_pct,
                "net_pnl_pct_diff": float(b.net_pnl_pct or 0.0) - float(a.net_pnl_pct or 0.0),
                "ref_favorable_excursion": a.favorable_excursion,
                "other_favorable_excursion": b.favorable_excursion,
                "favorable_excursion_diff": float(b.favorable_excursion or 0.0) - float(a.favorable_excursion or 0.0),
                "ref_adverse_excursion": a.adverse_excursion,
                "other_adverse_excursion": b.adverse_excursion,
                "adverse_excursion_diff": float(b.adverse_excursion or 0.0) - float(a.adverse_excursion or 0.0),
                "ref_cumulative_pnl": a.cumulative_pnl,
                "other_cumulative_pnl": b.cumulative_pnl,
                "cumulative_pnl_diff": float(b.cumulative_pnl or 0.0) - float(a.cumulative_pnl or 0.0),
            }

    extended_outcome_pass = (
        len(ref) == len(other)
        and position_value_match == compared
        and net_pnl_match == compared
        and net_pnl_pct_match == compared
        and favorable_excursion_match == compared
        and favorable_excursion_pct_match == compared
        and adverse_excursion_match == compared
        and adverse_excursion_pct_match == compared
        and cumulative_pnl_match == compared
        and cumulative_pnl_pct_match == compared
    )
    return {
        "extended_outcome_pass": extended_outcome_pass,
        "position_value_match_count": position_value_match,
        "net_pnl_match_count": net_pnl_match,
        "net_pnl_pct_match_count": net_pnl_pct_match,
        "favorable_excursion_match_count": favorable_excursion_match,
        "favorable_excursion_pct_match_count": favorable_excursion_pct_match,
        "adverse_excursion_match_count": adverse_excursion_match,
        "adverse_excursion_pct_match_count": adverse_excursion_pct_match,
        "cumulative_pnl_match_count": cumulative_pnl_match,
        "cumulative_pnl_pct_match_count": cumulative_pnl_pct_match,
        "first_outcome_mismatch": first_outcome_mismatch,
    }


def summarize_case(
    case_name: str,
    xlsx_path: Path,
    chart_tz_label: str,
    chart_tz: timezone,
    report_prefix: str,
    *,
    enable_python_research_overrides: bool,
) -> dict[str, Any]:
    xlsx_info = read_tv_workbook(xlsx_path, chart_tz)
    overrides = build_case_overrides(xlsx_info)
    override_json = write_case_override(case_name, overrides, report_prefix)
    python_only_overrides = build_python_research_overrides(case_name) if enable_python_research_overrides else {}
    python_only_override_json = None
    if python_only_overrides:
        python_only_override_json = write_case_override(
            case_name,
            python_only_overrides,
            f"{report_prefix}_python_only",
        )
    artifacts = run_case_parity(case_name, xlsx_info, chart_tz_label, override_json, python_only_override_json)

    tv_trades = xlsx_info["trades"]
    bars_df = load_bars_dataset(REPO_ROOT / "data" / "parity_input_from_pinets.csv")
    point_value = float(xlsx_info["props"].get("Point value", 1.0))
    initial_capital = float(xlsx_info["props"].get("Initial capital", 1_000_000.0))
    parity_summary = load_parity_report_summary(artifacts["parity_report_json"])
    pine_trades = load_trades(artifacts["pine_trades_csv"])
    python_trades = load_trades(artifacts["python_trades_csv"])
    if len(pine_trades) != parity_summary["pine_trades"]:
        if parity_summary["pine_trades"] == 0:
            pine_trades = []
        else:
            raise ValueError(
                "pine_trades.csv count does not match parity_compare report; stale artifact suspected "
                f"({len(pine_trades)} vs {parity_summary['pine_trades']})."
            )
    if len(python_trades) != parity_summary["python_trades"]:
        if parity_summary["python_trades"] == 0:
            python_trades = []
        else:
            raise ValueError(
                "python_trades.csv count does not match parity_compare report; stale artifact suspected "
                f"({len(python_trades)} vs {parity_summary['python_trades']})."
            )
    enrich_local_trades(pine_trades, bars_df=bars_df, initial_capital=initial_capital, point_value=point_value)
    enrich_local_trades(python_trades, bars_df=bars_df, initial_capital=initial_capital, point_value=point_value)
    tw_metrics = compute_trade_metrics(tv_trades, initial_capital=initial_capital)
    pine_metrics = compute_trade_metrics(pine_trades, initial_capital=initial_capital)
    python_metrics = compute_trade_metrics(python_trades, initial_capital=initial_capital)
    tw_vs_pine = compare_trades(tv_trades, pine_trades, **STRICT_TRADE_TOLERANCES)
    tw_vs_python = compare_trades(tv_trades, python_trades, **STRICT_TRADE_TOLERANCES)
    pine_vs_python = compare_trades(pine_trades, python_trades, **STRICT_TRADE_TOLERANCES)
    tw_vs_pine_soft = compare_trades(tv_trades, pine_trades, **SOFT_TRADE_TOLERANCES)
    tw_vs_python_soft = compare_trades(tv_trades, python_trades, **SOFT_TRADE_TOLERANCES)
    pine_vs_python_soft = compare_trades(pine_trades, python_trades, **SOFT_TRADE_TOLERANCES)
    tw_vs_pine_outcomes = compare_trade_outcomes(tv_trades, pine_trades, **STRICT_OUTCOME_TOLERANCES)
    tw_vs_python_outcomes = compare_trade_outcomes(tv_trades, python_trades, **STRICT_OUTCOME_TOLERANCES)
    pine_vs_python_outcomes = compare_trade_outcomes(pine_trades, python_trades, **STRICT_OUTCOME_TOLERANCES)
    tw_vs_pine_outcomes_soft = compare_trade_outcomes(tv_trades, pine_trades, **SOFT_OUTCOME_TOLERANCES)
    tw_vs_python_outcomes_soft = compare_trade_outcomes(tv_trades, python_trades, **SOFT_OUTCOME_TOLERANCES)
    pine_vs_python_outcomes_soft = compare_trade_outcomes(pine_trades, python_trades, **SOFT_OUTCOME_TOLERANCES)
    tw_trade_behavior = analyze_trade_list_behavior(xlsx_info["trade_list_df"], chart_tz)
    setting_effect = evaluate_setting_effect(overrides, tw_trade_behavior)
    broker_alignment = evaluate_broker_leverage_alignment(props=xlsx_info["props"], overrides=overrides)

    return {
        "case": case_name,
        "xlsx": xlsx_path.name,
        "chart_timezone": chart_tz_label,
        "override_json": str(override_json.relative_to(REPO_ROOT)),
        "override_mode": "export_properties_full_rebuild",
        "trade_strict_contract": STRICT_TRADE_TOLERANCES,
        "trade_soft_contract": SOFT_TRADE_TOLERANCES,
        "outcome_strict_contract": STRICT_OUTCOME_TOLERANCES,
        "outcome_soft_contract": SOFT_OUTCOME_TOLERANCES,
        "applied_overrides": overrides,
        "tv_broker_leverage": broker_alignment["tv_broker_leverage"],
        "tv_margin_long_pct": broker_alignment["tv_margin_long_pct"],
        "tv_margin_short_pct": broker_alignment["tv_margin_short_pct"],
        "expected_margin_pct_from_max_lev": broker_alignment["expected_margin_pct_from_max_lev"],
        "broker_leverage_alignment_status": broker_alignment["broker_leverage_alignment_status"],
        "broker_leverage_alignment_note": broker_alignment["broker_leverage_alignment_note"],
        "python_only_override_json": (
            str(python_only_override_json.relative_to(REPO_ROOT)) if python_only_override_json is not None else ""
        ),
        "python_only_overrides": python_only_overrides,
        "tw_trades": len(tv_trades),
        "pine_trades": len(pine_trades),
        "python_trades": len(python_trades),
        "tw_win_rate_pct": tw_metrics["win_rate_pct"],
        "pine_win_rate_pct": pine_metrics["win_rate_pct"],
        "python_win_rate_pct": python_metrics["win_rate_pct"],
        "tw_net_pnl": tw_metrics["net_pnl"],
        "pine_net_pnl": pine_metrics["net_pnl"],
        "python_net_pnl": python_metrics["net_pnl"],
        "tw_net_pnl_pct": tw_metrics["net_pnl_pct"],
        "pine_net_pnl_pct": pine_metrics["net_pnl_pct"],
        "python_net_pnl_pct": python_metrics["net_pnl_pct"],
        "tw_profit_factor": tw_metrics["profit_factor"],
        "pine_profit_factor": pine_metrics["profit_factor"],
        "python_profit_factor": python_metrics["profit_factor"],
        "tw_max_drawdown_pct": tw_metrics["max_drawdown_pct"],
        "pine_max_drawdown_pct": pine_metrics["max_drawdown_pct"],
        "python_max_drawdown_pct": python_metrics["max_drawdown_pct"],
        "tw_first_entry": str(tv_trades[0].entry_time) if tv_trades else "",
        "pine_first_entry": str(pine_trades[0].entry_time) if pine_trades else "",
        "python_first_entry": str(python_trades[0].entry_time) if python_trades else "",
        "tw_last_entry": str(tv_trades[-1].entry_time) if tv_trades else "",
        "pine_last_entry": str(pine_trades[-1].entry_time) if pine_trades else "",
        "python_last_entry": str(python_trades[-1].entry_time) if python_trades else "",
        "backtesting_start_utc": xlsx_info["backtesting_start"],
        "backtesting_end_utc": xlsx_info["backtesting_end"],
        "trading_start_utc": xlsx_info["trading_start"],
        "trading_end_utc": xlsx_info["trading_end"],
        "tw_vs_pine_strict": tw_vs_pine["strict_pass"],
        "tw_vs_python_strict": tw_vs_python["strict_pass"],
        "pine_vs_python_strict": pine_vs_python["strict_pass"],
        "tw_vs_pine_trade_soft_pass": tw_vs_pine_soft["strict_pass"],
        "tw_vs_python_trade_soft_pass": tw_vs_python_soft["strict_pass"],
        "pine_vs_python_trade_soft_pass": pine_vs_python_soft["strict_pass"],
        "tw_vs_pine_core_match": tw_vs_pine["core_match_count"],
        "tw_vs_python_core_match": tw_vs_python["core_match_count"],
        "pine_vs_python_core_match": pine_vs_python["core_match_count"],
        "tw_vs_pine_first_mismatch": tw_vs_pine["first_mismatch"],
        "tw_vs_python_first_mismatch": tw_vs_python["first_mismatch"],
        "pine_vs_python_first_mismatch": pine_vs_python["first_mismatch"],
        "tw_vs_pine_outcome_pass": tw_vs_pine_outcomes["extended_outcome_pass"],
        "tw_vs_python_outcome_pass": tw_vs_python_outcomes["extended_outcome_pass"],
        "pine_vs_python_outcome_pass": pine_vs_python_outcomes["extended_outcome_pass"],
        "tw_vs_pine_outcome_soft_pass": tw_vs_pine_outcomes_soft["extended_outcome_pass"],
        "tw_vs_python_outcome_soft_pass": tw_vs_python_outcomes_soft["extended_outcome_pass"],
        "pine_vs_python_outcome_soft_pass": pine_vs_python_outcomes_soft["extended_outcome_pass"],
        "tw_vs_pine_position_value_match": tw_vs_pine_outcomes["position_value_match_count"],
        "tw_vs_python_position_value_match": tw_vs_python_outcomes["position_value_match_count"],
        "pine_vs_python_position_value_match": pine_vs_python_outcomes["position_value_match_count"],
        "tw_vs_pine_net_pnl_match": tw_vs_pine_outcomes["net_pnl_match_count"],
        "tw_vs_python_net_pnl_match": tw_vs_python_outcomes["net_pnl_match_count"],
        "pine_vs_python_net_pnl_match": pine_vs_python_outcomes["net_pnl_match_count"],
        "tw_vs_pine_cumulative_pnl_match": tw_vs_pine_outcomes["cumulative_pnl_match_count"],
        "tw_vs_python_cumulative_pnl_match": tw_vs_python_outcomes["cumulative_pnl_match_count"],
        "pine_vs_python_cumulative_pnl_match": pine_vs_python_outcomes["cumulative_pnl_match_count"],
        "tw_vs_pine_favorable_excursion_match": tw_vs_pine_outcomes["favorable_excursion_match_count"],
        "tw_vs_python_favorable_excursion_match": tw_vs_python_outcomes["favorable_excursion_match_count"],
        "pine_vs_python_favorable_excursion_match": pine_vs_python_outcomes["favorable_excursion_match_count"],
        "tw_vs_pine_adverse_excursion_match": tw_vs_pine_outcomes["adverse_excursion_match_count"],
        "tw_vs_python_adverse_excursion_match": tw_vs_python_outcomes["adverse_excursion_match_count"],
        "pine_vs_python_adverse_excursion_match": pine_vs_python_outcomes["adverse_excursion_match_count"],
        "tw_vs_pine_first_outcome_mismatch": tw_vs_pine_outcomes["first_outcome_mismatch"],
        "tw_vs_python_first_outcome_mismatch": tw_vs_python_outcomes["first_outcome_mismatch"],
        "pine_vs_python_first_outcome_mismatch": pine_vs_python_outcomes["first_outcome_mismatch"],
        "tw_observed_max_open_long": tw_trade_behavior["observed_max_open_long"],
        "tw_observed_max_open_short": tw_trade_behavior["observed_max_open_short"],
        "tw_observed_max_open_same_side": tw_trade_behavior["observed_max_open_same_side"],
        "tw_stack_entry_events": tw_trade_behavior["stack_entry_events"],
        "tw_pyramiding_observed": tw_trade_behavior["pyramiding_observed"],
        "tw_first_stack_event": tw_trade_behavior["first_stack_event"],
        "setting_effect_status": setting_effect["setting_effect_status"],
        "setting_effect_note": setting_effect["setting_effect_note"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-start", type=int, default=1)
    parser.add_argument("--case-end", type=int, default=25)
    parser.add_argument("--xlsx-path", default="")
    parser.add_argument("--case-name", default="")
    parser.add_argument("--chart-timezone", default="UTC+3")
    parser.add_argument("--report-prefix", default="manual_tw_futures_audit")
    parser.add_argument(
        "--enable-python-research-overrides",
        action="store_true",
        help="Explicitly allow Python-only research overrides for selected legacy TW audit cases.",
    )
    args = parser.parse_args()

    chart_tz = parse_chart_timezone(args.chart_timezone)

    results: list[dict[str, Any]] = []
    if args.xlsx_path:
        xlsx_path = Path(args.xlsx_path).resolve()
        case_name = args.case_name.strip() or xlsx_path.parent.name or "case_custom"
        results.append(
            summarize_case(
                case_name,
                xlsx_path,
                args.chart_timezone,
                chart_tz,
                args.report_prefix,
                enable_python_research_overrides=args.enable_python_research_overrides,
            )
        )
    else:
        for case_num in range(args.case_start, args.case_end + 1):
            case_name = f"case_{case_num:03d}"
            case_dir = PARITY_ROOT / case_name
            try:
                xlsx_path = latest_xlsx(case_dir)
            except FileNotFoundError:
                continue
            results.append(
                summarize_case(
                    case_name,
                    xlsx_path,
                    args.chart_timezone,
                    chart_tz,
                    args.report_prefix,
                    enable_python_research_overrides=args.enable_python_research_overrides,
                )
            )

    json_path = REPORTS_DIR / f"{args.report_prefix}.json"
    csv_path = REPORTS_DIR / f"{args.report_prefix}.csv"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(results[0].keys()) if results else [])
        writer.writeheader()
        writer.writerows(results)

    print(f"Saved {json_path}")
    print(f"Saved {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

