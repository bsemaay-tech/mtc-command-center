#!/usr/bin/env python3
"""Build the Triage Worklist registry from 11_TRIAGE.

The 172-row worklist (`11_TRIAGE/2026-05-30_rejected_worklist.xlsx`) is a raw,
mostly-rejected candidate stage that lives upstream of the 46 matured QuantLens
strategies. Many rows were rejected only because they lacked a transcript; the
user has since added transcripts into `11_TRIAGE/strategies/stgNNN.md`.

This script reconciles the (stale) xlsx metadata with the **actual** transcript
files on disk and emits ``05_REGISTRY/TRIAGE_CANDIDATE_REGISTRY.json`` so the
worklist is visible in the Strategy Research Lab tab, with a live
``eligible_for_retriage`` flag.

Source of truth:
  * metadata  -> the xlsx worklist
  * transcript presence -> the actual stgNNN.md file size/content (overrides xlsx)

Usage:
    python 03_QUANTLENS/tools/build_triage_registry.py [--check]
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None

QUANTLENS_ROOT = Path(__file__).resolve().parents[1]
MCC_ROOT = QUANTLENS_ROOT.parent
TRIAGE_DIR = MCC_ROOT / "11_TRIAGE"
STRATEGIES_DIR = TRIAGE_DIR / "strategies"
WORKLIST_XLSX = TRIAGE_DIR / "2026-05-30_rejected_worklist.xlsx"
OUT = MCC_ROOT / "05_REGISTRY" / "TRIAGE_CANDIDATE_REGISTRY.json"

# A transcript-bearing md file is meaningfully larger than a stub header.
TRANSCRIPT_MIN_BYTES = 2000
RETRIAGE_QUALITIES = {"HIGH", "MEDIUM"}


def _transcript_present(md_file: str | None) -> tuple[bool, str | None, int]:
    if not md_file:
        return False, None, 0
    path = STRATEGIES_DIR / md_file
    if not path.exists():
        return False, None, 0
    size = path.stat().st_size
    rel = path.relative_to(MCC_ROOT).as_posix()
    return size >= TRANSCRIPT_MIN_BYTES, rel, size


def _read_worklist() -> list[dict[str, Any]]:
    if openpyxl is None or not WORKLIST_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(WORKLIST_XLSX, read_only=True, data_only=True)
    ws = wb["Worklist"] if "Worklist" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def g(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in rows[1:]:
        if not any(row):
            continue
        md_file = g(row, "md_file")
        present, rel, size = _transcript_present(md_file)
        quality = str(g(row, "source_quality") or "").upper()
        has_url = str(g(row, "has_source_url") or "").upper() == "YES"
        # Refresh coverage from actual disk state.
        if has_url and present:
            coverage = "HAS_BOTH"
        elif has_url:
            coverage = "HAS_URL_NO_TRANSCRIPT"
        elif present:
            coverage = "NO_URL_HAS_TRANSCRIPT"
        else:
            coverage = "NO_URL_NO_TRANSCRIPT"
        eligible = bool(present and quality in RETRIAGE_QUALITIES)
        out.append({
            "stg_code": g(row, "stg_code"),
            "md_file": md_file,
            "candidate_id": g(row, "candidate_id"),
            "title": g(row, "title"),
            "priority": g(row, "priority"),
            "source_quality": quality or None,
            "blocked_reason": g(row, "blocked_reason"),
            "source_url": g(row, "existing_source_url"),
            "has_source_url": has_url,
            "transcript_present": present,
            "transcript_path": rel,
            "transcript_bytes": size,
            "coverage_status_live": coverage,
            "coverage_status_xlsx": g(row, "coverage_status"),
            "recommended_next_step": g(row, "recommended_next_step"),
            "eligible_for_retriage": eligible,
            "user_notes": g(row, "user_notes"),
        })
    return out


def build() -> dict[str, Any]:
    candidates = _read_worklist()
    now = datetime.now(timezone.utc).isoformat()
    summary = {
        "total": len(candidates),
        "with_transcript": sum(1 for c in candidates if c["transcript_present"]),
        "without_transcript": sum(1 for c in candidates if not c["transcript_present"]),
        "high_quality": sum(1 for c in candidates if c["source_quality"] == "HIGH"),
        "eligible_for_retriage": sum(1 for c in candidates if c["eligible_for_retriage"]),
    }
    return {
        "schema_version": "1.0",
        "generated_at": now,
        "generator": "build_triage_registry.py",
        "source_worklist": WORKLIST_XLSX.relative_to(MCC_ROOT).as_posix(),
        "summary": summary,
        "candidates": candidates,
    }


def _strip_generated_at(text: str) -> str:
    return re.sub(r'"generated_at":\s*"[^"]*"', '"generated_at": "X"', text)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args(argv)
    payload = build()
    text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    if args.check:
        old = OUT.read_text(encoding="utf-8") if OUT.exists() else ""
        if _strip_generated_at(old) != _strip_generated_at(text):
            print("DRIFT: TRIAGE_CANDIDATE_REGISTRY.json differs")
            return 1
        return 0
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(text, encoding="utf-8")
    print(f"wrote {OUT.name}: {payload['summary']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
