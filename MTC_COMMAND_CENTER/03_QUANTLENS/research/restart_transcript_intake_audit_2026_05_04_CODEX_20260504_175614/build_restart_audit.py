from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


REPO_ROOT = Path(r"C:\LAB\tradingview-lab\01_MASTER TEMPLATE_V2")
QUANT_ROOT = REPO_ROOT / "06_QUANTLENS_LAB"
TRANSCRIPTS_DIR = QUANT_ROOT / "00_INBOX_REPORTS" / "Transcrips"
INTAKES_DIR = QUANT_ROOT / "00_INBOX_REPORTS" / "3 Mayıs"
OUTPUT_ROOT = Path(__file__).resolve().parent
CORRECTED_DIR = OUTPUT_ROOT / "corrected_intakes"
TRADER_WIKI_ROOT = QUANT_ROOT / "11_TRADER_WIKI"
LLM_WIKI_ROOT = QUANT_ROOT / "12_LLM_WIKI"
RUN_AT = datetime.now().isoformat(timespec="seconds")

TEXT_EXTS = {".md", ".txt", ".csv", ".json", ".jsonl", ".yml", ".yaml"}
YOUTUBE_RE = re.compile(r"(?:youtu\.be/|youtube\.com/watch\?v=)([A-Za-z0-9_-]{6,})")
URL_RE = re.compile(r"https?://[^\s)`>]+")
TIME_RE = re.compile(r"(?m)^\s*(?:\d{1,2}:)?\d{1,2}:\d{2}\b")


def rel(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve()))
    except Exception:
        return str(path)


def read_text(path: Path) -> str:
    if path.suffix.lower() not in TEXT_EXTS:
        return ""
    for encoding in ("utf-8-sig", "utf-8", "cp1254", "latin-1"):
        try:
            return path.read_text(encoding=encoding, errors="replace")
        except Exception:
            continue
    return ""


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8", newline="\n")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_title(value: str) -> str:
    value = value.lower()
    value = re.sub(r"https?://\S+", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    stop = {"the", "a", "an", "with", "and", "to", "of", "in", "for", "how", "report", "intake", "quantlens", "trading", "trade"}
    words = [w for w in value.split() if w not in stop]
    return " ".join(words)


def extract_video_id(text: str, filename: str = "") -> str:
    for source in (text, filename):
        match = YOUTUBE_RE.search(source)
        if match:
            return match.group(1)
        match = re.search(r"(?i)\bvideo id\b[^A-Za-z0-9_-]*`?([A-Za-z0-9_-]{6,})`?", source)
        if match:
            return match.group(1)
    match = re.search(r"(?<![A-Za-z0-9_-])([A-Za-z0-9_-]{11})(?![A-Za-z0-9_-])", filename)
    return match.group(1) if match else ""


def extract_url(text: str) -> str:
    match = URL_RE.search(text)
    return match.group(0).rstrip(".,") if match else ""


def extract_field(text: str, names: list[str]) -> str:
    for name in names:
        patterns = [
            rf"(?im)^\s*\*\*{re.escape(name)}\s*:\*\*\s*`?(.+?)`?\s*$",
            rf"(?im)^\s*-\s*\*\*{re.escape(name)}\s*:\*\*\s*`?(.+?)`?\s*$",
            rf"(?im)^\s*{re.escape(name)}\s*:\s*`?(.+?)`?\s*$",
            rf"(?im)^\s*\|\s*{re.escape(name)}\s*\|\s*`?(.+?)`?\s*\|",
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1).strip(" `")
    return ""


def md_table(rows: list[dict], cols: list[str], max_rows: int | None = None) -> str:
    shown = rows if max_rows is None else rows[:max_rows]
    out = ["|" + "|".join(cols) + "|", "|" + "|".join(["---"] * len(cols)) + "|"]
    for row in shown:
        vals = []
        for col in cols:
            vals.append(str(row.get(col, "")).replace("\n", " ")[:220])
        out.append("|" + "|".join(vals) + "|")
    if max_rows is not None and len(rows) > max_rows:
        out.append(f"\n_Only first {max_rows} of {len(rows)} rows shown._")
    return "\n".join(out) + "\n"


def write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def append_command(text: str) -> None:
    with (OUTPUT_ROOT / "COMMAND_LOG.txt").open("a", encoding="utf-8", newline="\n") as handle:
        handle.write(text + "\n")


def keywords_found(text: str, terms: list[str]) -> list[str]:
    lower = text.lower()
    return [term for term in terms if term.lower() in lower]


def first_snippet(text: str, terms: list[str]) -> tuple[str, str]:
    lower = text.lower()
    for term in terms:
        idx = lower.find(term.lower())
        if idx >= 0:
            start = max(0, idx - 90)
            end = min(len(text), idx + 190)
            stamp = ""
            prev = text[max(0, idx - 500):idx]
            stamps = TIME_RE.findall(prev)
            if stamps:
                stamp = stamps[-1].strip()
            snippet = re.sub(r"\s+", " ", text[start:end]).strip()
            return stamp, snippet[:320]
    return "", ""


@dataclass
class SourceDoc:
    kind: str
    path: Path
    text: str
    sha256: str
    size: int
    url: str
    video_id: str
    title: str
    norm: str


def scan_docs(root: Path, kind: str) -> list[SourceDoc]:
    docs = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        text = read_text(path)
        url = extract_url(text)
        video_id = extract_video_id(text, path.name)
        title = extract_field(text, ["Title", "Source Title", "Detected / File Title", "Uploaded Transcript File"])
        if not title:
            title = path.stem
        docs.append(SourceDoc(kind, path, text, sha256_file(path), path.stat().st_size, url, video_id, title, norm_title(title + " " + path.stem)))
    return docs


def classify_transcript(doc: SourceDoc) -> dict:
    text = doc.text
    lower = text.lower()
    timestamps = len(TIME_RE.findall(text))
    has_timestamps = timestamps >= 3
    word_count = len(re.findall(r"\w+", text))
    complete = "LIKELY_COMPLETE" if word_count > 1000 and has_timestamps else ("PARTIAL_OR_SHORT" if word_count > 120 else "LOW_CONTENT_OR_BINARY")
    sections = re.findall(r"(?m)^\s*(?:\d+\.\s*b[^\n]{0,80}|#{1,3}\s+.+|[A-Z][^\n]{8,90}:)\s*$", text)[:12]
    strategy_terms = keywords_found(lower, ["setup", "entry", "breakout", "pullback", "vcp", "vwap", "stage analysis", "stop", "position size", "risk", "sell rules", "swing", "day trader"])
    wisdom_terms = keywords_found(lower, ["process", "psychology", "discipline", "mistake", "lessons", "patience", "overtrading", "journal", "draw down", "mental"])
    tool_terms = keywords_found(lower, ["ai", "automation", "indicator", "workflow", "scanner", "spreadsheet"])
    if "interview" in lower or "podcast" in lower or "market wizard" in lower or "investing championship" in lower:
        transcript_type = "TRADER_INTERVIEW"
    elif len(strategy_terms) >= 4 and any(x in lower for x in ["entry", "setup", "stop", "breakout", "pullback"]):
        transcript_type = "STRATEGY_MECHANICAL"
    elif len(wisdom_terms) >= 3:
        transcript_type = "TRADER_WISDOM"
    elif len(tool_terms) >= 2:
        transcript_type = "TOOL_OR_AI_WORKFLOW"
    elif "market" in lower and "commentary" in lower:
        transcript_type = "MARKET_COMMENTARY"
    elif word_count < 200:
        transcript_type = "LOW_QUALITY_GENERIC"
    elif not any(x in lower for x in ["stock", "trade", "trading", "market", "risk"]):
        transcript_type = "NON_TRADING"
    else:
        transcript_type = "UNKNOWN"
    return {
        "transcript_path": rel(doc.path),
        "url": doc.url,
        "video_id": doc.video_id,
        "title": doc.title,
        "channel_or_source": extract_field(text, ["Primary Speaker", "Host / Context", "Source", "Channel"]) or "UNKNOWN",
        "duration": extract_field(text, ["Duration"]) or "UNKNOWN",
        "major_sections_or_timestamps": " | ".join([s.strip() for s in sections[:8]]) or ("timestamps_present" if has_timestamps else "none_detected"),
        "appears_complete": complete,
        "timestamps_present": "YES" if has_timestamps else "NO",
        "word_count": word_count,
        "transcript_type": transcript_type,
        "strategy_keywords": "; ".join(strategy_terms[:12]),
        "wisdom_keywords": "; ".join(wisdom_terms[:12]),
    }


def best_match(transcript: SourceDoc, intakes: list[SourceDoc]) -> tuple[SourceDoc | None, float, str]:
    if transcript.video_id:
        exact = [i for i in intakes if i.video_id == transcript.video_id]
        if exact:
            return exact[0], 1.0, "video_id"
    scores = []
    for intake in intakes:
        score = SequenceMatcher(None, transcript.norm, intake.norm).ratio()
        if transcript.url and transcript.url == intake.url:
            score = max(score, 0.98)
        scores.append((score, intake))
    if not scores:
        return None, 0.0, "none"
    score, intake = max(scores, key=lambda x: x[0])
    return (intake, score, "title_filename_similarity") if score >= 0.45 else (None, score, "low_similarity")


def match_all(transcripts: list[SourceDoc], intakes: list[SourceDoc]) -> list[dict]:
    rows = []
    intake_by_vid = defaultdict(list)
    for intake in intakes:
        if intake.video_id:
            intake_by_vid[intake.video_id].append(intake)
    for transcript in transcripts:
        matched, score, method = best_match(transcript, intakes)
        matches = intake_by_vid.get(transcript.video_id, []) if transcript.video_id else ([matched] if matched else [])
        if matches:
            for intake in matches:
                rows.append({
                    "transcript_path": rel(transcript.path),
                    "intake_path": rel(intake.path),
                    "transcript_video_id": transcript.video_id,
                    "intake_video_id": intake.video_id,
                    "transcript_title": transcript.title,
                    "intake_title": intake.title,
                    "match_method": "video_id" if transcript.video_id and transcript.video_id == intake.video_id else method,
                    "match_score": 1.0 if transcript.video_id and transcript.video_id == intake.video_id else round(score, 3),
                    "match_status": "MATCHED",
                })
        else:
            rows.append({
                "transcript_path": rel(transcript.path),
                "intake_path": "",
                "transcript_video_id": transcript.video_id,
                "intake_video_id": "",
                "transcript_title": transcript.title,
                "intake_title": "",
                "match_method": method,
                "match_score": round(score, 3),
                "match_status": "TRANSCRIPT_WITHOUT_INTAKE",
            })
    transcript_vids = {t.video_id for t in transcripts if t.video_id}
    for intake in intakes:
        if intake.video_id and intake.video_id in transcript_vids:
            continue
        matched_to = any(row["intake_path"] == rel(intake.path) for row in rows)
        if not matched_to:
            rows.append({
                "transcript_path": "",
                "intake_path": rel(intake.path),
                "transcript_video_id": "",
                "intake_video_id": intake.video_id,
                "transcript_title": "",
                "intake_title": intake.title,
                "match_method": "none",
                "match_score": 0,
                "match_status": "INTAKE_WITHOUT_TRANSCRIPT",
            })
    return rows


FAITHFULNESS_FIELDS = {
    "main_strategy_idea": ["strategy", "setup", "model", "edge", "playbook"],
    "market_asset_class": ["stock", "equity", "crypto", "futures", "options", "tqqq", "small cap", "low float"],
    "timeframe": ["weekly", "daily", "intraday", "5m", "1h", "time frame", "morning", "session"],
    "setup_conditions": ["setup", "base", "flag", "vcp", "pullback", "breakout", "relative strength", "volume"],
    "entry_rules": ["entry", "buy", "trigger", "breakout", "opening range", "retest"],
    "exit_rules": ["exit", "sell", "take profit", "profit", "trim"],
    "stop_risk_logic": ["stop", "risk", "loss", "drawdown", "invalidat"],
    "sizing_logic": ["size", "position", "exposure", "concentrat", "portfolio"],
    "filters_gates": ["filter", "gate", "market environment", "regime", "sector", "relative strength"],
    "data_requirements": ["data", "float", "volume", "earnings", "halt", "scanner", "session"],
    "warnings_caveats": ["warning", "caveat", "avoid", "don't", "do not", "mistake", "overtrading"],
    "trader_wisdom": ["wisdom", "lesson", "process", "patience", "discipline", "psychology", "journal"],
    "backtesting_lessons": ["backtest", "sample", "lookahead", "misleading", "data"],
}


def audit_pair(transcript: SourceDoc | None, intake: SourceDoc, duplicate_intake: bool) -> dict:
    if not transcript:
        return {
            "transcript_path": "",
            "intake_path": rel(intake.path),
            "video_id": intake.video_id,
            "title": intake.title,
            "intake_faithfulness": "WRONG_SOURCE_MATCH",
            "mismatch_summary": "No matching transcript found in transcript directory for this intake video_id/title.",
            "evidence": "INTAKE_WITHOUT_TRANSCRIPT",
            "missing_fields": "source transcript",
            "needs_corrected_copy": "YES",
        }
    t_lower = transcript.text.lower()
    i_lower = intake.text.lower()
    missing = []
    weak = []
    evidence_parts = []
    for field, terms in FAITHFULNESS_FIELDS.items():
        in_transcript = any(term in t_lower for term in terms)
        in_intake = any(term in i_lower for term in terms)
        if in_transcript and not in_intake:
            missing.append(field)
            stamp, snippet = first_snippet(transcript.text, terms)
            if snippet:
                evidence_parts.append(f"{field}@{stamp or 'no_timestamp'}: {snippet}")
        elif in_transcript and in_intake:
            weak.append(field)
    is_raw = len(re.findall(r"(?m)^\s*\d{1,2}:\d{2}", intake.text)) > 20 and "# " not in intake.text[:500]
    if is_raw:
        status = "RAW_TRANSCRIPT_NOT_INTAKE"
    elif duplicate_intake:
        status = "MOSTLY_FAITHFUL_MINOR_MISSING" if len(missing) <= 3 else "PARTIAL_MISSING_RULES"
    elif len(missing) >= 7:
        status = "NEEDS_REWRITE"
    elif {"entry_rules", "exit_rules", "stop_risk_logic"} & set(missing):
        status = "PARTIAL_MISSING_RULES"
    elif {"trader_wisdom", "warnings_caveats", "backtesting_lessons"} & set(missing):
        status = "PARTIAL_MISSING_IMPORTANT_WISDOM"
    elif len(missing) >= 3:
        status = "MOSTLY_FAITHFUL_MINOR_MISSING"
    else:
        status = "FAITHFUL"
    return {
        "transcript_path": rel(transcript.path),
        "intake_path": rel(intake.path),
        "video_id": transcript.video_id or intake.video_id,
        "title": transcript.title,
        "intake_faithfulness": status,
        "mismatch_summary": "Missing/weak capture: " + (", ".join(missing) if missing else "none detected by keyword audit"),
        "evidence": " || ".join(evidence_parts[:4]),
        "missing_fields": "; ".join(missing),
        "needs_corrected_copy": "YES" if status not in {"FAITHFUL", "MOSTLY_FAITHFUL_MINOR_MISSING"} else "NO",
    }


def classify_item(doc: SourceDoc | None, intake: SourceDoc | None, faithfulness: str, corrected_path: str = "") -> dict:
    source = doc or intake
    text = ((doc.text if doc else "") + "\n" + (intake.text if intake else "")).lower()
    title = source.title if source else "UNKNOWN"
    vid = source.video_id if source else ""
    if faithfulness in {"WRONG_SOURCE_MATCH", "RAW_TRANSCRIPT_NOT_INTAKE"}:
        classification = "DATA_BLOCKED"
    elif any(x in text for x in ["day trader", "day-trading", "intraday", "low float", "halt", "morning"]):
        classification = "DAY_TRADE_STRATEGY"
    elif any(x in text for x in ["weekly", "position trading", "tqqq", "longer-term"]):
        classification = "POSITION_TRADING_STRATEGY"
    elif any(x in text for x in ["swing", "pullback", "breakout", "vcp", "wedge", "flag"]):
        classification = "SWING_TRADE_STRATEGY"
    elif any(x in text for x in ["vwap", "avwap", "relative strength", "stage analysis", "canslim", "market regime", "sector"]):
        classification = "FILTER_MODULE"
    elif any(x in text for x in ["sell rule", "exit", "profit", "stop", "risk management", "position sizing", "sizing"]):
        classification = "SIZING_OR_RISK_MODULE" if any(x in text for x in ["risk", "sizing", "position size"]) else "EXIT_MODULE"
    elif any(x in text for x in ["psychology", "mental", "process", "discipline", "overtrading", "jared tendler"]):
        classification = "TRADER_WISDOM_ONLY"
    elif any(x in text for x in ["automation", "workflow", "scanner", "ai"]):
        classification = "PROCESS_OR_WORKFLOW_ONLY"
    else:
        classification = "REJECT_NOT_TESTABLE"
    if any(x in text[:3000] for x in ["low_value_generic_strategy", "reject_low_quality", "low_quality_generic"]):
        classification = "REJECT_LOW_QUALITY"
    source_quality = "HIGH_QUALITY_TRADER_INTERVIEW" if any(x in text for x in ["market wizard", "investing championship", "hedge fund", "champion"]) else ("MEDIUM_QUALITY_EDUCATIONAL" if "strategy" in text or "setup" in text else "UNKNOWN")
    rule_clarity = min(5, sum(1 for x in ["entry", "exit", "stop", "timeframe", "setup"] if x in text))
    testability = min(5, sum(1 for x in ["entry", "exit", "stop", "rule", "scan", "condition"] if x in text))
    credibility = 5 if source_quality == "HIGH_QUALITY_TRADER_INTERVIEW" else (3 if source_quality == "MEDIUM_QUALITY_EDUCATIONAL" else 1)
    data_availability = 2 if any(x in text for x in ["low float", "halt", "options", "earnings"]) else 4
    edge = min(5, 1 + sum(1 for x in ["relative strength", "volume", "risk", "trend", "breakout"] if x in text))
    mtc = 2 if classification in {"DAY_TRADE_STRATEGY", "TRADER_WISDOM_ONLY", "DATA_BLOCKED"} else (4 if classification in {"FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE"} else 3)
    visual = 5 if classification in {"DAY_TRADE_STRATEGY", "SWING_TRADE_STRATEGY", "POSITION_TRADING_STRATEGY"} and testability >= 3 else 2
    final = round(rule_clarity * 1.1 + testability * 1.2 + credibility + data_availability + edge + mtc + visual, 2)
    next_action = {
        "DAY_TRADE_STRATEGY": "MANUAL_VISUAL_REVIEW_ONLY_NO_BACKTEST_YET",
        "SWING_TRADE_STRATEGY": "MANUAL_VISUAL_REVIEW_ONLY_NO_BACKTEST_YET",
        "POSITION_TRADING_STRATEGY": "MANUAL_VISUAL_REVIEW_ONLY_NO_BACKTEST_YET",
        "FILTER_MODULE": "SPEC_AS_FILTER_IDEA_THEN_SANDBOX_LATER",
        "EXIT_MODULE": "SPEC_AS_EXIT_MODULE_THEN_SANDBOX_LATER",
        "SIZING_OR_RISK_MODULE": "KEEP_FOR_MTC_MONEY_MANAGEMENT_MAPPING",
        "TRADER_WISDOM_ONLY": "ADD_TO_WIKI_ONLY",
        "DATA_BLOCKED": "DO_NOT_TEST_UNTIL_SOURCE_FIXED",
        "REJECT_LOW_QUALITY": "REJECT_NO_TEST",
    }.get(classification, "NEEDS_MANUAL_REVIEW")
    return {
        "candidate_id": f"QLR_{vid or hashlib.sha1(title.encode()).hexdigest()[:10]}",
        "source_url": f"https://www.youtube.com/watch?v={vid}" if vid else (source.url if source else ""),
        "source_title": title,
        "transcript_path": rel(doc.path) if doc else "",
        "intake_path": rel(intake.path) if intake else "",
        "corrected_intake_path": corrected_path,
        "source_quality": source_quality,
        "rule_clarity_score": rule_clarity,
        "mechanical_testability_score": testability,
        "trader_credibility_context_score": credibility,
        "data_availability_score": data_availability,
        "expected_edge_plausibility_score": edge,
        "MTC_compatibility_score": mtc,
        "visual_review_priority": visual,
        "final_priority_score": final,
        "classification": classification,
        "next_action": next_action,
        "do_not_test_yet_reason": "No strategy backtests in this phase; needs isolated sandbox and visual review.",
    }


def extract_wisdom(doc: SourceDoc, intake: SourceDoc | None) -> list[dict]:
    buckets = [
        ("RISK_MANAGEMENT", ["stop loss", "risk", "drawdown", "loss", "position size", "exposure"]),
        ("EXECUTION", ["execute", "entry", "exit", "sell", "overtrading", "chase"]),
        ("BACKTESTING_AND_DATA", ["backtest", "data", "sample", "lookahead", "misleading"]),
        ("MARKET_STRUCTURE", ["stage", "trend", "relative strength", "volume", "sector", "regime"]),
        ("POSITION_SIZING_AND_PORTFOLIO", ["portfolio", "concentrat", "position sizing", "exposure"]),
        ("TRADER_PSYCHOLOGY", ["psychology", "discipline", "process", "patience", "mental", "promise"]),
        ("STRATEGY_DESIGN", ["setup", "edge", "playbook", "rules", "framework"]),
        ("THINGS_NOT_TO_AUTOMATE_YET", ["discretion", "feel", "judgment", "avoid", "don't", "do not"]),
    ]
    rows = []
    for category, terms in buckets:
        stamp, snippet = first_snippet(doc.text, terms)
        if not snippet:
            continue
        wid = f"WIS_{doc.video_id or hashlib.sha1(doc.title.encode()).hexdigest()[:8]}_{len(rows)+1:02d}"
        confidence = "HIGH" if stamp and len(snippet) > 80 else "MEDIUM"
        action = "DO_NOT_AUTOMATE" if category == "THINGS_NOT_TO_AUTOMATE_YET" else ("CONVERT_TO_RISK_MODULE_IDEA" if category in {"RISK_MANAGEMENT", "POSITION_SIZING_AND_PORTFOLIO"} else ("CONVERT_TO_FILTER_IDEA" if category == "MARKET_STRUCTURE" else "ADD_TO_WIKI_ONLY"))
        rows.append({
            "wisdom_id": wid,
            "trader_source_video": doc.title,
            "source_url": f"https://www.youtube.com/watch?v={doc.video_id}" if doc.video_id else doc.url,
            "source_transcript_path": rel(doc.path),
            "source_intake_path": rel(intake.path) if intake else "",
            "timestamp": stamp,
            "short_wisdom_statement": snippet,
            "why_it_matters": "Transcript evidence points to a process/risk/design lesson that should not be lost during strategy extraction.",
            "MTC_or_QuantLens_mapping": "Use as research constraint, checklist item, or future module idea; not proof of profitability.",
            "category": category,
            "confidence": confidence,
            "action": action,
        })
    return rows


def create_corrected_intake(row: dict, transcript: SourceDoc | None, intake: SourceDoc) -> str:
    CORRECTED_DIR.mkdir(parents=True, exist_ok=True)
    base = f"CORRECTED_{intake.video_id or hashlib.sha1(rel(intake.path).encode()).hexdigest()[:10]}_{Path(intake.path).stem}.md"
    out = CORRECTED_DIR / re.sub(r"[^A-Za-z0-9_.-]+", "_", base)
    classification = classify_item(transcript, intake, row["intake_faithfulness"])
    text = [
        "# Corrected QuantLens Intake Copy",
        "",
        f"- Source transcript path: `{rel(transcript.path) if transcript else 'UNKNOWN_NOT_FOUND'}`",
        f"- Source intake path: `{rel(intake.path)}`",
        f"- URL: `{classification['source_url']}`",
        f"- Video ID: `{classification['candidate_id'].replace('QLR_', '')}`",
        f"- Title: `{classification['source_title']}`",
        "",
        "## Transcript-Based Correction Summary",
        row["mismatch_summary"],
        "",
        "## Corrected Classification",
        f"- Classification: `{classification['classification']}`",
        f"- Next action: `{classification['next_action']}`",
        "- Production/Pine status: `NOT_PINE_READY_NOT_PROFIT_CLAIMED`",
        "",
        "## Mechanical Strategy Rules",
        "Only fields directly supported by transcript/intake keyword evidence are preserved. Any missing entry/exit/stop detail remains blocked for manual review.",
        f"- Rule clarity score: `{classification['rule_clarity_score']}/5`",
        f"- Mechanical testability score: `{classification['mechanical_testability_score']}/5`",
        "",
        "## Trader Wisdom",
        row["evidence"] or "No short transcript evidence snippet was detected by automated audit.",
        "",
        "## Risk / Execution Lessons",
        "Respect stops, avoid overtrading/chasing where transcript indicates those risks, and keep sizing/risk logic separate from signal extraction.",
        "",
        "## Backtesting / Data Lessons",
        "No backtest was run in this phase. Data blockers such as low-float equity, halts, options, earnings, session structure, or discretionary execution must be resolved before testing.",
        "",
        "## MTC Relevance",
        f"MTC compatibility score: `{classification['MTC_compatibility_score']}/5`. Reuse only through isolated sandbox mapping, not direct production integration.",
        "",
        "## Routing Decision",
        f"- Test/keep/reject decision: `{classification['next_action']}`",
        "",
        "## What Was Wrong Or Missing In Original Intake",
        f"- Faithfulness class: `{row['intake_faithfulness']}`",
        f"- Missing fields: `{row['missing_fields'] or 'none detected'}`",
    ]
    write_text(out, "\n".join(text) + "\n")
    return rel(out)


def safe_subfolder(root: Path, name: str) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    target = root / name
    if target.exists():
        target = root / f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    target.mkdir(parents=True, exist_ok=False)
    return target


def build_workbook(path: Path, tables: dict[str, tuple[list[dict], list[str]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fills = {
        "header": PatternFill("solid", fgColor="1F4E78"),
        "good": PatternFill("solid", fgColor="C6EFCE"),
        "warn": PatternFill("solid", fgColor="FFF2CC"),
        "bad": PatternFill("solid", fgColor="F4CCCC"),
    }
    for sheet_name, (rows, fields) in tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fills["header"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 55)
        for data_row in ws.iter_rows(min_row=2):
            joined = " ".join(str(c.value) for c in data_row if c.value)
            fill = None
            if any(x in joined for x in ["FAITHFUL", "MATCHED", "HIGH"]):
                fill = fills["good"]
            if any(x in joined for x in ["MISSING", "PARTIAL", "MEDIUM", "NEEDS"]):
                fill = fills["warn"]
            if any(x in joined for x in ["WRONG", "REJECT", "BLOCKED", "LOW_QUALITY"]):
                fill = fills["bad"]
            if fill:
                for cell in data_row:
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    append_command(f"{RUN_AT} python build_restart_audit.py")
    before_hashes = {}
    protected = [REPO_ROOT / "01_PINE" / "MTC_V2.pine", REPO_ROOT / "00_PYTHON" / "mtc_v2" / "core" / "runner.py"]
    for path in protected:
        before_hashes[rel(path)] = sha256_file(path) if path.exists() else "MISSING"
    original_hashes = {}
    for root in (TRANSCRIPTS_DIR, INTAKES_DIR):
        for p in root.rglob("*"):
            if p.is_file():
                original_hashes[rel(p)] = sha256_file(p)

    transcripts = scan_docs(TRANSCRIPTS_DIR, "transcript")
    intakes = scan_docs(INTAKES_DIR, "intake")
    match_rows = match_all(transcripts, intakes)
    transcript_by_vid = {t.video_id: t for t in transcripts if t.video_id}
    transcript_by_path = {rel(t.path): t for t in transcripts}
    intake_by_path = {rel(i.path): i for i in intakes}
    duplicate_vids_t = {vid for vid, n in Counter(t.video_id for t in transcripts if t.video_id).items() if n > 1}
    duplicate_vids_i = {vid for vid, n in Counter(i.video_id for i in intakes if i.video_id).items() if n > 1}

    transcript_inventory = [{
        "path": rel(t.path), "filename": t.path.name, "size_bytes": t.size, "sha256": t.sha256,
        "url": t.url, "video_id": t.video_id, "title": t.title,
        "duplicate_video_id": "YES" if t.video_id in duplicate_vids_t else "NO",
        "is_text_readable": "YES" if t.text else "NO",
    } for t in transcripts]
    intake_inventory = [{
        "path": rel(i.path), "filename": i.path.name, "size_bytes": i.size, "sha256": i.sha256,
        "url": i.url, "video_id": i.video_id, "title": i.title,
        "duplicate_video_id": "YES" if i.video_id in duplicate_vids_i else "NO",
        "raw_transcript_suspect": "YES" if len(TIME_RE.findall(i.text)) > 20 and "# " not in i.text[:500] else "NO",
        "is_text_readable": "YES" if i.text else "NO",
    } for i in intakes]
    quality_rows = [classify_transcript(t) for t in transcripts]

    intake_paths_by_vid = defaultdict(list)
    for i in intakes:
        intake_paths_by_vid[i.video_id].append(rel(i.path))
    faith_rows = []
    corrected_index = []
    corrected_by_intake = {}
    for intake in intakes:
        transcript = transcript_by_vid.get(intake.video_id)
        if not transcript:
            best_t = None
            best_score = 0
            for t in transcripts:
                score = SequenceMatcher(None, t.norm, intake.norm).ratio()
                if score > best_score:
                    best_t, best_score = t, score
            transcript = best_t if best_score >= 0.55 else None
        row = audit_pair(transcript, intake, duplicate_intake=len(intake_paths_by_vid[intake.video_id]) > 1 if intake.video_id else False)
        faith_rows.append(row)
        if row["needs_corrected_copy"] == "YES":
            corrected_path = create_corrected_intake(row, transcript, intake)
            corrected_by_intake[rel(intake.path)] = corrected_path
            corrected_index.append({
                "corrected_intake_path": corrected_path,
                "source_transcript_path": row["transcript_path"],
                "source_intake_path": row["intake_path"],
                "video_id": row["video_id"],
                "title": row["title"],
                "faithfulness": row["intake_faithfulness"],
                "reason": row["mismatch_summary"],
            })

    missing_rows = []
    for row in faith_rows:
        if row["missing_fields"] or row["intake_faithfulness"] in {"WRONG_SOURCE_MATCH", "RAW_TRANSCRIPT_NOT_INTAKE"}:
            missing_rows.append({
                "video_id": row["video_id"], "title": row["title"], "transcript_path": row["transcript_path"],
                "intake_path": row["intake_path"], "missing_or_wrong": row["missing_fields"] or row["intake_faithfulness"],
                "evidence": row["evidence"], "repair_required": row["needs_corrected_copy"],
            })

    classification_rows = []
    for row in faith_rows:
        intake = intake_by_path.get(row["intake_path"])
        transcript = transcript_by_path.get(row["transcript_path"])
        if not intake:
            continue
        classification_rows.append(classify_item(transcript, intake, row["intake_faithfulness"], corrected_by_intake.get(row["intake_path"], "")))
    seen_candidate = Counter(r["candidate_id"] for r in classification_rows)
    for r in classification_rows:
        if seen_candidate[r["candidate_id"]] > 1:
            r["classification"] = "DUPLICATE_OR_MERGED"
            r["next_action"] = "MERGE_WITH_CANONICAL_SOURCE_BEFORE_TESTING"

    wisdom_rows = []
    intake_for_vid = {i.video_id: i for i in intakes if i.video_id}
    for t in transcripts:
        wisdom_rows.extend(extract_wisdom(t, intake_for_vid.get(t.video_id)))
    wisdom_rows = wisdom_rows[: max(1, len(transcripts) * 4)]

    transcript_fields = ["path", "filename", "size_bytes", "sha256", "url", "video_id", "title", "duplicate_video_id", "is_text_readable"]
    intake_fields = ["path", "filename", "size_bytes", "sha256", "url", "video_id", "title", "duplicate_video_id", "raw_transcript_suspect", "is_text_readable"]
    match_fields = ["transcript_path", "intake_path", "transcript_video_id", "intake_video_id", "transcript_title", "intake_title", "match_method", "match_score", "match_status"]
    quality_fields = ["transcript_path", "url", "video_id", "title", "channel_or_source", "duration", "major_sections_or_timestamps", "appears_complete", "timestamps_present", "word_count", "transcript_type", "strategy_keywords", "wisdom_keywords"]
    faith_fields = ["transcript_path", "intake_path", "video_id", "title", "intake_faithfulness", "mismatch_summary", "evidence", "missing_fields", "needs_corrected_copy"]
    missing_fields = ["video_id", "title", "transcript_path", "intake_path", "missing_or_wrong", "evidence", "repair_required"]
    corrected_fields = ["corrected_intake_path", "source_transcript_path", "source_intake_path", "video_id", "title", "faithfulness", "reason"]
    class_fields = ["candidate_id", "source_url", "source_title", "transcript_path", "intake_path", "corrected_intake_path", "source_quality", "rule_clarity_score", "mechanical_testability_score", "trader_credibility_context_score", "data_availability_score", "expected_edge_plausibility_score", "MTC_compatibility_score", "visual_review_priority", "final_priority_score", "classification", "next_action", "do_not_test_yet_reason"]
    wisdom_fields = ["wisdom_id", "trader_source_video", "source_url", "source_transcript_path", "source_intake_path", "timestamp", "short_wisdom_statement", "why_it_matters", "MTC_or_QuantLens_mapping", "category", "confidence", "action"]

    csv_outputs = [
        ("TRANSCRIPT_INVENTORY.csv", transcript_inventory, transcript_fields),
        ("INTAKE_INVENTORY.csv", intake_inventory, intake_fields),
        ("TRANSCRIPT_INTAKE_MATCH_MAP.csv", match_rows, match_fields),
        ("TRANSCRIPT_QUALITY_AUDIT.csv", quality_rows, quality_fields),
        ("INTAKE_FAITHFULNESS_AUDIT.csv", faith_rows, faith_fields),
        ("MISSING_WISDOM_AND_RULES.csv", missing_rows, missing_fields),
        ("CORRECTED_INTAKE_INDEX.csv", corrected_index, corrected_fields),
        ("MASTER_CANDIDATE_CLASSIFICATION.csv", classification_rows, class_fields),
        ("TRADER_WISDOM_INDEX.csv", wisdom_rows, wisdom_fields),
    ]
    for name, rows, fields in csv_outputs:
        write_csv(OUTPUT_ROOT / name, rows, fields)

    subsets = {
        "DAY_TRADE_CANDIDATES": [r for r in classification_rows if r["classification"] == "DAY_TRADE_STRATEGY"],
        "SWING_TRADE_CANDIDATES": [r for r in classification_rows if r["classification"] == "SWING_TRADE_STRATEGY"],
        "POSITION_TRADING_CANDIDATES": [r for r in classification_rows if r["classification"] == "POSITION_TRADING_STRATEGY"],
        "FILTER_EXIT_SIZING_MODULES": [r for r in classification_rows if r["classification"] in {"FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE", "PORTFOLIO_OR_REGIME_MODULE"}],
        "WISDOM_ONLY_ITEMS": [r for r in classification_rows if r["classification"] == "TRADER_WISDOM_ONLY"],
        "REJECTED_OR_BLOCKED_ITEMS": [r for r in classification_rows if r["classification"] in {"DATA_BLOCKED", "REJECT_NOT_TESTABLE", "REJECT_LOW_QUALITY", "DUPLICATE_OR_MERGED"}],
    }
    for name, rows in subsets.items():
        write_csv(OUTPUT_ROOT / f"{name}.csv", rows, class_fields)
        write_text(OUTPUT_ROOT / f"{name}.md", f"# {name.replace('_', ' ').title()}\n\nCount: {len(rows)}\n\n" + md_table(rows, ["candidate_id", "source_title", "classification", "final_priority_score", "next_action"], 80))

    duplicate_report = [
        "# Duplicate And Missing Report",
        "",
        f"- Transcripts scanned: {len(transcripts)}",
        f"- Intakes scanned: {len(intakes)}",
        f"- Duplicate transcript video IDs: {len(duplicate_vids_t)}",
        f"- Duplicate intake video IDs: {len(duplicate_vids_i)}",
        f"- Transcript without intake rows: {sum(1 for r in match_rows if r['match_status'] == 'TRANSCRIPT_WITHOUT_INTAKE')}",
        f"- Intake without transcript rows: {sum(1 for r in match_rows if r['match_status'] == 'INTAKE_WITHOUT_TRANSCRIPT')}",
        "",
        "## Duplicate Intake Video IDs",
        ", ".join(sorted(duplicate_vids_i)) or "None",
        "",
        "## Duplicate Transcript Video IDs",
        ", ".join(sorted(duplicate_vids_t)) or "None",
    ]
    write_text(OUTPUT_ROOT / "DUPLICATE_AND_MISSING_REPORT.md", "\n".join(duplicate_report) + "\n")
    write_text(OUTPUT_ROOT / "DISCOVERY_REPORT.md", f"# Discovery Report\n\n- Transcript directory: `{TRANSCRIPTS_DIR}`\n- Intake directory: `{INTAKES_DIR}`\n- Transcripts scanned: {len(transcripts)}\n- Intakes scanned: {len(intakes)}\n- Matched rows: {sum(1 for r in match_rows if r['match_status']=='MATCHED')}\n- Intake without transcript: {sum(1 for r in match_rows if r['match_status']=='INTAKE_WITHOUT_TRANSCRIPT')}\n- Transcript without intake: {sum(1 for r in match_rows if r['match_status']=='TRANSCRIPT_WITHOUT_INTAKE')}\n\n" + md_table(match_rows, ["transcript_video_id", "intake_video_id", "match_status", "match_method", "transcript_title", "intake_title"], 120))
    write_text(OUTPUT_ROOT / "TRANSCRIPT_QUALITY_AUDIT.md", "# Transcript Quality Audit\n\n" + md_table(quality_rows, ["video_id", "title", "transcript_type", "appears_complete", "timestamps_present", "word_count"], 120))
    write_text(OUTPUT_ROOT / "INTAKE_FAITHFULNESS_AUDIT.md", "# Intake Faithfulness Audit\n\n" + md_table(faith_rows, ["video_id", "title", "intake_faithfulness", "missing_fields", "needs_corrected_copy"], 120))
    write_text(OUTPUT_ROOT / "MISCLASSIFICATION_REPORT.md", "# Misclassification Report\n\n" + md_table([r for r in faith_rows if r["needs_corrected_copy"] == "YES"], ["video_id", "title", "intake_faithfulness", "mismatch_summary"], 120))
    write_text(OUTPUT_ROOT / "CORRECTED_INTAKE_SUMMARY.md", "# Corrected Intake Summary\n\n" + md_table(corrected_index, ["video_id", "title", "faithfulness", "corrected_intake_path"], 120))
    write_text(OUTPUT_ROOT / "MASTER_CANDIDATE_CLASSIFICATION.md", "# Master Candidate Classification\n\nNo candidate is profitable or Pine-ready based on this phase.\n\n" + md_table(classification_rows, ["candidate_id", "source_title", "classification", "final_priority_score", "next_action"], 150))

    wiki_dir = safe_subfolder(TRADER_WIKI_ROOT, "2026_05_04_restart_import")
    category_files = {
        "RISK_MANAGEMENT": "01_RISK_MANAGEMENT_WISDOM.md",
        "EXECUTION": "02_EXECUTION_WISDOM.md",
        "BACKTESTING_AND_DATA": "03_BACKTESTING_AND_DATA_WISDOM.md",
        "MARKET_STRUCTURE": "04_MARKET_STRUCTURE_WISDOM.md",
        "POSITION_SIZING_AND_PORTFOLIO": "05_POSITION_SIZING_AND_PORTFOLIO_WISDOM.md",
        "TRADER_PSYCHOLOGY": "06_TRADER_PSYCHOLOGY.md",
        "STRATEGY_DESIGN": "07_STRATEGY_DESIGN_PRINCIPLES.md",
        "THINGS_NOT_TO_AUTOMATE_YET": "08_THINGS_NOT_TO_AUTOMATE_YET.md",
    }
    write_text(wiki_dir / "README.md", "# Trader Wiki Restart Import\n\nNon-destructive dated import from transcript audit. Wisdom is not a strategy claim.\n")
    write_text(wiki_dir / "00_INDEX.md", "# Trader Wiki Index\n\n" + "\n".join(f"- [{v}]({v})" for v in category_files.values()) + "\n")
    for cat, fname in category_files.items():
        rows = [r for r in wisdom_rows if r["category"] == cat]
        write_text(wiki_dir / fname, f"# {cat.replace('_', ' ').title()}\n\n" + md_table(rows, ["wisdom_id", "trader_source_video", "timestamp", "short_wisdom_statement", "confidence", "action"], 200))
    write_csv(wiki_dir / "09_SOURCE_MAP.csv", wisdom_rows, wisdom_fields)
    write_csv(OUTPUT_ROOT / "TRADER_WISDOM_INDEX.csv", wisdom_rows, wisdom_fields)
    write_text(OUTPUT_ROOT / "TRADER_WIKI_IMPORT_REPORT.md", f"# Trader Wiki Import Report\n\n- Wiki directory: `{rel(wiki_dir)}`\n- Wisdom items: {len(wisdom_rows)}\n- Dated non-destructive import: yes\n")

    llm_dir = safe_subfolder(LLM_WIKI_ROOT, "2026_05_04_restart_import")
    def jsonl(path: Path, rows: list[dict]) -> None:
        with path.open("w", encoding="utf-8", newline="\n") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    kb_rows = []
    for r in classification_rows:
        kb_rows.append({
            "record_type": "strategy_or_module_classification",
            "source_url": r["source_url"], "transcript_path": r["transcript_path"], "intake_path": r["intake_path"],
            "candidate_id": r["candidate_id"], "wisdom_id": "",
            "title": r["source_title"], "summary": f"{r['classification']} classified for later manual/sandbox review only.",
            "exact_rules_if_available": "See corrected intake or original intake; no Pine-ready rules promoted.",
            "ambiguity_flags": [r["do_not_test_yet_reason"]],
            "data_requirements": "Resolve source-specific market/timeframe/session requirements before testing.",
            "MTC_mapping": f"MTC compatibility score {r['MTC_compatibility_score']}/5; sandbox only.",
            "recommended_next_action": r["next_action"],
        })
    wisdom_json_rows = [{
        "record_type": "wisdom_item",
        "source_url": r["source_url"], "transcript_path": r["source_transcript_path"], "intake_path": r["source_intake_path"],
        "candidate_id": "", "wisdom_id": r["wisdom_id"], "title": r["trader_source_video"],
        "summary": r["short_wisdom_statement"], "exact_rules_if_available": "",
        "ambiguity_flags": ["wisdom_not_strategy"],
        "data_requirements": "N/A unless converted to a module idea later.",
        "MTC_mapping": r["MTC_or_QuantLens_mapping"], "recommended_next_action": r["action"],
    } for r in wisdom_rows]
    jsonl(llm_dir / "quantlens_knowledge_base.jsonl", kb_rows + wisdom_json_rows)
    jsonl(llm_dir / "quantlens_strategy_candidates.jsonl", [x for x in kb_rows if "STRATEGY" in x["summary"]])
    jsonl(llm_dir / "quantlens_wisdom_items.jsonl", wisdom_json_rows)
    jsonl(llm_dir / "quantlens_filter_exit_sizing_modules.jsonl", [x for x in kb_rows if any(k in x["summary"] for k in ["FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE"])])
    write_csv(llm_dir / "quantlens_source_map.csv", classification_rows, class_fields)
    write_text(llm_dir / "LLM_WIKI_README.md", "# LLM Wiki Restart Import\n\nSelf-contained JSONL records. No profitability, Pine-ready, or production integration claims.\n")

    manual_rows = []
    candidates_for_visual = [r for r in classification_rows if r["classification"] in {"DAY_TRADE_STRATEGY", "SWING_TRADE_STRATEGY", "POSITION_TRADING_STRATEGY", "FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE"}]
    candidates_for_visual.sort(key=lambda r: float(r["final_priority_score"]), reverse=True)
    for rank, r in enumerate(candidates_for_visual, 1):
        klass = r["classification"]
        market = "US equities likely" if any(x in r["source_title"].lower() for x in ["stock", "investing championship", "minervini", "tqqq", "low float"]) else "UNKNOWN_OR_CROSS_MARKET"
        tf = "weekly/daily" if klass == "POSITION_TRADING_STRATEGY" else ("intraday" if klass == "DAY_TRADE_STRATEGY" else "daily/1h visual first")
        manual_rows.append({
            "rank": rank,
            "candidate_id": r["candidate_id"],
            "candidate_name": r["source_title"],
            "strategy_class": klass,
            "market": market,
            "native_timeframe": tf,
            "recommended_first_test_asset_universe": "Manual chart set only; no backtest yet",
            "recommended_first_chart_timeframe_in_TradingView": tf,
            "what_visual_behavior_should_be_checked": "Confirm raw long/short pulse concept, repaint risk, entry/exit visibility, and whether rules are actually mechanical.",
            "required_Python_Pine_sandbox_later": "candidate_contract.yml + python_signal_model.py + standalone_pine_visual_review.pine",
            "MTC_money_management_applicability": f"MTC score {r['MTC_compatibility_score']}/5; reuse only through sandbox harness",
            "data_blockers": r["do_not_test_yet_reason"],
            "expected_difficulty": "HIGH" if r["data_availability_score"] <= 2 else "MEDIUM",
            "priority_rationale": f"Score {r['final_priority_score']}; visual priority {r['visual_review_priority']}/5; no profitability claim.",
        })
    manual_fields = ["rank", "candidate_id", "candidate_name", "strategy_class", "market", "native_timeframe", "recommended_first_test_asset_universe", "recommended_first_chart_timeframe_in_TradingView", "what_visual_behavior_should_be_checked", "required_Python_Pine_sandbox_later", "MTC_money_management_applicability", "data_blockers", "expected_difficulty", "priority_rationale"]
    write_csv(OUTPUT_ROOT / "MANUAL_VISUAL_TEST_QUEUE.csv", manual_rows, manual_fields)
    write_text(OUTPUT_ROOT / "MANUAL_VISUAL_TEST_QUEUE.md", "# Manual Visual Test Queue\n\nNo strategy implementation or backtest was run.\n\n" + md_table(manual_rows, ["rank", "candidate_id", "candidate_name", "strategy_class", "expected_difficulty", "priority_rationale"], 120))

    mtc_decision = """# MTC Sandbox Architecture Decision

1. Should all strategies be added to MTC_V2 at once?
No. Do not add all strategies to MTC_V2. This phase found research candidates, modules, wisdom, duplicates, and blockers; none are production-ready.

2. Should each be separate at first?
Yes. Create one isolated candidate sandbox per strategy/module first.

3. How can MTC_V2 money management be reused without polluting production MTC_V2?
Reuse it through a shared MTC-compatible sandbox harness that receives raw long/short pulses and applies direction toggles, position sizing, SL/TP, trailing/BE, time exit, exit-first behavior, and no-repaint discipline outside production MTC_V2.

4. What should the standalone Pine visual review harness look like?
One standalone Pine file per candidate that plots raw setup, long pulse, short pulse, invalidation, stop/target guides, and debug labels. It must not be merged into 01_PINE/MTC_V2.pine.

5. What should the Python parity harness look like?
candidate_contract.yml defines inputs/outputs; python_signal_model.py emits raw pulses; python_backtest_harness.py is sandbox-only; golden_cases.csv captures expected signal rows; tests verify deterministic signal behavior before broader research.

6. What files should be created later for each candidate?
candidate_contract.yml, candidate_rules.md, python_signal_model.py, python_backtest_harness.py, golden_cases.csv, tests/, standalone_pine_visual_review.pine, MTC_mapping.md, VISUAL_REVIEW_CHECKLIST.md.

7. What promotion gate is required before real MTC integration?
Manual visual review, source-faithful rule spec, Python/Pine signal parity, no-repaint checks, data availability checks, OOS/robustness research, and explicit owner approval. Only then can a candidate become an MTC producer/filter candidate.
"""
    write_text(OUTPUT_ROOT / "MTC_SANDBOX_ARCHITECTURE_DECISION.md", mtc_decision)

    workbook_path = OUTPUT_ROOT / "QUANTLENS_RESTART_CLASSIFICATION_WORKBOOK.xlsx"
    tables = {
        "README": ([{"field": "Scope", "value": "Transcript/intake restart audit; no strategy backtests; no Pine promotion."}, {"field": "Output Root", "value": str(OUTPUT_ROOT)}], ["field", "value"]),
        "Source Inventory": (transcript_inventory + intake_inventory, sorted(set(transcript_fields + intake_fields))),
        "Transcript Intake Match": (match_rows, match_fields),
        "Intake Faithfulness": (faith_rows, faith_fields),
        "Missing Wisdom": (missing_rows, missing_fields),
        "Corrected Intakes": (corrected_index, corrected_fields),
        "Master Classification": (classification_rows, class_fields),
        "Day Trade": (subsets["DAY_TRADE_CANDIDATES"], class_fields),
        "Swing Trade": (subsets["SWING_TRADE_CANDIDATES"], class_fields),
        "Position Trading": (subsets["POSITION_TRADING_CANDIDATES"], class_fields),
        "Filters Exits Sizing": (subsets["FILTER_EXIT_SIZING_MODULES"], class_fields),
        "Trader Wisdom": (wisdom_rows, wisdom_fields),
        "Data Blocked": ([r for r in classification_rows if r["classification"] == "DATA_BLOCKED"], class_fields),
        "Rejected": (subsets["REJECTED_OR_BLOCKED_ITEMS"], class_fields),
        "Manual Visual Test Queue": (manual_rows, manual_fields),
        "MTC Mapping": (classification_rows, class_fields),
    }
    build_workbook(workbook_path, tables)

    summary_counts = Counter(r["classification"] for r in classification_rows)
    final_report = f"""# Final Restart Audit Report

## 1. Executive summary
Restart audit completed from local transcript and intake files. No strategy was backtested, promoted to Pine, or claimed profitable.

## 2. Transcript count
{len(transcripts)}

## 3. Intake count
{len(intakes)}

## 4. Match success/failure
- Matched rows: {sum(1 for r in match_rows if r['match_status']=='MATCHED')}
- Transcript without intake: {sum(1 for r in match_rows if r['match_status']=='TRANSCRIPT_WITHOUT_INTAKE')}
- Intake without transcript: {sum(1 for r in match_rows if r['match_status']=='INTAKE_WITHOUT_TRANSCRIPT')}

## 5. Corrected intake count
{len(corrected_index)}

## 6. Major intake errors found
{md_table([r for r in faith_rows if r['needs_corrected_copy']=='YES'], ['video_id','title','intake_faithfulness','missing_fields'], 50)}

## 7. Wisdom recovered
{len(wisdom_rows)} wisdom rows routed to Trader Wiki.

## 8. Wiki files created
Trader Wiki import: `{rel(wiki_dir)}`

## 9. LLM wiki files created
LLM Wiki import: `{rel(llm_dir)}`

## 10. Classification summary
{dict(summary_counts)}

## 11. Day trade candidates summary
{len(subsets['DAY_TRADE_CANDIDATES'])} candidates; manual visual review only.

## 12. Swing trade candidates summary
{len(subsets['SWING_TRADE_CANDIDATES'])} candidates; manual visual review only.

## 13. Position trading candidates summary
{len(subsets['POSITION_TRADING_CANDIDATES'])} candidates; manual visual review only.

## 14. Filters/exits/sizing modules
{len(subsets['FILTER_EXIT_SIZING_MODULES'])} module candidates.

## 15. Rejected/data blocked items
{len(subsets['REJECTED_OR_BLOCKED_ITEMS'])} rejected, blocked, or duplicate/merged items.

## 16. Excel workbook path
`{rel(workbook_path)}`

## 17. Manual visual test queue
`{rel(OUTPUT_ROOT / 'MANUAL_VISUAL_TEST_QUEUE.csv')}`

## 18. MTC sandbox architecture decision
`{rel(OUTPUT_ROOT / 'MTC_SANDBOX_ARCHITECTURE_DECISION.md')}`

## 19. What Antigravity must audit next
Review corrected intakes against transcripts manually, resolve no-transcript intakes, choose one sandbox candidate, and verify source-specific data availability before any backtest.

## 20. Validation
See `VALIDATION_REPORT.md`.

## 21. Exact files created
See `FILES_CREATED.txt`.

## 22. Next prompt for Antigravity
Audit `{rel(OUTPUT_ROOT)}`. Verify corrected intakes against source transcripts, resolve DATA_BLOCKED rows, and select exactly one candidate for a standalone visual-review sandbox. Do not modify MTC_V2.pine or production runners.
"""
    write_text(OUTPUT_ROOT / "FINAL_RESTART_AUDIT_REPORT.md", final_report)

    files_created = sorted(rel(p) for p in OUTPUT_ROOT.rglob("*") if p.is_file())
    files_created += sorted(rel(p) for p in wiki_dir.rglob("*") if p.is_file())
    files_created += sorted(rel(p) for p in llm_dir.rglob("*") if p.is_file())
    write_text(OUTPUT_ROOT / "FILES_CREATED.txt", "\n".join(files_created) + "\n")

    validations = []
    try:
        subprocess.run(["python", "-m", "py_compile", str(Path(__file__))], cwd=REPO_ROOT, check=True, capture_output=True, text=True)
        validations.append({"check": "py_compile_new_python", "status": "PASS", "detail": rel(Path(__file__))})
    except subprocess.CalledProcessError as exc:
        validations.append({"check": "py_compile_new_python", "status": "FAIL", "detail": exc.stderr})
    try:
        wb2 = load_workbook(workbook_path, read_only=True)
        sheet_count = len(wb2.sheetnames)
        wb2.close()
        validations.append({"check": "excel_openpyxl_open", "status": "PASS", "detail": f"{sheet_count} sheets"})
    except Exception as exc:
        validations.append({"check": "excel_openpyxl_open", "status": "FAIL", "detail": str(exc)})
    csv_failures = []
    for p in OUTPUT_ROOT.rglob("*.csv"):
        try:
            with p.open("r", encoding="utf-8", newline="") as handle:
                list(csv.reader(handle))
        except Exception as exc:
            csv_failures.append(f"{rel(p)}: {exc}")
    validations.append({"check": "csv_readable", "status": "PASS" if not csv_failures else "FAIL", "detail": "; ".join(csv_failures[:5])})
    jsonl_failures = []
    for p in llm_dir.rglob("*.jsonl"):
        try:
            with p.open("r", encoding="utf-8") as handle:
                for line_no, line in enumerate(handle, 1):
                    if line.strip():
                        json.loads(line)
        except Exception as exc:
            jsonl_failures.append(f"{rel(p)}:{line_no}: {exc}")
    validations.append({"check": "jsonl_valid", "status": "PASS" if not jsonl_failures else "FAIL", "detail": "; ".join(jsonl_failures[:5])})
    after_original_hashes = {}
    for root in (TRANSCRIPTS_DIR, INTAKES_DIR):
        for p in root.rglob("*"):
            if p.is_file():
                after_original_hashes[rel(p)] = sha256_file(p)
    changed_originals = [k for k, v in original_hashes.items() if after_original_hashes.get(k) != v]
    validations.append({"check": "original_transcripts_intakes_unmodified", "status": "PASS" if not changed_originals else "FAIL", "detail": "; ".join(changed_originals[:10])})
    after_hashes = {rel(path): sha256_file(path) if path.exists() else "MISSING" for path in protected}
    pine_key = rel(protected[0])
    runner_key = rel(protected[1])
    validations.append({"check": "MTC_V2_pine_untouched", "status": "PASS" if before_hashes.get(pine_key) == after_hashes.get(pine_key) else "FAIL", "detail": f"before={before_hashes.get(pine_key)} after={after_hashes.get(pine_key)}"})
    validations.append({"check": "production_runner_untouched", "status": "PASS" if before_hashes.get(runner_key) == after_hashes.get(runner_key) else "FAIL", "detail": f"before={before_hashes.get(runner_key)} after={after_hashes.get(runner_key)}"})
    try:
        after_status = subprocess.run(["git", "status", "--short"], cwd=REPO_ROOT, check=True, capture_output=True, text=True).stdout
        write_text(OUTPUT_ROOT / "git_status_after.txt", after_status)
        validations.append({"check": "git_status_after_captured", "status": "PASS", "detail": "git_status_after.txt"})
    except Exception as exc:
        validations.append({"check": "git_status_after_captured", "status": "FAIL", "detail": str(exc)})
    write_csv(OUTPUT_ROOT / "VALIDATION_CHECKLIST.csv", validations, ["check", "status", "detail"])
    write_text(OUTPUT_ROOT / "VALIDATION_REPORT.md", "# Validation Report\n\n" + md_table(validations, ["check", "status", "detail"]))

    state = {
        "status": "COMPLETED_WITH_AUDIT_LIMITS",
        "run_at": RUN_AT,
        "output_root": str(OUTPUT_ROOT),
        "transcript_count": len(transcripts),
        "intake_count": len(intakes),
        "corrected_intake_count": len(corrected_index),
        "wisdom_item_count": len(wisdom_rows),
        "classification_summary": dict(summary_counts),
        "validation": validations,
        "no_backtests_run": True,
        "no_pine_promotion": True,
    }
    write_text(OUTPUT_ROOT / "STATE.json", json.dumps(state, indent=2, ensure_ascii=False) + "\n")
    write_text(OUTPUT_ROOT / "RUN_LOG.md", f"# Run Log\n\n- {RUN_AT}: initialized output root and captured git_status_before.txt.\n- Scanned transcripts: {len(transcripts)}.\n- Scanned intakes: {len(intakes)}.\n- Built match map, faithfulness audit, corrected intakes, Trader Wiki, LLM Wiki, workbook, final report.\n- Ran validation checks; see VALIDATION_REPORT.md.\n")
    write_text(OUTPUT_ROOT / "ERRORS_AND_RECOVERY.md", "# Errors And Recovery\n\nNo fatal errors. Automated faithfulness audit is keyword/evidence based and requires Antigravity manual review before testing.\n")


if __name__ == "__main__":
    main()
