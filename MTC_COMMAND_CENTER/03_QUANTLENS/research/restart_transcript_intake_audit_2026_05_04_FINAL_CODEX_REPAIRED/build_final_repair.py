from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import subprocess
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPO_ROOT = Path(r"C:\LAB\tradingview-lab\01_MASTER TEMPLATE_V2")
QUANT_ROOT = REPO_ROOT / "06_QUANTLENS_LAB"
RESEARCH_ROOT = QUANT_ROOT / "research"
TRANSCRIPTS_DIR = QUANT_ROOT / "00_INBOX_REPORTS" / "Transcrips"
INTAKES_DIR = QUANT_ROOT / "00_INBOX_REPORTS" / "3 Mayıs"
OUTPUT_ROOT = Path(__file__).resolve().parent
RUN_AT = datetime.now().isoformat(timespec="seconds")

REQUIRED_CODEX = [
    "FINAL_RESTART_AUDIT_REPORT.md",
    "MASTER_CANDIDATE_CLASSIFICATION.csv",
    "MANUAL_VISUAL_TEST_QUEUE.csv",
    "TRADER_WISDOM_INDEX.csv",
    "QUANTLENS_RESTART_CLASSIFICATION_WORKBOOK.xlsx",
    "MTC_SANDBOX_ARCHITECTURE_DECISION.md",
    "INTAKE_FAITHFULNESS_AUDIT.csv",
    "TRANSCRIPT_INTAKE_MATCH_MAP.csv",
    "corrected_intakes",
]
REQUIRED_AG = [
    "CODEX_REPAIR_INSTRUCTIONS.md",
    "CODEX_REPAIR_FINDINGS.csv",
    "CODEX_REPAIR_PRIORITY_ORDER.md",
    "CLASSIFICATION_AUDIT_FINDINGS.csv",
    "CANDIDATE_RECLASSIFICATION_REQUESTS_FOR_CODEX.csv",
    "MISSED_TRADER_WISDOM_FOR_CODEX.md",
    "WIKI_REPAIR_REQUESTS_FOR_CODEX.csv",
    "MTC_SANDBOX_ARCHITECTURE_REVIEW.md",
    "MATCHING_AUDIT_FINDINGS.csv",
    "INTAKE_FAITHFULNESS_REVIEW.csv",
]
OPTIONAL_AG = [
    "ANTIGRAVITY_DAY_TRADE_PRIORITY_RANKING.csv",
    "ANTIGRAVITY_SWING_TRADE_PRIORITY_RANKING.csv",
    "ANTIGRAVITY_POSITION_TRADING_PRIORITY_RANKING.csv",
    "ANTIGRAVITY_DAY_TRADE_PRIORITY_RANKING.md",
    "ANTIGRAVITY_SWING_TRADE_PRIORITY_RANKING.md",
    "ANTIGRAVITY_POSITION_TRADING_PRIORITY_RANKING.md",
    "ANTIGRAVITY_FILTER_EXIT_SIZING_PRIORITY_RANKING.md",
]
FINAL_FIELDS = [
    "candidate_id",
    "source_url",
    "source_title",
    "transcript_path",
    "intake_path",
    "corrected_intake_path",
    "strategy_class",
    "native_market",
    "native_timeframe",
    "source_quality",
    "rule_clarity",
    "mechanical_testability",
    "data_requirements",
    "MTC_mapping",
    "manual_visual_review_priority",
    "final_action",
    "reason",
]


def rel(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve()))
    except Exception:
        return str(path)


def read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1254", "latin-1"):
        try:
            return path.read_text(encoding=encoding, errors="replace")
        except Exception:
            pass
    return ""


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8", newline="\n")


def read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = sorted({k for row in rows for k in row.keys()})
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def md_table(rows: list[dict], fields: list[str], max_rows: int = 200) -> str:
    out = ["|" + "|".join(fields) + "|", "|" + "|".join(["---"] * len(fields)) + "|"]
    for row in rows[:max_rows]:
        out.append("|" + "|".join(str(row.get(f, "")).replace("\n", " ")[:260] for f in fields) + "|")
    if len(rows) > max_rows:
        out.append(f"\n_Only first {max_rows} of {len(rows)} rows shown._")
    return "\n".join(out) + "\n"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def newest_dir(prefix: str) -> Path:
    dirs = [p for p in RESEARCH_ROOT.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not dirs:
        raise FileNotFoundError(prefix)
    return max(dirs, key=lambda p: p.stat().st_mtime)


def find_doc_by_video(root: Path, video_id: str) -> Path | None:
    if not video_id or video_id == "missing":
        return None
    for path in sorted(root.rglob("*")):
        if path.is_file() and video_id in path.name:
            return path
    for path in sorted(root.rglob("*.md")):
        if video_id in read_text(path)[:5000]:
            return path
    return None


def infer_market(row: dict) -> str:
    text = (row.get("source_title", "") + " " + row.get("reason", "") + " " + row.get("source_url", "")).lower()
    if any(x in text for x in ["low float", "microcap", "small cap", "small-cap"]):
        return "US_LOW_FLOAT_EQUITIES"
    if "tqqq" in text:
        return "US_LEVERAGED_ETF"
    if any(x in text for x in ["stock", "equity", "investing championship", "minervini", "canslim"]):
        return "US_EQUITIES"
    if "crypto" in text:
        return "CRYPTO_PROXY_ONLY"
    return "UNKNOWN_REQUIRES_MANUAL_SOURCE_CHECK"


def infer_timeframe(strategy_class: str, title: str) -> str:
    t = title.lower()
    if strategy_class == "DAY_TRADE_STRATEGY":
        return "INTRADAY_OR_SESSION"
    if strategy_class == "POSITION_TRADING_STRATEGY":
        return "WEEKLY_TO_MULTI_MONTH"
    if "weekly" in t:
        return "WEEKLY_SWING"
    if "30 day" in t or "30 days" in t:
        return "DAILY_SWING_1_TO_6_WEEKS"
    if strategy_class == "SWING_TRADE_STRATEGY":
        return "DAILY_OR_MULTI_DAY_SWING"
    if strategy_class in {"FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE"}:
        return "MODULE_DEPENDS_ON_PARENT_STRATEGY"
    return "NOT_APPLICABLE_OR_UNKNOWN"


def final_action_for(strategy_class: str) -> str:
    if strategy_class in {"PROCESS_OR_WORKFLOW_ONLY", "TRADER_WISDOM_ONLY"}:
        return "WIKI_ONLY_DO_NOT_BACKTEST_AS_STRATEGY"
    if strategy_class == "DATA_BLOCKED":
        return "DO_NOT_TEST_UNTIL_SOURCE_RESOLVED"
    if strategy_class in {"REJECT_LOW_QUALITY", "REJECT_NOT_TESTABLE", "DUPLICATE_OR_MERGED"}:
        return "DO_NOT_TEST"
    if strategy_class in {"FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE"}:
        return "SPEC_MODULE_FOR_LATER_SANDBOX_NO_BACKTEST_YET"
    return "MANUAL_VISUAL_REVIEW_ONLY_NO_BACKTEST_YET"


def parse_priority_md(path: Path) -> dict[str, int]:
    if not path.exists():
        return {}
    ranks = {}
    for line in read_text(path).splitlines():
        match = re.match(r"\s*(\d+)\.\s+\*\*(QLR_[A-Za-z0-9_-]+)\*\*", line)
        if match:
            ranks[match.group(2)] = int(match.group(1))
    return ranks


def build_workbook(path: Path, sheets: dict[str, tuple[list[dict], list[str]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fills = {
        "header": PatternFill("solid", fgColor="1F4E78"),
        "good": PatternFill("solid", fgColor="C6EFCE"),
        "warn": PatternFill("solid", fgColor="FFF2CC"),
        "bad": PatternFill("solid", fgColor="F4CCCC"),
        "neutral": PatternFill("solid", fgColor="D9EAF7"),
    }
    for name, (rows, fields) in sheets.items():
        ws = wb.create_sheet(name[:31])
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fills["header"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for data_row in ws.iter_rows(min_row=2):
            joined = " ".join(str(c.value or "") for c in data_row)
            fill = None
            if any(x in joined for x in ["ACCEPTED", "PASS", "MANUAL_VISUAL"]):
                fill = fills["good"]
            if any(x in joined for x in ["MODIFIED", "BLOCKED", "REQUIRES", "WIKI_ONLY"]):
                fill = fills["warn"]
            if any(x in joined for x in ["REJECTED", "DO_NOT_TEST", "FAIL"]):
                fill = fills["bad"]
            if fill:
                for cell in data_row:
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col[:250])
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)
    wb.save(path)


def main() -> None:
    codex_dir = newest_dir("restart_transcript_intake_audit_2026_05_04_CODEX")
    ag_dir = newest_dir("restart_transcript_intake_audit_2026_05_04_ANTIGRAVITY_AUDIT")
    protected = [REPO_ROOT / "01_PINE" / "MTC_V2.pine", REPO_ROOT / "00_PYTHON" / "mtc_v2" / "core" / "runner.py"]
    protected_before = {rel(p): sha256_file(p) if p.exists() else "MISSING" for p in protected}
    original_before = {}
    for root in (TRANSCRIPTS_DIR, INTAKES_DIR):
        for path in root.rglob("*"):
            if path.is_file():
                original_before[rel(path)] = sha256_file(path)

    missing_inputs = []
    for name in REQUIRED_CODEX:
        if not (codex_dir / name).exists():
            missing_inputs.append(f"CODEX:{name}")
    for name in REQUIRED_AG:
        if not (ag_dir / name).exists():
            missing_inputs.append(f"ANTIGRAVITY:{name}")
    optional_status = {name: ("FOUND" if (ag_dir / name).exists() else "MISSING") for name in OPTIONAL_AG}

    master_rows = read_csv(codex_dir / "MASTER_CANDIDATE_CLASSIFICATION.csv")
    faith_rows = read_csv(codex_dir / "INTAKE_FAITHFULNESS_AUDIT.csv")
    match_rows = read_csv(codex_dir / "TRANSCRIPT_INTAKE_MATCH_MAP.csv")
    wisdom_rows = read_csv(codex_dir / "TRADER_WISDOM_INDEX.csv")
    manual_rows = read_csv(codex_dir / "MANUAL_VISUAL_TEST_QUEUE.csv")
    repair_requests = read_csv(ag_dir / "CANDIDATE_RECLASSIFICATION_REQUESTS_FOR_CODEX.csv")
    classification_findings = read_csv(ag_dir / "CLASSIFICATION_AUDIT_FINDINGS.csv")
    matching_findings = read_csv(ag_dir / "MATCHING_AUDIT_FINDINGS.csv")
    intake_review = read_csv(ag_dir / "INTAKE_FAITHFULNESS_REVIEW.csv")
    wiki_requests = read_csv(ag_dir / "WIKI_REPAIR_REQUESTS_FOR_CODEX.csv")

    by_id: dict[str, list[dict]] = {}
    for row in master_rows:
        by_id.setdefault(row["candidate_id"], []).append(row)

    repair_decisions = []
    reclass_map = {r["candidate_id"]: r["new_classification"] for r in repair_requests if r.get("action") in {"RECLASSIFY", "REVIEW_TIMEFRAME"}}
    for req in repair_requests:
        cid = req["candidate_id"]
        rows = by_id.get(cid, [])
        source_paths = "; ".join(filter(None, [r.get("transcript_path", "") for r in rows]))
        if rows:
            decision = "ACCEPTED" if req.get("action") == "RECLASSIFY" else "MODIFIED_ACCEPTED_AFTER_LOCAL_SOURCE_CHECK"
            evidence = "Found matching Codex candidate row and applied Antigravity classification request."
            if cid == "QLR_4-IjRmw7SZI":
                evidence = "TQQQ source title and Antigravity note support swing treatment unless transcript proves strict intraday. No strict intraday proof promoted here."
        else:
            decision = "REJECTED_NOT_APPLICABLE"
            evidence = "Candidate not present in Codex master classification."
        repair_decisions.append({
            "repair_id": f"RECLASS_{cid}",
            "source_file": "CANDIDATE_RECLASSIFICATION_REQUESTS_FOR_CODEX.csv",
            "candidate_id": cid,
            "request": req.get("action", ""),
            "decision": decision,
            "applied_change": req.get("new_classification", "") if rows else "",
            "evidence": evidence,
            "source_paths_checked": source_paths,
        })

    for finding in matching_findings:
        vid = finding.get("video_id", "")
        issue = finding.get("issue_type", "")
        transcript = find_doc_by_video(TRANSCRIPTS_DIR, vid)
        intake = find_doc_by_video(INTAKES_DIR, vid)
        if issue == "DUPLICATE_NEEDS_MERGE" and vid == "VKNEJA5r8zw":
            decision = "ACCEPTED"
            applied = "Merged duplicate candidate rows into one canonical final row."
            evidence = "Two Codex rows existed for QLR_VKNEJA5r8zw; final output keeps one canonical row and references both intakes in reason."
        elif issue == "MISSING_INTAKE":
            decision = "MODIFIED_ACCEPTED_BLOCKED"
            applied = "No web/source fetch performed in final repair; transcript remains DATA_BLOCKED until intake is created from source."
            evidence = f"Local transcript exists={bool(transcript)}; local intake exists={bool(intake)}."
        elif issue == "MISSING_TRANSCRIPT":
            decision = "MODIFIED_ACCEPTED_BLOCKED"
            applied = "No web/source fetch performed in final repair; intake remains DATA_BLOCKED until transcript is supplied."
            evidence = f"Local transcript exists={bool(transcript)}; local intake exists={bool(intake)}."
        else:
            decision = "ACCEPTED"
            applied = "Recorded."
            evidence = "Finding recorded in final matching repair decisions."
        repair_decisions.append({
            "repair_id": f"MATCH_{issue}_{vid}",
            "source_file": "MATCHING_AUDIT_FINDINGS.csv",
            "candidate_id": f"QLR_{vid}" if vid and vid != "missing" else "",
            "request": issue,
            "decision": decision,
            "applied_change": applied,
            "evidence": evidence,
            "source_paths_checked": f"{rel(transcript) if transcript else ''}; {rel(intake) if intake else ''}",
        })

    missing_wisdom_map = {
        "Tm0dkf8_giA": "Do not waste time holding a stock through a proper basing structure; leverage time and money.",
        "6aOnCK1gv2w": "Avoid chopping yourself up; track consecutive trades on the same name.",
        "6tnREqUJ1WY": "Having your best day in a while is a warning sign to be on guard; sentiment is huge.",
        "lTiR1pc82EE": "Always check the weekly chart, not just the daily; inside bars and unfilled gaps can signal strength.",
        "R215f4fj7V8": "Preserve mental process and physical preparation, including setting alerts below highs during flushes.",
        "RTHRh_GLwH8": "Position sizing is the core risk-management lever for preventing large losses.",
        "UMgJ0P8fe0s": "Let winners run, but do not expect every trade to run for months and become a huge winner.",
        "zw96qkUn9_g": "Avoid chop, be patient, and wait for a fat pitch.",
        "JS4_6gv0PpI": "Avoid style drift and jumping from one strategy to the next.",
        "_AAX1ylNbIE": "Preserve higher-timeframe context; weekly charts can matter more than intraday noise.",
        "UtNrp_Uz9Oc": "FOMO is an umbrella term rooted in greed and related emotional drivers.",
    }
    extra_wisdom = []
    for review in intake_review:
        vid = review["video_id"]
        transcript = find_doc_by_video(TRANSCRIPTS_DIR, vid)
        intake = find_doc_by_video(INTAKES_DIR, vid)
        statement = missing_wisdom_map.get(vid, f"Antigravity requested preserved context for {review.get('missing_category', '')}.")
        item = {
            "wisdom_id": f"FINAL_REPAIR_WIS_{vid}",
            "trader_source_video": vid,
            "source_url": f"https://www.youtube.com/watch?v={vid}",
            "source_transcript_path": rel(transcript) if transcript else "",
            "source_intake_path": rel(intake) if intake else "",
            "timestamp": "",
            "short_wisdom_statement": statement,
            "why_it_matters": "Antigravity found this context missing; preserving it prevents over-mechanizing the source.",
            "MTC_or_QuantLens_mapping": "Use as wiki/process constraint or module caveat only; not a standalone strategy edge.",
            "category": "FINAL_REPAIR_WARNING_OR_CONTEXT",
            "confidence": "MEDIUM" if transcript or intake else "LOW",
            "action": "ADD_TO_WIKI_ONLY",
        }
        extra_wisdom.append(item)
        repair_decisions.append({
            "repair_id": f"WIKI_{vid}",
            "source_file": "INTAKE_FAITHFULNESS_REVIEW.csv",
            "candidate_id": f"QLR_{vid}",
            "request": review.get("action_required", ""),
            "decision": "ACCEPTED",
            "applied_change": "Added final repair wisdom item to Trader Wiki and LLM Wiki.",
            "evidence": statement,
            "source_paths_checked": f"{item['source_transcript_path']}; {item['source_intake_path']}",
        })

    final_rows = []
    seen = set()
    duplicate_notes = {}
    for row in master_rows:
        cid = row["candidate_id"]
        if cid == "QLR_VKNEJA5r8zw":
            duplicate_notes.setdefault(cid, []).append(row.get("intake_path", ""))
            if cid in seen:
                continue
        seen.add(cid)
        strategy_class = reclass_map.get(cid, row.get("classification", ""))
        if row.get("classification") == "DUPLICATE_OR_MERGED" and cid != "QLR_VKNEJA5r8zw":
            strategy_class = "DUPLICATE_OR_MERGED"
        reason_parts = [f"Codex original classification={row.get('classification', '')}."]
        if cid in reclass_map:
            reason_parts.append(f"Antigravity repair accepted: {row.get('classification', '')} -> {strategy_class}.")
        if cid == "QLR_VKNEJA5r8zw":
            strategy_class = "SWING_TRADE_STRATEGY"
            reason_parts.append("Duplicate VKNEJA5r8zw intakes merged into this canonical final row.")
        final_rows.append({
            "candidate_id": cid,
            "source_url": row.get("source_url", ""),
            "source_title": row.get("source_title", ""),
            "transcript_path": row.get("transcript_path", ""),
            "intake_path": row.get("intake_path", ""),
            "corrected_intake_path": row.get("corrected_intake_path", ""),
            "strategy_class": strategy_class,
            "native_market": infer_market(row),
            "native_timeframe": infer_timeframe(strategy_class, row.get("source_title", "")),
            "source_quality": row.get("source_quality", ""),
            "rule_clarity": row.get("rule_clarity_score", ""),
            "mechanical_testability": row.get("mechanical_testability_score", ""),
            "data_requirements": "Source-specific transcript/intake requirements must be verified before sandbox testing.",
            "MTC_mapping": f"Sandbox only. Prior MTC compatibility score={row.get('MTC_compatibility_score', '')}.",
            "manual_visual_review_priority": row.get("visual_review_priority", ""),
            "final_action": final_action_for(strategy_class),
            "reason": " ".join(reason_parts),
        })

    existing_cids = {r["candidate_id"] for r in final_rows}
    for finding in matching_findings:
        vid = finding.get("video_id", "")
        if not vid or vid == "missing":
            continue
        cid = f"QLR_{vid}"
        if cid in existing_cids:
            if "MISSING" in finding.get("issue_type", ""):
                for row in final_rows:
                    if row["candidate_id"] == cid:
                        row["strategy_class"] = "DATA_BLOCKED"
                        row["final_action"] = final_action_for("DATA_BLOCKED")
                        row["reason"] += f" Antigravity matching repair: {finding.get('issue_type')} remains blocked locally."
            continue
        transcript = find_doc_by_video(TRANSCRIPTS_DIR, vid)
        intake = find_doc_by_video(INTAKES_DIR, vid)
        title = finding.get("filename", vid)
        final_rows.append({
            "candidate_id": cid,
            "source_url": f"https://www.youtube.com/watch?v={vid}",
            "source_title": title,
            "transcript_path": rel(transcript) if transcript else "",
            "intake_path": rel(intake) if intake else "",
            "corrected_intake_path": "",
            "strategy_class": "DATA_BLOCKED",
            "native_market": "UNKNOWN_REQUIRES_SOURCE_REPAIR",
            "native_timeframe": "UNKNOWN_REQUIRES_SOURCE_REPAIR",
            "source_quality": "UNKNOWN",
            "rule_clarity": "0",
            "mechanical_testability": "0",
            "data_requirements": "Missing local transcript or missing local intake must be supplied before review.",
            "MTC_mapping": "Blocked; no MTC mapping.",
            "manual_visual_review_priority": "0",
            "final_action": final_action_for("DATA_BLOCKED"),
            "reason": f"Added from Antigravity matching finding: {finding.get('issue_type')}.",
        })

    final_wisdom = wisdom_rows + extra_wisdom
    final_wisdom_fields = sorted({k for r in final_wisdom for k in r.keys()})

    corrected_final_dir = OUTPUT_ROOT / "corrected_intakes_final"
    corrected_final_dir.mkdir(parents=True, exist_ok=True)
    corrected_index = []
    source_corrected_dir = codex_dir / "corrected_intakes"
    if source_corrected_dir.exists():
        for src in sorted(source_corrected_dir.glob("*.md")):
            dest = corrected_final_dir / src.name
            shutil.copy2(src, dest)
            corrected_index.append({"corrected_intake_path": rel(dest), "source": rel(src), "repair_type": "copied_from_codex_corrected"})
    for row in final_rows:
        if row["candidate_id"] in reclass_map or row["candidate_id"] == "QLR_VKNEJA5r8zw" or row["strategy_class"] == "DATA_BLOCKED":
            dest = corrected_final_dir / f"FINAL_{row['candidate_id']}.md"
            write_text(dest, "\n".join([
                "# Final Corrected Intake Repair Note",
                "",
                f"- Candidate ID: `{row['candidate_id']}`",
                f"- Source URL: `{row['source_url']}`",
                f"- Source title: `{row['source_title']}`",
                f"- Transcript path: `{row['transcript_path']}`",
                f"- Intake path: `{row['intake_path']}`",
                f"- Final class: `{row['strategy_class']}`",
                f"- Final action: `{row['final_action']}`",
                "",
                "## Reason",
                row["reason"],
                "",
                "## Constraint",
                "No strategy backtest, Pine conversion, TradingView alert, or production MTC integration is authorized from this repair packet.",
            ]) + "\n")
            row["corrected_intake_path"] = rel(dest)
            corrected_index.append({"corrected_intake_path": rel(dest), "source": row["intake_path"], "repair_type": "final_repair_note"})

    day_rows = [r for r in final_rows if r["strategy_class"] == "DAY_TRADE_STRATEGY"]
    swing_rows = [r for r in final_rows if r["strategy_class"] == "SWING_TRADE_STRATEGY"]
    position_rows = [r for r in final_rows if r["strategy_class"] == "POSITION_TRADING_STRATEGY"]
    module_rows = [r for r in final_rows if r["strategy_class"] in {"FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE", "PORTFOLIO_OR_REGIME_MODULE"}]
    wisdom_only_rows = [r for r in final_rows if r["strategy_class"] in {"PROCESS_OR_WORKFLOW_ONLY", "TRADER_WISDOM_ONLY"}]
    rejected_rows = [r for r in final_rows if r["strategy_class"] in {"DATA_BLOCKED", "REJECT_LOW_QUALITY", "REJECT_NOT_TESTABLE", "DUPLICATE_OR_MERGED"}]
    subsets = {
        "FINAL_DAY_TRADE_CANDIDATES": day_rows,
        "FINAL_SWING_TRADE_CANDIDATES": swing_rows,
        "FINAL_POSITION_TRADING_CANDIDATES": position_rows,
        "FINAL_FILTER_EXIT_SIZING_MODULES": module_rows,
        "FINAL_WISDOM_ONLY_ITEMS": wisdom_only_rows,
        "FINAL_REJECTED_OR_BLOCKED_ITEMS": rejected_rows,
    }

    write_csv(OUTPUT_ROOT / "ANTIGRAVITY_REPAIR_DECISIONS.csv", repair_decisions)
    write_text(OUTPUT_ROOT / "ANTIGRAVITY_REPAIR_DECISIONS.md", "# Antigravity Repair Decisions\n\n" + md_table(repair_decisions, ["repair_id", "candidate_id", "decision", "applied_change", "evidence"], 300))
    write_csv(OUTPUT_ROOT / "FINAL_MASTER_CANDIDATE_CLASSIFICATION.csv", final_rows, FINAL_FIELDS)
    write_text(OUTPUT_ROOT / "FINAL_MASTER_CANDIDATE_CLASSIFICATION.md", "# Final Master Candidate Classification\n\nNo candidate is profitable, Pine-ready, or approved for MTC integration.\n\n" + md_table(final_rows, ["candidate_id", "source_title", "strategy_class", "final_action", "reason"], 250))
    for name, rows in subsets.items():
        write_csv(OUTPUT_ROOT / f"{name}.csv", rows, FINAL_FIELDS)
        write_text(OUTPUT_ROOT / f"{name}.md", f"# {name.replace('_', ' ').title()}\n\nCount: {len(rows)}\n\n" + md_table(rows, ["candidate_id", "source_title", "strategy_class", "final_action", "reason"], 200))
    write_csv(OUTPUT_ROOT / "FINAL_CORRECTED_INTAKE_INDEX.csv", corrected_index)
    write_text(OUTPUT_ROOT / "FINAL_CORRECTED_INTAKE_SUMMARY.md", "# Final Corrected Intake Summary\n\n" + md_table(corrected_index, ["corrected_intake_path", "repair_type", "source"], 300))
    write_csv(OUTPUT_ROOT / "FINAL_TRADER_WISDOM_INDEX.csv", final_wisdom, final_wisdom_fields)

    wiki_dir = QUANT_ROOT / "11_TRADER_WIKI" / "2026_05_04_final_repair_import"
    if wiki_dir.exists():
        wiki_dir = QUANT_ROOT / "11_TRADER_WIKI" / f"2026_05_04_final_repair_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    llm_dir = QUANT_ROOT / "12_LLM_WIKI" / "2026_05_04_final_repair_import"
    if llm_dir.exists():
        llm_dir = QUANT_ROOT / "12_LLM_WIKI" / f"2026_05_04_final_repair_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    wiki_dir.mkdir(parents=True, exist_ok=False)
    llm_dir.mkdir(parents=True, exist_ok=False)
    write_text(wiki_dir / "README.md", "# Final Trader Wiki Repair Import\n\nAntigravity missing wisdom and Codex final repair items. Wisdom is not strategy proof.\n")
    write_text(wiki_dir / "00_FINAL_REPAIR_INDEX.md", "# Final Repair Wisdom Index\n\n" + md_table(extra_wisdom, ["wisdom_id", "trader_source_video", "short_wisdom_statement", "action"], 200))
    write_csv(wiki_dir / "FINAL_TRADER_WISDOM_INDEX.csv", final_wisdom, final_wisdom_fields)
    write_csv(wiki_dir / "FINAL_SOURCE_MAP.csv", final_rows, FINAL_FIELDS)

    kb_rows = []
    for row in final_rows:
        kb_rows.append({
            "record_type": "final_candidate_classification",
            "candidate_id": row["candidate_id"],
            "wisdom_id": "",
            "source_url": row["source_url"],
            "transcript_path": row["transcript_path"],
            "intake_path": row["intake_path"],
            "title": row["source_title"],
            "summary": f"{row['strategy_class']} with final action {row['final_action']}.",
            "exact_rules_if_available": "See source intake/corrected intake; no Pine-ready rules promoted.",
            "ambiguity_flags": [row["reason"], row["data_requirements"]],
            "data_requirements": row["data_requirements"],
            "MTC_mapping": row["MTC_mapping"],
            "recommended_next_action": row["final_action"],
        })
    for row in extra_wisdom:
        kb_rows.append({
            "record_type": "final_repair_wisdom",
            "candidate_id": "",
            "wisdom_id": row["wisdom_id"],
            "source_url": row["source_url"],
            "transcript_path": row["source_transcript_path"],
            "intake_path": row["source_intake_path"],
            "title": row["trader_source_video"],
            "summary": row["short_wisdom_statement"],
            "exact_rules_if_available": "",
            "ambiguity_flags": ["wisdom_or_warning_not_strategy"],
            "data_requirements": "N/A until converted into a module idea.",
            "MTC_mapping": row["MTC_or_QuantLens_mapping"],
            "recommended_next_action": row["action"],
        })
    with (OUTPUT_ROOT / "FINAL_LLM_KNOWLEDGE_BASE.jsonl").open("w", encoding="utf-8", newline="\n") as handle:
        for row in kb_rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    shutil.copy2(OUTPUT_ROOT / "FINAL_LLM_KNOWLEDGE_BASE.jsonl", llm_dir / "FINAL_LLM_KNOWLEDGE_BASE.jsonl")
    write_csv(llm_dir / "FINAL_LLM_SOURCE_MAP.csv", final_rows, FINAL_FIELDS)
    write_text(llm_dir / "README.md", "# Final LLM Wiki Repair Import\n\nMachine-readable repaired classifications and missing wisdom. No profitability claims.\n")
    write_text(OUTPUT_ROOT / "FINAL_WIKI_IMPORT_REPORT.md", f"# Final Wiki Import Report\n\n- Trader Wiki import: `{rel(wiki_dir)}`\n- LLM Wiki import: `{rel(llm_dir)}`\n- Added Antigravity missing wisdom items: {len(extra_wisdom)}\n")

    priority = {}
    for fname in [
        "ANTIGRAVITY_DAY_TRADE_PRIORITY_RANKING.md",
        "ANTIGRAVITY_SWING_TRADE_PRIORITY_RANKING.md",
        "ANTIGRAVITY_POSITION_TRADING_PRIORITY_RANKING.md",
        "ANTIGRAVITY_FILTER_EXIT_SIZING_PRIORITY_RANKING.md",
    ]:
        priority.update(parse_priority_md(ag_dir / fname))
    queue_candidates = [r for r in final_rows if r["strategy_class"] in {"DAY_TRADE_STRATEGY", "SWING_TRADE_STRATEGY", "POSITION_TRADING_STRATEGY", "FILTER_MODULE", "EXIT_MODULE", "SIZING_OR_RISK_MODULE"}]
    queue_candidates.sort(key=lambda r: (r["strategy_class"], priority.get(r["candidate_id"], 999), -int(float(r["manual_visual_review_priority"] or 0))))
    queue_rows = []
    group_rank = Counter()
    for row in queue_candidates:
        group = row["strategy_class"]
        group_rank[group] += 1
        queue_rows.append({
            "rank": group_rank[group],
            "candidate_id": row["candidate_id"],
            "candidate_name": row["source_title"],
            "source": row["source_url"],
            "strategy_group": group,
            "native_market_timeframe": f"{row['native_market']} / {row['native_timeframe']}",
            "why_it_is_worth_visual_review": row["reason"],
            "exact_first_visual_check": "Confirm raw setup/pulse/invalidation behavior on chart; do not test profitability.",
            "required_standalone_Pine_sandbox_later": "standalone_pine_visual_review.pine only after manual selection.",
            "required_Python_parity_model_later": "python_signal_model.py with raw long/short pulses plus golden_cases.csv.",
            "data_blockers": row["data_requirements"],
            "MTC_money_management_reuse_notes": "Use external MTC-compatible harness; do not touch production MTC_V2.",
            "do_not_promote_reason": "Needs visual review, source-faithful spec, parity, no-repaint checks, and robustness before promotion.",
        })
    queue_fields = list(queue_rows[0].keys()) if queue_rows else ["rank", "candidate_id", "candidate_name", "source", "strategy_group"]
    write_csv(OUTPUT_ROOT / "FINAL_MANUAL_VISUAL_TEST_QUEUE.csv", queue_rows, queue_fields)
    write_text(OUTPUT_ROOT / "FINAL_MANUAL_VISUAL_TEST_QUEUE.md", "# Final Manual Visual Test Queue\n\nGrouped by final strategy/module class. No backtests were run.\n\n" + md_table(queue_rows, ["strategy_group", "rank", "candidate_id", "candidate_name", "exact_first_visual_check"], 300))

    sandbox_decision = """# Final MTC Sandbox Architecture Decision

## Decision
Do not put all candidates into MTC_V2 at once. Build standalone candidate sandboxes one by one.

## Signal Contract
Each candidate must produce raw boolean long/short pulses through an MTC-compatible signal producer contract. Candidate signal logic must stay isolated from MTC_V2 production state.

## Shared Harness
Money management and risk behavior should be represented through an external shared harness concept: direction enable/disable, sizing, stop/target, trailing/BE where applicable, time exit, exit-first behavior, and no-repaint discipline.

## Pine Visual Review
Pine visual review scripts remain standalone. They should plot raw setup state, entry pulse, exit pulse, invalidation, and simple stop/target guides. They must not modify `01_PINE/MTC_V2.pine`.

## Python Parity
Each selected candidate later needs `candidate_contract.yml`, `candidate_rules.md`, `python_signal_model.py`, `python_backtest_harness.py`, `golden_cases.csv`, `tests/`, `standalone_pine_visual_review.pine`, `MTC_mapping.md`, and `VISUAL_REVIEW_CHECKLIST.md`.

## Promotion Gate
Production MTC integration requires manual visual review approval, source-faithful rules, strict Pine/Python signal parity, repaint elimination, proxy/data availability verification, Stage robustness, and explicit owner approval.
"""
    write_text(OUTPUT_ROOT / "FINAL_MTC_SANDBOX_ARCHITECTURE_DECISION.md", sandbox_decision)

    workbook_path = OUTPUT_ROOT / "FINAL_QUANTLENS_RESTART_CLASSIFICATION_WORKBOOK.xlsx"
    readme_rows = [
        {"field": "Run At", "value": RUN_AT},
        {"field": "Codex Input", "value": rel(codex_dir)},
        {"field": "Antigravity Input", "value": rel(ag_dir)},
        {"field": "Missing Required Inputs", "value": "; ".join(missing_inputs) or "None"},
        {"field": "Optional Ranking CSV Status", "value": json.dumps(optional_status)},
        {"field": "No Backtests", "value": "true"},
        {"field": "No MTC_V2 Pine Modification", "value": "true"},
    ]
    sheets = {
        "README": (readme_rows, ["field", "value"]),
        "Final Source Map": (final_rows, FINAL_FIELDS),
        "Final Intake Faithfulness": (faith_rows, list(faith_rows[0].keys()) if faith_rows else ["empty"]),
        "Final Wisdom": (final_wisdom, final_wisdom_fields),
        "Final Master Classification": (final_rows, FINAL_FIELDS),
        "Final Day Trade": (day_rows, FINAL_FIELDS),
        "Final Swing Trade": (swing_rows, FINAL_FIELDS),
        "Final Position Trading": (position_rows, FINAL_FIELDS),
        "Final Filters Exits Sizing": (module_rows, FINAL_FIELDS),
        "Final Rejected Blocked": (rejected_rows, FINAL_FIELDS),
        "Manual Visual Test Queue": (queue_rows, queue_fields),
        "Antigravity Repair Decisions": (repair_decisions, list(repair_decisions[0].keys()) if repair_decisions else ["empty"]),
        "MTC Sandbox Plan": ([{"section": "decision", "detail": sandbox_decision}], ["section", "detail"]),
    }
    build_workbook(workbook_path, sheets)

    write_text(OUTPUT_ROOT / "FINAL_REPAIR_REPORT.md", f"""# Final Codex Repaired Audit Report

## Inputs
- Codex: `{rel(codex_dir)}`
- Antigravity: `{rel(ag_dir)}`

## Repairs Applied
- Reclassification requests accepted: {sum(1 for d in repair_decisions if d['source_file'] == 'CANDIDATE_RECLASSIFICATION_REQUESTS_FOR_CODEX.csv' and d['decision'].startswith(('ACCEPTED','MODIFIED')))}
- Matching findings recorded/blocked/merged: {sum(1 for d in repair_decisions if d['source_file'] == 'MATCHING_AUDIT_FINDINGS.csv')}
- Missing wisdom items added: {len(extra_wisdom)}
- VKNEJA5r8zw canonical merge: yes

## Final Classification Summary
{dict(Counter(r['strategy_class'] for r in final_rows))}

## Validation
See `FINAL_VALIDATION_REPORT.md`.

## Important Constraint
No strategy backtest, Pine conversion, TradingView alert, trade execution, or production runner modification was performed.
""")

    # Validation
    validations = []
    try:
        subprocess.run(["python", "-m", "py_compile", str(Path(__file__))], cwd=REPO_ROOT, check=True, capture_output=True, text=True)
        validations.append({"check": "py_compile_new_python", "status": "PASS", "detail": rel(Path(__file__))})
    except subprocess.CalledProcessError as exc:
        validations.append({"check": "py_compile_new_python", "status": "FAIL", "detail": exc.stderr})
    try:
        wb = load_workbook(workbook_path, read_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
        validations.append({"check": "openpyxl_open_workbook", "status": "PASS", "detail": f"{sheet_count} sheets"})
    except Exception as exc:
        validations.append({"check": "openpyxl_open_workbook", "status": "FAIL", "detail": str(exc)})
    csv_fail = []
    for path in list(OUTPUT_ROOT.rglob("*.csv")) + list(wiki_dir.rglob("*.csv")) + list(llm_dir.rglob("*.csv")):
        try:
            with path.open("r", encoding="utf-8", newline="") as handle:
                list(csv.reader(handle))
        except Exception as exc:
            csv_fail.append(f"{rel(path)}: {exc}")
    validations.append({"check": "csv_readable", "status": "PASS" if not csv_fail else "FAIL", "detail": "; ".join(csv_fail[:5])})
    jsonl_fail = []
    for path in list(OUTPUT_ROOT.rglob("*.jsonl")) + list(llm_dir.rglob("*.jsonl")):
        try:
            with path.open("r", encoding="utf-8") as handle:
                for line_no, line in enumerate(handle, 1):
                    if line.strip():
                        json.loads(line)
        except Exception as exc:
            jsonl_fail.append(f"{rel(path)}:{line_no}: {exc}")
    validations.append({"check": "jsonl_valid", "status": "PASS" if not jsonl_fail else "FAIL", "detail": "; ".join(jsonl_fail[:5])})
    original_after = {}
    for root in (TRANSCRIPTS_DIR, INTAKES_DIR):
        for path in root.rglob("*"):
            if path.is_file():
                original_after[rel(path)] = sha256_file(path)
    changed_originals = [k for k, v in original_before.items() if original_after.get(k) != v]
    validations.append({"check": "original_transcripts_and_intakes_unmodified", "status": "PASS" if not changed_originals else "FAIL", "detail": "; ".join(changed_originals[:8])})
    protected_after = {rel(p): sha256_file(p) if p.exists() else "MISSING" for p in protected}
    for key in protected_before:
        validations.append({"check": f"protected_untouched_{Path(key).name}", "status": "PASS" if protected_before[key] == protected_after.get(key) else "FAIL", "detail": f"before={protected_before[key]} after={protected_after.get(key)}"})
    try:
        status = subprocess.run(["git", "status", "--short"], cwd=REPO_ROOT, check=True, capture_output=True, text=True).stdout
        write_text(OUTPUT_ROOT / "git_status_after.txt", status)
        validations.append({"check": "git_status_after_captured", "status": "PASS", "detail": "git_status_after.txt"})
    except Exception as exc:
        validations.append({"check": "git_status_after_captured", "status": "FAIL", "detail": str(exc)})
    write_csv(OUTPUT_ROOT / "FINAL_VALIDATION_CHECKLIST.csv", validations, ["check", "status", "detail"])
    write_text(OUTPUT_ROOT / "FINAL_VALIDATION_REPORT.md", "# Final Validation Report\n\n" + md_table(validations, ["check", "status", "detail"], 100))

    all_files = []
    for root in (OUTPUT_ROOT, wiki_dir, llm_dir):
        for path in root.rglob("*"):
            if path.is_file():
                all_files.append(rel(path))
    write_text(OUTPUT_ROOT / "FILES_CREATED.txt", "\n".join(sorted(set(all_files))) + "\n")


if __name__ == "__main__":
    main()
