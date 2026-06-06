"""Triage worklist generator for MCC candidate audit.

Produces an XLSX of all REJECTED-quality and missing-coverage rows AND
a per-strategy .md file under strategies/ that the user fills in with
verified YouTube URLs and transcripts.

stg_code (Stg001, Stg002, ...) is persisted in strategies/_stg_code_map.json
so re-runs keep the same code for the same candidate and never overwrite
existing .md files the user has been editing.

Read-only with respect to MTC_v2 and QuantLens. Only writes inside 11_TRIAGE/.
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

THIS = Path(__file__).resolve()
STRATEGIES_DIR = THIS.parent / "strategies"
CODE_MAP_PATH = STRATEGIES_DIR / "_stg_code_map.json"
TRANSCRIPT_EMBED_LIMIT = 250_000  # bytes; larger transcripts are linked, not embedded

API_ROOT = THIS.parent.parent / "08_DASHBOARD_APP" / "apps" / "api"
sys.path.insert(0, str(API_ROOT))

from mcc_readonly.audit_reader import build_candidate_audit  # noqa: E402
from mcc_readonly.paths import default_mcc_root, default_quantlens_root  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


def classify_coverage(row: dict) -> str:
    has_url = bool(row.get("has_source_url"))
    has_tx = bool(row.get("has_transcript"))
    if not has_url and not has_tx:
        return "NO_URL_NO_TRANSCRIPT"
    if has_url and not has_tx:
        return "HAS_URL_NO_TRANSCRIPT"
    if not has_url and has_tx:
        return "HAS_TRANSCRIPT_NO_URL"
    return "HAS_BOTH"


def priority(row: dict) -> str:
    quality = (row.get("source_quality") or "").upper()
    blocked = (row.get("blocked_reason") or "").lower()
    if quality == "REJECTED" or "wiki-only" in blocked:
        return "P1_REJECTED"
    if not row.get("has_source_url") or not row.get("has_transcript"):
        return "P2_MISSING_COVERAGE"
    return "P3_COMPLETE"


COVERAGE_FILL = {
    "NO_URL_NO_TRANSCRIPT": "FFC7CE",
    "HAS_URL_NO_TRANSCRIPT": "FFEB9C",
    "HAS_TRANSCRIPT_NO_URL": "FFEB9C",
    "HAS_BOTH": "C6EFCE",
}

COLUMNS = [
    ("stg_code", 12),
    ("md_file", 22),
    ("priority", 22),
    ("coverage_status", 24),
    ("candidate_id", 38),
    ("title", 50),
    ("source_quality", 14),
    ("blocked_reason", 28),
    ("has_source_url", 14),
    ("has_transcript", 14),
    ("existing_source_url", 50),
    ("existing_transcript_path", 50),
    ("intake_source_file", 60),
    ("recommended_next_step", 28),
    ("user_notes", 30),
]


def load_code_map() -> dict[str, str]:
    if CODE_MAP_PATH.exists():
        try:
            return json.loads(CODE_MAP_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_code_map(code_map: dict[str, str]) -> None:
    STRATEGIES_DIR.mkdir(parents=True, exist_ok=True)
    CODE_MAP_PATH.write_text(json.dumps(code_map, indent=2, sort_keys=True), encoding="utf-8")


def assign_codes(selected: list[tuple[str, dict]], code_map: dict[str, str]) -> dict[str, str]:
    used = set(code_map.values())
    if used:
        max_n = max(int(c[3:]) for c in used if c.startswith("Stg") and c[3:].isdigit())
    else:
        max_n = 0
    for _p, row in selected:
        cid = row.get("id") or ""
        if not cid or cid in code_map:
            continue
        max_n += 1
        code_map[cid] = f"Stg{max_n:03d}"
    return code_map


def write_md_files(selected: list[tuple[str, dict]], code_map: dict[str, str], mcc_root: Path) -> int:
    STRATEGIES_DIR.mkdir(parents=True, exist_ok=True)
    qlab_root = default_quantlens_root(mcc_root)
    written = 0
    for _p, row in selected:
        cid = row.get("id") or ""
        code = code_map.get(cid)
        if not code:
            continue
        md_path = STRATEGIES_DIR / f"{code.lower()}.md"
        if md_path.exists():
            continue
        title = (row.get("name") or "").strip()
        url = (row.get("source_url") or "").strip()
        transcript_rel = (row.get("transcript_path") or "").strip()
        body = [
            f"# {code} — {cid}",
            "",
            f"Video name: {title}",
            f"Video Url: {url}",
            "",
        ]
        if transcript_rel:
            transcript_full = qlab_root / Path(transcript_rel.replace("\\", "/"))
            body.append("## Transcript")
            body.append("")
            try:
                if transcript_full.exists():
                    size = transcript_full.stat().st_size
                    if size <= TRANSCRIPT_EMBED_LIMIT:
                        body.append(transcript_full.read_text(encoding="utf-8", errors="replace"))
                    else:
                        body.append(f"(transcript {size:,} bytes — see {transcript_rel})")
                else:
                    body.append(f"(transcript path on record: {transcript_rel})")
            except Exception as exc:
                body.append(f"(could not read transcript: {exc})")
        md_path.write_text("\n".join(body), encoding="utf-8")
        written += 1
    return written


def select_rows(audit_rows):
    coverage_order = {
        "NO_URL_NO_TRANSCRIPT": 0,
        "HAS_URL_NO_TRANSCRIPT": 1,
        "HAS_TRANSCRIPT_NO_URL": 2,
        "HAS_BOTH": 3,
    }
    chosen = []
    for r in audit_rows:
        p = priority(r)
        if p in ("P1_REJECTED", "P2_MISSING_COVERAGE"):
            chosen.append((p, r))
    chosen.sort(key=lambda pr: (pr[0], coverage_order.get(classify_coverage(pr[1]), 9), pr[1].get("id", "")))
    return chosen


def write_xlsx(selected, audit, code_map: dict[str, str], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worklist"

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="2F5061")
    header_align = Alignment(horizontal="left", vertical="center")
    body_font = Font(name="Arial")

    for i, (name, width) in enumerate(COLUMNS, start=1):
        col = get_column_letter(i)
        c = ws[f"{col}1"]
        c.value = name
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        ws.column_dimensions[col].width = width

    for r_idx, (p, row) in enumerate(selected, start=2):
        cov = classify_coverage(row)
        sr = row.get("source_record") or {}
        intake = sr.get("source_file") or sr.get("relative_source_path") or ""
        cid = row.get("id", "")
        code = code_map.get(cid, "")
        md_file = f"{code.lower()}.md" if code else ""
        values = [
            code,
            md_file,
            p,
            cov,
            cid,
            (row.get("name") or "")[:200],
            row.get("source_quality", ""),
            row.get("blocked_reason", ""),
            "YES" if row.get("has_source_url") else "NO",
            "YES" if row.get("has_transcript") else "NO",
            row.get("source_url", ""),
            row.get("transcript_path", ""),
            intake,
            row.get("recommended_next_pipeline_step", ""),
            "",
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
        fill_color = COVERAGE_FILL.get(cov)
        if fill_color:
            ws.cell(row=r_idx, column=4).fill = PatternFill("solid", start_color=fill_color)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    summary = audit.get("summary", {}) or {}
    ws2["A1"] = "MCC Triage Worklist Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    base = [
        ("Generated (date)", date.today().isoformat()),
        ("Total audit rows", summary.get("total_rows")),
        ("Eligible for backtest", summary.get("eligible_for_backtest_rows")),
        ("Blocked rows", summary.get("blocked_rows")),
        ("Source material rows", summary.get("source_material_rows")),
        ("Worklist rows in this file", len(selected)),
    ]
    quality = summary.get("source_quality_counts", {}) or {}
    blocked = summary.get("blocked_reason_counts", {}) or {}
    rows_iter = list(base)
    for k in ("HIGH", "MEDIUM", "LOW", "REJECTED"):
        rows_iter.append((f"  source_quality {k}", quality.get(k, 0)))
    for k, v in sorted(blocked.items()):
        rows_iter.append((f"  blocked_reason '{k}'", v))

    for i, (label, val) in enumerate(rows_iter, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial")
        ws2.cell(row=i, column=2, value=val).font = Font(bold=True, name="Arial")
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 14

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> None:
    mcc_root = Path(default_mcc_root()).resolve()
    audit = build_candidate_audit()
    rows = audit.get("rows", []) or []
    selected = select_rows(rows)

    code_map = load_code_map()
    assign_codes(selected, code_map)
    save_code_map(code_map)
    written_md = write_md_files(selected, code_map, mcc_root)

    today = date.today().isoformat()
    out = THIS.parent / f"{today}_rejected_worklist.xlsx"
    write_xlsx(selected, audit, code_map, out)
    print(f"OK {len(selected)} rows -> {out}")
    print(f"OK {written_md} new .md files in {STRATEGIES_DIR} (existing files preserved)")
    print(f"OK stg_code map -> {CODE_MAP_PATH}")


if __name__ == "__main__":
    main()
